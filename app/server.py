from __future__ import annotations

import calendar
import json
import base64
import csv
import io
import hashlib
import hmac
import mimetypes
import os
import re
import secrets
import struct
import subprocess
import tempfile
import threading
import time
import math
import unicodedata
import zipfile
import qrcode
from qrcode.image.svg import SvgPathImage
from datetime import date, datetime, timedelta, timezone
from http import HTTPStatus
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from http.cookies import SimpleCookie
from urllib.parse import parse_qs, quote, urlencode, urlparse
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen
from openpyxl import load_workbook
from db_backend import IntegrityError, configure as configure_database, connect

try:
    from cryptography.exceptions import InvalidTag
    from cryptography.hazmat.primitives.ciphers.aead import AESGCM
except ImportError:  # dependência de SEC-011 ainda não instalada neste ambiente
    InvalidTag = ValueError
    AESGCM = None


ROOT = Path(__file__).resolve().parent
PUBLIC = ROOT / "public"
DATA = ROOT / "data"
DATABASE_URL = os.environ.get("DATABASE_URL", "").strip()
configure_database(DATABASE_URL)
N8N_CONFIG_PATH = DATA / "n8n-crm-config.json"
CRM_MEDIA_DIR = DATA / "crm-media"
SEED_PATH = DATA / "patients.seed.json"
ADMIN_ROUTE = os.environ.get("ADMIN_ROUTE", "/central-gestao-iea-6x9p4m2k")
CRC_ROUTE = "/central-crc"
STANDARD_PATIENT_STATUSES = {"Consulta", "Controle", "Tratamento", "Inativo"}
RELEASE_ID = os.environ.get("APP_RELEASE_ID") or datetime.utcnow().strftime("%Y%m%d%H%M%S")
MAX_BODY_BYTES = 60 * 1024 * 1024  # 60 MB: acomoda mídia de CRM em base64 (limite decodificado de 40 MB)
TOTP_ENC_PREFIX = "encv1:"
INTEGRATION_SECRET_PREFIX = "secv1:"
INTEGRATION_SECRET_AAD = b"iea-integration-secret-v1"
_app_secret_key_raw = os.environ.get("APP_SECRET_KEY", "").strip()
APP_SECRET_KEY = base64.b64decode(_app_secret_key_raw) if _app_secret_key_raw else None
INTEGRATION_TOKEN = os.environ.get("INTEGRATION_TOKEN", "")
EVOLUTION_WEBHOOK_TOKEN = os.environ.get("EVOLUTION_WEBHOOK_TOKEN", "")
EVOLUTION_API_URL = os.environ.get("EVOLUTION_API_URL", "").rstrip("/")
EVOLUTION_API_KEY = os.environ.get("EVOLUTION_API_KEY", "")
N8N_INTERNAL_URL = os.environ.get("N8N_INTERNAL_URL", "http://n8n-czmx-n8n-1:5678").rstrip("/")
PUBLIC_APP_URL = os.environ.get("PUBLIC_APP_URL", "https://dentistas.automacaocentraliea.me").rstrip("/")
CLINIC_UTC_OFFSET_HOURS = int(os.environ.get("CLINIC_UTC_OFFSET_HOURS", "-4"))
CLINIC_TIMEZONE = timezone(timedelta(hours=CLINIC_UTC_OFFSET_HOURS))
# O container roda em UTC por padrão. SQLite usa o fuso do processo quando recebe
# o modificador `localtime`, então padronizamos toda a aplicação no horário de Cuiabá.
os.environ.setdefault("TZ", "GMT+4")
if hasattr(time, "tzset"):
    time.tzset()
EVOLUTION_HISTORY_CUTOFF = "2026-07-20 00:00:00"
EVOLUTION_CHAT_SYNC_INTERVAL = max(30, int(os.environ.get("EVOLUTION_CHAT_SYNC_INTERVAL", "60")))
WEBHOOK_PAYLOAD_RETENTION_DAYS = max(7, int(os.environ.get("WEBHOOK_PAYLOAD_RETENTION_DAYS", "90")))
SECURITY_EVENT_RETENTION_DAYS = max(90, int(os.environ.get("SECURITY_EVENT_RETENTION_DAYS", "365")))
API_SYNC_LOCK = threading.Lock()
API_SYNC_THREADS: dict[int, threading.Thread] = {}
EVOLUTION_HISTORY_SYNC_LOCK = threading.Lock()
EVOLUTION_CHAT_SYNC_LOCK = threading.Lock()
EVOLUTION_CHAT_SYNC_STATUS = {
    "running": False,
    "last_started_at": None,
    "last_finished_at": None,
    "instances_synced": 0,
    "unread_conversations": 0,
    "pending_conversations": 0,
    "unmatched_conversations": 0,
    "errors": [],
}
EVOLUTION_HISTORY_SYNC_STATUS = {
    "running": False,
    "instances_total": 0,
    "instances_done": 0,
    "contacts": 0,
    "conversations": 0,
    "messages": 0,
    "older_messages_removed": 0,
    "since": EVOLUTION_HISTORY_CUTOFF[:10],
    "phase": "Aguardando sincronização",
    "errors": [],
    "started_at": None,
    "finished_at": None,
}


def encrypt_integration_secret(value: str | None) -> str | None:
    """Encrypt a persisted integration credential while keeping empty values intact."""
    if not value or value.startswith(INTEGRATION_SECRET_PREFIX):
        return value
    if not APP_SECRET_KEY or AESGCM is None:
        raise RuntimeError("APP_SECRET_KEY ausente: não é possível proteger a credencial da integração")
    nonce = secrets.token_bytes(12)
    ciphertext = AESGCM(APP_SECRET_KEY).encrypt(
        nonce, value.encode("utf-8"), INTEGRATION_SECRET_AAD
    )
    return INTEGRATION_SECRET_PREFIX + base64.b64encode(nonce + ciphertext).decode("ascii")


def decrypt_integration_secret(value: str | None) -> str | None:
    """Decrypt a credential only at its point of use; accept plaintext solely for migration."""
    if not value or not value.startswith(INTEGRATION_SECRET_PREFIX):
        return value
    if not APP_SECRET_KEY or AESGCM is None:
        raise RuntimeError("APP_SECRET_KEY ausente: não é possível decifrar a credencial da integração")
    try:
        raw = base64.b64decode(value[len(INTEGRATION_SECRET_PREFIX):], validate=True)
        if len(raw) <= 12:
            raise ValueError("credencial cifrada incompleta")
        plaintext = AESGCM(APP_SECRET_KEY).decrypt(
            raw[:12], raw[12:], INTEGRATION_SECRET_AAD
        )
        return plaintext.decode("utf-8")
    except (ValueError, UnicodeDecodeError, InvalidTag) as error:
        raise RuntimeError("Credencial de integração cifrada inválida") from error


def migrate_integration_secrets(db) -> int:
    """Upgrade legacy plaintext credentials after the schema is available."""
    if not APP_SECRET_KEY or AESGCM is None:
        return 0
    sensitive_columns = (
        ("integration_configs", "api_token"),
        ("api_integrations", "api_token"),
        ("api_integration_backups", "api_token"),
        ("crm_n8n_config", "api_token"),
        ("crm_channels", "evolution_api_key"),
    )
    key_columns = {
        "integration_configs": "name",
        "api_integrations": "id",
        "api_integration_backups": "integration_id",
        "crm_n8n_config": "id",
        "crm_channels": "id",
    }
    migrated = 0
    for table, column in sensitive_columns:
        key_column = key_columns[table]
        rows = db.execute(
            f"SELECT {key_column} AS secret_row_key,{column} AS secret_value "
            f"FROM {table} WHERE {column} IS NOT NULL AND {column}<>'' AND {column} NOT LIKE ?",
            (INTEGRATION_SECRET_PREFIX + "%",),
        ).fetchall()
        for row in rows:
            db.execute(
                f"UPDATE {table} SET {column}=? WHERE {key_column}=?",
                (encrypt_integration_secret(str(row["secret_value"])), row["secret_row_key"]),
            )
            migrated += 1
    return migrated


def migrate_n8n_config_file() -> bool:
    if not APP_SECRET_KEY or AESGCM is None or not N8N_CONFIG_PATH.exists():
        return False
    try:
        stored = json.loads(N8N_CONFIG_PATH.read_text(encoding="utf-8"))
    except (OSError, ValueError, TypeError):
        return False
    token = str(stored.get("api_token") or "") if isinstance(stored, dict) else ""
    if not token or token.startswith(INTEGRATION_SECRET_PREFIX):
        return False
    encrypted_token = encrypt_integration_secret(token)
    stored = {**stored, "api_token": encrypted_token}
    temporary_path = N8N_CONFIG_PATH.with_suffix(".tmp")
    temporary_path.write_text(
        json.dumps(stored, ensure_ascii=False, separators=(",", ":")), encoding="utf-8"
    )
    try:
        os.chmod(temporary_path, 0o600)
    except OSError:
        pass
    os.replace(temporary_path, N8N_CONFIG_PATH)
    return True
# A tela de integrações é aberta e reconstruída várias vezes pelo CRM. Manter a
# última auditoria válida em memória evita que uma troca de aba deixe a tela
# dependente de uma nova consulta externa ao n8n.
N8N_OVERVIEW_CACHE_LOCK = threading.Lock()
N8N_OVERVIEW_CACHE: dict[str, object] = {
    "payload": None,
    "updated_at": 0.0,
}
CRM_PROFILE_PHOTO_MISS_LOCK = threading.Lock()
CRM_PROFILE_PHOTO_MISS: dict[int, float] = {}

CRM_FEATURE_KEYS = (
    "inbox", "queue", "funnel", "management", "contacts", "campaigns",
    "integrations", "settings",
)
CRM_WORKSPACE_FEATURES = ("inbox", "queue", "funnel")

CRM_GOAL_METRICS = {
    "first_consultations": "Primeiras consultas",
    "recoveries": "Recuperação de pacientes",
    "attendances": "Atendimentos",
}
CRM_PATIENT_TYPES = {"Primeira consulta", "Retorno s/ Tratamento"}
CRM_MONTH_NAMES = (
    "", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
)


def crm_goal_month_bounds(month_value: str) -> tuple[date, date]:
    """Return the first and last day for a validated YYYY-MM goal period."""
    if not re.fullmatch(r"\d{4}-\d{2}", month_value or ""):
        raise ValueError("Mês inválido")
    first = datetime.strptime(month_value + "-01", "%Y-%m-%d").date()
    last = date(first.year, first.month, calendar.monthrange(first.year, first.month)[1])
    return first, last


def crm_open_days_remaining(reference: date, month_start: date) -> int:
    """Count open clinic days (Monday through Saturday) still available."""
    month_end = date(month_start.year, month_start.month, calendar.monthrange(month_start.year, month_start.month)[1])
    if reference > month_end:
        return 0
    cursor = max(reference, month_start)
    return sum(1 for offset in range((month_end - cursor).days + 1)
               if (cursor + timedelta(days=offset)).weekday() < 6)


def crm_goal_progress(target: int, realized: int, remaining_days: int) -> dict:
    target = max(0, int(target or 0))
    realized = max(0, int(realized or 0))
    gap = max(0, target - realized)
    return {
        "target": target,
        "realized": realized,
        "percentage": round(realized / target * 100, 1) if target else 0,
        "gap": gap,
        "required_per_open_day": math.ceil(gap / remaining_days) if gap and remaining_days else 0,
        "reached": bool(target and realized >= target),
    }


def convert_crm_audio_to_ogg(audio_bytes: bytes) -> bytes:
    """Normalize browser recordings to the WhatsApp voice-note format."""
    try:
        result = subprocess.run(
            [
                "ffmpeg", "-hide_banner", "-loglevel", "error",
                "-i", "pipe:0", "-vn", "-map_metadata", "-1",
                "-c:a", "libopus", "-application", "voip",
                "-b:a", "64k", "-ar", "48000", "-ac", "1",
                "-f", "ogg", "pipe:1",
            ],
            input=audio_bytes,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            timeout=45,
            check=False,
        )
    except FileNotFoundError as error:
        raise RuntimeError("Conversor de áudio não instalado no servidor.") from error
    except subprocess.TimeoutExpired as error:
        raise RuntimeError("A conversão do áudio excedeu o tempo permitido.") from error

    if result.returncode != 0 or not result.stdout.startswith(b"OggS"):
        detail = result.stderr.decode("utf-8", errors="replace").strip()[-300:]
        raise RuntimeError(f"Não foi possível preparar o áudio gravado. {detail}".strip())
    return result.stdout


def crm_audio_duration_seconds(audio_bytes: bytes) -> float | None:
    """Read the real audio duration without trusting a browser-side timer."""
    try:
        with tempfile.NamedTemporaryFile(suffix=".ogg") as temporary:
            temporary.write(audio_bytes)
            temporary.flush()
            result = subprocess.run(
                [
                    "ffprobe", "-v", "error", "-show_entries", "format=duration",
                    "-of", "default=noprint_wrappers=1:nokey=1", temporary.name,
                ],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                timeout=15,
                check=False,
            )
        duration = float(result.stdout.decode("ascii", errors="ignore").strip() or 0)
        return round(duration, 2) if duration > 0 else None
    except (FileNotFoundError, subprocess.TimeoutExpired, TypeError, ValueError):
        return None


def migrate_crm_timezone(db) -> None:
    timezone_migration = "crm_timezone_cuiaba_v1"
    if db.execute("SELECT 1 FROM app_migrations WHERE migration_key=?", (timezone_migration,)).fetchone():
        return
    # Os registros anteriores eram persistidos em UTC pelo container, mas a
    # interface os interpretava como horário local. Corrige uma única vez.
    for table, columns in {
        "crm_messages": ("message_at", "created_at"),
        "crm_conversations": (
            "last_message_at", "resolved_at", "assigned_at", "queue_entered_at", "first_response_at", "created_at", "updated_at"
        ),
        "crm_webhook_events": ("received_at", "processed_at"),
        "crm_channels": ("last_event_at", "created_at", "updated_at"),
        "crm_contacts": ("created_at", "updated_at"),
    }.items():
        for column in columns:
            db.execute(f"UPDATE {table} SET {column}=datetime({column}, '-4 hours') WHERE {column} IS NOT NULL")
    db.execute("INSERT INTO app_migrations(migration_key) VALUES(?)", (timezone_migration,))


def ensure_crm_permission_constraints(db) -> None:
    """Normaliza valores legados e valida integralmente as permissões no PostgreSQL."""
    if getattr(db, "backend", "") != "postgres":
        return
    db.execute("""UPDATE users SET
        crm_channel_scope_enabled=CASE WHEN crm_channel_scope_enabled IN (0,1) THEN crm_channel_scope_enabled ELSE 0 END,
        crm_feature_scope_enabled=CASE WHEN crm_feature_scope_enabled IN (0,1) THEN crm_feature_scope_enabled ELSE 0 END,
        crm_operational_agent=CASE WHEN crm_operational_agent IN (0,1) THEN crm_operational_agent ELSE 0 END,
        crm_manage_automation=CASE WHEN crm_manage_automation IN (0,1) THEN crm_manage_automation ELSE 0 END,
        crm_access_level=CASE WHEN crm_access_level IN ('attendant','admin') THEN crm_access_level ELSE 'attendant' END""")
    db.execute("""UPDATE crm_user_channels SET
        can_reply=CASE WHEN can_reply IN (0,1) THEN can_reply ELSE 0 END,
        can_manage_automation=CASE WHEN can_manage_automation IN (0,1) THEN can_manage_automation ELSE 0 END""")
    db.execute("""DELETE FROM crm_user_features WHERE feature_key NOT IN
        ('inbox','queue','funnel','management','contacts','campaigns','integrations','settings')""")
    constraints = (
        ("users", "users_crm_channel_scope_bool", "crm_channel_scope_enabled IN (0,1)"),
        ("users", "users_crm_feature_scope_bool", "crm_feature_scope_enabled IN (0,1)"),
        ("users", "users_crm_operational_agent_bool", "crm_operational_agent IN (0,1)"),
        ("users", "users_crm_manage_automation_bool", "crm_manage_automation IN (0,1)"),
        ("users", "users_crm_access_level_valid", "crm_access_level IN ('attendant','admin')"),
        ("crm_user_channels", "crm_user_channels_reply_bool", "can_reply IN (0,1)"),
        ("crm_user_channels", "crm_user_channels_automation_bool", "can_manage_automation IN (0,1)"),
        ("crm_user_features", "crm_user_features_key_valid",
         "feature_key IN ('inbox','queue','funnel','management','contacts','campaigns','integrations','settings')"),
    )
    for table, constraint_name, expression in constraints:
        exists = db.execute("SELECT 1 FROM pg_constraint WHERE conname=?", (constraint_name,)).fetchone()
        if not exists:
            db.execute(f"ALTER TABLE {table} ADD CONSTRAINT {constraint_name} CHECK ({expression}) NOT VALID")
        validated = db.execute(
            "SELECT convalidated FROM pg_constraint WHERE conname=?",
            (constraint_name,),
        ).fetchone()
        if validated and not bool(validated["convalidated"]):
            db.execute(f"ALTER TABLE {table} VALIDATE CONSTRAINT {constraint_name}")


def cleanup_retention_data(db) -> None:
    """Reduz PII técnica sem apagar o histórico clínico e operacional necessário."""
    webhook_cutoff = f"-{WEBHOOK_PAYLOAD_RETENTION_DAYS} days"
    security_cutoff = f"-{SECURITY_EVENT_RETENTION_DAYS} days"
    redacted_payload = '{"retention":"redacted"}'
    db.execute("DELETE FROM import_batches WHERE expires_at <= datetime('now')")
    db.execute("DELETE FROM login_challenges WHERE expires_at <= datetime('now')")
    db.execute("DELETE FROM login_attempts WHERE attempted_at < datetime('now', ?)", (security_cutoff,))
    db.execute("DELETE FROM security_events WHERE created_at < datetime('now', ?)", (security_cutoff,))
    db.execute(
        "UPDATE crm_webhook_events SET payload_json=? WHERE received_at < datetime('now', ?) AND payload_json<>?",
        (redacted_payload, webhook_cutoff, redacted_payload),
    )
    db.execute(
        "UPDATE crm_automation_events SET payload_json=? WHERE received_at < datetime('now', ?) AND payload_json<>?",
        (redacted_payload, webhook_cutoff, redacted_payload),
    )


def initialize_database() -> None:
    DATA.mkdir(parents=True, exist_ok=True)
    CRM_MEDIA_DIR.mkdir(parents=True, exist_ok=True)
    schema = (ROOT / "schema.sql").read_text(encoding="utf-8")
    with connect() as db:
        db.executescript(schema)
        login_challenge_columns = {row[1] for row in db.execute("PRAGMA table_info(login_challenges)").fetchall()}
        if "attempts" not in login_challenge_columns:
            db.execute("ALTER TABLE login_challenges ADD COLUMN attempts INTEGER NOT NULL DEFAULT 0")
        if APP_SECRET_KEY and AESGCM is not None:
            legacy_secrets = db.execute(
                "SELECT id, two_factor_secret FROM users WHERE two_factor_secret IS NOT NULL AND two_factor_secret NOT LIKE ?",
                (TOTP_ENC_PREFIX + "%",),
            ).fetchall()
            for legacy_row in legacy_secrets:
                encrypted = ClinicHandler.encrypt_totp_secret(legacy_row["two_factor_secret"])
                db.execute("UPDATE users SET two_factor_secret=? WHERE id=?", (encrypted, legacy_row["id"]))
            migrate_integration_secrets(db)
        crm_conversation_columns = {row[1] for row in db.execute("PRAGMA table_info(crm_conversations)").fetchall()}
        for column, definition in {
            "pipeline_stage": "TEXT NOT NULL DEFAULT 'Novo'",
            "internal_note": "TEXT NOT NULL DEFAULT ''",
            "assigned_at": "TEXT",
            "queue_entered_at": "TEXT",
            "first_response_at": "TEXT",
            "resolved_by_user_id": "INTEGER",
            "last_direction": "TEXT",
            "automation_state": "TEXT NOT NULL DEFAULT 'manual'",
            "automation_flow": "TEXT",
            "automation_turns": "INTEGER NOT NULL DEFAULT 0",
            "handoff_reason": "TEXT",
            "resolution_reason": "TEXT",
            "scheduled_return_at": "TEXT",
            "reopened_at": "TEXT",
        }.items():
            if column not in crm_conversation_columns:
                db.execute(f"ALTER TABLE crm_conversations ADD COLUMN {column} {definition}")
        crm_contact_columns = {row[1] for row in db.execute("PRAGMA table_info(crm_contacts)").fetchall()}
        if "is_internal" not in crm_contact_columns:
            db.execute("ALTER TABLE crm_contacts ADD COLUMN is_internal INTEGER NOT NULL DEFAULT 0")
        crm_channel_columns = {row[1] for row in db.execute("PRAGMA table_info(crm_channels)").fetchall()}
        if "sync_enabled" not in crm_channel_columns:
            db.execute("ALTER TABLE crm_channels ADD COLUMN sync_enabled INTEGER NOT NULL DEFAULT 1")
        if "sync_from_date" not in crm_channel_columns:
            db.execute("ALTER TABLE crm_channels ADD COLUMN sync_from_date TEXT NOT NULL DEFAULT '2026-07-20'")
        if "sla_minutes" not in crm_channel_columns:
            db.execute("ALTER TABLE crm_channels ADD COLUMN sla_minutes INTEGER NOT NULL DEFAULT 60")
        user_columns = {row[1] for row in db.execute("PRAGMA table_info(users)").fetchall()}
        if "crm_channel_scope_enabled" not in user_columns:
            db.execute("ALTER TABLE users ADD COLUMN crm_channel_scope_enabled INTEGER NOT NULL DEFAULT 0")
        if "crm_feature_scope_enabled" not in user_columns:
            db.execute("ALTER TABLE users ADD COLUMN crm_feature_scope_enabled INTEGER NOT NULL DEFAULT 0")
        if "crm_operational_agent" not in user_columns:
            db.execute("ALTER TABLE users ADD COLUMN crm_operational_agent INTEGER NOT NULL DEFAULT 1")
        if "crm_manage_automation" not in user_columns:
            db.execute("ALTER TABLE users ADD COLUMN crm_manage_automation INTEGER NOT NULL DEFAULT 0")
            db.execute("""UPDATE users SET crm_manage_automation=1
                          WHERE EXISTS (SELECT 1 FROM crm_user_channels permission
                                        WHERE permission.user_id=users.id
                                          AND permission.can_manage_automation=1)""")
        if "crm_access_level" not in user_columns:
            db.execute("ALTER TABLE users ADD COLUMN crm_access_level TEXT NOT NULL DEFAULT 'attendant'")
            db.execute("""UPDATE users SET crm_access_level='admin'
                          WHERE access_role='crc'
                            AND lower(email) IN ('matheuscrc@instituto.local','melocrc@instituto.local')""")
        db.execute("""UPDATE users SET crm_channel_scope_enabled=0,crm_feature_scope_enabled=0,
                         crm_manage_automation=1,crm_operational_agent=0
                      WHERE access_role='crc' AND crm_access_level='admin'""")
        if "service_sector" not in user_columns:
            db.execute("ALTER TABLE users ADD COLUMN service_sector TEXT NOT NULL DEFAULT ''")
        db.execute("UPDATE users SET service_sector='CRC' WHERE access_role='crc' AND TRIM(COALESCE(service_sector,''))=''")
        ensure_crm_permission_constraints(db)
        crm_resolution_columns = {
            row[1] for row in db.execute("PRAGMA table_info(crm_service_resolutions)").fetchall()
        }
        if "patient_type" not in crm_resolution_columns:
            db.execute("ALTER TABLE crm_service_resolutions ADD COLUMN patient_type TEXT")
        if "is_recovery" not in crm_resolution_columns:
            db.execute("ALTER TABLE crm_service_resolutions ADD COLUMN is_recovery INTEGER NOT NULL DEFAULT 0")
        db.execute("""CREATE INDEX IF NOT EXISTS idx_crm_service_resolutions_goals
                      ON crm_service_resolutions(resolved_by_user_id,patient_type,is_recovery,outcome,resolved_at DESC)""")
        crm_message_columns = {row[1] for row in db.execute("PRAGMA table_info(crm_messages)").fetchall()}
        for column, definition in {
            "author_type": "TEXT NOT NULL DEFAULT 'unknown'",
            "author_label": "TEXT",
            "source_channel": "TEXT",
            "duration_seconds": "REAL",
        }.items():
            if column not in crm_message_columns:
                db.execute(f"ALTER TABLE crm_messages ADD COLUMN {column} {definition}")
        crm_n8n_setting_columns = {
            row[1] for row in db.execute("PRAGMA table_info(crm_n8n_workflow_settings)").fetchall()
        }
        for column, definition in {
            "source_label": "TEXT NOT NULL DEFAULT ''",
            "channel_label": "TEXT NOT NULL DEFAULT ''",
        }.items():
            if column not in crm_n8n_setting_columns:
                db.execute(f"ALTER TABLE crm_n8n_workflow_settings ADD COLUMN {column} {definition}")
        # Bancos criados antes do rastreamento de campanhas não possuem o nome
        # do fluxo em cada evento. A coluna permite agrupar os dados reais da
        # campanha sem perder o histórico já armazenado.
        crm_n8n_event_columns = {
            row[1] for row in db.execute("PRAGMA table_info(crm_n8n_patient_events)").fetchall()
        }
        if "flow_name" not in crm_n8n_event_columns:
            db.execute("ALTER TABLE crm_n8n_patient_events ADD COLUMN flow_name TEXT")
        if "appointment_source" not in crm_n8n_event_columns:
            db.execute("ALTER TABLE crm_n8n_patient_events ADD COLUMN appointment_source TEXT")
        db.execute("""UPDATE crm_messages SET author_type=CASE
                      WHEN direction='inbound' THEN 'patient'
                      WHEN sent_by_user_id IS NOT NULL THEN 'human'
                      ELSE 'external' END
                      WHERE author_type IS NULL OR author_type='unknown'""")
        db.execute("""UPDATE crm_messages SET author_label=CASE
                      WHEN direction='inbound' THEN COALESCE(NULLIF(sender_name,''),'Paciente')
                      WHEN sent_by_user_id IS NOT NULL THEN COALESCE((SELECT name FROM users WHERE users.id=crm_messages.sent_by_user_id),'Atendente')
                      ELSE 'Enviado fora do CRM' END
                      WHERE author_label IS NULL OR TRIM(author_label)=''""")
        db.execute("""UPDATE crm_conversations SET assigned_user_id=NULL,assigned_at=NULL,
                      queue_entered_at=NULL,first_response_at=NULL,unread_count=0,pipeline_stage='Novo'
                      WHERE contact_id IN (SELECT id FROM crm_contacts WHERE is_internal=1)""")
        db.execute("""UPDATE crm_conversations
                      SET last_direction=(
                          SELECT direction FROM crm_messages
                          WHERE crm_messages.conversation_id=crm_conversations.id
                          ORDER BY datetime(message_at) DESC,id DESC LIMIT 1
                      )
                      WHERE last_direction IS NULL""")
        db.execute("""UPDATE crm_conversations
                      SET queue_entered_at=COALESCE(last_message_at,created_at)
                      WHERE queue_entered_at IS NULL AND status<>'Resolvida'
                        AND assigned_user_id IS NULL AND unread_count>0 AND last_direction='inbound'""")
        migrate_crm_timezone(db)
        db.executemany(
            "INSERT OR IGNORE INTO crm_tags(name,color) VALUES(?,?)",
            [
                ("Campanha", "#7c5cff"),
                ("Retorno", "#f59e0b"),
                ("Lead novo", "#25d366"),
                ("Prioridade", "#ef4444"),
            ],
        )
        followup_columns = {row[1] for row in db.execute("PRAGMA table_info(patient_followup)").fetchall()}
        if "resolved_at" not in followup_columns:
            db.execute("ALTER TABLE patient_followup ADD COLUMN resolved_at TEXT")
        if "custom_status" not in followup_columns:
            db.execute("ALTER TABLE patient_followup ADD COLUMN custom_status TEXT")
        if "next_appointment_type" not in followup_columns:
            db.execute("ALTER TABLE patient_followup ADD COLUMN next_appointment_type TEXT")
        for column, definition in {
            "crc_status": "TEXT",
            "crc_started_at": "TEXT",
            "crc_completed_at": "TEXT",
        }.items():
            if column not in followup_columns:
                db.execute(f"ALTER TABLE patient_followup ADD COLUMN {column} {definition}")
        db.execute("""
            UPDATE patient_followup
            SET crc_status = 'Aguardando contato'
            WHERE resolved_at IS NOT NULL AND (crc_status IS NULL OR crc_status = '')
        """)
        db.execute("UPDATE patient_followup SET custom_status=NULL WHERE custom_status IS NOT NULL")
        db.execute("""UPDATE patient_followup
                      SET next_appointment_type='Agendado'
                      WHERE next_appointment IS NOT NULL
                        AND COALESCE(next_appointment_type, '') NOT IN ('Agendado', 'Programado')""")
        db.execute("UPDATE patient_followup SET next_appointment_type=NULL WHERE next_appointment IS NULL")
        db.execute("""
            UPDATE patient_followup
               SET crc_status='Aguardando contato',
                   crc_started_at=NULL,
                   crc_completed_at=NULL
             WHERE next_appointment IS NOT NULL
               AND next_appointment_type='Programado'
               AND COALESCE(crc_status, '') NOT IN ('Aguardando contato', 'Em atendimento', 'Jornada compartilhada')
        """)
        old_air_flow = "Profilaxia c/ Air Flow"
        child_air_flow = "Profilaxia c/ Air Flow Criança"
        db.execute(
            """UPDATE patient_followup
               SET next_action = REPLACE(next_action, ?, ?)
               WHERE INSTR(next_action, ?) > 0 AND INSTR(next_action, ?) = 0""",
            (old_air_flow, child_air_flow, old_air_flow, child_air_flow),
        )
        db.execute("DELETE FROM action_templates WHERE description = ?", (old_air_flow,))
        db.executemany(
            "INSERT OR IGNORE INTO action_templates (description) VALUES (?)",
            [(child_air_flow,), ("Profilaxia c/ Air Flow Adulto",), ("Reabilitação",)],
        )
        db.execute("UPDATE patient_followup SET next_action = REPLACE(next_action, 'Never', 'Niver') WHERE INSTR(next_action, 'Never') > 0")
        db.execute("DELETE FROM action_templates WHERE description = 'Never'")
        db.execute("INSERT OR IGNORE INTO action_templates (description) VALUES ('Niver')")
        db.executemany(
            "INSERT OR IGNORE INTO patient_statuses (name) VALUES (?)",
            [(status,) for status in sorted(STANDARD_PATIENT_STATUSES)],
        )
        db.execute("""
            INSERT OR IGNORE INTO crc_export_queue
            (patient_id, export_key, patient_name, phone, last_visit, professional_name, observation_text, status)
            SELECT p.id,
                   'crc-' || p.id || '-' || COALESCE(f.last_visit, 'sem-data'),
                   p.name, p.phone, f.last_visit, pr.name,
                   COALESCE((SELECT note FROM patient_visit_notes vn WHERE vn.patient_id=p.id ORDER BY datetime(vn.created_at) DESC, vn.id DESC LIMIT 1), ''),
                   'Pendente'
            FROM patients p
            JOIN patient_followup f ON f.patient_id=p.id
            JOIN patient_assignments pa ON pa.patient_id=p.id AND pa.is_primary=1
            JOIN professionals pr ON pr.id=pa.professional_id
            WHERE date(f.resolved_at)=date('now','localtime')
        """)
        # Integrações são gerenciadas exclusivamente pelo usuário. A inicialização
        # nunca deve criar, fundir ou apagar cartões e credenciais já cadastrados.
        api_integration_columns = {row[1] for row in db.execute("PRAGMA table_info(api_integrations)").fetchall()}
        for column, definition in {"subscriber_id": "TEXT", "sync_interval_seconds": "INTEGER NOT NULL DEFAULT 60", "last_sync_at": "TEXT", "last_sync_status": "TEXT", "last_sync_message": "TEXT", "last_sync_count": "INTEGER NOT NULL DEFAULT 0"}.items():
            if column not in api_integration_columns:
                db.execute(f"ALTER TABLE api_integrations ADD COLUMN {column} {definition}")
        api_sync_log_columns = {row[1] for row in db.execute("PRAGMA table_info(api_sync_logs)").fetchall()}
        if "phone_result" not in api_sync_log_columns:
            db.execute("ALTER TABLE api_sync_logs ADD COLUMN phone_result TEXT")
        integration_phone_rows = db.execute("""SELECT DISTINCT p.id, p.phone FROM patients p
                                                JOIN patient_events event ON event.patient_id=p.id
                                                WHERE event.event_type='Integração Clinicorp' AND p.phone IS NOT NULL""").fetchall()
        for integration_phone in integration_phone_rows:
            normalized_phone = re.sub(r"\D", "", integration_phone["phone"] or "")
            if normalized_phone.startswith("55") and len(normalized_phone) in {12, 13}:
                normalized_phone = normalized_phone[2:]
            if normalized_phone:
                db.execute("UPDATE patients SET phone=? WHERE id=?", (normalized_phone, integration_phone["id"]))
        procedure_columns = {row[1] for row in db.execute("PRAGMA table_info(procedures)").fetchall()}
        if "discount_cents" not in procedure_columns:
            db.execute("ALTER TABLE procedures ADD COLUMN discount_cents INTEGER NOT NULL DEFAULT 0")
        relationship_columns = {row[1] for row in db.execute("PRAGMA table_info(patient_relationships)").fetchall()}
        if "connection" not in relationship_columns:
            db.execute("ALTER TABLE patient_relationships ADD COLUMN connection TEXT")
        visit_note_columns = {row[1] for row in db.execute("PRAGMA table_info(patient_visit_notes)").fetchall()}
        if "author_name" not in visit_note_columns:
            db.execute("ALTER TABLE patient_visit_notes ADD COLUMN author_name TEXT")
        if "updated_at" not in visit_note_columns:
            db.execute("ALTER TABLE patient_visit_notes ADD COLUMN updated_at TEXT")
        assignment_columns = {row[1] for row in db.execute("PRAGMA table_info(patient_assignments)").fetchall()}
        for column, definition in {
            "journey_status": "TEXT NOT NULL DEFAULT 'Ativo'",
            "origin_professional_id": "INTEGER",
            "forward_reason": "TEXT",
            "completed_at": "TEXT",
            "stage_status": "TEXT NOT NULL DEFAULT 'Aguardando início'",
            "stage_note": "TEXT",
            "stage_updated_at": "TEXT",
        }.items():
            if column not in assignment_columns:
                db.execute(f"ALTER TABLE patient_assignments ADD COLUMN {column} {definition}")
        user_columns = {row[1] for row in db.execute("PRAGMA table_info(users)").fetchall()}
        for column, definition in {
            "password_hash": "TEXT",
            "password_salt": "TEXT",
            "must_change_password": "INTEGER NOT NULL DEFAULT 1",
            "permissions_json": "TEXT NOT NULL DEFAULT '{}'",
            "last_login_at": "TEXT",
            "two_factor_secret": "TEXT",
            "two_factor_enabled": "INTEGER NOT NULL DEFAULT 0",
            "two_factor_enrolled_at": "TEXT",
            "two_factor_exempt": "INTEGER NOT NULL DEFAULT 0",
            "linked_professional_id": "INTEGER",
        }.items():
            if column not in user_columns:
                db.execute(f"ALTER TABLE users ADD COLUMN {column} {definition}")
        users_sql = (db.execute("SELECT sql FROM sqlite_master WHERE type='table' AND name='users'").fetchone()[0] or "").lower()
        if "'asb'" not in users_sql:
            db.commit()
            db.execute("PRAGMA foreign_keys = OFF")
            db.execute("""CREATE TABLE users_crc_migration (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                professional_id INTEGER,
                name TEXT NOT NULL,
                email TEXT NOT NULL UNIQUE,
                access_role TEXT NOT NULL CHECK(access_role IN ('owner', 'professional', 'admin', 'crc', 'asb')),
                linked_professional_id INTEGER,
                active INTEGER NOT NULL DEFAULT 1,
                password_hash TEXT, password_salt TEXT,
                must_change_password INTEGER NOT NULL DEFAULT 1,
                permissions_json TEXT NOT NULL DEFAULT '{}', last_login_at TEXT,
                two_factor_secret TEXT, two_factor_enabled INTEGER NOT NULL DEFAULT 0, two_factor_enrolled_at TEXT,
                two_factor_exempt INTEGER NOT NULL DEFAULT 0,
                crm_channel_scope_enabled INTEGER NOT NULL DEFAULT 0,
                crm_feature_scope_enabled INTEGER NOT NULL DEFAULT 0,
                crm_operational_agent INTEGER NOT NULL DEFAULT 1,
                crm_manage_automation INTEGER NOT NULL DEFAULT 0,
                crm_access_level TEXT NOT NULL DEFAULT 'attendant',
                service_sector TEXT NOT NULL DEFAULT '',
                FOREIGN KEY (professional_id) REFERENCES professionals(id), FOREIGN KEY (linked_professional_id) REFERENCES professionals(id))""")
            db.execute("""INSERT INTO users_crc_migration
                (id, professional_id, name, email, access_role, linked_professional_id, active, password_hash, password_salt, must_change_password, permissions_json, last_login_at, two_factor_secret, two_factor_enabled, two_factor_enrolled_at, two_factor_exempt, crm_channel_scope_enabled, crm_feature_scope_enabled, crm_operational_agent, crm_manage_automation, crm_access_level, service_sector)
                SELECT id, professional_id, name, email, access_role, linked_professional_id, active, password_hash, password_salt, must_change_password, permissions_json, last_login_at, two_factor_secret, two_factor_enabled, two_factor_enrolled_at, two_factor_exempt, crm_channel_scope_enabled, crm_feature_scope_enabled, crm_operational_agent, crm_manage_automation, crm_access_level, service_sector FROM users""")
            db.execute("DROP TABLE users")
            db.execute("ALTER TABLE users_crc_migration RENAME TO users")
            db.commit()
            db.execute("PRAGMA foreign_keys = ON")
        session_columns = {row[1] for row in db.execute("PRAGMA table_info(auth_sessions)").fetchall()}
        if "active_professional_id" not in session_columns:
            db.execute("ALTER TABLE auth_sessions ADD COLUMN active_professional_id INTEGER")
        db.execute("""INSERT OR IGNORE INTO asb_professional_links (user_id, professional_id)
                      SELECT id, linked_professional_id FROM users WHERE access_role='asb' AND linked_professional_id IS NOT NULL""")
        professional_columns = {row[1] for row in db.execute("PRAGMA table_info(professionals)").fetchall()}
        for column, definition in {"photo_data": "TEXT", "photo_mime": "TEXT"}.items():
            if column not in professional_columns:
                db.execute(f"ALTER TABLE professionals ADD COLUMN {column} {definition}")
        db.execute("""CREATE TABLE IF NOT EXISTS password_reset_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER NOT NULL, requested_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            status TEXT NOT NULL DEFAULT 'pending', completed_at TEXT, completed_by_user_id INTEGER,
            FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE,
            FOREIGN KEY (completed_by_user_id) REFERENCES users(id))""")
        db.execute("""
            INSERT INTO procedures (patient_id, name, value_cents, stage)
            SELECT f.patient_id, f.procedure_name, 0, 'Indicado'
            FROM patient_followup f
            WHERE f.procedure_name IS NOT NULL AND TRIM(f.procedure_name) != ''
              AND NOT EXISTS (SELECT 1 FROM procedures pr WHERE pr.patient_id = f.patient_id)
        """)
        db.executemany(
            "INSERT OR IGNORE INTO specialties (code, name) VALUES (?, ?)",
            [
                ("ORTODONTIA", "Ortodontia"),
                ("ODONTOPEDIATRIA", "Odontopediatria"),
                ("OUTROS", "Outros"),
            ],
        )
        cleanup_retention_data(db)
        migrate_n8n_config_file()
        office_id = db.execute(
            "INSERT OR IGNORE INTO offices (code, name) VALUES ('CONSULTORIO_3', 'Consultório 3') RETURNING id"
        ).fetchone()
        if office_id:
            office_id = office_id[0]
        else:
            office_id = db.execute("SELECT id FROM offices WHERE code = 'CONSULTORIO_3'").fetchone()[0]
        owner = db.execute("SELECT id FROM professionals WHERE is_owner = 1 ORDER BY id LIMIT 1").fetchone()
        if owner:
            db.execute("""
                INSERT INTO professional_offices (professional_id, office_id, is_responsible)
                VALUES (?, ?, 1)
                ON CONFLICT(professional_id, office_id) DO UPDATE SET is_responsible = 1
            """, (owner[0], office_id))
        count = db.execute("SELECT COUNT(*) FROM patients").fetchone()[0]
        if count:
            return
        professional_id = db.execute(
            "INSERT INTO professionals (name, role, is_owner) VALUES (?, ?, ?)",
            ("Dra. Dulce", "Cirurgiã-dentista", 1),
        ).lastrowid
        db.execute("""
            INSERT INTO professional_offices (professional_id, office_id, is_responsible)
            VALUES (?, ?, 1)
            ON CONFLICT(professional_id, office_id) DO UPDATE SET is_responsible = 1
        """, (professional_id, office_id))
        db.execute(
            "INSERT INTO users (professional_id, name, email, access_role) VALUES (?, ?, ?, ?)",
            (professional_id, "Dra. Dulce", "dulce@instituto.local", "owner"),
        )
        seed = json.loads(SEED_PATH.read_text(encoding="utf-8"))
        for patient in seed:
            patient_id = db.execute(
                """
                INSERT INTO patients
                (external_id, name, phone, reference, status, notes, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                """,
                (
                    patient["external_id"], patient["name"], patient["phone"],
                    patient["reference"], patient["status"], patient["notes"],
                ),
            ).lastrowid
            db.execute(
                "INSERT INTO patient_assignments (patient_id, professional_id, is_primary) VALUES (?, ?, 1)",
                (patient_id, professional_id),
            )
            db.execute(
                """
                INSERT INTO patient_followup
                (patient_id, last_visit, next_appointment, next_appointment_type, procedure_name, last_contact, next_action)
                VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    patient_id, patient["last_visit"], patient["next_appointment"],
                    "Agendado" if patient["next_appointment"] else None,
                    patient["procedure"], patient["last_contact"], patient["next_action"],
                ),
            )


def patient_select() -> str:
    return """
        SELECT p.id, p.external_id, p.name, p.phone, p.reference, p.status, p.notes,
               f.last_visit, f.next_appointment,
               CASE WHEN f.next_appointment IS NOT NULL THEN COALESCE(f.next_appointment_type, 'Agendado') END AS next_appointment_type,
               f.procedure_name AS procedure,
               f.last_contact, f.next_action, f.resolved_at, f.crc_status, f.crc_started_at, f.crc_completed_at,
               CASE WHEN date(f.resolved_at) = date('now', 'localtime') THEN 1 ELSE 0 END AS is_resolved_today,
               pr.name AS professional,
               (SELECT GROUP_CONCAT(px.name, ', ') FROM procedures px
                 WHERE px.patient_id = p.id AND px.stage != 'Concluído') AS procedure_summary,
               (SELECT COUNT(*) FROM procedures px WHERE px.patient_id = p.id) AS procedure_count,
               (SELECT COALESCE(SUM(GREATEST(px.value_cents - px.discount_cents, 0)), 0) FROM procedures px
                 WHERE px.patient_id = p.id AND px.stage != 'Concluído') AS potential_value_cents,
               COALESCE((SELECT cp.health_change FROM patient_clinical_profile cp WHERE cp.patient_id = p.id), 0) AS has_health_change,
               CASE WHEN f.next_appointment IS NOT NULL AND date(f.next_appointment) >= date('now', 'localtime') THEN 1 ELSE 0 END AS is_scheduled,
               CAST(julianday('now', 'localtime') - julianday(CASE WHEN f.next_appointment IS NOT NULL AND date(f.next_appointment) < date('now', 'localtime') THEN f.next_appointment ELSE f.last_visit END) AS INTEGER) AS days_away
               ,COALESCE((SELECT q.status FROM crc_export_queue q WHERE q.patient_id=p.id ORDER BY q.id DESC LIMIT 1), 'Não enviado') AS export_status
               ,COALESCE((SELECT q.message_created FROM crc_export_queue q WHERE q.patient_id=p.id ORDER BY q.id DESC LIMIT 1), '') AS message_created
               ,(SELECT COUNT(*) FROM patient_assignments shared_pa
                  WHERE shared_pa.patient_id=p.id AND shared_pa.is_primary=0 AND shared_pa.journey_status='Ativo') AS journey_professional_count
               ,COALESCE((SELECT GROUP_CONCAT(shared_pr.name || '|' || COALESCE(shared_pa.stage_status, 'Aguardando início'), '||')
                  FROM patient_assignments shared_pa
                  JOIN professionals shared_pr ON shared_pr.id=shared_pa.professional_id
                  WHERE shared_pa.patient_id=p.id AND shared_pa.is_primary=0 AND shared_pa.journey_status='Ativo'), '') AS journey_summary
               ,(SELECT pel.user_id FROM patient_edit_locks pel
                  WHERE pel.patient_id=p.id AND datetime(pel.expires_at) > datetime('now', 'localtime')
                  LIMIT 1) AS edit_lock_user_id
               ,(SELECT u.name FROM patient_edit_locks pel
                  JOIN users u ON u.id=pel.user_id
                  WHERE pel.patient_id=p.id AND datetime(pel.expires_at) > datetime('now', 'localtime')
                  LIMIT 1) AS edit_lock_user_name
               ,(SELECT pel.expires_at FROM patient_edit_locks pel
                  WHERE pel.patient_id=p.id AND datetime(pel.expires_at) > datetime('now', 'localtime')
                  LIMIT 1) AS edit_lock_expires_at
        FROM patients p
        JOIN patient_followup f ON f.patient_id = p.id
        JOIN patient_assignments pa ON pa.patient_id = p.id AND pa.is_primary = 1
        JOIN professionals pr ON pr.id = pa.professional_id
    """


class ClinicHandler(SimpleHTTPRequestHandler):
    server_version = "InstitutoAyubPilot/1.0"

    def setup(self) -> None:
        super().setup()
        self.request_id = secrets.token_hex(12)
        self.request_started_at = time.monotonic()

    @staticmethod
    def redact_log_value(value) -> str:
        text = str(value)
        text = re.sub(
            r"(?i)([?&](?:token|webhook_key|key|secret|api_key)=)[^&\s]+",
            r"\1[REDACTED]",
            text,
        )
        return re.sub(
            r'(?i)("?(?:api_token|api_key|secret|authorization)"?\s*[:=]\s*"?)[^",\s]+',
            r"\1[REDACTED]",
            text,
        )

    def log_message(self, fmt: str, *args) -> None:
        safe_args = tuple(self.redact_log_value(value) for value in args)
        started_at = getattr(self, "request_started_at", time.monotonic())
        event = {
            "timestamp": datetime.now(timezone.utc).isoformat(timespec="milliseconds"),
            "level": "info",
            "event": "http_request",
            "request_id": getattr(self, "request_id", "unavailable"),
            "method": getattr(self, "command", None),
            "path": self.redact_log_value(getattr(self, "path", "")),
            "remote_ip": (getattr(self, "client_address", (None,))[0]),
            "duration_ms": round((time.monotonic() - started_at) * 1000),
            "message": self.redact_log_value(fmt % safe_args),
        }
        print(json.dumps(event, ensure_ascii=False, separators=(",", ":")), flush=True)

    def end_headers(self) -> None:
        self.send_header("X-Request-ID", getattr(self, "request_id", "unavailable"))
        super().end_headers()

    def send_security_headers(self, allow_bundled_ui: bool = False, allow_same_origin_frame: bool = False, bundled_ui_nonce: str | None = None) -> None:
        self.send_header("Strict-Transport-Security", "max-age=31536000; includeSubDomains")
        frame_ancestors = "'self'" if allow_same_origin_frame else "'none'"
        if allow_bundled_ui:
            # 'unsafe-eval' continua necessário: o bundle roda Babel standalone
            # no navegador para transpilar JSX em tempo real (ver crm-whatsapp.html).
            # 'unsafe-inline' foi substituído por nonce por requisição — os
            # <script> inline do bundle carregam o nonce correspondente (SEC-008).
            script_src = f"'self' 'nonce-{bundled_ui_nonce}' 'unsafe-eval' blob:" if bundled_ui_nonce else "'self' 'unsafe-inline' 'unsafe-eval' blob:"
            self.send_header("Content-Security-Policy", f"default-src 'self' blob: data:; script-src {script_src}; style-src 'self' 'unsafe-inline' blob: https://fonts.googleapis.com; font-src 'self' blob: data: https://fonts.gstatic.com; img-src 'self' blob: data:; media-src 'self' blob: data:; connect-src 'self'; frame-ancestors {frame_ancestors}; base-uri 'self'; form-action 'self'; upgrade-insecure-requests; block-all-mixed-content")
        else:
            self.send_header("Content-Security-Policy", f"default-src 'self'; script-src 'self'; style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; font-src 'self' https://fonts.gstatic.com data:; img-src 'self' data:; connect-src 'self'; frame-ancestors {frame_ancestors}; base-uri 'self'; form-action 'self'; upgrade-insecure-requests; block-all-mixed-content")
        self.send_header("X-Content-Type-Options", "nosniff")
        self.send_header("Referrer-Policy", "strict-origin-when-cross-origin")
        self.send_header("Permissions-Policy", "camera=(), microphone=(self), geolocation=()")

    def send_json(self, payload, status=HTTPStatus.OK) -> None:
        body = json.dumps(payload, ensure_ascii=False, default=str).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.send_security_headers()
        self.end_headers()
        self.wfile.write(body)

    def send_json_with_cookie(self, payload, cookie: str, status=HTTPStatus.OK) -> None:
        body = json.dumps(payload, ensure_ascii=False, default=str).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.send_header("Set-Cookie", cookie)
        self.send_security_headers()
        self.end_headers()
        self.wfile.write(body)

    def current_user(self):
        raw_cookie = self.headers.get("Cookie", "")
        cookies = SimpleCookie()
        cookies.load(raw_cookie)
        session = cookies.get("iea_session")
        if not session:
            return None
        token_hash = hashlib.sha256(session.value.encode("utf-8")).hexdigest()
        with connect() as db:
            row = db.execute("""
                SELECT u.id, u.professional_id, u.linked_professional_id, u.name, u.email, u.access_role, u.crm_access_level, u.service_sector, u.permissions_json, pr.photo_data,
                       linked_pr.name AS linked_professional_name,
                       u.must_change_password, u.active, u.two_factor_enabled, u.two_factor_secret, u.two_factor_exempt,
                       (SELECT GROUP_CONCAT(o.name, ', ') FROM professional_offices po JOIN offices o ON o.id=po.office_id WHERE po.professional_id=u.professional_id AND o.active=1) AS offices,
                       (SELECT GROUP_CONCAT(s.name, ', ') FROM professional_specialties ps JOIN specialties s ON s.id=ps.specialty_id WHERE ps.professional_id=u.professional_id AND s.active=1) AS specialties
                FROM auth_sessions s JOIN users u ON u.id = s.user_id
                LEFT JOIN professionals pr ON pr.id = u.professional_id
                LEFT JOIN professionals linked_pr ON linked_pr.id = u.linked_professional_id
                WHERE s.token_hash = ? AND s.expires_at > datetime('now') AND u.active = 1
            """, (token_hash,)).fetchone()
            if row:
                db.execute("UPDATE auth_sessions SET last_seen_at=CURRENT_TIMESTAMP WHERE token_hash=?", (token_hash,))
        return dict(row) if row else None

    def require_auth(self, roles=None):
        user = self.current_user()
        if not user:
            self.send_json({"error": "Autenticação necessária"}, HTTPStatus.UNAUTHORIZED)
            return None
        if roles and user["access_role"] not in roles:
            self.send_json({"error": "Você não tem permissão para esta ação"}, HTTPStatus.FORBIDDEN)
            return None
        return user

    @staticmethod
    def password_digest(password: str, salt: str) -> str:
        return hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), 600_000).hex()

    def request_ip(self) -> str:
        # Nginx anexa o IP observado ao fim da cadeia. Usar o primeiro valor
        # permitiria que um cliente forjasse o IP empregado em limites e auditoria.
        forwarded_chain = [
            item.strip() for item in self.headers.get("X-Forwarded-For", "").split(",") if item.strip()
        ]
        forwarded = forwarded_chain[-1] if forwarded_chain else ""
        return forwarded or self.client_address[0]

    @staticmethod
    def totp_code(secret: str, timestamp: int | None = None) -> str:
        # TOTP usa Unix time absoluto. `datetime.utcnow().timestamp()` trata o
        # datetime ingênuo como horário local e desloca o código quando o
        # processo está em UTC-4.
        timestamp = int(time.time()) if timestamp is None else int(timestamp)
        key = base64.b32decode(secret.upper() + "=" * (-len(secret) % 8))
        digest = hmac.new(key, struct.pack(">Q", timestamp // 30), hashlib.sha1).digest()
        offset = digest[-1] & 0x0F
        value = (struct.unpack(">I", digest[offset:offset + 4])[0] & 0x7FFFFFFF) % 1_000_000
        return f"{value:06d}"

    def valid_totp(self, secret: str | None, code: str) -> bool:
        if not secret or not re.fullmatch(r"\d{6}", code or ""):
            return False
        now = int(time.time())
        return any(hmac.compare_digest(self.totp_code(secret, now + step * 30), code) for step in (-1, 0, 1))

    @staticmethod
    def new_totp_secret() -> str:
        return base64.b32encode(secrets.token_bytes(20)).decode("ascii").rstrip("=")

    @staticmethod
    def encrypt_totp_secret(secret: str) -> str:
        if not secret:
            return secret
        if not APP_SECRET_KEY or AESGCM is None:
            raise RuntimeError("Criptografia do 2FA indisponível. Configure APP_SECRET_KEY antes de ativar o recurso.")
        aesgcm = AESGCM(APP_SECRET_KEY)
        nonce = secrets.token_bytes(12)
        ciphertext = aesgcm.encrypt(nonce, secret.encode("utf-8"), None)
        return TOTP_ENC_PREFIX + base64.b64encode(nonce + ciphertext).decode("ascii")

    @staticmethod
    def decrypt_totp_secret(value: str | None) -> str | None:
        if not value or not value.startswith(TOTP_ENC_PREFIX):
            return value  # texto claro legado (pré-SEC-011) ou vazio
        if not APP_SECRET_KEY or AESGCM is None:
            raise RuntimeError("APP_SECRET_KEY ausente: não é possível decifrar o segredo TOTP armazenado")
        raw = base64.b64decode(value[len(TOTP_ENC_PREFIX):])
        aesgcm = AESGCM(APP_SECRET_KEY)
        return aesgcm.decrypt(raw[:12], raw[12:], None).decode("utf-8")

    @staticmethod
    def otp_uri(email: str, secret: str) -> str:
        from urllib.parse import quote
        return f"otpauth://totp/Instituto%20Eduardo%20Ayub:{quote(email)}?secret={secret}&issuer=Instituto%20Eduardo%20Ayub&algorithm=SHA1&digits=6&period=30"

    @staticmethod
    def record_security_event(db, event_type: str, ip_address: str, user_id=None, detail=None) -> None:
        db.execute("INSERT INTO security_events (user_id, event_type, detail, ip_address) VALUES (?, ?, ?, ?)", (user_id, event_type, detail, ip_address))

    def issue_session(self, db, user_id: int) -> str:
        token = secrets.token_urlsafe(48)
        token_hash = hashlib.sha256(token.encode("utf-8")).hexdigest()
        db.execute("DELETE FROM auth_sessions WHERE expires_at <= datetime('now')")
        db.execute("INSERT INTO auth_sessions (user_id, token_hash, expires_at) VALUES (?, ?, ?)", (user_id, token_hash, (datetime.utcnow() + timedelta(hours=12)).strftime("%Y-%m-%d %H:%M:%S")))
        return token

    @staticmethod
    def session_cookie(token: str) -> str:
        return f"iea_session={token}; HttpOnly; SameSite=Strict; Path=/; Max-Age=43200" + ("; Secure" if os.environ.get("SESSION_SECURE") == "1" else "")

    def auth_payload(self, user: dict) -> dict:
        permissions = json.loads(user["permissions_json"] or "{}")
        return {"authenticated": True, "user": {
            "id": user["id"], "name": user["name"], "email": user["email"],
            "role": user["access_role"], "professional_id": user["professional_id"],
            "crm_access_level": user.get("crm_access_level") or "attendant",
            "linked_professional_id": user.get("linked_professional_id"),
            "portfolio_professional_id": (
                user.get("linked_professional_id")
                if user["access_role"] == "asb" and user.get("linked_professional_id")
                else user["professional_id"]
            ),
            "portfolio_professional_name": (
                user.get("linked_professional_name")
                if user["access_role"] == "asb" and user.get("linked_professional_id")
                else user["name"]
            ),
            "service_sector": user.get("service_sector") or ("CRC" if user["access_role"] == "crc" else ""),
            "offices": user.get("offices"), "specialties": user.get("specialties"),
            "photo_url": f"/api/professionals/{user['professional_id']}/photo" if user.get("photo_data") else ("/assets/dra-dulce.jpeg" if user["access_role"] == "owner" else None),
            "must_change_password": bool(user["must_change_password"]),
            "two_factor_enabled": bool(user.get("two_factor_enabled")),
            "two_factor_exempt": bool(user.get("two_factor_exempt")),
            "permissions": permissions,
            "can_admin_portal": self.can_admin_portal(user),
            "can_manage_crm": self.can_manage_crm(user),
            "admin_path": ADMIN_ROUTE if self.can_admin_portal(user) else None,
        }}

    @staticmethod
    def can_admin_portal(user: dict) -> bool:
        if user["access_role"] == "owner":
            return True
        try:
            permissions = json.loads(user.get("permissions_json") or "{}")
        except (TypeError, json.JSONDecodeError):
            permissions = {}
        return user["access_role"] == "admin" and permissions.get("admin_portal", True) is not False

    @classmethod
    def can_manage_crm(cls, user: dict) -> bool:
        """Administra somente o CRM, sem herdar o painel administrativo geral."""
        return cls.can_admin_portal(user) or (
            user.get("access_role") == "crc" and user.get("crm_access_level") == "admin"
        )

    @staticmethod
    def can_manage_crc_fields(user: dict) -> bool:
        return user["access_role"] == "crc"

    def send_svg(self, svg: bytes) -> None:
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", "image/svg+xml; charset=utf-8")
        self.send_header("Content-Length", str(len(svg)))
        self.send_header("Cache-Control", "no-store")
        self.send_security_headers()
        self.end_headers()
        self.wfile.write(svg)

    def send_image(self, image: bytes, mime_type: str) -> None:
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", mime_type)
        self.send_header("Content-Length", str(len(image)))
        self.send_header("Cache-Control", "private, max-age=3600")
        self.send_security_headers()
        self.end_headers()
        self.wfile.write(image)

    def get_two_factor_qr(self, user: dict) -> None:
        with connect() as db:
            row = db.execute("SELECT two_factor_secret FROM users WHERE id=?", (user["id"],)).fetchone()
        if not row or not row["two_factor_secret"]:
            return self.send_json({"error": "Prepare a autenticação em duas etapas antes de gerar o QR Code."}, HTTPStatus.CONFLICT)
        qr = qrcode.make(self.otp_uri(user["email"], self.decrypt_totp_secret(row["two_factor_secret"])), image_factory=SvgPathImage, box_size=8, border=3)
        self.send_svg(qr.to_string())

    def get_professional_photo(self, professional_id: int, user: dict) -> None:
        with connect() as db:
            row = db.execute("SELECT name, photo_data, photo_mime FROM professionals WHERE id=?", (professional_id,)).fetchone()
        if not row:
            return self.send_json({"error": "Foto não encontrada"}, HTTPStatus.NOT_FOUND)
        # A foto institucional da Dra. Dulce usa sempre o arquivo mestre em alta resolução.
        if "dulc" in str(row["name"] or "").casefold():
            source = PUBLIC / "assets" / "dra-dulce.jpeg"
            if source.exists():
                return self.send_image(source.read_bytes(), "image/jpeg")
        if not row["photo_data"]:
            return self.send_json({"error": "Foto não encontrada"}, HTTPStatus.NOT_FOUND)
        self.send_image(base64.b64decode(row["photo_data"]), row["photo_mime"] or "image/jpeg")

    def get_crm_contact_profile_photo(self, contact_id: int) -> None:
        """Serve fotos do WhatsApp pelo pr\u00f3prio CRM, sem expor a URL externa ao navegador."""
        if not self.require_crc_access():
            return
        if not self.require_crm_any_feature((*CRM_WORKSPACE_FEATURES, "contacts")):
            return
        scope_sql, scope_params = self.crm_channel_scope_clause("ch")
        with connect() as db:
            row = db.execute("""SELECT ct.profile_picture_url,ct.phone,ch.instance_name
                FROM crm_contacts ct
                LEFT JOIN crm_conversations cv ON cv.contact_id=ct.id
                LEFT JOIN crm_channels ch ON ch.id=cv.channel_id
                WHERE ct.id=? AND (COALESCE(ct.is_internal,0)=1 OR {scope_sql})
                ORDER BY datetime(cv.last_message_at) DESC,cv.id DESC LIMIT 1""".format(scope_sql=scope_sql),
                (contact_id, *scope_params)).fetchone()
        picture_url = str(row["profile_picture_url"] or "").strip() if row else ""
        if not picture_url and row and row["phone"] and row["instance_name"]:
            with CRM_PROFILE_PHOTO_MISS_LOCK:
                retry_after = CRM_PROFILE_PHOTO_MISS.get(contact_id, 0)
            if time.time() >= retry_after:
                try:
                    payload = self.evolution_api_request(
                        f"/chat/fetchProfilePictureUrl/{quote(str(row['instance_name']), safe='')}",
                        "POST", {"number": self.crm_phone(row["phone"])}
                    )
                    candidates = [payload]
                    if isinstance(payload, dict):
                        candidates.extend([payload.get("data"), payload.get("response")])
                    for candidate in candidates:
                        if not isinstance(candidate, dict):
                            continue
                        found = candidate.get("profilePictureUrl") or candidate.get("profilePicUrl") or candidate.get("picture") or candidate.get("url")
                        if found:
                            picture_url = str(found).strip()
                            break
                    if picture_url:
                        with connect() as db:
                            db.execute("UPDATE crm_contacts SET profile_picture_url=?,updated_at=datetime('now','localtime') WHERE id=?", (picture_url, contact_id))
                    else:
                        with CRM_PROFILE_PHOTO_MISS_LOCK:
                            CRM_PROFILE_PHOTO_MISS[contact_id] = time.time() + 6 * 3600
                except RuntimeError:
                    with CRM_PROFILE_PHOTO_MISS_LOCK:
                        CRM_PROFILE_PHOTO_MISS[contact_id] = time.time() + 6 * 3600
        parsed = urlparse(picture_url)
        host = (parsed.hostname or "").casefold()
        allowed_hosts = ("whatsapp.net", "fbcdn.net", "facebook.com")
        if parsed.scheme != "https" or not host or not any(host == suffix or host.endswith("." + suffix) for suffix in allowed_hosts):
            return self.send_json({"error": "Foto indispon\u00edvel"}, HTTPStatus.NOT_FOUND)
        try:
            request = Request(picture_url, headers={"User-Agent": "Instituto-IEA-CRM/1.0", "Accept": "image/*"})
            with urlopen(request, timeout=8) as response:
                content_type = str(response.headers.get_content_type() or "").lower()
                image = response.read(4 * 1024 * 1024 + 1)
            if not content_type.startswith("image/") or not image or len(image) > 4 * 1024 * 1024:
                raise ValueError("Resposta de foto inv\u00e1lida")
            return self.send_image(image, content_type)
        except (HTTPError, URLError, TimeoutError, ValueError, OSError):
            return self.send_json({"error": "Foto indispon\u00edvel"}, HTTPStatus.NOT_FOUND)

    def require_patient_access(self, patient_id: int, user: dict) -> bool:
        if user["access_role"] in {"admin", "crc"}:
            return True
        portfolio_id = user.get("linked_professional_id") if user["access_role"] == "asb" else user["professional_id"]
        with connect() as db:
            assigned = db.execute("SELECT 1 FROM patient_assignments WHERE patient_id=? AND professional_id=?", (patient_id, portfolio_id)).fetchone()
        if assigned:
            return True
        self.send_json({"error": "Este paciente não pertence à sua carteira"}, HTTPStatus.FORBIDDEN)
        return False

    @staticmethod
    def active_patient_edit_lock(db, patient_id: int):
        return db.execute(
            """
            SELECT pel.user_id, pel.expires_at, u.name AS user_name
            FROM patient_edit_locks pel
            JOIN users u ON u.id=pel.user_id
            WHERE pel.patient_id=? AND datetime(pel.expires_at) > datetime('now', 'localtime')
            """,
            (patient_id,),
        ).fetchone()

    def patient_edit_lock_conflict(self, db, patient_id: int) -> bool:
        lock = self.active_patient_edit_lock(db, patient_id)
        if lock and int(lock["user_id"]) != int(self.authenticated_user["id"]):
            self.send_json(
                {
                    "error": f"Este prontuário está em atendimento por {lock['user_name']}.",
                    "code": "patient_in_use",
                    "locked_by": lock["user_name"],
                    "expires_at": lock["expires_at"],
                },
                HTTPStatus.CONFLICT,
            )
            return True
        return False

    def acquire_patient_edit_lock(self, patient_id: int) -> None:
        if not self.require_patient_access(patient_id, self.authenticated_user):
            return
        with connect() as db:
            db.execute(
                "DELETE FROM patient_edit_locks WHERE datetime(expires_at) <= datetime('now', 'localtime')"
            )
            lock = self.active_patient_edit_lock(db, patient_id)
            if lock and int(lock["user_id"]) != int(self.authenticated_user["id"]):
                return self.send_json({
                    "acquired": False,
                    "locked_by_user_id": lock["user_id"],
                    "locked_by": lock["user_name"],
                    "expires_at": lock["expires_at"],
                })
            db.execute(
                """
                INSERT INTO patient_edit_locks
                    (patient_id, user_id, acquired_at, heartbeat_at, expires_at)
                VALUES
                    (?, ?, datetime('now','localtime'), datetime('now','localtime'), datetime('now','localtime','+2 minutes'))
                ON CONFLICT(patient_id) DO UPDATE SET
                    user_id=excluded.user_id,
                    heartbeat_at=excluded.heartbeat_at,
                    expires_at=excluded.expires_at
                """,
                (patient_id, self.authenticated_user["id"]),
            )
            lock = self.active_patient_edit_lock(db, patient_id)
        self.send_json({
            "acquired": True,
            "locked_by_user_id": self.authenticated_user["id"],
            "locked_by": self.authenticated_user["name"],
            "expires_at": lock["expires_at"] if lock else None,
        })

    def release_patient_edit_lock(self, patient_id: int) -> None:
        with connect() as db:
            result = db.execute(
                "DELETE FROM patient_edit_locks WHERE patient_id=? AND user_id=?",
                (patient_id, self.authenticated_user["id"]),
            )
        self.send_json({"released": bool(result.rowcount)})

    @staticmethod
    def portfolio_professional_id(user: dict) -> int | None:
        if user["access_role"] == "asb" and user.get("linked_professional_id"):
            return int(user["linked_professional_id"])
        return int(user["professional_id"]) if user.get("professional_id") else None

    def read_json(self) -> dict:
        length = int(self.headers.get("Content-Length", "0"))
        if length > MAX_BODY_BYTES:
            self.send_json({"error": "Corpo da requisição excede o limite permitido."}, HTTPStatus.REQUEST_ENTITY_TOO_LARGE)
            raise ConnectionAbortedError("payload too large")
        return json.loads(self.rfile.read(length).decode("utf-8")) if length else {}

    def get_health(self) -> None:
        try:
            with connect() as db:
                db.execute("SELECT 1").fetchone()
        except Exception:
            return self.send_json(
                {"status": "degraded", "database": "unavailable", "release": RELEASE_ID},
                HTTPStatus.SERVICE_UNAVAILABLE,
            )
        self.send_json({"status": "ok", "database": "ok", "api_version": 5, "release": RELEASE_ID})

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/api/health":
            return self.get_health()
        if parsed.path == "/api/release":
            return self.send_json({"release": RELEASE_ID})
        if parsed.path == "/api/auth/status":
            user = self.current_user()
            return self.send_json(self.auth_payload(user) if user else {"authenticated": False})
        if parsed.path == "/api/auth/2fa/qr":
            user = self.require_auth()
            return self.get_two_factor_qr(user) if user else None
        photo_match = re.fullmatch(r"/api/professionals/(\d+)/photo", parsed.path)
        if photo_match:
            user = self.require_auth()
            return self.get_professional_photo(int(photo_match.group(1)), user) if user else None
        if parsed.path.startswith("/api/"):
            self.authenticated_user = self.require_auth()
            if not self.authenticated_user:
                return
        crm_contact_photo_match = re.fullmatch(r"/api/crm/contacts/(\d+)/profile-photo", parsed.path)
        if crm_contact_photo_match:
            return self.get_crm_contact_profile_photo(int(crm_contact_photo_match.group(1)))
        crm_admin_endpoint = parsed.path == "/api/admin/crm-channel-access"
        if parsed.path.startswith("/api/admin/") and not (
            self.can_admin_portal(self.authenticated_user)
            or (crm_admin_endpoint and self.can_manage_crm(self.authenticated_user))
        ):
            return self.send_json({"error": "Acesso administrativo necessário"}, HTTPStatus.FORBIDDEN)
        if parsed.path == "/api/admin":
            if not self.can_admin_portal(self.authenticated_user):
                return self.send_json({"error": "Acesso administrativo necessário"}, HTTPStatus.FORBIDDEN)
            return self.get_admin_overview()
        if parsed.path == "/api/admin/crm-channel-access":
            if not self.can_manage_crm(self.authenticated_user):
                return self.send_json({"error": "Acesso de administrador do CRM necessário"}, HTTPStatus.FORBIDDEN)
            return self.get_admin_crm_channel_access()
        if parsed.path == "/api/admin/audit":
            if not self.can_admin_portal(self.authenticated_user):
                return self.send_json({"error": "Acesso administrativo necessário"}, HTTPStatus.FORBIDDEN)
            return self.get_admin_audit(parse_qs(parsed.query))
        if parsed.path == "/api/admin/integrations/clinicorp":
            if not self.can_admin_portal(self.authenticated_user):
                return self.send_json({"error": "Acesso administrativo necessário"}, HTTPStatus.FORBIDDEN)
            return self.get_clinicorp_config()
        if parsed.path == "/api/admin/apis":
            if not self.can_admin_portal(self.authenticated_user):
                return self.send_json({"error": "Acesso administrativo necessário"}, HTTPStatus.FORBIDDEN)
            return self.get_api_integrations()
        api_sync_status_match = re.fullmatch(r"/api/admin/apis/(\d+)/sync-status", parsed.path)
        if api_sync_status_match:
            if not self.can_admin_portal(self.authenticated_user):
                return self.send_json({"error": "Acesso administrativo necessário"}, HTTPStatus.FORBIDDEN)
            return self.get_api_sync_status(int(api_sync_status_match.group(1)))
        if parsed.path == "/api/specialties":
            return self.get_specialties()
        if parsed.path == "/api/dashboard":
            return self.get_dashboard()
        if parsed.path == "/api/filters":
            return self.get_filters()
        if parsed.path == "/api/crm/contacts":
            return self.get_crm_contacts()
        if parsed.path == "/api/crm/channels":
            return self.get_crm_channels()
        if parsed.path == "/api/crm/evolution/instances":
            return self.get_evolution_instances()
        if parsed.path == "/api/crm/evolution/config":
            return self.get_evolution_config()
        if parsed.path == "/api/crm/evolution/sync":
            return self.get_evolution_sync_status()
        if parsed.path == "/api/crm/conversations":
            return self.get_crm_conversations(parse_qs(parsed.query))
        if parsed.path == "/api/crm/tags":
            return self.get_crm_tags()
        if parsed.path == "/api/crm/metrics":
            return self.get_crm_metrics(parse_qs(parsed.query))
        if parsed.path == "/api/crm/goals":
            return self.get_crm_goals(parse_qs(parsed.query))
        if parsed.path == "/api/crm/resolution-reports":
            return self.get_crm_resolution_reports(parse_qs(parsed.query))
        if parsed.path == "/api/crm/patient-control":
            return self.get_crm_patient_control(parse_qs(parsed.query))
        if parsed.path == "/api/crm/quick-replies":
            return self.get_crm_quick_replies()
        if parsed.path == "/api/crm/integrations/health":
            return self.get_crm_integration_health()
        if parsed.path == "/api/crm/n8n/config":
            return self.get_crm_n8n_config()
        if parsed.path == "/api/crm/n8n/overview":
            return self.get_crm_n8n_overview(parse_qs(parsed.query))
        if parsed.path == "/api/crm/n8n/runs":
            return self.get_crm_n8n_runs(parse_qs(parsed.query))
        if parsed.path == "/api/crm/n8n/patient-events":
            return self.get_crm_n8n_patient_events(parse_qs(parsed.query))
        if parsed.path == "/api/crm/campaigns":
            return self.get_crm_campaigns(parse_qs(parsed.query))
        if parsed.path == "/api/crm/n8n/versions":
            return self.get_crm_n8n_versions()
        if parsed.path == "/api/crm/n8n/callback-keys":
            return self.get_crm_n8n_callback_keys()
        n8n_workflow_detail_match = re.fullmatch(r"/api/crm/n8n/workflows/([^/]+)", parsed.path)
        if n8n_workflow_detail_match:
            return self.get_crm_n8n_workflow_detail(n8n_workflow_detail_match.group(1))
        n8n_run_detail_match = re.fullmatch(r"/api/crm/n8n/runs/(\d+)", parsed.path)
        if n8n_run_detail_match:
            return self.get_crm_n8n_run_detail(int(n8n_run_detail_match.group(1)))
        if parsed.path == "/api/crm/agents":
            return self.get_crm_agents()
        if parsed.path == "/api/crm/permissions":
            return self.get_crm_permissions()
        crm_messages_match = re.fullmatch(r"/api/crm/conversations/(\d+)/messages", parsed.path)
        if crm_messages_match:
            return self.get_crm_messages(int(crm_messages_match.group(1)), parse_qs(parsed.query))
        crm_timeline_match = re.fullmatch(r"/api/crm/conversations/(\d+)/timeline", parsed.path)
        if crm_timeline_match:
            return self.get_crm_conversation_timeline(int(crm_timeline_match.group(1)))
        crm_message_media_match = re.fullmatch(r"/api/crm/messages/(\d+)/media", parsed.path)
        if crm_message_media_match:
            return self.get_crm_message_media(int(crm_message_media_match.group(1)))
        crm_media_match = re.fullmatch(r"/api/crm/media/([a-f0-9]{32}\.(?:webm|ogg|oga|mp4|m4a|jpg|jpeg|png|webp|pdf|doc|docx|xls|xlsx|ppt|pptx|txt|zip|bin))", parsed.path)
        if crm_media_match:
            return self.get_crm_media(crm_media_match.group(1))
        if parsed.path == "/api/crm/webhook-events":
            return self.get_crm_webhook_events()
        if parsed.path == "/api/crm/automation-events":
            return self.get_crm_automation_events()
        if parsed.path == "/api/procedure-catalog":
            return self.get_procedure_catalog()
        if parsed.path == "/api/action-templates":
            return self.get_action_templates()
        if parsed.path == "/api/relationship-directory":
            return self.get_relationship_directory(parse_qs(parsed.query))
        if parsed.path == "/api/daily-log":
            return self.get_daily_log(parse_qs(parsed.query))
        if parsed.path == "/api/journey/professionals":
            return self.get_journey_professionals()
        if parsed.path == "/api/journeys":
            return self.get_journeys()
        if parsed.path == "/api/patients":
            return self.get_patients(parse_qs(parsed.query))
        match = re.fullmatch(r"/api/patients/(\d+)", parsed.path)
        if match:
            return self.get_patient(int(match.group(1)))
        if parsed.path in {"", "/", ADMIN_ROUTE, f"{ADMIN_ROUTE}/", CRC_ROUTE, f"{CRC_ROUTE}/"} or parsed.path.startswith(f"{CRC_ROUTE}/"):
            user = self.current_user()
            if not user:
                self.send_response(HTTPStatus.FOUND)
                self.send_header("Location", "/login")
                self.end_headers()
                return
            if not user.get("two_factor_enabled") and not user.get("two_factor_exempt"):
                self.send_response(HTTPStatus.FOUND)
                self.send_header("Location", "/two-factor-setup")
                self.end_headers()
                return
            if parsed.path.startswith(ADMIN_ROUTE) and not self.can_admin_portal(user):
                self.send_response(HTTPStatus.FOUND)
                self.send_header("Location", "/")
                self.end_headers()
                return
            if parsed.path.startswith(CRC_ROUTE) and user["access_role"] != "crc":
                self.send_response(HTTPStatus.FOUND)
                self.send_header("Location", "/")
                self.end_headers()
                return
        return self.serve_static(parsed.path)

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/api/integrations/evolution/webhook":
            return self.receive_evolution_webhook(self.read_json(), parse_qs(parsed.query))
        if parsed.path == "/api/integrations/crm/handoff":
            return self.receive_crm_handoff(self.read_json(), parse_qs(parsed.query))
        if parsed.path == "/api/integrations/crm/automation-event":
            return self.receive_crm_automation_event(self.read_json(), parse_qs(parsed.query))
        if parsed.path == "/api/auth/login":
            return self.login(self.read_json())
        if parsed.path == "/api/auth/2fa/verify":
            return self.verify_two_factor_login(self.read_json())
        if parsed.path == "/api/auth/2fa/setup":
            user = self.require_auth()
            return self.begin_two_factor_setup(user) if user else None
        if parsed.path == "/api/auth/2fa/confirm":
            user = self.require_auth()
            return self.confirm_two_factor_setup(user, self.read_json()) if user else None
        if parsed.path == "/api/auth/setup":
            return self.setup_owner(self.read_json())
        if parsed.path == "/api/auth/request-password-reset":
            return self.request_password_reset(self.read_json())
        if parsed.path == "/api/auth/logout":
            return self.logout()
        if parsed.path == "/api/auth/change-password":
            user = self.require_auth()
            return self.change_password(user, self.read_json()) if user else None
        if parsed.path == "/api/integrations/crc-export/claim":
            return self.claim_crc_exports(self.read_json())
        if parsed.path == "/api/integrations/crc-export/ack":
            return self.ack_crc_export(self.read_json())
        if parsed.path == "/api/integrations/crm/ai-message":
            return self.mark_crm_ai_message(self.read_json(), parse_qs(parsed.query))
        if parsed.path.startswith("/api/"):
            self.authenticated_user = self.require_auth()
            if not self.authenticated_user:
                return
        crm_admin_endpoint = parsed.path in {
            "/api/admin/crm-channel-access", "/api/admin/crm-channel-access/test"
        }
        if parsed.path.startswith("/api/admin/") and not (
            self.can_admin_portal(self.authenticated_user)
            or (crm_admin_endpoint and self.can_manage_crm(self.authenticated_user))
        ):
            return self.send_json({"error": "Acesso administrativo necessário"}, HTTPStatus.FORBIDDEN)
        if parsed.path == "/api/admin/crm-channel-access":
            return self.save_admin_crm_channel_access(self.read_json())
        if parsed.path == "/api/admin/crm-channel-access/test":
            return self.test_admin_crm_channel_access(self.read_json())
        if parsed.path == "/api/crm/channels":
            return self.save_crm_channel(self.read_json())
        if parsed.path == "/api/crm/tags":
            return self.create_crm_tag(self.read_json())
        if parsed.path == "/api/crm/quick-replies":
            return self.create_crm_quick_reply(self.read_json())
        if parsed.path == "/api/crm/goals":
            return self.save_crm_goals(self.read_json())
        if parsed.path == "/api/crm/evolution/connect":
            return self.connect_evolution_instance(self.read_json())
        if parsed.path == "/api/crm/evolution/config":
            return self.save_evolution_config(self.read_json())
        if parsed.path == "/api/crm/evolution/sync":
            return self.start_evolution_history_sync()
        if parsed.path == "/api/crm/contacts/cleanup":
            return self.cleanup_crm_imported_contacts()
        if parsed.path == "/api/crm/conversations":
            return self.start_crm_conversation(self.read_json())
        crm_send_match = re.fullmatch(r"/api/crm/conversations/(\d+)/messages", parsed.path)
        if crm_send_match:
            return self.send_crm_message(int(crm_send_match.group(1)), self.read_json())
        crm_resolve_match = re.fullmatch(r"/api/crm/conversations/(\d+)/resolve", parsed.path)
        if crm_resolve_match:
            return self.resolve_crm_conversation(int(crm_resolve_match.group(1)), self.read_json())
        crm_claim_match = re.fullmatch(r"/api/crm/conversations/(\d+)/claim", parsed.path)
        if crm_claim_match:
            return self.claim_crm_conversation(int(crm_claim_match.group(1)))
        if parsed.path == "/api/crm/n8n/config":
            return self.save_crm_n8n_config(self.read_json())
        if parsed.path == "/api/crm/n8n/callback-keys":
            return self.create_crm_n8n_callback_key(self.read_json())
        n8n_callback_revoke_match = re.fullmatch(r"/api/crm/n8n/callback-keys/(\d+)/revoke", parsed.path)
        if n8n_callback_revoke_match:
            return self.revoke_crm_n8n_callback_key(int(n8n_callback_revoke_match.group(1)))
        n8n_workflow_settings_match = re.fullmatch(r"/api/crm/n8n/workflows/([^/]+)/settings", parsed.path)
        if n8n_workflow_settings_match:
            return self.save_crm_n8n_workflow_settings(
                n8n_workflow_settings_match.group(1),
                self.read_json(),
            )
        n8n_workflow_run_match = re.fullmatch(r"/api/crm/n8n/workflows/([^/]+)/run", parsed.path)
        if n8n_workflow_run_match:
            return self.run_crm_n8n_workflow(
                n8n_workflow_run_match.group(1),
                self.read_json(),
            )
        n8n_workflow_action_match = re.fullmatch(r"/api/crm/n8n/workflows/([^/]+)/(activate|deactivate)", parsed.path)
        if n8n_workflow_action_match:
            return self.change_crm_n8n_workflow(
                n8n_workflow_action_match.group(1),
                n8n_workflow_action_match.group(2),
            )
        n8n_workflow_restore_match = re.fullmatch(
            r"/api/crm/n8n/workflows/([^/]+)/versions/(\d+)/restore",
            parsed.path,
        )
        if n8n_workflow_restore_match:
            return self.restore_crm_n8n_workflow_version(
                n8n_workflow_restore_match.group(1),
                int(n8n_workflow_restore_match.group(2)),
            )
        if self.authenticated_user["access_role"] == "crc":
            return self.send_json({"error": "A Central CRC é somente para acompanhamento."}, HTTPStatus.FORBIDDEN)
        if parsed.path == "/api/patients":
            return self.create_patient(self.read_json())
        if parsed.path == "/api/procedure-catalog":
            if not self.can_admin_portal(self.authenticated_user):
                return self.send_json({"error": "Acesso administrativo necessário"}, HTTPStatus.FORBIDDEN)
            return self.create_catalog_procedure(self.read_json())
        if parsed.path == "/api/admin/professionals":
            if not self.can_admin_portal(self.authenticated_user):
                return self.send_json({"error": "Acesso administrativo necessário"}, HTTPStatus.FORBIDDEN)
            return self.create_admin_professional(self.read_json())
        if parsed.path == "/api/admin/integrations/clinicorp":
            if not self.can_admin_portal(self.authenticated_user):
                return self.send_json({"error": "Acesso administrativo necessário"}, HTTPStatus.FORBIDDEN)
            return self.save_clinicorp_config(self.read_json())
        if parsed.path == "/api/admin/apis":
            if not self.can_admin_portal(self.authenticated_user):
                return self.send_json({"error": "Acesso administrativo necessário"}, HTTPStatus.FORBIDDEN)
            return self.save_api_integration(self.read_json())
        api_sync_match = re.fullmatch(r"/api/admin/apis/(\d+)/sync", parsed.path)
        if api_sync_match:
            if not self.can_admin_portal(self.authenticated_user):
                return self.send_json({"error": "Acesso administrativo necessário"}, HTTPStatus.FORBIDDEN)
            return self.start_api_phone_sync(int(api_sync_match.group(1)), self.read_json())
        api_retry_failures_match = re.fullmatch(r"/api/admin/apis/(\d+)/retry-failures", parsed.path)
        if api_retry_failures_match:
            if not self.can_admin_portal(self.authenticated_user):
                return self.send_json({"error": "Acesso administrativo necessário"}, HTTPStatus.FORBIDDEN)
            return self.retry_api_sync_failures(int(api_retry_failures_match.group(1)))
        api_revert_match = re.fullmatch(r"/api/admin/apis/(\d+)/revert", parsed.path)
        if api_revert_match:
            if not self.can_admin_portal(self.authenticated_user):
                return self.send_json({"error": "Acesso administrativo necessário"}, HTTPStatus.FORBIDDEN)
            return self.revert_api_integration(int(api_revert_match.group(1)))
        photo_match = re.fullmatch(r"/api/admin/professionals/(\d+)/photo", parsed.path)
        if photo_match:
            if not self.can_admin_portal(self.authenticated_user):
                return self.send_json({"error": "Acesso administrativo necessário"}, HTTPStatus.FORBIDDEN)
            return self.update_professional_photo(int(photo_match.group(1)), self.read_json())
        password_match = re.fullmatch(r"/api/admin/professionals/(\d+)/reset-password", parsed.path)
        if password_match:
            if not self.can_admin_portal(self.authenticated_user):
                return self.send_json({"error": "Acesso administrativo necessário"}, HTTPStatus.FORBIDDEN)
            return self.reset_professional_password(int(password_match.group(1)), self.read_json())
        if parsed.path == "/api/admin/specialties":
            return self.create_admin_specialty(self.read_json())
        if parsed.path == "/api/admin/offices":
            return self.create_admin_office(self.read_json())
        if parsed.path == "/api/admin/import/patients":
            return self.import_admin_patients(self.read_json())
        if parsed.path == "/api/admin/import/appointments/audit":
            if not self.can_admin_portal(self.authenticated_user):
                return self.send_json({"error": "Acesso administrativo necessário"}, HTTPStatus.FORBIDDEN)
            return self.appointment_audit(self.read_json())
        if parsed.path == "/api/admin/import/appointments/validate":
            if not self.can_admin_portal(self.authenticated_user):
                return self.send_json({"error": "Acesso administrativo necessário"}, HTTPStatus.FORBIDDEN)
            return self.appointment_validate(self.read_json())
        if parsed.path == "/api/admin/import/appointments/confirm":
            if not self.can_admin_portal(self.authenticated_user):
                return self.send_json({"error": "Acesso administrativo necessário"}, HTTPStatus.FORBIDDEN)
            return self.appointment_confirm(self.read_json())
        if parsed.path == "/api/admin/import/relationship-directory/audit":
            if not self.can_admin_portal(self.authenticated_user):
                return self.send_json({"error": "Acesso administrativo necessário"}, HTTPStatus.FORBIDDEN)
            return self.relationship_directory_audit(self.read_json())
        if parsed.path == "/api/admin/import/relationship-directory/confirm":
            if not self.can_admin_portal(self.authenticated_user):
                return self.send_json({"error": "Acesso administrativo necessário"}, HTTPStatus.FORBIDDEN)
            return self.relationship_directory_confirm(self.read_json())
        if parsed.path == "/api/admin/import/scheduled/audit":
            if not self.can_admin_portal(self.authenticated_user):
                return self.send_json({"error": "Acesso administrativo necessário"}, HTTPStatus.FORBIDDEN)
            return self.scheduled_appointment_audit(self.read_json())
        if parsed.path == "/api/admin/import/scheduled/validate":
            if not self.can_admin_portal(self.authenticated_user):
                return self.send_json({"error": "Acesso administrativo necessário"}, HTTPStatus.FORBIDDEN)
            return self.scheduled_appointment_validate(self.read_json())
        if parsed.path == "/api/admin/import/scheduled/confirm":
            if not self.can_admin_portal(self.authenticated_user):
                return self.send_json({"error": "Acesso administrativo necessário"}, HTTPStatus.FORBIDDEN)
            return self.scheduled_appointment_confirm(self.read_json())
        reset_match = re.fullmatch(r"/api/admin/password-resets/(\d+)/complete", parsed.path)
        if reset_match:
            return self.complete_password_reset(int(reset_match.group(1)), self.read_json())
        observation_match = re.fullmatch(r"/api/patients/(\d+)/observations", parsed.path)
        if observation_match:
            return self.add_visit_observation(int(observation_match.group(1)), self.read_json())
        edit_lock_match = re.fullmatch(r"/api/patients/(\d+)/edit-lock", parsed.path)
        if edit_lock_match:
            return self.acquire_patient_edit_lock(int(edit_lock_match.group(1)))
        forward_match = re.fullmatch(r"/api/patients/(\d+)/forward", parsed.path)
        if forward_match:
            return self.forward_patient(int(forward_match.group(1)), self.read_json())
        journey_stage_match = re.fullmatch(r"/api/patients/(\d+)/journey-stage", parsed.path)
        if journey_stage_match:
            return self.save_journey_stage(int(journey_stage_match.group(1)), self.read_json())
        reopen_match = re.fullmatch(r"/api/patients/(\d+)/reopen-with-password", parsed.path)
        if reopen_match:
            return self.reopen_patient_with_password(int(reopen_match.group(1)), self.read_json())
        match = re.fullmatch(r"/api/patients/(\d+)/resolve", parsed.path)
        if match:
            return self.resolve_patient(int(match.group(1)), self.read_json())
        return self.send_json({"error": "Rota não encontrada"}, HTTPStatus.NOT_FOUND)

    def do_DELETE(self) -> None:
        parsed = urlparse(self.path)
        self.authenticated_user = self.require_auth()
        if not self.authenticated_user:
            return
        crm_quick_reply_match = re.fullmatch(r"/api/crm/quick-replies/(\d+)", parsed.path)
        if crm_quick_reply_match:
            return self.delete_crm_quick_reply(int(crm_quick_reply_match.group(1)))
        edit_lock_match = re.fullmatch(r"/api/patients/(\d+)/edit-lock", parsed.path)
        if edit_lock_match:
            return self.release_patient_edit_lock(int(edit_lock_match.group(1)))
        if self.authenticated_user["access_role"] == "crc":
            return self.send_json({"error": "A Central CRC é somente para acompanhamento."}, HTTPStatus.FORBIDDEN)
        if parsed.path.startswith("/api/admin/") and not self.can_admin_portal(self.authenticated_user):
            return self.send_json({"error": "Acesso administrativo necessário"}, HTTPStatus.FORBIDDEN)
        if parsed.path.startswith("/api/procedure-catalog/") and not self.can_admin_portal(self.authenticated_user):
            return self.send_json({"error": "Acesso administrativo necessário"}, HTTPStatus.FORBIDDEN)
        observation_match = re.fullmatch(r"/api/patients/(\d+)/observations/(\d+)", parsed.path)
        if observation_match:
            return self.delete_visit_observation(int(observation_match.group(1)), int(observation_match.group(2)))
        journey_match = re.fullmatch(r"/api/patients/(\d+)/journey/(\d+)", parsed.path)
        if journey_match:
            return self.delete_journey_assignment(int(journey_match.group(1)), int(journey_match.group(2)))
        patient_match = re.fullmatch(r"/api/patients/(\d+)", parsed.path)
        if patient_match:
            return self.delete_patient(int(patient_match.group(1)), self.read_json())
        specialty_match = re.fullmatch(r"/api/admin/specialties/(\d+)", parsed.path)
        if specialty_match:
            return self.delete_admin_structure("specialties", "professional_specialties", "specialty_id", int(specialty_match.group(1)))
        office_match = re.fullmatch(r"/api/admin/offices/(\d+)", parsed.path)
        if office_match:
            return self.delete_admin_structure("offices", "professional_offices", "office_id", int(office_match.group(1)))
        api_match = re.fullmatch(r"/api/admin/apis/(\d+)", parsed.path)
        if api_match:
            return self.delete_api_integration(int(api_match.group(1)))
        match = re.fullmatch(r"/api/procedure-catalog/(\d+)", parsed.path)
        if not match:
            return self.send_json({"error": "Rota não encontrada"}, HTTPStatus.NOT_FOUND)
        procedure_id = int(match.group(1))
        with connect() as db:
            exists = db.execute("SELECT id FROM procedure_catalog WHERE id = ?", (procedure_id,)).fetchone()
            if not exists:
                return self.send_json({"error": "Procedimento não encontrado"}, HTTPStatus.NOT_FOUND)
            db.execute("DELETE FROM procedure_catalog WHERE id = ?", (procedure_id,))
        self.send_json({"deleted": True, "id": procedure_id})

    def delete_api_integration(self, integration_id: int) -> None:
        with connect() as db:
            exists = db.execute("SELECT id FROM api_integrations WHERE id=?", (integration_id,)).fetchone()
            if not exists:
                return self.send_json({"error": "API não encontrada"}, HTTPStatus.NOT_FOUND)
            db.execute("DELETE FROM api_integrations WHERE id=?", (integration_id,))
        self.send_json({"deleted": True, "id": integration_id})

    def delete_patient(self, patient_id: int, payload: dict) -> None:
        if self.authenticated_user["access_role"] not in {"owner", "professional", "asb", "admin"}:
            return self.send_json({"error": "Seu nível de acesso não permite excluir pacientes."}, HTTPStatus.FORBIDDEN)
        password = str(payload.get("password") or "")
        if not password:
            return self.send_json({"error": "Confirme sua senha para excluir o paciente."}, HTTPStatus.BAD_REQUEST)
        with connect() as db:
            if self.patient_edit_lock_conflict(db, patient_id):
                return
            user = db.execute(
                "SELECT password_hash, password_salt FROM users WHERE id=? AND active=1",
                (self.authenticated_user["id"],),
            ).fetchone()
            password_valid = (
                user
                and user["password_hash"]
                and user["password_salt"]
                and hmac.compare_digest(
                    self.password_digest(password, user["password_salt"]),
                    user["password_hash"],
                )
            )
            if not password_valid:
                self.record_security_event(
                    db, "patient_delete_denied", self.request_ip(),
                    self.authenticated_user["id"], f"Paciente {patient_id}: senha inválida",
                )
                return self.send_json({"error": "Senha inválida."}, HTTPStatus.UNAUTHORIZED)

            patient = db.execute("""
                SELECT p.id, p.name,
                       (SELECT pr.name
                          FROM patient_assignments pa
                          JOIN professionals pr ON pr.id=pa.professional_id
                         WHERE pa.patient_id=p.id AND pa.is_primary=1
                         ORDER BY pa.professional_id LIMIT 1) AS professional_name,
                       CASE WHEN date(f.resolved_at) = date('now', 'localtime') THEN 1 ELSE 0 END AS locked
                FROM patients p
                JOIN patient_followup f ON f.patient_id = p.id
                WHERE p.id = ?
            """, (patient_id,)).fetchone()
            if not patient:
                return self.send_json({"error": "Paciente não encontrado"}, HTTPStatus.NOT_FOUND)

            if self.authenticated_user["access_role"] != "admin":
                portfolio_id = (
                    self.authenticated_user.get("linked_professional_id")
                    if self.authenticated_user["access_role"] == "asb"
                    else self.authenticated_user["professional_id"]
                )
                owns_patient = db.execute(
                    """SELECT 1 FROM patient_assignments
                       WHERE patient_id=? AND professional_id=? AND is_primary=1""",
                    (patient_id, portfolio_id),
                ).fetchone()
                if not owns_patient:
                    self.record_security_event(
                        db, "patient_delete_denied", self.request_ip(),
                        self.authenticated_user["id"], f"Paciente {patient_id}: fora da carteira principal",
                    )
                    return self.send_json(
                        {"error": "Somente a carteira principal deste paciente pode excluí-lo."},
                        HTTPStatus.FORBIDDEN,
                    )

            if patient["locked"]:
                return self.send_json(
                    {"error": "Paciente verificado está bloqueado. Reabra o acompanhamento antes de excluir."},
                    HTTPStatus.CONFLICT,
                )

            db.execute(
                """INSERT INTO patient_deletion_audit
                   (patient_id, patient_name, professional_name, deleted_by_user_id,
                    deleted_by_name, deleted_by_role, ip_address, deleted_at)
                   VALUES (?, ?, ?, ?, ?, ?, ?, datetime('now','localtime'))""",
                (
                    patient_id, patient["name"], patient["professional_name"],
                    self.authenticated_user["id"], self.authenticated_user["name"],
                    self.authenticated_user["access_role"], self.request_ip(),
                ),
            )
            self.record_security_event(
                db, "patient_deleted", self.request_ip(), self.authenticated_user["id"],
                f"Paciente {patient_id} - {patient['name']} excluído por {self.authenticated_user['name']}",
            )
            db.execute("DELETE FROM daily_resolutions WHERE patient_id = ?", (patient_id,))
            db.execute("DELETE FROM patient_events WHERE patient_id = ?", (patient_id,))
            db.execute("DELETE FROM procedures WHERE patient_id = ?", (patient_id,))
            db.execute("DELETE FROM patient_relationships WHERE patient_id = ?", (patient_id,))
            db.execute("DELETE FROM patient_clinical_profile WHERE patient_id = ?", (patient_id,))
            db.execute("DELETE FROM patient_followup WHERE patient_id = ?", (patient_id,))
            db.execute("DELETE FROM patient_assignments WHERE patient_id = ?", (patient_id,))
            db.execute("DELETE FROM patients WHERE id = ?", (patient_id,))
        self.send_json({
            "deleted": True,
            "id": patient_id,
            "patient_name": patient["name"],
            "deleted_by": self.authenticated_user["name"],
        })

    def do_PATCH(self) -> None:
        parsed = urlparse(self.path)
        self.authenticated_user = self.require_auth()
        if not self.authenticated_user:
            return
        crm_conversation_match = re.fullmatch(r"/api/crm/conversations/(\d+)", parsed.path)
        if crm_conversation_match:
            return self.update_crm_conversation(int(crm_conversation_match.group(1)), self.read_json())
        crm_contact_match = re.fullmatch(r"/api/crm/contacts/(\d+)", parsed.path)
        if crm_contact_match:
            return self.update_crm_contact(int(crm_contact_match.group(1)), self.read_json())
        crm_channel_match = re.fullmatch(r"/api/crm/channels/(\d+)", parsed.path)
        if crm_channel_match:
            return self.update_crm_channel(int(crm_channel_match.group(1)), self.read_json())
        crm_quick_reply_match = re.fullmatch(r"/api/crm/quick-replies/(\d+)", parsed.path)
        if crm_quick_reply_match:
            return self.update_crm_quick_reply(int(crm_quick_reply_match.group(1)), self.read_json())
        crc_patient_match = re.fullmatch(r"/api/patients/(\d+)/crc", parsed.path)
        if crc_patient_match:
            return self.update_crc_patient(int(crc_patient_match.group(1)), self.read_json())
        observation_match = re.fullmatch(r"/api/patients/(\d+)/observations/(\d+)", parsed.path)
        if observation_match:
            return self.update_visit_observation(int(observation_match.group(1)), int(observation_match.group(2)), self.read_json())
        journey_match = re.fullmatch(r"/api/patients/(\d+)/journey/(\d+)", parsed.path)
        if journey_match:
            return self.update_journey_assignment(int(journey_match.group(1)), int(journey_match.group(2)), self.read_json())
        if self.authenticated_user["access_role"] == "crc":
            return self.send_json({"error": "A CRC pode atualizar apenas procedimentos e contato/referência."}, HTTPStatus.FORBIDDEN)
        if (parsed.path.startswith("/api/admin/") or parsed.path.startswith("/api/procedure-catalog/")) and not self.can_admin_portal(self.authenticated_user):
            return self.send_json({"error": "Acesso administrativo necessário"}, HTTPStatus.FORBIDDEN)
        if parsed.path == "/api/admin/integrations/clinicorp":
            return self.save_clinicorp_config(self.read_json())
        catalog_match = re.fullmatch(r"/api/procedure-catalog/(\d+)", parsed.path)
        if catalog_match:
            return self.update_catalog_procedure(int(catalog_match.group(1)), self.read_json())
        admin_professional_match = re.fullmatch(r"/api/admin/professionals/(\d+)", parsed.path)
        if admin_professional_match:
            return self.update_admin_professional(int(admin_professional_match.group(1)), self.read_json())
        specialty_match = re.fullmatch(r"/api/admin/specialties/(\d+)", parsed.path)
        if specialty_match:
            return self.update_admin_structure("specialties", int(specialty_match.group(1)), self.read_json())
        office_match = re.fullmatch(r"/api/admin/offices/(\d+)", parsed.path)
        if office_match:
            return self.update_admin_structure("offices", int(office_match.group(1)), self.read_json())
        match = re.fullmatch(r"/api/patients/(\d+)", parsed.path)
        if not match:
            return self.send_json({"error": "Rota não encontrada"}, HTTPStatus.NOT_FOUND)
        return self.update_patient(int(match.group(1)), self.read_json())

    def login(self, payload: dict) -> None:
        email = str(payload.get("email") or "").strip().lower()
        password = str(payload.get("password") or "")
        ip_address = self.request_ip()
        if not email or not password:
            return self.send_json({"error": "Informe e-mail e senha"}, HTTPStatus.BAD_REQUEST)
        with connect() as db:
            db.execute("DELETE FROM login_attempts WHERE attempted_at < datetime('now', '-24 hours')")
            attempts = db.execute("SELECT COUNT(*) FROM login_attempts WHERE email=? AND ip_address=? AND successful=0 AND attempted_at > datetime('now', '-15 minutes')", (email, ip_address)).fetchone()[0]
            if attempts >= 5:
                self.record_security_event(db, "login_blocked", ip_address, detail="Muitas tentativas para a mesma conta")
                return self.send_json({"error": "Muitas tentativas. Aguarde 15 minutos antes de tentar novamente."}, HTTPStatus.TOO_MANY_REQUESTS)
            row = db.execute("SELECT * FROM users WHERE lower(email)=? AND active=1", (email,)).fetchone()
            if not row or not row["password_hash"] or not row["password_salt"]:
                # Computa um digest "descartável" de custo equivalente ao caminho de sucesso,
                # para não vazar a existência da conta por diferença de latência (timing side-channel).
                self.password_digest(password, secrets.token_hex(16))
                db.execute("INSERT INTO login_attempts (email, ip_address) VALUES (?, ?)", (email, ip_address))
                self.record_security_event(db, "login_failed", ip_address, detail="Credenciais inválidas")
                return self.send_json({"error": "E-mail ou senha inválidos"}, HTTPStatus.UNAUTHORIZED)
            digest = self.password_digest(password, row["password_salt"])
            if not hmac.compare_digest(digest, row["password_hash"]):
                db.execute("INSERT INTO login_attempts (email, ip_address) VALUES (?, ?)", (email, ip_address))
                self.record_security_event(db, "login_failed", ip_address, row["id"], "Senha inválida")
                return self.send_json({"error": "E-mail ou senha inválidos"}, HTTPStatus.UNAUTHORIZED)
            db.execute("INSERT INTO login_attempts (email, ip_address, successful) VALUES (?, ?, 1)", (email, ip_address))
            db.execute("DELETE FROM login_attempts WHERE email=? AND ip_address=? AND successful=0", (email, ip_address))
            if row["two_factor_enabled"]:
                challenge = secrets.token_urlsafe(32)
                db.execute("DELETE FROM login_challenges WHERE expires_at <= datetime('now') OR user_id=?", (row["id"],))
                db.execute("INSERT INTO login_challenges (user_id, token_hash, expires_at) VALUES (?, ?, ?)", (row["id"], hashlib.sha256(challenge.encode("utf-8")).hexdigest(), (datetime.utcnow() + timedelta(minutes=5)).strftime("%Y-%m-%d %H:%M:%S")))
                self.record_security_event(db, "login_password_verified", ip_address, row["id"])
                return self.send_json({"two_factor_required": True, "challenge": challenge, "email": row["email"]})
            token = self.issue_session(db, row["id"])
            db.execute("UPDATE users SET last_login_at=CURRENT_TIMESTAMP WHERE id=?", (row["id"],))
            if row["two_factor_exempt"]:
                self.record_security_event(db, "login_success_2fa_temporarily_exempt", ip_address, row["id"])
                return self.send_json_with_cookie(self.auth_payload(dict(row)), self.session_cookie(token))
            self.record_security_event(db, "login_pending_2fa_enrollment", ip_address, row["id"])
        self.send_json_with_cookie({"two_factor_setup_required": True}, self.session_cookie(token))

    def begin_two_factor_setup(self, user: dict) -> None:
        if user.get("two_factor_enabled"):
            return self.send_json({"error": "A autenticação em duas etapas já está ativa."}, HTTPStatus.CONFLICT)
        secret = self.new_totp_secret()
        try:
            encrypted_secret = self.encrypt_totp_secret(secret)
        except RuntimeError:
            return self.send_json(
                {"error": "A proteção do 2FA ainda não está configurada. Contate o administrador."},
                HTTPStatus.SERVICE_UNAVAILABLE,
            )
        with connect() as db:
            db.execute("UPDATE users SET two_factor_secret=? WHERE id=?", (encrypted_secret, user["id"]))
        self.send_json({"secret": secret, "otpauth_uri": self.otp_uri(user["email"], secret), "issuer": "Instituto Eduardo Ayub"})

    def confirm_two_factor_setup(self, user: dict, payload: dict) -> None:
        code = str(payload.get("code") or "").replace(" ", "")
        with connect() as db:
            row = db.execute("SELECT two_factor_secret FROM users WHERE id=?", (user["id"],)).fetchone()
            if not row or not self.valid_totp(self.decrypt_totp_secret(row["two_factor_secret"]), code):
                return self.send_json({"error": "Código inválido. Confira o aplicativo autenticador e tente novamente."}, HTTPStatus.UNAUTHORIZED)
            db.execute("UPDATE users SET two_factor_enabled=1, two_factor_enrolled_at=CURRENT_TIMESTAMP WHERE id=?", (user["id"],))
            self.record_security_event(db, "two_factor_enabled", self.request_ip(), user["id"])
        self.send_json({"enabled": True})

    def verify_two_factor_login(self, payload: dict) -> None:
        challenge = str(payload.get("challenge") or "")
        code = str(payload.get("code") or "").replace(" ", "")
        if not challenge or not re.fullmatch(r"\d{6}", code):
            return self.send_json({"error": "Informe o código de 6 dígitos."}, HTTPStatus.BAD_REQUEST)
        with connect() as db:
            row = db.execute("""SELECT u.*,c.expires_at AS challenge_expires_at, c.attempts AS challenge_attempts
                              FROM login_challenges c JOIN users u ON u.id=c.user_id
                              WHERE c.token_hash=? AND u.active=1""",
                             (hashlib.sha256(challenge.encode("utf-8")).hexdigest(),)).fetchone()
            if not row:
                return self.send_json({"error": "Sessão de verificação inválida. Faça login novamente."}, HTTPStatus.UNAUTHORIZED)
            if row["challenge_expires_at"] <= datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"):
                db.execute("DELETE FROM login_challenges WHERE user_id=?", (row["id"],))
                self.record_security_event(db, "two_factor_challenge_expired", self.request_ip(), row["id"])
                return self.send_json({"error": "A sessão de verificação expirou. Faça login novamente."}, HTTPStatus.UNAUTHORIZED)
            if row["challenge_attempts"] >= 5:
                db.execute("DELETE FROM login_challenges WHERE user_id=?", (row["id"],))
                self.record_security_event(db, "two_factor_blocked", self.request_ip(), row["id"])
                return self.send_json({"error": "Muitas tentativas. Faça login novamente."}, HTTPStatus.TOO_MANY_REQUESTS)
            if not self.valid_totp(self.decrypt_totp_secret(row["two_factor_secret"]), code):
                db.execute("UPDATE login_challenges SET attempts=attempts+1 WHERE user_id=?", (row["id"],))
                self.record_security_event(db, "two_factor_failed", self.request_ip(), row["id"])
                return self.send_json({"error": "Código inválido. Use o código atual exibido no autenticador."}, HTTPStatus.UNAUTHORIZED)
            token = self.issue_session(db, row["id"])
            db.execute("DELETE FROM login_challenges WHERE user_id=?", (row["id"],))
            db.execute("UPDATE users SET last_login_at=CURRENT_TIMESTAMP WHERE id=?", (row["id"],))
            self.record_security_event(db, "login_success", self.request_ip(), row["id"])
        self.send_json_with_cookie(self.auth_payload(dict(row)), self.session_cookie(token))

    def setup_owner(self, payload: dict) -> None:
        setup_token = os.environ.get("AUTH_SETUP_TOKEN", "")
        if not setup_token:
            return self.send_json({"error": "Configuração de ativação ausente. Contate o administrador."}, HTTPStatus.SERVICE_UNAVAILABLE)
        if not hmac.compare_digest(str(payload.get("setup_token") or ""), setup_token):
            return self.send_json({"error": "Código de ativação inválido"}, HTTPStatus.FORBIDDEN)
        password = str(payload.get("password") or "")
        if len(password) < 10:
            return self.send_json({"error": "A senha precisa ter ao menos 10 caracteres"}, HTTPStatus.BAD_REQUEST)
        with connect() as db:
            owner = db.execute("SELECT * FROM users WHERE access_role='owner' AND active=1 ORDER BY id LIMIT 1").fetchone()
            if not owner:
                return self.send_json({"error": "Nenhuma proprietária foi configurada"}, HTTPStatus.CONFLICT)
            if owner["password_hash"]:
                return self.send_json({"error": "A conta proprietária já foi ativada"}, HTTPStatus.CONFLICT)
            salt = secrets.token_hex(16)
            db.execute("UPDATE users SET password_hash=?, password_salt=?, must_change_password=0 WHERE id=?", (self.password_digest(password, salt), salt, owner["id"]))
        self.send_json({"setup": True, "email": owner["email"]})

    def logout(self) -> None:
        cookies = SimpleCookie(); cookies.load(self.headers.get("Cookie", ""))
        token = cookies.get("iea_session")
        if token:
            with connect() as db:
                db.execute("DELETE FROM auth_sessions WHERE token_hash=?", (hashlib.sha256(token.value.encode("utf-8")).hexdigest(),))
        self.send_json_with_cookie({"authenticated": False}, "iea_session=; HttpOnly; SameSite=Strict; Path=/; Max-Age=0")

    def change_password(self, user: dict, payload: dict) -> None:
        password = str(payload.get("password") or "")
        if len(password) < 10:
            return self.send_json({"error": "A senha precisa ter ao menos 10 caracteres"}, HTTPStatus.BAD_REQUEST)
        salt = secrets.token_hex(16)
        cookies = SimpleCookie()
        cookies.load(self.headers.get("Cookie", ""))
        current_session = cookies.get("iea_session")
        current_token_hash = (
            hashlib.sha256(current_session.value.encode("utf-8")).hexdigest()
            if current_session else ""
        )
        with connect() as db:
            db.execute("UPDATE users SET password_hash=?, password_salt=?, must_change_password=0 WHERE id=?", (self.password_digest(password, salt), salt, user["id"]))
            if current_token_hash:
                db.execute(
                    "DELETE FROM auth_sessions WHERE user_id=? AND token_hash<>?",
                    (user["id"], current_token_hash),
                )
            else:
                db.execute("DELETE FROM auth_sessions WHERE user_id=?", (user["id"],))
            self.record_security_event(db, "password_changed", self.request_ip(), user["id"], "Demais sessões revogadas")
        self.send_json({"changed": True})

    def request_password_reset(self, payload: dict) -> None:
        email = str(payload.get("email") or "").strip().lower()
        if not email:
            return self.send_json({"error": "Informe seu e-mail"}, HTTPStatus.BAD_REQUEST)
        with connect() as db:
            user = db.execute("SELECT id FROM users WHERE lower(email)=? AND active=1", (email,)).fetchone()
            if user:
                existing = db.execute("SELECT id FROM password_reset_requests WHERE user_id=? AND status='pending'", (user["id"],)).fetchone()
                if not existing:
                    db.execute("INSERT INTO password_reset_requests (user_id) VALUES (?)", (user["id"],))
        self.send_json({"requested": True, "message": "Se o e-mail estiver cadastrado, a solicitação foi enviada para a central."})

    def complete_password_reset(self, request_id: int, payload: dict) -> None:
        if not self.can_admin_portal(self.authenticated_user):
            return self.send_json({"error": "Acesso administrativo necessário"}, HTTPStatus.FORBIDDEN)
        temporary_password = str(payload.get("temporary_password") or "")
        if len(temporary_password) < 10:
            return self.send_json({"error": "Defina uma senha temporária com ao menos 10 caracteres"}, HTTPStatus.BAD_REQUEST)
        with connect() as db:
            request = db.execute("SELECT user_id FROM password_reset_requests WHERE id=? AND status='pending'", (request_id,)).fetchone()
            if not request:
                return self.send_json({"error": "Solicitação não encontrada ou já atendida"}, HTTPStatus.NOT_FOUND)
            salt = secrets.token_hex(16)
            db.execute("UPDATE users SET password_hash=?, password_salt=?, must_change_password=1 WHERE id=?", (self.password_digest(temporary_password, salt), salt, request["user_id"]))
            db.execute("DELETE FROM auth_sessions WHERE user_id=?", (request["user_id"],))
            db.execute("UPDATE password_reset_requests SET status='completed', completed_at=CURRENT_TIMESTAMP, completed_by_user_id=? WHERE id=?", (self.authenticated_user["id"], request_id))
        self.send_json({"completed": True})

    def get_clinicorp_config(self) -> None:
        with connect() as db:
            row = db.execute("SELECT api_base_url, subscriber_id, api_user, api_token, updated_at FROM integration_configs WHERE name='clinicorp'").fetchone()
        if not row:
            return self.send_json({"configured": False, "api_base_url": "", "subscriber_id": "", "api_user": "", "updated_at": None})
        data = dict(row)
        self.send_json({"configured": bool(data["api_token"]), "api_base_url": data["api_base_url"] or "", "subscriber_id": data["subscriber_id"] or "", "api_user": data["api_user"] or "", "updated_at": data["updated_at"]})

    @staticmethod
    def valid_clinicorp_url(value: str) -> bool:
        parsed = urlparse(str(value or "").strip())
        host = (parsed.hostname or "").rstrip(".").lower()
        try:
            port = parsed.port
        except ValueError:
            return False
        return bool(
            parsed.scheme == "https"
            and host
            and (host == "clinicorp.com" or host.endswith(".clinicorp.com"))
            and parsed.username is None
            and parsed.password is None
            and port in (None, 443)
        )

    def save_clinicorp_config(self, payload: dict) -> None:
        api_base_url_input = str(payload.get("api_base_url") or "").strip()
        parsed_base_url = urlparse(api_base_url_input)
        api_base_url = (
            f"{parsed_base_url.scheme}://{parsed_base_url.netloc}"
            if parsed_base_url.scheme in {"http", "https"} and parsed_base_url.netloc
            else ""
        )
        subscriber_id = str(payload.get("subscriber_id") or "").strip()
        api_user = str(payload.get("api_user") or "").strip()
        api_token = str(payload.get("api_token") or "").strip()
        if not self.valid_clinicorp_url(api_base_url) or not subscriber_id or not api_user:
            return self.send_json({"error": "Informe uma URL HTTPS oficial da Clinicorp, ID do assinante e usuário API."}, HTTPStatus.BAD_REQUEST)
        with connect() as db:
            existing = db.execute("SELECT api_token FROM integration_configs WHERE name='clinicorp'").fetchone()
            if not api_token and not existing:
                return self.send_json({"error": "Informe o token API na primeira configuração."}, HTTPStatus.BAD_REQUEST)
            token = encrypt_integration_secret(api_token) if api_token else existing["api_token"]
            db.execute("""INSERT INTO integration_configs (name, api_base_url, subscriber_id, api_user, api_token, updated_at, updated_by)
                         VALUES ('clinicorp', ?, ?, ?, ?, CURRENT_TIMESTAMP, ?)
                         ON CONFLICT(name) DO UPDATE SET api_base_url=excluded.api_base_url, subscriber_id=excluded.subscriber_id,
                         api_user=excluded.api_user, api_token=excluded.api_token, updated_at=CURRENT_TIMESTAMP, updated_by=excluded.updated_by""",
                       (api_base_url, subscriber_id, api_user, token, self.authenticated_user["id"]))
        self.send_json({"configured": True})

    def get_api_integrations(self) -> None:
        with connect() as db:
            rows = db.execute("""SELECT ai.id, ai.name, ai.description, ai.api_base_url, ai.subscriber_id, ai.api_user, ai.active,
                                      ai.sync_interval_seconds, ai.last_sync_at, ai.last_sync_status, ai.last_sync_message,
                                      ai.last_sync_count, ai.updated_at,
                                      CASE WHEN ai.api_token IS NOT NULL AND ai.api_token != '' THEN 1 ELSE 0 END AS token_configured,
                                      CASE WHEN backup.integration_id IS NOT NULL THEN 1 ELSE 0 END AS has_backup
                               FROM api_integrations ai
                               LEFT JOIN api_integration_backups backup ON backup.integration_id=ai.id
                               ORDER BY ai.name COLLATE NOCASE""").fetchall()
        self.send_json({"items": [dict(row) for row in rows]})

    def save_api_integration(self, payload: dict) -> None:
        name = str(payload.get("name") or "").strip()
        description = str(payload.get("description") or "").strip()
        api_base_url = str(payload.get("api_base_url") or "").strip().rstrip("/")
        subscriber_id = str(payload.get("subscriber_id") or "").strip()
        api_user = str(payload.get("api_user") or "").strip()
        api_token = str(payload.get("api_token") or "").strip()
        active = 1 if payload.get("active", True) else 0
        try:
            sync_interval_seconds = max(10, min(3600, int(payload.get("sync_interval_seconds") or 60)))
        except (TypeError, ValueError):
            return self.send_json({"error": "Informe um intervalo de sincronização válido."}, HTTPStatus.BAD_REQUEST)
        if not name or not description:
            return self.send_json({"error": "Informe o nome e a finalidade da API."}, HTTPStatus.BAD_REQUEST)
        if name.casefold().startswith("clinicorp") and not self.valid_clinicorp_url(api_base_url):
            return self.send_json({"error": "A integração Clinicorp aceita somente um endereço HTTPS oficial da Clinicorp."}, HTTPStatus.BAD_REQUEST)
        with connect() as db:
            existing = db.execute("SELECT id, api_token FROM api_integrations WHERE lower(name)=lower(?)", (name,)).fetchone()
            if existing:
                current = db.execute("SELECT * FROM api_integrations WHERE id=?", (existing["id"],)).fetchone()
                db.execute("""INSERT INTO api_integration_backups
                              (integration_id, name, description, api_base_url, subscriber_id, api_user, api_token, active, sync_interval_seconds, saved_at)
                              VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                              ON CONFLICT(integration_id) DO UPDATE SET name=excluded.name, description=excluded.description,
                              api_base_url=excluded.api_base_url, subscriber_id=excluded.subscriber_id, api_user=excluded.api_user,
                              api_token=excluded.api_token, active=excluded.active, sync_interval_seconds=excluded.sync_interval_seconds,
                              saved_at=CURRENT_TIMESTAMP""",
                           (current["id"], current["name"], current["description"], current["api_base_url"], current["subscriber_id"], current["api_user"], current["api_token"], current["active"], current["sync_interval_seconds"]))
                stored_token = encrypt_integration_secret(api_token) if api_token else existing["api_token"]
                db.execute("UPDATE api_integrations SET description=?, api_base_url=?, subscriber_id=?, api_user=?, api_token=?, active=?, sync_interval_seconds=?, updated_at=CURRENT_TIMESTAMP, updated_by=? WHERE id=?", (description, api_base_url or None, subscriber_id or None, api_user or None, stored_token, active, sync_interval_seconds, self.authenticated_user["id"], existing["id"]))
            else:
                stored_token = encrypt_integration_secret(api_token) if api_token else None
                db.execute("INSERT INTO api_integrations (name, description, api_base_url, subscriber_id, api_user, api_token, active, sync_interval_seconds, updated_by) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)", (name, description, api_base_url or None, subscriber_id or None, api_user or None, stored_token, active, sync_interval_seconds, self.authenticated_user["id"]))
        self.send_json({"saved": True})

    def revert_api_integration(self, integration_id: int) -> None:
        with connect() as db:
            backup = db.execute("SELECT * FROM api_integration_backups WHERE integration_id=?", (integration_id,)).fetchone()
            current = db.execute("SELECT id FROM api_integrations WHERE id=?", (integration_id,)).fetchone()
            if not current:
                return self.send_json({"error": "API não encontrada."}, HTTPStatus.NOT_FOUND)
            if not backup:
                return self.send_json({"error": "Ainda não existe uma alteração anterior para reverter."}, HTTPStatus.BAD_REQUEST)
            db.execute("""UPDATE api_integrations SET name=?, description=?, api_base_url=?, subscriber_id=?, api_user=?, api_token=?, active=?, sync_interval_seconds=?, updated_at=CURRENT_TIMESTAMP, updated_by=? WHERE id=?""",
                       (backup["name"], backup["description"], backup["api_base_url"], backup["subscriber_id"], backup["api_user"], backup["api_token"], backup["active"], backup["sync_interval_seconds"], self.authenticated_user["id"], integration_id))
            db.execute("DELETE FROM api_integration_backups WHERE integration_id=?", (integration_id,))
        self.send_json({"reverted": True})

    def start_api_phone_sync(self, integration_id: int, payload: dict) -> None:
        """Starts an intentionally paced Clinicorp phone synchronization.

        The Clinicorp account caps requests at 360/hour, therefore the interval is
        clamped to at least 10 seconds and work occurs outside the HTTP request.
        """
        test_only = bool(payload.get("test_only"))
        with connect() as db:
            integration = db.execute("SELECT * FROM api_integrations WHERE id=?", (integration_id,)).fetchone()
            if not integration:
                return self.send_json({"error": "API não encontrada."}, HTTPStatus.NOT_FOUND)
            config = dict(integration)
            if not config["active"]:
                return self.send_json({"error": "Ative a API antes de sincronizar."}, HTTPStatus.BAD_REQUEST)
            if not config["name"].strip().lower().startswith("clinicorp"):
                return self.send_json({"error": "A sincronização automática de telefone está disponível somente para a API Clinicorp."}, HTTPStatus.BAD_REQUEST)
            if not self.valid_clinicorp_url(config["api_base_url"]):
                return self.send_json({"error": "A URL salva não pertence ao domínio oficial da Clinicorp."}, HTTPStatus.BAD_REQUEST)
            if not config["api_base_url"] or not config["api_user"] or not config["api_token"]:
                return self.send_json({"error": "Complete URL base, usuário API e token antes de sincronizar."}, HTTPStatus.BAD_REQUEST)
            # Clinicorp identifies the account with the subscriber id. In some
            # installations it is the same value as the API user, so that fallback
            # is supported deliberately and remains visible in the configuration.
            if not config["subscriber_id"]:
                config["subscriber_id"] = config["api_user"]
            candidates = db.execute("""SELECT patient.name FROM patients patient
                                      WHERE (patient.phone IS NULL OR TRIM(patient.phone)='') AND TRIM(patient.name)!=''
                                        AND NOT EXISTS (SELECT 1 FROM api_sync_logs failed
                                                        WHERE failed.integration_id=? AND failed.patient_id=patient.id
                                                          AND failed.status IN ('Falhou','Sem telefone'))
                                      LIMIT 500""", (integration_id,)).fetchall()
            if not candidates:
                return self.send_json({"error": "Nenhum paciente sem telefone foi encontrado."}, HTTPStatus.BAD_REQUEST)
        with API_SYNC_LOCK:
            running = API_SYNC_THREADS.get(integration_id)
            if running and running.is_alive():
                return self.send_json({"error": "A sincronização desta API já está em andamento."}, HTTPStatus.CONFLICT)
            with connect() as db:
                run_id = db.execute("""INSERT INTO api_sync_runs (integration_id, started_at, status, message)
                                      VALUES (?, datetime('now','localtime'), 'Em andamento', 'Sincronização iniciada.')""", (integration_id,)).lastrowid
            worker = threading.Thread(target=self.run_clinicorp_phone_sync, args=(integration_id, test_only, run_id), daemon=True, name=f"clinicorp-sync-{integration_id}")
            API_SYNC_THREADS[integration_id] = worker
            with connect() as db:
                db.execute("UPDATE api_integrations SET last_sync_status='Em andamento', last_sync_message='Consultando telefones em segundo plano.', last_sync_count=0 WHERE id=?", (integration_id,))
            worker.start()
        self.send_json({"started": True, "test_only": test_only, "message": "Teste iniciado." if test_only else "Sincronização iniciada em segundo plano."})

    def get_api_sync_status(self, integration_id: int) -> None:
        with connect() as db:
            integration = db.execute("SELECT id, name, last_sync_at, last_sync_status, last_sync_message, last_sync_count FROM api_integrations WHERE id=?", (integration_id,)).fetchone()
            if not integration:
                return self.send_json({"error": "API não encontrada."}, HTTPStatus.NOT_FOUND)
            logs = db.execute("""SELECT log.id, log.external_id, log.phone_result, log.status, log.detail, log.created_at, log.updated_at, patient.name AS patient_name
                               FROM api_sync_logs log
                               LEFT JOIN patients patient ON patient.id=log.patient_id
                               WHERE log.integration_id=? ORDER BY log.id DESC LIMIT 100""", (integration_id,)).fetchall()
            failures = db.execute("""SELECT log.id, log.external_id, log.status, log.detail, log.created_at, log.updated_at,
                                            patient.name AS patient_name
                                     FROM api_sync_logs log
                                     JOIN patients patient ON patient.id=log.patient_id
                                     WHERE log.integration_id=? AND log.status IN ('Falhou','Sem telefone')
                                       AND (patient.phone IS NULL OR TRIM(patient.phone)='')
                                       AND log.id=(SELECT MAX(latest.id) FROM api_sync_logs latest
                                                   WHERE latest.integration_id=log.integration_id
                                                     AND latest.patient_id=log.patient_id
                                                     AND latest.status IN ('Falhou','Sem telefone'))
                                     ORDER BY log.id DESC""", (integration_id,)).fetchall()
            runs = db.execute("""SELECT id, started_at, finished_at, status, attempted_count, updated_count, failed_count, message
                                  FROM api_sync_runs WHERE integration_id=? ORDER BY id DESC LIMIT 30""", (integration_id,)).fetchall()
        with API_SYNC_LOCK:
            worker = API_SYNC_THREADS.get(integration_id)
            running = bool(worker and worker.is_alive())
        self.send_json({"integration": dict(integration), "running": running, "logs": [dict(log) for log in logs],
                        "failures": [dict(row) for row in failures], "failure_count": len(failures),
                        "runs": [dict(row) for row in runs]})

    def retry_api_sync_failures(self, integration_id: int) -> None:
        with connect() as db:
            integration = db.execute("SELECT id FROM api_integrations WHERE id=?", (integration_id,)).fetchone()
            if not integration:
                return self.send_json({"error": "API não encontrada."}, HTTPStatus.NOT_FOUND)
            result = db.execute("DELETE FROM api_sync_logs WHERE integration_id=? AND status IN ('Falhou','Sem telefone')", (integration_id,))
        self.send_json({"released": result.rowcount, "message": f"{result.rowcount} falha(s) liberada(s) para nova tentativa."})

    @staticmethod
    def clinicorp_id_from_name(name: str) -> str | None:
        match = re.search(r"\((\d+)\)\s*$", name or "")
        return match.group(1) if match else None

    @staticmethod
    def clinicorp_clean_name(name: str) -> str:
        return re.sub(r"\s*\(\d+\)\s*$", "", name or "").strip()

    @staticmethod
    def clinicorp_name_key(name: str) -> str:
        plain = unicodedata.normalize("NFKD", name or "").encode("ascii", "ignore").decode("ascii")
        return re.sub(r"\s+", " ", plain).strip().casefold()

    @classmethod
    def clinicorp_patient(cls, config: dict, patient_name: str) -> dict:
        if not cls.valid_clinicorp_url(config.get("api_base_url")):
            raise RuntimeError("A URL da Clinicorp não pertence ao domínio oficial permitido.")
        parsed_url = urlparse(str(config["api_base_url"]))
        if parsed_url.netloc.endswith("clinicorp.com"):
            base_url = f"{parsed_url.scheme or 'https'}://{parsed_url.netloc}/rest/v1"
        else:
            base_url = str(config["api_base_url"]).rstrip("/")
        clean_name = cls.clinicorp_clean_name(patient_name)
        query = urlencode({"subscriber_id": config["subscriber_id"], "Name": clean_name})
        credentials = base64.b64encode(f"{config['api_user']}:{config['api_token']}".encode("utf-8")).decode("ascii")
        request = Request(f"{base_url}/patient/get?{query}", headers={"Accept": "application/json", "Authorization": f"Basic {credentials}", "User-Agent": "IEA-CRM/1.0"})
        try:
            with urlopen(request, timeout=20) as response:
                payload = json.loads(response.read().decode("utf-8"))
        except HTTPError as error:
            body = error.read(600).decode("utf-8", "replace")
            raise RuntimeError(f"Clinicorp respondeu {error.code}: {body}") from error
        except URLError as error:
            raise RuntimeError(f"Não foi possível conectar à Clinicorp: {error.reason}") from error
        items = payload if isinstance(payload, list) else [payload]
        exact = [item for item in items if isinstance(item, dict) and cls.clinicorp_name_key(item.get("Name")) == cls.clinicorp_name_key(clean_name)]
        if not exact:
            raise RuntimeError(f"Clinicorp não encontrou correspondência exata para '{clean_name}'.")
        if len(exact) > 1:
            raise RuntimeError(f"Clinicorp retornou {len(exact)} pacientes com o nome '{clean_name}'. Preenchimento bloqueado por segurança.")
        return exact[0]

    def run_clinicorp_phone_sync(self, integration_id: int, test_only: bool = False, run_id: int | None = None) -> None:
        updated = attempted = without_phone = 0
        status = "Concluída"
        message = "Nenhum novo telefone encontrado."
        attempted_patient_ids: set[int] = set()
        try:
            while True:
                with connect() as db:
                    config_row = db.execute("SELECT * FROM api_integrations WHERE id=?", (integration_id,)).fetchone()
                    if not config_row or not config_row["active"]:
                        status, message = "Interrompida", "A API foi desativada durante a sincronização."
                        break
                    config = dict(config_row)
                    config["api_token"] = decrypt_integration_secret(config.get("api_token"))
                    config["subscriber_id"] = config["subscriber_id"] or config["api_user"]
                    candidates = db.execute("""SELECT patient.id, patient.name FROM patients patient
                                              WHERE (patient.phone IS NULL OR TRIM(patient.phone)='') AND TRIM(patient.name)!=''
                                                AND NOT EXISTS (SELECT 1 FROM api_sync_logs failed
                                                                WHERE failed.integration_id=? AND failed.patient_id=patient.id
                                                                  AND failed.status IN ('Falhou','Sem telefone'))
                                              ORDER BY patient.id LIMIT 3000""", (integration_id,)).fetchall()
                candidate = next((row for row in candidates if row["id"] not in attempted_patient_ids), None)
                if not candidate:
                    message = f"{updated} telefone(s) preenchido(s); não há mais pacientes elegíveis nesta execução."
                    break
                row = candidate
                imported_reference = self.clinicorp_id_from_name(row["name"])
                attempted_patient_ids.add(row["id"])
                attempted += 1
                with connect() as db:
                    log_id = db.execute("INSERT INTO api_sync_logs (integration_id, patient_id, external_id, status, detail) VALUES (?, ?, ?, 'Consultando', 'Busca exata por nome enviada à Clinicorp.')", (integration_id, row["id"], imported_reference)).lastrowid
                try:
                    clinicorp_patient = self.clinicorp_patient(config, row["name"])
                    phone = re.sub(r"\D", "", str(clinicorp_patient.get("Phone") or ""))
                    if phone.startswith("55") and len(phone) in {12, 13}:
                        phone = phone[2:]
                    clinicorp_patient_id = str(clinicorp_patient.get("PatientId") or "").strip()
                except RuntimeError as error:
                    authentication_error = "respondeu 401" in str(error) or "respondeu 403" in str(error)
                    with connect() as db:
                        db.execute("UPDATE api_sync_logs SET status=?, detail=?, updated_at=datetime('now','localtime') WHERE id=?", ("Erro de autenticação" if authentication_error else "Falhou", str(error), log_id))
                    without_phone += 1
                    message = str(error)
                    if authentication_error:
                        status = "Falhou"
                        break
                    if test_only:
                        status = "Teste concluído com erro"
                        break
                    time.sleep(max(10, min(3600, int(config.get("sync_interval_seconds") or 60))))
                    continue
                with connect() as db:
                    db.execute("UPDATE patients SET external_id=?, phone=CASE WHEN ?<>'' THEN ? ELSE phone END, updated_at=CURRENT_TIMESTAMP WHERE name=? AND (phone IS NULL OR TRIM(phone)='')", (clinicorp_patient_id or imported_reference, phone, phone, row["name"]))
                    db.execute("INSERT INTO patient_events (patient_id, event_type, description) VALUES (?, 'Integração Clinicorp', ?)", (row["id"], "Telefone preenchido automaticamente pela Clinicorp." if phone else "Clinicorp não retornou telefone para este paciente."))
                    db.execute("UPDATE api_sync_logs SET external_id=?, phone_result=?, status=?, detail=?, updated_at=datetime('now','localtime') WHERE id=?", (clinicorp_patient_id or imported_reference, phone or None, "Atualizado" if phone else "Sem telefone", "Nome validado e telefone preenchido com sucesso." if phone else "Nome validado, mas a Clinicorp não retornou telefone.", log_id))
                    db.execute("UPDATE api_integrations SET last_sync_count=? WHERE id=?", (updated + (1 if phone else 0), integration_id))
                if phone:
                    updated += 1
                    message = f"{updated} telefone(s) preenchido(s)."
                else:
                    without_phone += 1
                if test_only or attempted >= 2000:
                    break
                time.sleep(max(10, min(3600, int(config.get("sync_interval_seconds") or 60))))
        except Exception as error:
            status, message = "Falhou", f"Erro inesperado: {error}"
        finally:
            with connect() as db:
                db.execute("UPDATE api_integrations SET last_sync_at=datetime('now','localtime'), last_sync_status=?, last_sync_message=?, last_sync_count=? WHERE id=?", (status, message, updated, integration_id))
                if run_id:
                    db.execute("""UPDATE api_sync_runs SET finished_at=datetime('now','localtime'), status=?, attempted_count=?,
                                  updated_count=?, failed_count=?, message=? WHERE id=?""",
                               (status, attempted, updated, without_phone, message, run_id))
            with API_SYNC_LOCK:
                API_SYNC_THREADS.pop(integration_id, None)

    def get_admin_overview(self) -> None:
        with connect() as db:
            summary = db.execute("""
                SELECT
                    (SELECT COUNT(*) FROM patients) AS patients,
                    (SELECT COUNT(*) FROM professionals WHERE active = 1) AS professionals,
                    (SELECT COUNT(*) FROM users WHERE active = 1) AS active_accesses,
                    (SELECT COUNT(*) FROM offices WHERE active = 1) AS offices,
                    (SELECT COUNT(*) FROM specialties WHERE active = 1) AS specialties,
                    (SELECT COUNT(*) FROM patients p WHERE NOT EXISTS (SELECT 1 FROM patient_assignments pa WHERE pa.patient_id = p.id)) AS unassigned,
                    (SELECT COALESCE(SUM(GREATEST(value_cents - discount_cents, 0)), 0) FROM procedures WHERE stage != 'Concluído') AS potential_value_cents
            """).fetchone()
            professionals = db.execute("""
                SELECT pr.id, pr.name, pr.role, pr.is_owner, pr.active, CASE WHEN pr.photo_data IS NOT NULL THEN 1 ELSE 0 END AS has_photo,
                       u.id AS user_id,u.email, u.access_role, u.linked_professional_id,
                       linked_pr.name AS linked_professional_name,
                       COALESCE(u.active, 0) AS access_active,
                       COALESCE(u.service_sector,'') AS service_sector,
                       COALESCE(u.crm_access_level,'attendant') AS crm_access_level,
                       COALESCE(u.crm_channel_scope_enabled,0) AS crm_channel_scope_enabled,
                       (SELECT GROUP_CONCAT(CAST(cuc.channel_id AS TEXT)) FROM crm_user_channels cuc WHERE cuc.user_id=u.id) AS crm_channel_ids,
                       COALESCE(u.crm_manage_automation,0) AS crm_can_manage_automation,
                       (SELECT GROUP_CONCAT(s.name, ', ') FROM professional_specialties ps JOIN specialties s ON s.id = ps.specialty_id WHERE ps.professional_id = pr.id) AS specialties,
                       (SELECT GROUP_CONCAT(o.name, ', ') FROM professional_offices po JOIN offices o ON o.id = po.office_id WHERE po.professional_id = pr.id) AS offices,
                       (SELECT ps.specialty_id FROM professional_specialties ps WHERE ps.professional_id = pr.id ORDER BY ps.is_primary DESC, ps.specialty_id LIMIT 1) AS specialty_id,
                       (SELECT po.office_id FROM professional_offices po WHERE po.professional_id = pr.id ORDER BY po.is_responsible DESC, po.office_id LIMIT 1) AS office_id,
                       (SELECT po.is_responsible FROM professional_offices po WHERE po.professional_id = pr.id ORDER BY po.is_responsible DESC LIMIT 1) AS office_responsible,
                       (SELECT COUNT(*) FROM patient_assignments pa WHERE pa.professional_id = pr.id AND pa.is_primary = 1) AS patient_count
                FROM professionals pr
                LEFT JOIN users u ON u.professional_id = pr.id
                LEFT JOIN professionals linked_pr ON linked_pr.id = u.linked_professional_id
                ORDER BY pr.is_owner DESC, pr.active DESC, pr.name COLLATE NOCASE
            """).fetchall()
            specialties = db.execute("""
                SELECT s.id, s.code, s.name, s.active,
                       (SELECT COUNT(*) FROM professional_specialties ps WHERE ps.specialty_id = s.id) AS professional_count
                FROM specialties s ORDER BY s.active DESC, s.name COLLATE NOCASE
            """).fetchall()
            offices = db.execute("""
                SELECT o.id, o.code, o.name, o.active,
                       (SELECT COUNT(*) FROM professional_offices po WHERE po.office_id = o.id) AS professional_count,
                       (SELECT GROUP_CONCAT(pr.name, ', ') FROM professional_offices po JOIN professionals pr ON pr.id = po.professional_id WHERE po.office_id = o.id AND po.is_responsible = 1) AS responsibles
                FROM offices o ORDER BY o.active DESC, o.name COLLATE NOCASE
            """).fetchall()
            reset_requests = db.execute("""
                SELECT r.id, r.requested_at, u.name, u.email
                FROM password_reset_requests r JOIN users u ON u.id=r.user_id
                WHERE r.status='pending' ORDER BY r.requested_at ASC
            """).fetchall()
            crm_channels = db.execute("""SELECT id,instance_name,display_name,phone,active,sync_enabled,connection_status
                                          FROM crm_channels ORDER BY active DESC,display_name COLLATE NOCASE""").fetchall()
        self.send_json({
            "summary": dict(summary),
            "professionals": [dict(row) for row in professionals],
            "specialties": [dict(row) for row in specialties],
            "offices": [dict(row) for row in offices],
            "password_reset_requests": [dict(row) for row in reset_requests],
            "crm_channels": [dict(row) for row in crm_channels],
        })

    def get_admin_crm_channel_access(self) -> None:
        with connect() as db:
            users = db.execute("""SELECT u.id,u.name,u.email,u.active,u.crm_access_level,u.crm_channel_scope_enabled,u.crm_feature_scope_enabled,u.crm_operational_agent,u.crm_manage_automation,
                       COALESCE(GROUP_CONCAT(CAST(cuc.channel_id AS TEXT)),'') AS channel_ids,
                       COALESCE(u.crm_manage_automation,0) AS can_manage_automation,
                       COALESCE((SELECT GROUP_CONCAT(cuf.feature_key) FROM crm_user_features cuf WHERE cuf.user_id=u.id),'') AS feature_keys
                       FROM users u LEFT JOIN crm_user_channels cuc ON cuc.user_id=u.id
                       WHERE u.access_role='crc' GROUP BY u.id ORDER BY u.name COLLATE NOCASE""").fetchall()
            channels = db.execute("""SELECT id,instance_name,display_name,phone,active,sync_enabled,connection_status
                                      FROM crm_channels ORDER BY display_name COLLATE NOCASE""").fetchall()
        self.send_json({"users":[dict(row) for row in users],"channels":[dict(row) for row in channels]})

    def save_admin_crm_channel_access(self, payload: dict) -> None:
        try:
            user_id = int(payload.get("user_id") or 0)
            raw_channel_ids = payload.get("channel_ids") or []
            channel_ids = list(dict.fromkeys(int(value) for value in raw_channel_ids))
            feature_keys = list(dict.fromkeys(
                str(value).strip().lower() for value in (payload.get("feature_keys") or []) if str(value).strip()
            ))
        except (TypeError, ValueError):
            return self.send_json({"error": "Usuário ou canais inválidos."}, HTTPStatus.BAD_REQUEST)
        if not user_id:
            return self.send_json({"error": "Selecione o usuário do CRC."}, HTTPStatus.BAD_REQUEST)
        if set(feature_keys) - set(CRM_FEATURE_KEYS):
            return self.send_json({"error": "Uma das telas informadas não existe."}, HTTPStatus.BAD_REQUEST)

        scope_enabled = 1 if payload.get("scope_enabled") is True else 0
        feature_scope_enabled = 1 if payload.get("feature_scope_enabled") is True else 0
        can_manage_automation = 1 if payload.get("can_manage_automation") is True else 0
        operational_agent = 1 if payload.get("operational_agent") is True else 0
        with connect() as db:
            user = db.execute(
                "SELECT id,name,crm_access_level,crm_channel_scope_enabled,crm_feature_scope_enabled,crm_operational_agent,crm_manage_automation "
                "FROM users WHERE id=? AND access_role='crc'", (user_id,)
            ).fetchone()
            if not user:
                return self.send_json({"error": "Usuário CRC não encontrado."}, HTTPStatus.NOT_FOUND)
            if "operational_agent" not in payload:
                operational_agent = 1 if user["crm_operational_agent"] else 0
            is_crm_admin = user["crm_access_level"] == "admin"
            if is_crm_admin:
                scope_enabled = 0
                feature_scope_enabled = 0
                can_manage_automation = 1
                channel_ids = []
                feature_keys = []
            if channel_ids:
                placeholders = ",".join("?" for _ in channel_ids)
                valid = {row["id"] for row in db.execute(
                    f"SELECT id FROM crm_channels WHERE id IN ({placeholders})", channel_ids
                ).fetchall()}
                if valid != set(channel_ids):
                    return self.send_json({"error": "Um dos canais informados não existe."}, HTTPStatus.BAD_REQUEST)

            before_channel_ids = [int(row["channel_id"]) for row in db.execute(
                "SELECT channel_id FROM crm_user_channels WHERE user_id=? ORDER BY channel_id", (user_id,)
            ).fetchall()]
            before_feature_keys = [str(row["feature_key"]) for row in db.execute(
                "SELECT feature_key FROM crm_user_features WHERE user_id=? ORDER BY feature_key", (user_id,)
            ).fetchall()]
            before = {
                "channel_scope_enabled": bool(user["crm_channel_scope_enabled"]),
                "channel_ids": before_channel_ids,
                "feature_scope_enabled": bool(user["crm_feature_scope_enabled"]),
                "feature_keys": before_feature_keys,
                "operational_agent": bool(user["crm_operational_agent"]),
                "can_manage_automation": bool(user["crm_manage_automation"]),
                "crm_access_level": user["crm_access_level"],
            }
            after = {
                "channel_scope_enabled": bool(scope_enabled),
                "channel_ids": channel_ids,
                "can_manage_automation": bool(can_manage_automation),
                "feature_scope_enabled": bool(feature_scope_enabled),
                "feature_keys": feature_keys,
                "operational_agent": bool(operational_agent),
                "crm_access_level": user["crm_access_level"],
            }

            db.execute("DELETE FROM crm_user_channels WHERE user_id=?", (user_id,))
            db.executemany(
                "INSERT INTO crm_user_channels(user_id,channel_id,can_reply,can_manage_automation) VALUES(?,?,?,?)",
                [(user_id, channel_id, 1, can_manage_automation) for channel_id in channel_ids],
            )
            db.execute(
                "UPDATE users SET crm_channel_scope_enabled=?,crm_manage_automation=?,crm_operational_agent=? WHERE id=?",
                (scope_enabled, can_manage_automation, operational_agent, user_id),
            )
            db.execute("DELETE FROM crm_user_features WHERE user_id=?", (user_id,))
            db.executemany(
                "INSERT INTO crm_user_features(user_id,feature_key) VALUES(?,?)",
                [(user_id, key) for key in feature_keys],
            )
            db.execute("UPDATE users SET crm_feature_scope_enabled=? WHERE id=?", (feature_scope_enabled, user_id))
            db.execute(
                "INSERT INTO crm_permission_audit "
                "(changed_by_user_id,target_user_id,before_json,after_json,ip_address) VALUES(?,?,?,?,?)",
                (
                    self.authenticated_user["id"], user_id,
                    json.dumps(before, ensure_ascii=False, sort_keys=True),
                    json.dumps(after, ensure_ascii=False, sort_keys=True), self.request_ip(),
                ),
            )
            self.record_security_event(
                db, "crm_permissions_updated", self.request_ip(), self.authenticated_user["id"],
                f"Permissões de {user['name']} (usuário {user_id}) atualizadas.",
            )
        self.send_json({
            "updated": True, "user_id": user_id, "channel_ids": channel_ids,
            "feature_keys": feature_keys, "operational_agent": bool(operational_agent),
            "can_manage_automation": bool(can_manage_automation),
        })

    def test_admin_crm_channel_access(self, payload: dict) -> None:
        try:
            user_id = int(payload.get("user_id") or 0)
            channel_id = int(payload.get("channel_id") or 0)
        except (TypeError, ValueError):
            return self.send_json({"error": "Atendente ou canal invÃ¡lido."}, HTTPStatus.BAD_REQUEST)
        with connect() as db:
            user = db.execute("SELECT id,name,email,active FROM users WHERE id=? AND access_role='crc'", (user_id,)).fetchone()
            channel = db.execute("SELECT id,display_name,instance_name,active,sync_enabled FROM crm_channels WHERE id=?", (channel_id,)).fetchone()
            if not user or not channel:
                return self.send_json({"error": "Atendente ou canal nÃ£o encontrado."}, HTTPStatus.NOT_FOUND)
            can_view = self.crm_channel_allowed(db, channel_id, user_id=user_id)
            can_reply = self.crm_channel_allowed(db, channel_id, "reply", user_id=user_id)
            can_manage = self.crm_channel_allowed(db, channel_id, "automation", user_id=user_id)
        feature_permissions = {key: self.crm_feature_allowed(key, user_id=user_id) for key in CRM_FEATURE_KEYS}
        transfer_allowed = bool(user["active"] and channel["active"] and channel["sync_enabled"] and can_reply)
        self.send_json({
            "user": dict(user), "channel": dict(channel), "can_view": can_view,
            "can_reply": can_reply, "can_manage_automation": can_manage,
            "feature_permissions": feature_permissions,
            "transfer_allowed": transfer_allowed,
            "message": ("TransferÃªncia liberada para esta atendente neste nÃºmero."
                        if transfer_allowed else "TransferÃªncia bloqueada pela permissÃ£o ou pelo estado do canal."),
        })

    @staticmethod
    def permission_change_summary(before: dict, after: dict) -> str:
        labels = {
            "channel_scope_enabled": "Restrição por canal",
            "feature_scope_enabled": "Personalização de telas",
            "can_manage_automation": "Supervisão de automações",
            "operational_agent": "Participação nos atendimentos",
        }
        changes = []
        for key, label in labels.items():
            if before.get(key) != after.get(key):
                old = "Sim" if before.get(key) else "Não"
                new = "Sim" if after.get(key) else "Não"
                changes.append(f"{label}: {old} → {new}")
        for key, label in (("channel_ids", "Canais"), ("feature_keys", "Telas")):
            old_values = before.get(key) or []
            new_values = after.get(key) or []
            if old_values != new_values:
                changes.append(f"{label}: {len(old_values)} → {len(new_values)}")
        return " · ".join(changes) or "Configuração salva sem alteração efetiva"

    def get_admin_audit(self, query: dict) -> None:
        try:
            limit = min(200, max(10, int(query.get("limit", ["80"])[0])))
        except ValueError:
            limit = 80
        with connect() as db:
            rows = db.execute("""
                SELECT id, event_type, description, created_at, patient_name,
                       details_before, details_after, ip_address
                FROM (
                    SELECT pe.id AS id, pe.event_type AS event_type,
                           pe.description AS description, pe.created_at AS created_at,
                           p.name AS patient_name, NULL AS details_before,
                           NULL AS details_after, NULL AS ip_address
                    FROM patient_events pe
                    JOIN patients p ON p.id = pe.patient_id
                    UNION ALL
                    SELECT 1000000000 + pda.id AS id,
                           'Paciente excluído' AS event_type,
                           'Excluído por ' || pda.deleted_by_name ||
                           ' (' || pda.deleted_by_role || ')' ||
                           CASE WHEN pda.professional_name IS NOT NULL
                                THEN ' · Carteira: ' || pda.professional_name ELSE '' END AS description,
                           pda.deleted_at AS created_at,
                           pda.patient_name AS patient_name, NULL AS details_before,
                           NULL AS details_after, NULL AS ip_address
                    FROM patient_deletion_audit pda
                    UNION ALL
                    SELECT 2000000000 + cpa.id AS id,
                           'Permissões do CRM alteradas' AS event_type,
                           'Alterado por ' || COALESCE(actor.name,'Usuário removido') ||
                           ' para ' || COALESCE(target.name,'Usuário removido') AS description,
                           cpa.created_at AS created_at,
                           NULL AS patient_name, cpa.before_json AS details_before,
                           cpa.after_json AS details_after, cpa.ip_address AS ip_address
                    FROM crm_permission_audit cpa
                    LEFT JOIN users actor ON actor.id=cpa.changed_by_user_id
                    LEFT JOIN users target ON target.id=cpa.target_user_id
                )
                ORDER BY created_at DESC, id DESC LIMIT ?
            """, (limit,)).fetchall()
        items = []
        for row in rows:
            item = dict(row)
            if item.get("details_before") or item.get("details_after"):
                try:
                    before = json.loads(item.pop("details_before") or "{}")
                    after = json.loads(item.pop("details_after") or "{}")
                    item["change_summary"] = self.permission_change_summary(before, after)
                except (TypeError, json.JSONDecodeError):
                    item["change_summary"] = "Detalhes de permissão indisponíveis"
            else:
                item.pop("details_before", None)
                item.pop("details_after", None)
                item["change_summary"] = ""
            items.append(item)
        self.send_json({"items": items})

    def normalized_code(self, value: str) -> str:
        plain = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
        return re.sub(r"[^A-Z0-9]+", "_", plain.upper()).strip("_")

    def import_key(self, value) -> str:
        return self.normalized_code(str(value or ""))

    @staticmethod
    def import_date(value):
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        text = str(value or "").strip()
        for pattern in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
            try:
                return datetime.strptime(text, pattern).date().isoformat()
            except ValueError:
                pass
        return None

    def store_typed_import_batch(self, kind: str, items: list[dict]) -> str:
        token = secrets.token_urlsafe(32)
        payload = {"kind": kind, "items": items}
        with connect() as db:
            db.execute("DELETE FROM import_batches WHERE expires_at <= datetime('now')")
            db.execute(
                "INSERT INTO import_batches (token_hash, created_by_user_id, payload_json, expires_at) VALUES (?, ?, ?, ?)",
                (
                    hashlib.sha256(token.encode()).hexdigest(),
                    self.authenticated_user["id"],
                    json.dumps(payload, ensure_ascii=False),
                    (datetime.utcnow() + timedelta(hours=2)).strftime("%Y-%m-%d %H:%M:%S"),
                ),
            )
        return token

    def get_typed_import_batch(self, token: str, kind: str) -> list[dict] | None:
        payload = self.get_import_batch(token)
        if not isinstance(payload, dict) or payload.get("kind") != kind:
            return None
        items = payload.get("items")
        return items if isinstance(items, list) else None

    def read_tabular_files(self, files: list) -> tuple[list[tuple[str, dict]], list[str]]:
        result, errors = [], []
        for item in files[:30]:
            name = str(item.get("name") or "planilha").strip()
            try:
                raw = base64.b64decode(str(item.get("data") or ""), validate=True)
                if len(raw) > 5 * 1024 * 1024:
                    raise ValueError("arquivo maior que 5 MB")
                if name.lower().endswith(".csv"):
                    matrix = list(csv.reader(io.StringIO(raw.decode("utf-8-sig"))))
                else:
                    book = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
                    matrix = list(book.active.iter_rows(values_only=True))
                if not matrix:
                    raise ValueError("planilha vazia")
                headers = [self.import_key(value) for value in matrix[0]]
                for values in matrix[1:]:
                    row = {
                        header: (values[index] if index < len(values) else None)
                        for index, header in enumerate(headers) if header
                    }
                    if any(value not in (None, "") for value in row.values()):
                        result.append((name, row))
            except Exception as error:
                errors.append(f"{name}: {error}")
        return result, errors

    def patient_import_name_issue(self, patient_name: str) -> tuple[str, str] | None:
        name = str(patient_name or "").strip()
        normalized = self.import_key(name)
        operational_tokens = {
            "COMPROMISSO", "EVENTO", "BLOQUEIO", "REUNIAO", "ALMOCO",
            "FERIADO", "INTERVALO", "AUSENCIA", "FOLGA", "CURSO",
        }
        operational_phrases = {"NAO_AGENDAR", "SEM_ATENDIMENTO", "HORARIO_BLOQUEADO"}
        name_tokens = set(normalized.split("_"))
        if name_tokens.intersection(operational_tokens) or any(term in normalized for term in operational_phrases):
            return ("bloqueado", "marcação operacional/agenda, não é paciente")
        if not re.search(r"[A-Za-zÀ-ÿ]{2,}", name):
            return ("bloqueado", "não possui um nome de pessoa válido")
        if not re.search(r"\(\s*\d+\s*\)\s*$", name):
            return ("revisao", "não possui código do paciente entre parênteses no final")
        return None

    def get_relationship_directory(self, query: dict) -> None:
        search = str((query.get("search") or [""])[0]).strip()
        if len(search) < 2:
            return self.send_json({"items": []})
        normalized = self.import_key(search)
        limit = min(max(int((query.get("limit") or ["8"])[0]), 1), 20)
        with connect() as db:
            rows = db.execute(
                """
                SELECT id, name
                FROM relationship_name_directory
                WHERE active=1 AND normalized_name LIKE ?
                ORDER BY CASE WHEN normalized_name LIKE ? THEN 0 ELSE 1 END, name
                LIMIT ?
                """,
                (f"%{normalized}%", f"{normalized}%", limit),
            ).fetchall()
        self.send_json({"items": [dict(row) for row in rows]})

    def relationship_directory_audit(self, payload: dict) -> None:
        files = payload.get("files") or []
        if not isinstance(files, list) or not files:
            return self.send_json({"error": "Selecione a lista de nomes."}, HTTPStatus.BAD_REQUEST)
        rows, errors = self.read_tabular_files(files)
        consolidated = {}
        for source, row in rows:
            name = str(row.get("PACIENTE") or row.get("NOME") or row.get("NOME_PACIENTE") or "").strip()
            key = self.import_key(name)
            if name and key:
                consolidated[key] = {"name": name, "source": source}
        items = sorted(consolidated.values(), key=lambda item: self.import_key(item["name"]))
        if not items:
            errors.append("Nenhum nome encontrado. Use a coluna Paciente ou Nome.")
        token = self.store_typed_import_batch("relationship_directory", items)
        self.send_json({
            "token": token,
            "read": len(rows),
            "valid": len(items),
            "duplicates_removed": max(0, len(rows) - len(items)),
            "errors": errors,
            "preview": items[:10],
        })

    def relationship_directory_confirm(self, payload: dict) -> None:
        token = str(payload.get("token") or "")
        items = self.get_typed_import_batch(token, "relationship_directory")
        if items is None:
            return self.send_json({"error": "Auditoria expirada. Rode novamente."}, HTTPStatus.GONE)
        inserted = updated = 0
        with connect() as db:
            for item in items:
                key = self.import_key(item["name"])
                exists = db.execute(
                    "SELECT id FROM relationship_name_directory WHERE normalized_name=?",
                    (key,),
                ).fetchone()
                if exists:
                    db.execute(
                        "UPDATE relationship_name_directory SET name=?, source=?, active=1, updated_at=CURRENT_TIMESTAMP WHERE id=?",
                        (item["name"], item.get("source"), exists["id"]),
                    )
                    updated += 1
                else:
                    db.execute(
                        "INSERT INTO relationship_name_directory (name, normalized_name, source) VALUES (?, ?, ?)",
                        (item["name"], key, item.get("source")),
                    )
                    inserted += 1
            db.execute("DELETE FROM import_batches WHERE token_hash=?", (hashlib.sha256(token.encode()).hexdigest(),))
        self.send_json({"inserted": inserted, "updated": updated, "total": len(items)})

    def scheduled_appointment_audit(self, payload: dict) -> None:
        files = payload.get("files") or []
        if not isinstance(files, list) or not files:
            return self.send_json({"error": "Selecione a lista de agendados."}, HTTPStatus.BAD_REQUEST)
        rows, errors = self.read_tabular_files(files)
        consolidated, ignored = {}, 0
        for source, row in rows:
            patient = str(row.get("PACIENTE") or row.get("NOME") or row.get("NOME_PACIENTE") or "").strip()
            professional = str(row.get("PROFISSIONAL") or row.get("DENTISTA") or row.get("RESPONSAVEL") or "").strip()
            status = str(row.get("STATUS") or row.get("SITUACAO") or "").strip()
            raw_date = row.get("DATA_AGENDAMENTO") or row.get("PROXIMA_CONSULTA") or row.get("AGENDAMENTO") or row.get("DATA")
            appointment = self.import_date(raw_date)
            if self.import_key(status) != "AGENDADO":
                ignored += 1
                continue
            issue = self.patient_import_name_issue(patient)
            if issue:
                level, reason = issue
                errors.append(f"{source}: {patient or 'nome vazio'} · {level}: {reason}.")
                continue
            if not patient or not professional or not appointment:
                errors.append(f"{source}: linha agendada sem Paciente, Profissional ou Data válida.")
                continue
            key = (self.import_key(patient), self.import_key(professional))
            item = {"patient": patient, "professional": professional, "date": appointment, "status": "Agendado", "source": source}
            if key not in consolidated or appointment > consolidated[key]["date"]:
                consolidated[key] = item
        items = sorted(consolidated.values(), key=lambda item: (self.import_key(item["professional"]), self.import_key(item["patient"])))
        token = self.store_typed_import_batch("scheduled_appointments", items)
        self.send_json({
            "token": token, "read": len(rows), "scheduled": len(items), "ignored": ignored,
            "duplicates_removed": max(0, len(rows) - ignored - len(items)), "errors": errors, "preview": items[:10],
        })

    def scheduled_appointment_matches(self, db, items: list[dict]) -> tuple[dict, list[str]]:
        professionals = {self.import_key(row["name"]): row["id"] for row in db.execute("SELECT id, name FROM professionals WHERE active=1").fetchall()}
        assignments = {}
        for row in db.execute("""
            SELECT p.id, p.name, pa.professional_id, f.last_visit,
                   f.next_appointment, f.next_appointment_type
            FROM patients p
            JOIN patient_assignments pa ON pa.patient_id=p.id AND pa.is_primary=1
            JOIN patient_followup f ON f.patient_id=p.id
        """).fetchall():
            key = (self.import_key(row["name"]), row["professional_id"])
            candidate = dict(row)
            current = assignments.get(key)
            candidate_order = (candidate.get("last_visit") or "", candidate.get("next_appointment") or "", candidate["id"])
            current_order = (current.get("last_visit") or "", current.get("next_appointment") or "", current["id"]) if current else None
            if current is None or candidate_order > current_order:
                assignments[key] = candidate
        errors, matches = [], {}
        for index, item in enumerate(items):
            professional_id = professionals.get(self.import_key(item["professional"]))
            if not professional_id:
                errors.append(f"Profissional não encontrado: {item['professional']}")
                continue
            patient = assignments.get((self.import_key(item["patient"]), professional_id))
            if not patient:
                errors.append(f"Paciente não encontrado na carteira de {item['professional']}: {item['patient']}")
                continue
            matches[index] = patient
        return matches, sorted(set(errors))

    @staticmethod
    def scheduled_appointment_decision(patient: dict, incoming_date: str) -> str:
        current_date = patient.get("next_appointment")
        current_type = patient.get("next_appointment_type")
        if not current_date:
            return "Adicionar agendamento"
        if incoming_date > current_date:
            return "Atualizar para data mais recente"
        if incoming_date == current_date and current_type != "Agendado":
            return "Confirmar como Agendado"
        if incoming_date == current_date:
            return "Manter: data já cadastrada"
        return "Manter: sistema possui data mais recente"

    def scheduled_appointment_validate(self, payload: dict) -> None:
        items = self.get_typed_import_batch(str(payload.get("token") or ""), "scheduled_appointments")
        if items is None:
            return self.send_json({"error": "Auditoria expirada. Rode novamente."}, HTTPStatus.GONE)
        with connect() as db:
            matches, errors = self.scheduled_appointment_matches(db, items)
        preview = []
        for index, item in enumerate(items[:10]):
            patient = matches.get(index)
            preview.append({
                **item,
                "current_date": patient.get("next_appointment") if patient else None,
                "decision": self.scheduled_appointment_decision(patient, item["date"]) if patient else "Bloqueado",
            })
        self.send_json({"total": len(items), "matched": len(matches), "errors": errors, "can_confirm": not errors and bool(items), "preview": preview})

    def scheduled_appointment_confirm(self, payload: dict) -> None:
        token = str(payload.get("token") or "")
        items = self.get_typed_import_batch(token, "scheduled_appointments")
        if items is None:
            return self.send_json({"error": "Auditoria expirada. Rode novamente."}, HTTPStatus.GONE)
        updated = unchanged = 0
        with connect() as db:
            matches, errors = self.scheduled_appointment_matches(db, items)
            if errors:
                return self.send_json({"error": "Corrija a validação antes de confirmar.", "details": errors}, HTTPStatus.CONFLICT)
            for index, item in enumerate(items):
                patient = matches[index]
                decision = self.scheduled_appointment_decision(patient, item["date"])
                if decision.startswith("Manter:"):
                    unchanged += 1
                    continue
                db.execute(
                    "UPDATE patient_followup SET next_appointment=?, next_appointment_type='Agendado' WHERE patient_id=?",
                    (item["date"] if decision != "Confirmar como Agendado" else patient.get("next_appointment"), patient["id"]),
                )
                db.execute(
                    "INSERT INTO patient_events (patient_id, event_type, description) VALUES (?, 'Importação', ?)",
                    (patient["id"], f"Próxima consulta atualizada para {item['date']} pela planilha {item['source']} · {self.authenticated_user['name']}"),
                )
                updated += 1
            db.execute("DELETE FROM import_batches WHERE token_hash=?", (hashlib.sha256(token.encode()).hexdigest(),))
        self.send_json({"updated": updated, "unchanged": unchanged, "total": len(items)})

    def read_appointment_rows(self, files: list) -> tuple[list[dict], list[str]]:
        rows, errors = [], []
        for item in files[:30]:
            name = str(item.get("name") or "planilha").strip()
            encoded = str(item.get("data") or "")
            try:
                raw = base64.b64decode(encoded, validate=True)
                if len(raw) > 5 * 1024 * 1024:
                    raise ValueError("arquivo maior que 5 MB")
                if name.lower().endswith(".csv"):
                    matrix = list(csv.reader(io.StringIO(raw.decode("utf-8-sig"))))
                else:
                    book = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
                    matrix = list(book.active.iter_rows(values_only=True))
                if not matrix:
                    raise ValueError("planilha vazia")
                headers = {self.import_key(value): index for index, value in enumerate(matrix[0])}
                required = {"DATA", "PACIENTE", "CATEGORIA", "PROFISSIONAL"}
                if not required.issubset(headers):
                    raise ValueError("colunas obrigatórias: Data, Paciente, Categoria e Profissional")
                for values in matrix[1:]:
                    def value(column): return values[headers[column]] if headers[column] < len(values) else None
                    appointment = self.import_date(value("DATA"))
                    patient, professional = str(value("PACIENTE") or "").strip(), str(value("PROFISSIONAL") or "").strip()
                    if appointment and patient and professional:
                        issue = self.patient_import_name_issue(patient)
                        if issue:
                            level, reason = issue
                            errors.append(f"{name}: {patient} · {level}: {reason}.")
                            continue
                        raw_status = str(value("STATUS") or "").strip() if "STATUS" in headers else ""
                        rows.append({"date": appointment, "patient": patient, "category": str(value("CATEGORIA") or "").strip(), "status": raw_status, "professional": professional, "source": name})
            except Exception as error:
                errors.append(f"{name}: {error}")
        return rows, errors

    def appointment_audit(self, payload: dict) -> None:
        files = payload.get("files") or []
        if not isinstance(files, list) or not files:
            return self.send_json({"error": "Selecione ao menos uma planilha."}, HTTPStatus.BAD_REQUEST)
        rows, errors = self.read_appointment_rows(files)
        consolidated = {}
        for row in rows:
            key = (self.import_key(row["patient"]), self.import_key(row["professional"]))
            if key not in consolidated or row["date"] > consolidated[key]["date"]:
                consolidated[key] = row
        items = sorted(consolidated.values(), key=lambda item: (self.import_key(item["professional"]), self.import_key(item["patient"])))
        token = secrets.token_urlsafe(32)
        with connect() as db:
            db.execute("DELETE FROM import_batches WHERE expires_at <= datetime('now')")
            db.execute("INSERT INTO import_batches (token_hash, created_by_user_id, payload_json, expires_at) VALUES (?, ?, ?, ?)", (hashlib.sha256(token.encode()).hexdigest(), self.authenticated_user["id"], json.dumps(items, ensure_ascii=False), (datetime.utcnow() + timedelta(hours=2)).strftime("%Y-%m-%d %H:%M:%S")))
        self.send_json({"token": token, "read": len(rows), "consolidated": len(items), "duplicates_removed": len(rows) - len(items), "errors": errors, "preview": items[:10]})

    def get_import_batch(self, token: str):
        with connect() as db:
            row = db.execute("SELECT payload_json FROM import_batches WHERE token_hash=? AND created_by_user_id=? AND expires_at > datetime('now')", (hashlib.sha256(token.encode()).hexdigest(), self.authenticated_user["id"])).fetchone()
        return json.loads(row["payload_json"]) if row else None

    def appointment_validate(self, payload: dict) -> None:
        items = self.get_import_batch(str(payload.get("token") or ""))
        if items is None:
            return self.send_json({"error": "Auditoria expirada. Rode a auditoria novamente."}, HTTPStatus.GONE)
        with connect() as db:
            professionals = {self.import_key(row["name"]): row["id"] for row in db.execute("SELECT id, name FROM professionals WHERE active=1").fetchall()}
        missing = sorted({item["professional"] for item in items if self.import_key(item["professional"]) not in professionals})
        self.send_json({"valid": len(items) - (len(items) if missing else 0), "total": len(items), "missing_professionals": missing, "can_confirm": not missing, "preview": items[:10]})

    def appointment_confirm(self, payload: dict) -> None:
        token = str(payload.get("token") or "")
        items = self.get_import_batch(token)
        if items is None:
            return self.send_json({"error": "Auditoria expirada. Rode a auditoria novamente."}, HTTPStatus.GONE)
        imported = updated = unchanged = 0
        with connect() as db:
            professionals = {self.import_key(row["name"]): row["id"] for row in db.execute("SELECT id, name FROM professionals WHERE active=1").fetchall()}
            missing = sorted({item["professional"] for item in items if self.import_key(item["professional"]) not in professionals})
            if missing:
                return self.send_json({"error": "Cadastre primeiro: " + ", ".join(missing)}, HTTPStatus.CONFLICT)
            assignments = {}
            for row in db.execute("""
                SELECT p.id, p.name, pa.professional_id, f.last_visit
                FROM patients p
                JOIN patient_assignments pa ON pa.patient_id=p.id AND pa.is_primary=1
                JOIN patient_followup f ON f.patient_id=p.id
            """).fetchall():
                key = (self.import_key(row["name"]), row["professional_id"])
                current = assignments.get(key)
                if not current or (row["last_visit"] or "") > (current["last_visit"] or ""):
                    assignments[key] = {"id": row["id"], "last_visit": row["last_visit"]}
            for item in items:
                professional_id = professionals[self.import_key(item["professional"])]
                key = (self.import_key(item["patient"]), professional_id)
                existing = assignments.get(key)
                requested_status = str(item.get("status") or "").strip()
                status = requested_status or ("Tratamento" if "TRATAMENTO" in self.import_key(item["category"]) else "Consulta")
                base_status, custom_status = self.patient_status_values(status)
                self.ensure_patient_status(db, status)
                if existing:
                    patient_id = existing["id"]
                    if existing["last_visit"] and item["date"] <= existing["last_visit"]:
                        unchanged += 1
                        continue
                    db.execute("UPDATE patients SET status=?, updated_at=CURRENT_TIMESTAMP WHERE id=?", (base_status, patient_id))
                    db.execute("UPDATE patient_followup SET last_visit=?, custom_status=? WHERE patient_id=?", (item["date"], custom_status, patient_id))
                    db.execute("INSERT INTO patient_events (patient_id, event_type, description) VALUES (?, 'Importação', ?)", (patient_id, f"Última consulta atualizada pela planilha: {item['source']}"))
                    existing["last_visit"] = item["date"]
                    updated += 1
                else:
                    patient_id = db.execute("INSERT INTO patients (name, status) VALUES (?, ?)", (item["patient"], base_status)).lastrowid
                    db.execute("INSERT INTO patient_assignments (patient_id, professional_id, is_primary) VALUES (?, ?, 1)", (patient_id, professional_id))
                    db.execute("INSERT INTO patient_followup (patient_id, last_visit, custom_status) VALUES (?, ?, ?)", (patient_id, item["date"], custom_status))
                    db.execute("INSERT INTO patient_events (patient_id, event_type, description) VALUES (?, 'Importação', ?)", (patient_id, f"Importado da planilha: {item['source']}"))
                    assignments[key] = {"id": patient_id, "last_visit": item["date"]}
                    imported += 1
            db.execute("DELETE FROM import_batches WHERE token_hash=?", (hashlib.sha256(token.encode()).hexdigest(),))
        self.send_json({"imported": imported, "updated": updated, "unchanged": unchanged, "total": len(items)})

    def update_professional_photo(self, professional_id: int, payload: dict) -> None:
        image = str(payload.get("image") or "")
        match = re.fullmatch(r"data:(image/(?:jpeg|png|webp));base64,([A-Za-z0-9+/=]+)", image)
        if not match:
            return self.send_json({"error": "Envie uma imagem JPG, PNG ou WEBP válida."}, HTTPStatus.BAD_REQUEST)
        try:
            binary = base64.b64decode(match.group(2), validate=True)
        except ValueError:
            return self.send_json({"error": "Não foi possível ler a imagem."}, HTTPStatus.BAD_REQUEST)
        if len(binary) > 2 * 1024 * 1024:
            return self.send_json({"error": "A foto deve ter no máximo 2 MB."}, HTTPStatus.BAD_REQUEST)
        with connect() as db:
            if not db.execute("SELECT id FROM professionals WHERE id=?", (professional_id,)).fetchone():
                return self.send_json({"error": "Profissional não encontrado"}, HTTPStatus.NOT_FOUND)
            db.execute("UPDATE professionals SET photo_data=?, photo_mime=? WHERE id=?", (match.group(2), match.group(1), professional_id))
        self.send_json({"updated": True, "photo_url": f"/api/professionals/{professional_id}/photo"})

    @staticmethod
    def replace_crm_user_channels(
        db, user_id: int, access_role: str, crm_access_level: str, payload: dict
    ) -> None:
        if access_role != "crc":
            db.execute("DELETE FROM crm_user_channels WHERE user_id=?", (user_id,))
            db.execute("DELETE FROM crm_user_features WHERE user_id=?", (user_id,))
            db.execute(
                "UPDATE users SET crm_access_level='attendant',crm_channel_scope_enabled=0,crm_feature_scope_enabled=0,crm_manage_automation=0,crm_operational_agent=0 WHERE id=?",
                (user_id,),
            )
            return
        db.execute(
            "UPDATE users SET crm_access_level=?,crm_operational_agent=? WHERE id=?",
            (crm_access_level, 0 if crm_access_level == "admin" else 1, user_id),
        )
        if crm_access_level == "admin":
            db.execute("DELETE FROM crm_user_channels WHERE user_id=?", (user_id,))
            db.execute("DELETE FROM crm_user_features WHERE user_id=?", (user_id,))
            db.execute(
                "UPDATE users SET crm_channel_scope_enabled=0,crm_feature_scope_enabled=0,crm_manage_automation=1 WHERE id=?",
                (user_id,),
            )
            return
        if "crm_channel_ids" not in payload:
            return
        values = payload.get("crm_channel_ids") or []
        channel_ids = list(dict.fromkeys(int(value) for value in values if str(value).isdigit()))
        if channel_ids:
            placeholders = ",".join("?" for _ in channel_ids)
            valid_count = db.execute(f"SELECT COUNT(*) FROM crm_channels WHERE id IN ({placeholders})", channel_ids).fetchone()[0]
            if valid_count != len(channel_ids):
                raise ValueError("Um dos canais selecionados não existe")
        db.execute("DELETE FROM crm_user_channels WHERE user_id=?", (user_id,))
        can_manage = 1 if payload.get("crm_can_manage_automation") else 0
        db.executemany("""INSERT INTO crm_user_channels(user_id,channel_id,can_reply,can_manage_automation)
                           VALUES(?,?,1,?)""", [(user_id, channel_id, can_manage) for channel_id in channel_ids])
        db.execute(
            "UPDATE users SET crm_channel_scope_enabled=1,crm_manage_automation=? WHERE id=?",
            (can_manage, user_id),
        )

    def create_admin_professional(self, payload: dict) -> None:
        name = str(payload.get("name") or "").strip()
        email = str(payload.get("email") or "").strip().lower()
        role = str(payload.get("role") or "").strip()
        access_role = str(payload.get("access_role") or "professional").strip()
        crm_access_level = str(payload.get("crm_access_level") or "attendant").strip().lower()
        service_sector = str(payload.get("service_sector") or ("CRC" if access_role == "crc" else "")).strip()
        temporary_password = str(payload.get("temporary_password") or "")
        if not name or not email:
            return self.send_json({"error": "Nome e e-mail são obrigatórios"}, HTTPStatus.BAD_REQUEST)
        if access_role not in {"owner", "professional", "admin", "crc", "asb"}:
            return self.send_json({"error": "Nível de acesso inválido"}, HTTPStatus.BAD_REQUEST)
        if crm_access_level not in {"attendant", "admin"}:
            return self.send_json({"error": "Grau de acesso do CRM inválido"}, HTTPStatus.BAD_REQUEST)
        if service_sector not in {"", "CRC", "Recepção"}:
            return self.send_json({"error": "Setor de atendimento inválido"}, HTTPStatus.BAD_REQUEST)
        if access_role != "crc":
            service_sector = ""
            crm_access_level = "attendant"
        linked_professional_id = int(payload.get("linked_professional_id") or 0) or None
        if access_role == "asb" and not linked_professional_id:
            return self.send_json({"error": "Selecione o dentista responsável pela ASB."}, HTTPStatus.BAD_REQUEST)
        if len(temporary_password) < 10:
            return self.send_json({"error": "Defina uma senha temporária com ao menos 10 caracteres"}, HTTPStatus.BAD_REQUEST)
        try:
            with connect() as db:
                email_in_use = db.execute("SELECT name FROM users WHERE lower(email) = ?", (email,)).fetchone()
                if email_in_use:
                    return self.send_json({"error": f"O e-mail {email} já está vinculado ao acesso de {email_in_use['name']}. Use outro e-mail ou edite esse acesso existente."}, HTTPStatus.CONFLICT)
                professional_id = db.execute(
                    "INSERT INTO professionals (name, role, is_owner, active) VALUES (?, ?, 0, 1)",
                    (name, role),
                ).lastrowid
                salt = secrets.token_hex(16)
                user_id = db.execute(
                    "INSERT INTO users (professional_id, name, email, access_role, crm_access_level, linked_professional_id, active, password_hash, password_salt, must_change_password, service_sector) VALUES (?, ?, ?, ?, ?, ?, 1, ?, ?, 1, ?)",
                    (professional_id, name, email, access_role, crm_access_level, linked_professional_id, self.password_digest(temporary_password, salt), salt, service_sector),
                ).lastrowid
                if access_role == "asb" and linked_professional_id:
                    db.execute(
                        "INSERT OR IGNORE INTO asb_professional_links (user_id, professional_id) VALUES (?, ?)",
                        (user_id, linked_professional_id),
                    )
                self.replace_crm_user_channels(db, user_id, access_role, crm_access_level, payload)
                specialty_id = int(payload.get("specialty_id") or 0)
                office_id = int(payload.get("office_id") or 0)
                if specialty_id:
                    db.execute("INSERT INTO professional_specialties (professional_id, specialty_id, is_primary) VALUES (?, ?, 1)", (professional_id, specialty_id))
                if office_id:
                    db.execute("INSERT INTO professional_offices (professional_id, office_id, is_responsible) VALUES (?, ?, ?)", (professional_id, office_id, 1 if payload.get("office_responsible") else 0))
        except (ValueError, TypeError):
            return self.send_json({"error": "Especialidade ou consultório inválido"}, HTTPStatus.BAD_REQUEST)
        except IntegrityError:
            return self.send_json({"error": "Já existe um acesso com esse e-mail"}, HTTPStatus.CONFLICT)
        self.send_json({"created": True, "id": professional_id}, HTTPStatus.CREATED)

    def update_admin_professional(self, professional_id: int, payload: dict) -> None:
        name = str(payload.get("name") or "").strip()
        email = str(payload.get("email") or "").strip().lower()
        access_role = str(payload.get("access_role") or "professional").strip()
        crm_access_level = str(payload.get("crm_access_level") or "attendant").strip().lower()
        service_sector = str(payload.get("service_sector") or ("CRC" if access_role == "crc" else "")).strip()
        if not name or not email or access_role not in {"owner", "professional", "admin", "crc", "asb"}:
            return self.send_json({"error": "Dados do acesso inválidos"}, HTTPStatus.BAD_REQUEST)
        if crm_access_level not in {"attendant", "admin"}:
            return self.send_json({"error": "Grau de acesso do CRM inválido"}, HTTPStatus.BAD_REQUEST)
        if service_sector not in {"", "CRC", "Recepção"}:
            return self.send_json({"error": "Setor de atendimento inválido"}, HTTPStatus.BAD_REQUEST)
        if access_role != "crc":
            service_sector = ""
            crm_access_level = "attendant"
        linked_professional_id = int(payload.get("linked_professional_id") or 0) or None
        if access_role == "asb" and not linked_professional_id:
            return self.send_json({"error": "Selecione o dentista responsável pela ASB."}, HTTPStatus.BAD_REQUEST)
        try:
            with connect() as db:
                exists = db.execute("SELECT id, is_owner FROM professionals WHERE id = ?", (professional_id,)).fetchone()
                if not exists:
                    return self.send_json({"error": "Profissional não encontrado"}, HTTPStatus.NOT_FOUND)
                email_in_use = db.execute("SELECT professional_id, name FROM users WHERE lower(email) = ?", (email,)).fetchone()
                if email_in_use and email_in_use["professional_id"] != professional_id:
                    return self.send_json({"error": f"O e-mail {email} já está vinculado ao acesso de {email_in_use['name']}."}, HTTPStatus.CONFLICT)
                active = 1 if exists["is_owner"] else (1 if payload.get("active", True) else 0)
                db.execute("UPDATE professionals SET name=?, role=?, active=? WHERE id=?", (name, str(payload.get("role") or "").strip(), active, professional_id))
                user = db.execute("SELECT id FROM users WHERE professional_id = ?", (professional_id,)).fetchone()
                if user:
                    db.execute("UPDATE users SET name=?, email=?, access_role=?, crm_access_level=?, linked_professional_id=?, active=?, service_sector=? WHERE professional_id=?", (name, email, access_role, crm_access_level, linked_professional_id, active, service_sector, professional_id))
                    user_id = user["id"]
                else:
                    user_id = db.execute("INSERT INTO users (professional_id, name, email, access_role, crm_access_level, linked_professional_id, active, service_sector) VALUES (?, ?, ?, ?, ?, ?, ?, ?)", (professional_id, name, email, access_role, crm_access_level, linked_professional_id, active, service_sector)).lastrowid
                db.execute("DELETE FROM asb_professional_links WHERE user_id=?", (user_id,))
                if access_role == "asb" and linked_professional_id:
                    db.execute(
                        "INSERT INTO asb_professional_links (user_id, professional_id) VALUES (?, ?)",
                        (user_id, linked_professional_id),
                    )
                self.replace_crm_user_channels(db, user_id, access_role, crm_access_level, payload)
                specialty_id = int(payload.get("specialty_id") or 0)
                office_id = int(payload.get("office_id") or 0)
                db.execute("DELETE FROM professional_specialties WHERE professional_id = ?", (professional_id,))
                if specialty_id:
                    db.execute("INSERT INTO professional_specialties (professional_id, specialty_id, is_primary) VALUES (?, ?, 1)", (professional_id, specialty_id))
                db.execute("DELETE FROM professional_offices WHERE professional_id = ?", (professional_id,))
                if office_id:
                    db.execute("INSERT INTO professional_offices (professional_id, office_id, is_responsible) VALUES (?, ?, ?)", (professional_id, office_id, 1 if payload.get("office_responsible") else 0))
        except (ValueError, TypeError, IntegrityError):
            return self.send_json({"error": "Não foi possível atualizar. Verifique e-mail, especialidade e consultório."}, HTTPStatus.CONFLICT)
        self.send_json({"updated": True, "id": professional_id})

    def reset_professional_password(self, professional_id: int, payload: dict) -> None:
        temporary_password = str(payload.get("temporary_password") or "")
        if len(temporary_password) < 10:
            return self.send_json({"error": "Defina uma senha temporária com ao menos 10 caracteres"}, HTTPStatus.BAD_REQUEST)
        with connect() as db:
            user = db.execute("SELECT id FROM users WHERE professional_id = ?", (professional_id,)).fetchone()
            if not user:
                return self.send_json({"error": "Este profissional ainda não possui acesso cadastrado"}, HTTPStatus.NOT_FOUND)
            salt = secrets.token_hex(16)
            db.execute("UPDATE users SET password_hash=?, password_salt=?, must_change_password=1 WHERE id=?", (self.password_digest(temporary_password, salt), salt, user["id"]))
            db.execute("DELETE FROM auth_sessions WHERE user_id=?", (user["id"],))
            self.record_security_event(db, "admin_password_reset", self.request_ip(), self.authenticated_user["id"], f"Profissional {professional_id}")
        self.send_json({"reset": True})

    def create_admin_specialty(self, payload: dict) -> None:
        name = str(payload.get("name") or "").strip()
        if not name:
            return self.send_json({"error": "Informe o nome da especialidade"}, HTTPStatus.BAD_REQUEST)
        try:
            with connect() as db:
                item_id = db.execute("INSERT INTO specialties (code, name) VALUES (?, ?)", (self.normalized_code(name), name)).lastrowid
        except IntegrityError:
            return self.send_json({"error": "Essa especialidade já está cadastrada"}, HTTPStatus.CONFLICT)
        self.send_json({"created": True, "id": item_id}, HTTPStatus.CREATED)

    def create_admin_office(self, payload: dict) -> None:
        name = str(payload.get("name") or "").strip()
        if not name:
            return self.send_json({"error": "Informe o nome do consultório"}, HTTPStatus.BAD_REQUEST)
        try:
            with connect() as db:
                item_id = db.execute("INSERT INTO offices (code, name) VALUES (?, ?)", (self.normalized_code(name), name)).lastrowid
        except IntegrityError:
            return self.send_json({"error": "Esse consultório já está cadastrado"}, HTTPStatus.CONFLICT)
        self.send_json({"created": True, "id": item_id}, HTTPStatus.CREATED)

    def update_admin_structure(self, table: str, item_id: int, payload: dict) -> None:
        name = str(payload.get("name") or "").strip()
        active = 1 if payload.get("active", True) else 0
        if not name:
            return self.send_json({"error": "Informe um nome"}, HTTPStatus.BAD_REQUEST)
        try:
            with connect() as db:
                exists = db.execute(f"SELECT id FROM {table} WHERE id=?", (item_id,)).fetchone()
                if not exists:
                    return self.send_json({"error": "Item não encontrado"}, HTTPStatus.NOT_FOUND)
                db.execute(f"UPDATE {table} SET name=?, code=?, active=? WHERE id=?", (name, self.normalized_code(name), active, item_id))
        except IntegrityError:
            return self.send_json({"error": "Já existe um item com esse nome"}, HTTPStatus.CONFLICT)
        self.send_json({"updated": True, "id": item_id})

    def delete_admin_structure(self, table: str, link_table: str, link_column: str, item_id: int) -> None:
        with connect() as db:
            linked = db.execute(f"SELECT COUNT(*) FROM {link_table} WHERE {link_column}=?", (item_id,)).fetchone()[0]
            if linked:
                return self.send_json({"error": f"Não é possível excluir: há {linked} profissional(is) vinculado(s). Remova ou altere os vínculos antes."}, HTTPStatus.CONFLICT)
            exists = db.execute(f"SELECT id FROM {table} WHERE id=?", (item_id,)).fetchone()
            if not exists:
                return self.send_json({"error": "Item não encontrado"}, HTTPStatus.NOT_FOUND)
            db.execute(f"DELETE FROM {table} WHERE id=?", (item_id,))
        self.send_json({"deleted": True, "id": item_id})

    def import_admin_patients(self, payload: dict) -> None:
        rows = payload.get("rows")
        dry_run = bool(payload.get("dry_run", True))
        if not isinstance(rows, list) or not rows:
            return self.send_json({"error": "Nenhum registro foi enviado"}, HTTPStatus.BAD_REQUEST)
        if len(rows) > 5000:
            return self.send_json({"error": "O limite é de 5.000 pacientes por importação"}, HTTPStatus.BAD_REQUEST)
        created = duplicates = 0
        errors = []
        with connect() as db:
            owner_id = db.execute("SELECT id FROM professionals WHERE is_owner = 1 ORDER BY id LIMIT 1").fetchone()[0]
            next_external = db.execute("SELECT COALESCE(MAX(CAST(external_id AS INTEGER)), 0) + 1 FROM patients").fetchone()[0]
            for index, raw in enumerate(rows, start=2):
                item = raw if isinstance(raw, dict) else {}
                name = str(item.get("name") or item.get("nome") or "").strip()
                last_visit = str(item.get("last_visit") or item.get("ultima_consulta") or "").strip()
                status = str(item.get("status") or "Consulta").strip().title()
                if not name or not re.fullmatch(r"\d{4}-\d{2}-\d{2}", last_visit):
                    errors.append({"row": index, "message": "Nome ou última consulta inválida"})
                    continue
                if status not in {"Consulta", "Controle", "Tratamento", "Inativo"}:
                    errors.append({"row": index, "message": "Status inválido"})
                    continue
                duplicate = db.execute("""
                    SELECT p.id FROM patients p JOIN patient_followup f ON f.patient_id=p.id
                    WHERE lower(trim(p.name))=lower(trim(?)) AND f.last_visit=?
                """, (name, last_visit)).fetchone()
                if duplicate:
                    duplicates += 1
                    continue
                professional_id = owner_id
                professional_email = str(item.get("professional_email") or item.get("email_profissional") or "").strip().lower()
                if professional_email:
                    target = db.execute("SELECT professional_id FROM users WHERE lower(email)=? AND active=1", (professional_email,)).fetchone()
                    if not target:
                        errors.append({"row": index, "message": "Profissional não encontrado"})
                        continue
                    professional_id = target[0]
                created += 1
                if dry_run:
                    continue
                patient_id = db.execute(
                    "INSERT INTO patients (external_id, name, phone, reference, status, notes) VALUES (?, ?, ?, ?, ?, ?)",
                    (next_external, name, item.get("phone") or item.get("telefone") or None, item.get("reference") or item.get("referencia") or None, status, item.get("notes") or item.get("observacoes") or None),
                ).lastrowid
                next_external += 1
                db.execute("INSERT INTO patient_assignments (patient_id, professional_id, is_primary) VALUES (?, ?, 1)", (patient_id, professional_id))
                next_appointment = item.get("next_appointment") or item.get("proxima_consulta") or None
                db.execute("INSERT INTO patient_followup (patient_id, last_visit, next_appointment, next_appointment_type, next_action) VALUES (?, ?, ?, ?, ?)", (patient_id, last_visit, next_appointment, "Agendado" if next_appointment else None, item.get("next_action") or item.get("proxima_acao") or None))
                db.execute("INSERT INTO patient_events (patient_id, event_type, description) VALUES (?, 'Importação', 'Paciente importado pelo Painel Administrativo')", (patient_id,))
        self.send_json({"dry_run": dry_run, "received": len(rows), "valid": created, "duplicates": duplicates, "errors": errors, "imported": 0 if dry_run else created})

    def get_dashboard(self) -> None:
        scope = ""
        params = []
        need_contact_condition = """p.status != 'Inativo' AND (f.next_appointment IS NULL OR date(f.next_appointment) < date('now', 'localtime'))
                              AND julianday('now', 'localtime') - julianday(CASE WHEN f.next_appointment IS NOT NULL AND date(f.next_appointment) < date('now', 'localtime') THEN f.next_appointment ELSE f.last_visit END) >= 60
                              AND (f.resolved_at IS NULL OR date(f.resolved_at) != date('now', 'localtime'))"""
        if self.authenticated_user["access_role"] not in {"admin", "crc"}:
            scope = " WHERE EXISTS (SELECT 1 FROM patient_assignments own_pa WHERE own_pa.patient_id = p.id AND own_pa.professional_id = ? AND own_pa.journey_status = 'Ativo')"
            params.append(self.authenticated_user["professional_id"])
        elif self.authenticated_user["access_role"] == "crc":
            scope = " WHERE f.crc_status IN ('Aguardando contato', 'Em atendimento', 'Jornada compartilhada')"
            need_contact_condition = """p.status != 'Inativo' AND (f.next_appointment IS NULL OR date(f.next_appointment) < date('now', 'localtime'))
                                      AND julianday('now', 'localtime') - julianday(CASE WHEN f.next_appointment IS NOT NULL AND date(f.next_appointment) < date('now', 'localtime') THEN f.next_appointment ELSE f.last_visit END) >= 60"""
        with connect() as db:
            row = db.execute("""
                SELECT
                    COUNT(*) AS total,
                    SUM(CASE WHEN p.status != 'Inativo' AND (f.next_appointment IS NULL OR date(f.next_appointment) < date('now', 'localtime')) THEN 1 ELSE 0 END) AS without_schedule,
                    SUM(CASE WHEN EXISTS (SELECT 1 FROM procedures px WHERE px.patient_id = p.id AND px.stage != 'Concluído') THEN 1 ELSE 0 END) AS open_treatments,
                    SUM(CASE WHEN p.status = 'Inativo' THEN 1 ELSE 0 END) AS inactive,
                    SUM(CASE WHEN """ + need_contact_condition + """ THEN 1 ELSE 0 END) AS need_contact,
                    SUM(CASE WHEN date(f.resolved_at) = date('now', 'localtime') THEN 1 ELSE 0 END) AS resolved_today
                    ,COALESCE(SUM((SELECT SUM(GREATEST(px.value_cents - px.discount_cents, 0)) FROM procedures px WHERE px.patient_id = p.id AND px.stage != 'Concluído')), 0) AS potential_value_cents
                FROM patients p
                JOIN patient_followup f ON f.patient_id = p.id
                JOIN patient_assignments pa ON pa.patient_id = p.id AND pa.is_primary = 1
            """ + scope, params).fetchone()
        self.send_json(dict(row))

    def get_daily_log(self, query: dict) -> None:
        try:
            days = min(365, max(7, int(query.get("days", ["30"])[0])))
        except ValueError:
            days = 30
        assignment_join = ""
        professional_filter = ""
        params = []
        if self.authenticated_user["access_role"] not in {"admin", "crc"}:
            assignment_join = " JOIN patient_assignments pa ON pa.patient_id = dr.patient_id AND pa.is_primary = 1"
            professional_filter = " AND pa.professional_id = ?"
            portfolio_id = (
                self.authenticated_user.get("linked_professional_id")
                if self.authenticated_user["access_role"] == "asb"
                else self.authenticated_user["professional_id"]
            )
            params.append(portfolio_id)
        with connect() as db:
            summary = db.execute("""
                SELECT
                    SUM(CASE WHEN resolution_date = CAST(date('now', 'localtime') AS TEXT) AND reopened_at IS NULL THEN 1 ELSE 0 END) AS today,
                    SUM(CASE WHEN resolution_date = CAST(date('now', 'localtime', '-1 day') AS TEXT) AND reopened_at IS NULL THEN 1 ELSE 0 END) AS yesterday,
                    SUM(CASE WHEN resolution_date >= CAST(date('now', 'localtime', '-6 days') AS TEXT) AND reopened_at IS NULL THEN 1 ELSE 0 END) AS last_7_days
                FROM daily_resolutions dr
            """ + assignment_join + " WHERE 1 = 1" + professional_filter, params).fetchone()
            rows = db.execute("""
                SELECT dr.resolution_date,
                       COUNT(*) AS total,
                       GROUP_CONCAT(CAST(p.id AS TEXT) || '##' || p.name || '##' || strftime('%H:%M', dr.resolved_at), '||') AS patient_names
                FROM daily_resolutions dr
                JOIN patients p ON p.id = dr.patient_id
            """ + assignment_join + """
                WHERE dr.reopened_at IS NULL
                  AND dr.resolution_date >= CAST(date('now', 'localtime', ?) AS TEXT)
            """ + professional_filter + """
                GROUP BY dr.resolution_date
                ORDER BY dr.resolution_date DESC
            """, [f"-{days - 1} days", *params]).fetchall()
        records = []
        for row in rows:
            records.append({
                "date": row["resolution_date"],
                "total": row["total"],
                "patients": [
                    {"id": parts[0], "name": parts[1] if len(parts) > 1 else "Paciente", "time": parts[2] if len(parts) > 2 else "—"}
                    for parts in (item.split("##", 2) for item in (row["patient_names"].split("||") if row["patient_names"] else []))
                ],
            })
        self.send_json({
            "summary": {"today": summary["today"] or 0, "yesterday": summary["yesterday"] or 0, "last_7_days": summary["last_7_days"] or 0},
            "records": records,
        })

    def get_procedure_catalog(self) -> None:
        with connect() as db:
            rows = db.execute(
                "SELECT id, name, default_value_cents, active, created_at, updated_at FROM procedure_catalog ORDER BY active DESC, name COLLATE NOCASE"
            ).fetchall()
        self.send_json({"items": [dict(row) for row in rows]})

    def get_specialties(self) -> None:
        with connect() as db:
            rows = db.execute(
                "SELECT id, code, name, active FROM specialties ORDER BY active DESC, name COLLATE NOCASE"
            ).fetchall()
        self.send_json({"items": [dict(row) for row in rows]})

    def get_action_templates(self) -> None:
        with connect() as db:
            rows = db.execute("SELECT id, description FROM action_templates ORDER BY description COLLATE NOCASE").fetchall()
        self.send_json({"items": [dict(row) for row in rows]})

    def validate_catalog_payload(self, payload: dict) -> tuple[str, int, int]:
        name = str(payload.get("name", "")).strip()
        if not name:
            raise ValueError("Informe o nome do procedimento")
        try:
            value = max(0, int(payload.get("default_value_cents") or 0))
        except (TypeError, ValueError):
            raise ValueError("Valor padrão inválido")
        return name, value, 1 if payload.get("active", True) else 0

    def create_catalog_procedure(self, payload: dict) -> None:
        try:
            name, value, active = self.validate_catalog_payload(payload)
            with connect() as db:
                procedure_id = db.execute(
                    "INSERT INTO procedure_catalog (name, default_value_cents, active) VALUES (?, ?, ?)",
                    (name, value, active),
                ).lastrowid
                row = db.execute("SELECT id, name, default_value_cents, active FROM procedure_catalog WHERE id = ?", (procedure_id,)).fetchone()
        except ValueError as error:
            return self.send_json({"error": str(error)}, HTTPStatus.BAD_REQUEST)
        except IntegrityError:
            return self.send_json({"error": "Já existe um procedimento com esse nome"}, HTTPStatus.CONFLICT)
        self.send_json(dict(row), HTTPStatus.CREATED)

    def update_catalog_procedure(self, procedure_id: int, payload: dict) -> None:
        try:
            name, value, active = self.validate_catalog_payload(payload)
            with connect() as db:
                exists = db.execute("SELECT id FROM procedure_catalog WHERE id = ?", (procedure_id,)).fetchone()
                if not exists:
                    return self.send_json({"error": "Procedimento não encontrado"}, HTTPStatus.NOT_FOUND)
                db.execute(
                    "UPDATE procedure_catalog SET name = ?, default_value_cents = ?, active = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                    (name, value, active, procedure_id),
                )
                row = db.execute("SELECT id, name, default_value_cents, active FROM procedure_catalog WHERE id = ?", (procedure_id,)).fetchone()
        except ValueError as error:
            return self.send_json({"error": str(error)}, HTTPStatus.BAD_REQUEST)
        except IntegrityError:
            return self.send_json({"error": "Já existe um procedimento com esse nome"}, HTTPStatus.CONFLICT)
        self.send_json(dict(row))

    def get_filters(self) -> None:
        with connect() as db:
            if self.authenticated_user["access_role"] not in {"admin", "crc"}:
                rows = db.execute("SELECT DISTINCT substr(f.last_visit, 1, 7) FROM patient_followup f JOIN patient_assignments pa ON pa.patient_id=f.patient_id WHERE pa.professional_id=? AND f.last_visit IS NOT NULL ORDER BY 1 DESC", (self.authenticated_user["professional_id"],)).fetchall()
            else:
                rows = db.execute("SELECT DISTINCT substr(last_visit, 1, 7) FROM patient_followup WHERE last_visit IS NOT NULL ORDER BY 1 DESC").fetchall()
            months = [row[0] for row in rows if row[0]]
            statuses = ["Consulta", "Controle", "Tratamento", "Inativo"]
            professionals = [dict(row) for row in db.execute("SELECT id, name, CASE WHEN photo_data IS NOT NULL THEN 1 ELSE 0 END AS has_photo FROM professionals WHERE active=1 ORDER BY name COLLATE NOCASE").fetchall()] if self.authenticated_user["access_role"] in {"admin", "crc"} else []
        self.send_json({"months": months, "statuses": statuses, "professionals": professionals})

    def get_crm_contacts(self) -> None:
        if self.authenticated_user["access_role"] != "crc":
            return self.send_json({"error": "Acesso exclusivo da Central CRC."}, HTTPStatus.FORBIDDEN)
        if not self.require_crm_feature("contacts"):
            return
        scope_sql, scope_params = self.crm_channel_scope_clause("ch")
        with connect() as db:
            crm_rows = db.execute("""SELECT ct.id,ct.patient_id,ct.name,ct.phone,ct.is_internal,ct.profile_picture_url,
                                      GROUP_CONCAT(DISTINCT ch.display_name) AS channels
                                      FROM crm_contacts ct
                                      LEFT JOIN crm_conversations cv ON cv.contact_id=ct.id
                                      LEFT JOIN crm_channels ch ON ch.id=cv.channel_id AND {scope_sql}
                                      WHERE COALESCE(ct.is_internal,0)=1
                                         OR ch.id IS NOT NULL
                                      GROUP BY ct.id ORDER BY ct.name COLLATE NOCASE""".format(scope_sql=scope_sql), scope_params).fetchall()
        contacts, seen_phones = [], set()
        channels_by_phone = {self.crm_phone(row["phone"]): [item.strip() for item in str(row["channels"] or "").split(",") if item.strip()] for row in crm_rows}
        internal_by_phone = {self.crm_phone(row["phone"]): bool(row["is_internal"]) for row in crm_rows}
        # A rota de foto faz a consulta sob demanda na Evolution quando ainda
        # não há cache. Assim o navegador nunca recebe nem armazena a URL
        # externa do WhatsApp e o avatar pode aparecer assim que existir.
        photo_by_phone = {self.crm_phone(row["phone"]): f"/api/crm/contacts/{row['id']}/profile-photo" for row in crm_rows if self.crm_phone(row["phone"])}
        for row in crm_rows:
            digits = self.crm_phone(row["phone"])
            if not digits or digits in seen_phones:
                continue
            seen_phones.add(digits)
            local_number = f"({digits[:2]}) {digits[2:7]}-{digits[7:]}" if len(digits) == 11 else f"({digits[:2]}) {digits[2:6]}-{digits[6:]}"
            contacts.append({"id": f"crm-{row['id']}", "name": row["name"], "phone": local_number,
                             "professional": "Novo contato", "stage": "Novo contato", "is_new": True,
                             "is_internal": bool(row["is_internal"]),
                             "channels": channels_by_phone.get(digits, []),
                             "profile_picture_url": photo_by_phone.get(digits, "")})
        contacts.sort(key=lambda item: (not item["is_new"], item["name"].casefold()))
        self.send_json({"items": contacts, "total": len(contacts)})

    def cleanup_crm_imported_contacts(self) -> None:
        """Remove contacts imported only from WhatsApp without operational value.

        A contact is preserved whenever it belongs to a clinical patient, is an
        internal team member, or has at least one CRM conversation. This keeps
        all history intact while removing the passive Evolution address book.
        """
        if not self.require_crc_access():
            return
        if not self.require_crm_feature("contacts"):
            return
        with connect() as db:
            removable = db.execute("""SELECT COUNT(*) AS total
                                       FROM crm_contacts ct
                                      WHERE COALESCE(ct.is_internal,0)=0
                                        AND NOT EXISTS (
                                            SELECT 1 FROM crm_conversations cv
                                             WHERE cv.contact_id=ct.id
                                        )""").fetchone()
            removed = db.execute("""DELETE FROM crm_contacts ct
                                    WHERE COALESCE(ct.is_internal,0)=0
                                      AND NOT EXISTS (
                                          SELECT 1 FROM crm_conversations cv
                                           WHERE cv.contact_id=ct.id
                                      )""").rowcount
        self.send_json({
            "removed": int(removed or 0),
            "eligible": int(removable["total"] or 0) if removable else 0,
            "preserved": "Pacientes clínicos, contatos internos e conversas com histórico foram preservados.",
        })

    @staticmethod
    def crm_phone(value) -> str:
        digits = re.sub(r"\D", "", str(value or ""))
        if digits.startswith("55") and len(digits) in {12, 13}:
            digits = digits[2:]
        return digits if len(digits) in {10, 11} else ""

    # CRM_PATIENT_OWNERSHIP_LOCK_V16
    @staticmethod
    def crm_contact_active_owner(db, contact_id: int):
        """Return the first active owner for a patient across every channel."""
        return db.execute(
            """SELECT cv.id AS conversation_id,cv.assigned_user_id,u.name AS assigned_to
                 FROM crm_conversations cv
                 JOIN users u ON u.id=cv.assigned_user_id
                WHERE cv.contact_id=? AND cv.status<>'Resolvida'
                  AND cv.assigned_user_id IS NOT NULL
                ORDER BY COALESCE(cv.assigned_at,cv.updated_at),cv.id
                LIMIT 1""",
            (contact_id,),
        ).fetchone()

    def reject_crm_contact_owner_conflict(self, owner) -> None:
        self.send_json({
            "error": f"Este paciente já está em atendimento com {owner['assigned_to'] or 'outra atendente'}. Solicite a transferência antes de acessar.",
            "code": "PATIENT_ASSIGNED_TO_ANOTHER_USER",
            "conversation_id": int(owner["conversation_id"]),
            "assigned_to": owner["assigned_to"],
        }, HTTPStatus.CONFLICT)

    def require_crc_access(self) -> bool:
        if self.authenticated_user and self.authenticated_user["access_role"] == "crc":
            return True
        self.send_json({"error": "Acesso exclusivo da Central CRC."}, HTTPStatus.FORBIDDEN)
        return False

    def crm_feature_allowed(self, feature_key: str, user_id: int | None = None) -> bool:
        if feature_key not in CRM_FEATURE_KEYS:
            return False
        user_id = int(user_id or self.authenticated_user["id"])
        with connect() as db:
            user = db.execute("SELECT access_role,crm_access_level,crm_feature_scope_enabled FROM users WHERE id=?", (user_id,)).fetchone()
            if not user:
                return False
            if user["access_role"] in {"admin", "owner"} or user["crm_access_level"] == "admin" or not user["crm_feature_scope_enabled"]:
                return True
            return bool(db.execute("SELECT 1 FROM crm_user_features WHERE user_id=? AND feature_key=?", (user_id, feature_key)).fetchone())

    def crm_any_feature_allowed(self, feature_keys, user_id: int | None = None) -> bool:
        keys = tuple(dict.fromkeys(key for key in feature_keys if key in CRM_FEATURE_KEYS))
        if not keys:
            return False
        user_id = int(user_id or self.authenticated_user["id"])
        with connect() as db:
            user = db.execute(
                "SELECT access_role,crm_access_level,crm_feature_scope_enabled FROM users WHERE id=?",
                (user_id,),
            ).fetchone()
            if not user:
                return False
            if user["access_role"] in {"admin", "owner"} or user["crm_access_level"] == "admin" or not user["crm_feature_scope_enabled"]:
                return True
            placeholders = ",".join("?" for _ in keys)
            return bool(db.execute(
                f"SELECT 1 FROM crm_user_features WHERE user_id=? AND feature_key IN ({placeholders}) LIMIT 1",
                (user_id, *keys),
            ).fetchone())

    def get_crm_permissions(self) -> None:
        if not self.require_crc_access():
            return
        user_id = int(self.authenticated_user["id"])
        with connect() as db:
            user = db.execute("SELECT crm_access_level,crm_feature_scope_enabled FROM users WHERE id=?", (user_id,)).fetchone()
            crm_access_level = (user["crm_access_level"] if user else "attendant") or "attendant"
            restricted = bool(user and user["crm_feature_scope_enabled"] and crm_access_level != "admin")
            if restricted:
                selected = {row["feature_key"] for row in db.execute(
                    "SELECT feature_key FROM crm_user_features WHERE user_id=?", (user_id,)
                ).fetchall()}
                allowed = [key for key in CRM_FEATURE_KEYS if key in selected]
            else:
                allowed = list(CRM_FEATURE_KEYS)
        self.send_json({
            "allowed_features": allowed,
            "feature_scope_enabled": restricted,
            "crm_access_level": crm_access_level,
            "can_manage_crm": self.can_manage_crm(self.authenticated_user),
        })

    def require_crm_feature(self, feature_key: str) -> bool:
        if self.crm_feature_allowed(feature_key):
            return True
        self.send_json({"error": "Esta tela nÃ£o estÃ¡ liberada para o seu acesso."}, HTTPStatus.FORBIDDEN)
        return False

    def require_crm_any_feature(self, feature_keys) -> bool:
        if self.crm_any_feature_allowed(feature_keys):
            return True
        self.send_json({"error": "Esta funcionalidade não está liberada para o seu acesso."}, HTTPStatus.FORBIDDEN)
        return False

    @staticmethod
    def crm_conversation_view_feature(view: str) -> str:
        if view == "queue":
            return "queue"
        if view == "operational":
            return "funnel"
        return "inbox"

    def crm_channel_id_scope_clause(self, channel_expression: str, capability: str | None = None) -> tuple[str, list]:
        user_id = int(self.authenticated_user["id"])
        permission = ""
        if capability == "reply":
            permission = " AND cuc.can_reply=1"
        elif capability == "automation":
            permission = " AND cuc.can_manage_automation=1"
        return (
            f"(COALESCE((SELECT crm_channel_scope_enabled FROM users WHERE id=?),0)=0 "
            f"OR EXISTS (SELECT 1 FROM crm_user_channels cuc WHERE cuc.user_id=? "
            f"AND cuc.channel_id={channel_expression}{permission}))",
            [user_id, user_id],
        )

    def crm_event_channel_scope_clause(self, event_alias: str = "e") -> tuple[str, list]:
        user_id = int(self.authenticated_user["id"])
        return (
            f"(COALESCE((SELECT crm_channel_scope_enabled FROM users WHERE id=?),0)=0 OR EXISTS ("
            f"SELECT 1 FROM crm_user_channels scoped_access "
            f"LEFT JOIN crm_conversations scoped_cv ON scoped_cv.id={event_alias}.conversation_id "
            f"LEFT JOIN crm_channels scoped_ch ON scoped_ch.id=scoped_access.channel_id "
            f"WHERE scoped_access.user_id=? AND (scoped_access.channel_id=scoped_cv.channel_id "
            f"OR LOWER(COALESCE(scoped_ch.instance_name,''))=LOWER(COALESCE({event_alias}.channel_name,'')) "
            f"OR LOWER(COALESCE(scoped_ch.display_name,''))=LOWER(COALESCE({event_alias}.channel_name,'')))))",
            [user_id, user_id],
        )

    def crm_channel_scope_clause(self, channel_alias: str = "ch", capability: str | None = None) -> tuple[str, list]:
        """Escopo SQL retrocompatível: usuários antigos continuam vendo todos os canais.

        Quando o administrador ativa o escopo, somente os canais explicitamente
        vinculados ao usuário ficam disponíveis. A permissão é aplicada no backend.
        """
        return self.crm_channel_id_scope_clause(f"{channel_alias}.id", capability)

    def crm_channel_allowed(self, db, channel_id: int, capability: str | None = None, user_id: int | None = None) -> bool:
        user_id = int(user_id or self.authenticated_user["id"])
        user = db.execute(
            "SELECT crm_channel_scope_enabled,crm_manage_automation FROM users WHERE id=?",
            (user_id,),
        ).fetchone()
        if not user:
            return False
        if capability == "automation" and not user["crm_manage_automation"]:
            return False
        if not user["crm_channel_scope_enabled"]:
            return True
        permission = ""
        if capability == "reply": permission = " AND can_reply=1"
        elif capability == "automation": permission = " AND can_manage_automation=1"
        return bool(db.execute(
            f"SELECT 1 FROM crm_user_channels WHERE user_id=? AND channel_id=?{permission}",
            (user_id, channel_id),
        ).fetchone())

    def crm_can_manage_automation(self, db, user_id: int | None = None) -> bool:
        user_id = int(user_id or self.authenticated_user["id"])
        user = db.execute("SELECT crm_manage_automation FROM users WHERE id=?", (user_id,)).fetchone()
        return bool(user and user["crm_manage_automation"])

    def get_crm_channels(self) -> None:
        if not self.require_crc_access(): return
        if not self.require_crm_any_feature(CRM_FEATURE_KEYS):
            return
        scope_sql, scope_params = self.crm_channel_scope_clause("ch")
        with connect() as db:
            rows = db.execute(f"""SELECT ch.id,ch.instance_name,ch.display_name,ch.phone,ch.active,ch.sync_enabled,ch.sync_from_date,ch.sla_minutes,
                                ch.connection_status,ch.last_event_at,
                                ch.evolution_base_url IS NOT NULL AS configured,
                                COALESCE(cuc.can_reply,CASE WHEN u.crm_channel_scope_enabled=0 THEN 1 ELSE 0 END) AS can_reply,
                                COALESCE(cuc.can_manage_automation,CASE WHEN u.crm_channel_scope_enabled=0 THEN 1 ELSE 0 END) AS can_manage_automation
                                FROM crm_channels ch
                                JOIN users u ON u.id=?
                                LEFT JOIN crm_user_channels cuc ON cuc.channel_id=ch.id AND cuc.user_id=u.id
                                WHERE {scope_sql}
                                ORDER BY ch.active DESC,ch.display_name COLLATE NOCASE""",
                              [self.authenticated_user["id"], *scope_params]).fetchall()
        self.send_json({"items": [dict(row) for row in rows]})

    def update_crm_channel(self, channel_id: int, payload: dict) -> None:
        if not self.require_crc_access():
            return
        if not self.require_crm_feature("integrations"):
            return
        with connect() as db:
            if not self.crm_channel_allowed(db, channel_id, "automation"):
                return self.send_json({"error": "Você não possui permissão para configurar este canal."}, HTTPStatus.FORBIDDEN)
        updates, params = [], []
        if "sync_enabled" in payload:
            updates.append("sync_enabled=?")
            params.append(1 if payload.get("sync_enabled") else 0)
        if "sync_from_date" in payload:
            sync_from_date = str(payload.get("sync_from_date") or "").strip()
            try:
                parsed_date = datetime.strptime(sync_from_date, "%Y-%m-%d")
            except ValueError:
                return self.send_json({"error": "Informe uma data válida para a sincronização."}, HTTPStatus.BAD_REQUEST)
            if parsed_date.date() > datetime.now().date():
                return self.send_json({"error": "A data inicial não pode estar no futuro."}, HTTPStatus.BAD_REQUEST)
            updates.append("sync_from_date=?")
            params.append(sync_from_date)
        if "sla_minutes" in payload:
            try:
                sla_minutes = int(payload.get("sla_minutes") or 0)
            except (TypeError, ValueError):
                sla_minutes = 0
            if sla_minutes < 1 or sla_minutes > 1440:
                return self.send_json({"error": "O SLA deve ficar entre 1 e 1.440 minutos."}, HTTPStatus.BAD_REQUEST)
            updates.append("sla_minutes=?")
            params.append(sla_minutes)
        if not updates:
            return self.send_json({"error": "Nenhuma configuração de sincronização foi informada."}, HTTPStatus.BAD_REQUEST)
        params.append(channel_id)
        with connect() as db:
            changed = db.execute(f"UPDATE crm_channels SET {','.join(updates)},updated_at=datetime('now','localtime') WHERE id=?", params).rowcount
            if not changed:
                return self.send_json({"error": "Canal não encontrado."}, HTTPStatus.NOT_FOUND)
            row = db.execute("""SELECT id,instance_name,display_name,phone,active,sync_enabled,sync_from_date,sla_minutes,
                                connection_status,last_event_at FROM crm_channels WHERE id=?""", (channel_id,)).fetchone()
        self.send_json(dict(row))

    @staticmethod
    def evolution_credentials() -> tuple[str, str]:
        with connect() as db:
            row = db.execute("SELECT api_base_url,api_token FROM integration_configs WHERE name='evolution_crm'").fetchone()
        if row and row["api_base_url"] and row["api_token"]:
            return str(row["api_base_url"]).rstrip("/"), str(decrypt_integration_secret(row["api_token"]) or "")
        return EVOLUTION_API_URL, EVOLUTION_API_KEY

    def evolution_api_request(self, path: str, method: str = "GET", payload: dict | None = None,
                              base_url: str | None = None, api_key: str | None = None) -> dict | list:
        configured_url, configured_key = self.evolution_credentials()
        base_url = (base_url or configured_url).rstrip("/")
        api_key = api_key or configured_key
        if not base_url or not api_key:
            raise RuntimeError("A Evolution API ainda não está configurada no servidor.")
        body = json.dumps(payload).encode("utf-8") if payload is not None else None
        request = Request(f"{base_url}{path}", data=body, method=method,
                          headers={"apikey": api_key, "Content-Type": "application/json"})
        try:
            with urlopen(request, timeout=30) as response:
                raw = response.read().decode("utf-8")
                if not raw:
                    return {}
                try:
                    return json.loads(raw)
                except json.JSONDecodeError as error:
                    raise RuntimeError(
                        "A Evolution retornou uma resposta inválida. Use somente o domínio da API, sem /manager ou outras rotas."
                    ) from error
        except HTTPError as error:
            detail = error.read().decode(errors="replace")
            raise RuntimeError(f"Evolution respondeu {error.code}: {detail[:300]}") from error
        except (URLError, TimeoutError) as error:
            raise RuntimeError(f"Não foi possível conectar à Evolution: {error}") from error

    def crm_evolution_connection_state(self, instance_name: str, base_url: str, api_key: str) -> str:
        instance_path = quote(str(instance_name), safe="")
        result = self.evolution_api_request(
            f"/instance/connectionState/{instance_path}", base_url=base_url, api_key=api_key
        )
        instance = result.get("instance") if isinstance(result, dict) else None
        state = (instance or {}).get("state") if isinstance(instance, dict) else None
        if not state and isinstance(result, dict):
            state = result.get("state") or result.get("connectionStatus")
        return str(state or "unknown").strip().lower()

    @staticmethod
    def evolution_instance_summary(item: dict) -> dict:
        instance = item.get("instance") if isinstance(item.get("instance"), dict) else item
        return {
            "name": instance.get("instanceName") or instance.get("name"),
            "state": instance.get("connectionStatus") or instance.get("state") or item.get("connectionStatus") or "desconhecido",
            "connected": bool(instance.get("ownerJid") or item.get("ownerJid")) or
                         str(instance.get("connectionStatus") or instance.get("state") or item.get("connectionStatus") or "").lower() == "open",
        }

    @staticmethod
    def evolution_unread_count(value) -> int:
        if isinstance(value, dict):
            value = value.get("low") or value.get("value") or 0
        try:
            return max(0, int(value or 0))
        except (TypeError, ValueError):
            return 0

    @staticmethod
    def evolution_chat_direction(chat: dict) -> str:
        last_message = chat.get("lastMessage") if isinstance(chat.get("lastMessage"), dict) else {}
        key = last_message.get("key") if isinstance(last_message.get("key"), dict) else {}
        from_me = key.get("fromMe") if "fromMe" in key else last_message.get("fromMe")
        return "outbound" if bool(from_me) else "inbound"

    def sync_evolution_chat_state(self) -> dict:
        """Reconcilia contadores da Evolution sem importar novamente o histórico."""
        global EVOLUTION_CHAT_SYNC_STATUS
        with EVOLUTION_CHAT_SYNC_LOCK:
            if EVOLUTION_CHAT_SYNC_STATUS["running"]:
                return dict(EVOLUTION_CHAT_SYNC_STATUS)
            EVOLUTION_CHAT_SYNC_STATUS = {
                "running": True,
                "last_started_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "last_finished_at": EVOLUTION_CHAT_SYNC_STATUS.get("last_finished_at"),
                "instances_synced": 0,
                "unread_conversations": 0,
                "pending_conversations": 0,
                "unmatched_conversations": 0,
                "errors": [],
            }
        try:
            payload = self.evolution_api_request("/instance/fetchInstances")
            instances = payload if isinstance(payload, list) else payload.get("instances") or payload.get("data") or []
            for raw_instance in instances if isinstance(instances, list) else []:
                if not isinstance(raw_instance, dict):
                    continue
                summary = self.evolution_instance_summary(raw_instance)
                instance_name = str(summary.get("name") or "").strip()
                if not instance_name or not summary.get("connected"):
                    continue
                try:
                    chat_payload = self.evolution_api_request(
                        f"/chat/findChats/{quote(instance_name, safe='')}", "POST", {"where": {}}
                    )
                    chats = self.evolution_list(chat_payload, "chats", "records", "items")
                    unread_total = 0
                    pending_total = 0
                    with connect() as db:
                        channel = db.execute(
                            "SELECT id FROM crm_channels WHERE instance_name=?", (instance_name,)
                        ).fetchone()
                        if not channel:
                            continue
                        channel_id = channel["id"]
                        # A Evolution é a fonte de verdade do contador. Limpa o
                        # estado antigo somente depois que a consulta do canal funcionou.
                        db.execute("UPDATE crm_conversations SET unread_count=0 WHERE channel_id=?", (channel_id,))
                        for chat in chats:
                            if not isinstance(chat, dict):
                                continue
                            unread = self.evolution_unread_count(chat.get("unreadCount"))
                            if unread <= 0:
                                continue
                            last_message = chat.get("lastMessage") if isinstance(chat.get("lastMessage"), dict) else {}
                            key = last_message.get("key") if isinstance(last_message.get("key"), dict) else {}
                            remote = str(chat.get("remoteJid") or chat.get("id") or key.get("remoteJid") or "")
                            if not remote or remote.endswith("@g.us") or remote.endswith("@broadcast"):
                                continue
                            remote_alt = str(key.get("remoteJidAlt") or "")
                            phone = self.crm_phone(remote.split("@")[0]) or self.crm_phone(remote_alt.split("@")[0])
                            external_id = str(key.get("id") or last_message.get("id") or "").strip()
                            existing_conversation = None
                            if external_id:
                                existing_conversation = db.execute("""SELECT cv.id,cv.status,cv.resolved_at FROM crm_messages m
                                    JOIN crm_conversations cv ON cv.id=m.conversation_id
                                    WHERE m.external_message_id=? AND cv.channel_id=?""",
                                    (external_id, channel_id)).fetchone()
                            if not phone and not existing_conversation:
                                with EVOLUTION_CHAT_SYNC_LOCK:
                                    EVOLUTION_CHAT_SYNC_STATUS["unmatched_conversations"] += 1
                                continue
                            direction = self.evolution_chat_direction(chat)
                            last_at = self.evolution_message_time(
                                last_message.get("messageTimestamp") or last_message.get("timestamp") or chat.get("updatedAt")
                            )
                            inbound_pending = direction == "inbound"
                            if existing_conversation:
                                # A Evolution pode manter unreadCount > 0 mesmo depois
                                # de a CRC resolver o atendimento. A conversa só deve
                                # reabrir se a última entrada for posterior à resolução.
                                stale_after_resolution = (
                                    existing_conversation["status"] == "Resolvida"
                                    and direction == "inbound"
                                    and bool(existing_conversation["resolved_at"])
                                    and last_at <= existing_conversation["resolved_at"]
                                )
                                if stale_after_resolution:
                                    db.execute("""UPDATE crm_conversations SET unread_count=0,
                                        updated_at=datetime('now','localtime') WHERE id=?""",
                                        (existing_conversation["id"],))
                                    continue
                                db.execute("""UPDATE crm_conversations SET
                                    unread_count=CASE WHEN assigned_user_id IS NULL THEN ? ELSE 0 END,
                                    last_direction=?,last_message_at=?,
                                    queue_entered_at=CASE
                                        WHEN status='Resolvida' AND ?='inbound' THEN ?
                                        WHEN queue_entered_at IS NULL AND assigned_user_id IS NULL AND ?='inbound' THEN ?
                                        ELSE queue_entered_at END,
                                    status=CASE WHEN ?='inbound' THEN 'Aberta' ELSE status END,
                                    pipeline_stage=CASE WHEN status='Resolvida' AND ?='inbound' THEN 'Novo' ELSE pipeline_stage END,
                                    assigned_user_id=CASE WHEN status='Resolvida' AND ?='inbound' THEN NULL ELSE assigned_user_id END,
                                    assigned_at=CASE WHEN status='Resolvida' AND ?='inbound' THEN NULL ELSE assigned_at END,
                                    resolved_at=CASE WHEN ?='inbound' THEN NULL ELSE resolved_at END,
                                    resolved_by_user_id=CASE WHEN status='Resolvida' AND ?='inbound' THEN NULL ELSE resolved_by_user_id END,
                                    updated_at=datetime('now','localtime') WHERE id=?""",
                                    (unread, direction, last_at, direction, last_at, direction, last_at,
                                     direction, direction, direction, direction,
                                     direction, direction, existing_conversation["id"]))
                            else:
                                name = str(chat.get("pushName") or phone).strip()
                                db.execute("""INSERT INTO crm_contacts(name,phone) VALUES(?,?)
                                              ON CONFLICT(phone) DO UPDATE SET
                                              name=CASE WHEN crm_contacts.name=crm_contacts.phone THEN excluded.name ELSE crm_contacts.name END,
                                              updated_at=datetime('now','localtime')""", (name, phone))
                                contact_id = db.execute("SELECT id FROM crm_contacts WHERE phone=?", (phone,)).fetchone()["id"]
                                db.execute("""INSERT INTO crm_conversations
                                    (channel_id,contact_id,status,pipeline_stage,unread_count,last_direction,last_message_at,queue_entered_at)
                                    VALUES(?,?,'Aberta','Novo',?,?,?,CASE WHEN ?='inbound' THEN ? ELSE NULL END)
                                    ON CONFLICT(channel_id,contact_id) DO UPDATE SET
                                    unread_count=CASE
                                        WHEN crm_conversations.status='Resolvida' AND crm_conversations.resolved_at IS NOT NULL
                                             AND datetime(excluded.last_message_at)<=datetime(crm_conversations.resolved_at) THEN 0
                                        WHEN crm_conversations.assigned_user_id IS NULL THEN excluded.unread_count ELSE 0 END,
                                    last_direction=CASE WHEN crm_conversations.last_message_at IS NULL OR datetime(excluded.last_message_at)>datetime(crm_conversations.last_message_at)
                                                        THEN excluded.last_direction ELSE crm_conversations.last_direction END,
                                    last_message_at=CASE WHEN crm_conversations.last_message_at IS NULL OR datetime(excluded.last_message_at)>datetime(crm_conversations.last_message_at)
                                                        THEN excluded.last_message_at ELSE crm_conversations.last_message_at END,
                                    queue_entered_at=CASE
                                        WHEN crm_conversations.status='Resolvida' AND excluded.last_direction='inbound'
                                             AND (crm_conversations.resolved_at IS NULL OR datetime(excluded.last_message_at)>datetime(crm_conversations.resolved_at))
                                        THEN excluded.last_message_at
                                        WHEN crm_conversations.queue_entered_at IS NULL AND crm_conversations.assigned_user_id IS NULL
                                             AND excluded.last_direction='inbound' THEN excluded.last_message_at
                                        ELSE crm_conversations.queue_entered_at END,
                                    status=CASE WHEN excluded.last_direction='inbound' AND
                                        (crm_conversations.status<>'Resolvida' OR crm_conversations.resolved_at IS NULL OR
                                         datetime(excluded.last_message_at)>datetime(crm_conversations.resolved_at))
                                        THEN 'Aberta' ELSE crm_conversations.status END,
                                    pipeline_stage=CASE WHEN crm_conversations.status='Resolvida' AND excluded.last_direction='inbound'
                                        AND (crm_conversations.resolved_at IS NULL OR datetime(excluded.last_message_at)>datetime(crm_conversations.resolved_at))
                                        THEN 'Novo' ELSE crm_conversations.pipeline_stage END,
                                    assigned_user_id=CASE WHEN crm_conversations.status='Resolvida' AND excluded.last_direction='inbound'
                                        AND (crm_conversations.resolved_at IS NULL OR datetime(excluded.last_message_at)>datetime(crm_conversations.resolved_at))
                                        THEN NULL ELSE crm_conversations.assigned_user_id END,
                                    assigned_at=CASE WHEN crm_conversations.status='Resolvida' AND excluded.last_direction='inbound'
                                        AND (crm_conversations.resolved_at IS NULL OR datetime(excluded.last_message_at)>datetime(crm_conversations.resolved_at))
                                        THEN NULL ELSE crm_conversations.assigned_at END,
                                    resolved_at=CASE WHEN excluded.last_direction='inbound' AND
                                        (crm_conversations.status<>'Resolvida' OR crm_conversations.resolved_at IS NULL OR
                                         datetime(excluded.last_message_at)>datetime(crm_conversations.resolved_at))
                                        THEN NULL ELSE crm_conversations.resolved_at END,
                                    resolved_by_user_id=CASE WHEN crm_conversations.status='Resolvida' AND excluded.last_direction='inbound'
                                        AND (crm_conversations.resolved_at IS NULL OR datetime(excluded.last_message_at)>datetime(crm_conversations.resolved_at))
                                        THEN NULL ELSE crm_conversations.resolved_by_user_id END,
                                    updated_at=datetime('now','localtime')""",
                                           (channel_id, contact_id, unread, direction, last_at, direction, last_at))
                            if existing_conversation:
                                conversation_state = db.execute(
                                    "SELECT unread_count,last_direction,status FROM crm_conversations WHERE id=?",
                                    (existing_conversation["id"],),
                                ).fetchone()
                            else:
                                conversation_state = db.execute(
                                    "SELECT unread_count,last_direction,status FROM crm_conversations WHERE channel_id=? AND contact_id=?",
                                    (channel_id, contact_id),
                                ).fetchone()
                            if conversation_state and conversation_state["unread_count"] > 0 and conversation_state["status"] != "Resolvida":
                                unread_total += 1
                                pending_total += 1 if conversation_state["last_direction"] == "inbound" else 0
                    with EVOLUTION_CHAT_SYNC_LOCK:
                        EVOLUTION_CHAT_SYNC_STATUS["instances_synced"] += 1
                        EVOLUTION_CHAT_SYNC_STATUS["unread_conversations"] += unread_total
                        EVOLUTION_CHAT_SYNC_STATUS["pending_conversations"] += pending_total
                except Exception as error:
                    with EVOLUTION_CHAT_SYNC_LOCK:
                        EVOLUTION_CHAT_SYNC_STATUS["errors"].append({
                            "instance": instance_name, "error": str(error)[:240]
                        })
        except Exception as error:
            with EVOLUTION_CHAT_SYNC_LOCK:
                EVOLUTION_CHAT_SYNC_STATUS["errors"].append({"instance": "Evolution", "error": str(error)[:240]})
        finally:
            with EVOLUTION_CHAT_SYNC_LOCK:
                EVOLUTION_CHAT_SYNC_STATUS["running"] = False
                EVOLUTION_CHAT_SYNC_STATUS["last_finished_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                return dict(EVOLUTION_CHAT_SYNC_STATUS)

    def get_evolution_instances(self) -> None:
        if not self.require_crm_feature("integrations"): return
        if not self.require_crc_access(): return
        try:
            payload = self.evolution_api_request("/instance/fetchInstances")
            items = payload if isinstance(payload, list) else payload.get("instances") or payload.get("data") or []
            if isinstance(items, dict): items = list(items.values())
            self.send_json({"items": [self.evolution_instance_summary(item) for item in items if isinstance(item, dict)]})
        except RuntimeError as error:
            self.send_json({"error": str(error)}, HTTPStatus.BAD_GATEWAY)

    def get_evolution_config(self) -> None:
        if not self.require_crm_feature("integrations"): return
        if not self.require_crc_access(): return
        with connect() as db:
            row = db.execute("SELECT api_base_url,api_token,updated_at FROM integration_configs WHERE name='evolution_crm'").fetchone()
        self.send_json({
            "configured": bool(row and row["api_base_url"] and row["api_token"]),
            "api_base_url": str(row["api_base_url"] or "") if row else "",
            "token_configured": bool(row and row["api_token"]),
            "updated_at": row["updated_at"] if row else None,
        })

    def save_evolution_config(self, payload: dict) -> None:
        if not self.require_crm_feature("integrations"): return
        if not self.require_crc_access(): return
        with connect() as db:
            if not self.crm_can_manage_automation(db):
                return self.send_json({"error": "Somente um responsÃ¡vel autorizado pode alterar a Evolution."}, HTTPStatus.FORBIDDEN)
        base_url = str(payload.get("api_base_url") or "").strip().rstrip("/")
        api_key = str(payload.get("api_key") or "").strip()
        parsed_base_url = urlparse(base_url)
        if parsed_base_url.scheme != "https" or not parsed_base_url.netloc:
            return self.send_json({"error": "Informe a URL HTTPS da Evolution API."}, HTTPStatus.BAD_REQUEST)
        # The manager copies deep links such as /manager/instance/.../dashboard.
        # API calls must always target the origin only.
        base_url = f"{parsed_base_url.scheme}://{parsed_base_url.netloc}"
        with connect() as db:
            existing = db.execute("SELECT api_token FROM integration_configs WHERE name='evolution_crm'").fetchone()
        api_key = api_key or (
            str(decrypt_integration_secret(existing["api_token"]) or "")
            if existing and existing["api_token"] else ""
        )
        if not api_key:
            return self.send_json({"error": "Informe a chave global da Evolution API."}, HTTPStatus.BAD_REQUEST)
        try:
            instances = self.evolution_api_request("/instance/fetchInstances", base_url=base_url, api_key=api_key)
        except RuntimeError as error:
            return self.send_json({"error": f"A conexão não foi salva: {error}"}, HTTPStatus.BAD_GATEWAY)
        items = instances if isinstance(instances, list) else instances.get("instances") or instances.get("data") or []
        with connect() as db:
            db.execute("""INSERT INTO integration_configs(name,api_base_url,api_token,updated_at,updated_by)
                          VALUES('evolution_crm',?,?,datetime('now','localtime'),?)
                          ON CONFLICT(name) DO UPDATE SET api_base_url=excluded.api_base_url,api_token=excluded.api_token,
                          updated_at=datetime('now','localtime'),updated_by=excluded.updated_by""",
                       (base_url, encrypt_integration_secret(api_key), self.authenticated_user["id"]))
            for item in items if isinstance(items, list) else []:
                if not isinstance(item, dict): continue
                summary = self.evolution_instance_summary(item)
                if not summary["name"]: continue
                db.execute("""INSERT INTO crm_channels(instance_name,display_name,active,connection_status)
                              VALUES(?,?,1,?) ON CONFLICT(instance_name) DO UPDATE SET active=1,
                              connection_status=excluded.connection_status,updated_at=datetime('now','localtime')""",
                           (summary["name"], summary["name"], "Conectado" if summary["connected"] else "Desconectado"))
        self.send_json({"saved": True, "configured": True, "instances_found": len(items) if hasattr(items, "__len__") else 0})

    @staticmethod
    def evolution_list(payload, *keys: str) -> list:
        if isinstance(payload, list):
            return payload
        current = payload
        for _ in range(3):
            if not isinstance(current, dict):
                break
            for key in keys:
                value = current.get(key)
                if isinstance(value, list):
                    return value
                if isinstance(value, dict):
                    for nested_key in ("records", "items", "rows", "data"):
                        nested = value.get(nested_key)
                        if isinstance(nested, list):
                            return nested
            for nested_key in ("data", "messages", "contacts", "chats"):
                nested = current.get(nested_key)
                if isinstance(nested, list):
                    return nested
                if isinstance(nested, dict):
                    current = nested
                    break
            else:
                break
        return []

    @staticmethod
    def evolution_message_details(record: dict) -> tuple[str, str, str | None, str | None, float | None]:
        message = record.get("message") if isinstance(record.get("message"), dict) else {}
        body = str(message.get("conversation") or (message.get("extendedTextMessage") or {}).get("text") or record.get("text") or "")
        kind, media_url, mime_type, duration_seconds = "text", None, None, None
        for candidate, media_kind in (("imageMessage", "image"), ("audioMessage", "audio"),
                                      ("videoMessage", "video"), ("documentMessage", "document"),
                                      ("stickerMessage", "sticker")):
            media = message.get(candidate)
            if isinstance(media, dict):
                kind = media_kind
                body = str(media.get("caption") or media.get("fileName") or f"[{media_kind}]")
                media_url = media.get("url")
                mime_type = media.get("mimetype")
                if media_kind == "audio":
                    try:
                        duration_seconds = float(media.get("seconds") or 0) or None
                    except (TypeError, ValueError):
                        duration_seconds = None
                break
        return body, kind, media_url, mime_type, duration_seconds

    @staticmethod
    def evolution_message_time(value) -> str:
        if isinstance(value, dict):
            value = value.get("low") or value.get("value")
        try:
            number = int(value)
            if number > 10_000_000_000:
                number //= 1000
            return datetime.fromtimestamp(number, tz=CLINIC_TIMEZONE).strftime("%Y-%m-%d %H:%M:%S")
        except (TypeError, ValueError, OSError):
            pass
        text_value = str(value or "").strip()
        if text_value:
            try:
                parsed = datetime.fromisoformat(text_value.replace("Z", "+00:00"))
                if parsed.tzinfo is None:
                    parsed = parsed.replace(tzinfo=timezone.utc)
                return parsed.astimezone(CLINIC_TIMEZONE).strftime("%Y-%m-%d %H:%M:%S")
            except ValueError:
                pass
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def configure_existing_evolution_webhook(self, instance_name: str) -> None:
        if not EVOLUTION_WEBHOOK_TOKEN:
            return
        webhook_url = f"{PUBLIC_APP_URL}/api/integrations/evolution/webhook?webhook_key={EVOLUTION_WEBHOOK_TOKEN}"
        self.evolution_api_request(f"/webhook/set/{quote(instance_name, safe='')}", "POST", {"webhook": {
            "enabled": True,
            "url": webhook_url,
            "webhookByEvents": False,
            "webhookBase64": False,
            "events": ["MESSAGES_UPSERT", "MESSAGES_UPDATE", "CONNECTION_UPDATE"],
        }})

    def import_evolution_messages(self, instance_name: str, records: list, contact_names: dict[str, str]) -> tuple[int, int, int]:
        imported_contacts, imported_conversations, imported_messages = 0, 0, 0
        with connect() as db:
            channel = db.execute("SELECT id FROM crm_channels WHERE instance_name=?", (instance_name,)).fetchone()
            if not channel:
                return 0, 0, 0
            channel_id = channel["id"]
            patient_by_phone = {}
            for patient in db.execute("SELECT id,name,phone FROM patients WHERE phone IS NOT NULL AND TRIM(phone)<>''").fetchall():
                normalized = self.crm_phone(patient["phone"])
                if normalized:
                    patient_by_phone[normalized] = (patient["id"], patient["name"])

            # Fase 1: normaliza os registros e descarta os inválidos antes de
            # tocar o banco (evita 1 SELECT de deduplicação por mensagem).
            parsed = []
            for record in records:
                if not isinstance(record, dict):
                    continue
                key = record.get("key") if isinstance(record.get("key"), dict) else {}
                remote = str(key.get("remoteJid") or record.get("remoteJid") or record.get("jid") or "")
                remote_alt = str(key.get("remoteJidAlt") or "")
                if not remote or remote.endswith("@g.us") or remote.endswith("@broadcast"):
                    continue
                phone = self.crm_phone(remote.split("@")[0]) or self.crm_phone(remote_alt.split("@")[0])
                if not phone:
                    continue
                external_id = str(key.get("id") or record.get("id") or "").strip()
                if not external_id:
                    continue
                parsed.append((record, key, remote, phone, external_id))
            if not parsed:
                return 0, 0, 0

            external_ids = list({item[4] for item in parsed})
            placeholders = ",".join("?" for _ in external_ids)
            existing_ids = {
                row["external_message_id"]
                for row in db.execute(
                    f"SELECT external_message_id FROM crm_messages WHERE external_message_id IN ({placeholders})",
                    external_ids,
                ).fetchall()
            }
            attributions = {
                row["external_message_id"]: (row["author_type"], row["author_label"])
                for row in db.execute(
                    f"SELECT external_message_id, author_type, author_label FROM crm_message_attributions "
                    f"WHERE external_message_id IN ({placeholders})",
                    external_ids,
                ).fetchall()
            }

            for record, key, remote, phone, external_id in parsed:
                if external_id in existing_ids:
                    continue
                patient_id, name = patient_by_phone.get(phone, (None, None))
                name = name or contact_names.get(remote) or str(record.get("pushName") or phone).strip()
                existing_contact = db.execute("SELECT id FROM crm_contacts WHERE phone=?", (phone,)).fetchone()
                contact_id = db.execute("""INSERT INTO crm_contacts(patient_id,name,phone) VALUES(?,?,?)
                              ON CONFLICT(phone) DO UPDATE SET patient_id=COALESCE(excluded.patient_id,crm_contacts.patient_id),
                              name=CASE WHEN excluded.patient_id IS NOT NULL OR crm_contacts.name=crm_contacts.phone THEN excluded.name ELSE crm_contacts.name END,
                              updated_at=datetime('now','localtime')
                              RETURNING id""", (patient_id, name, phone)).fetchone()["id"]
                if not existing_contact:
                    imported_contacts += 1
                existing_conversation = db.execute("SELECT id FROM crm_conversations WHERE channel_id=? AND contact_id=?", (channel_id, contact_id)).fetchone()
                message_at = self.evolution_message_time(record.get("messageTimestamp") or record.get("timestamp") or record.get("createdAt"))
                from_me = bool(key.get("fromMe") if "fromMe" in key else record.get("fromMe"))
                direction = "outbound" if from_me else "inbound"
                conversation_id = db.execute("""INSERT INTO crm_conversations(channel_id,contact_id,status,last_message_at,last_direction,unread_count)
                              VALUES(?,?,'Aberta',?,?,0) ON CONFLICT(channel_id,contact_id) DO UPDATE SET
                              last_direction=CASE WHEN crm_conversations.last_message_at IS NULL OR datetime(excluded.last_message_at)>datetime(crm_conversations.last_message_at)
                                                  THEN excluded.last_direction ELSE crm_conversations.last_direction END,
                              last_message_at=CASE WHEN crm_conversations.last_message_at IS NULL OR datetime(excluded.last_message_at)>datetime(crm_conversations.last_message_at)
                                                   THEN excluded.last_message_at ELSE crm_conversations.last_message_at END,
                              updated_at=datetime('now','localtime')
                              RETURNING id""", (channel_id, contact_id, message_at, direction)).fetchone()["id"]
                if not existing_conversation:
                    imported_conversations += 1
                body, kind, media_url, mime_type, duration_seconds = self.evolution_message_details(record)
                attribution = attributions.get(external_id)
                author_type = attribution[0] if attribution else ("external" if from_me else "patient")
                author_label = attribution[1] if attribution else ("Enviado fora do CRM" if from_me else name)
                source_channel = str(record.get("source") or "evolution").strip()[:40]
                result = db.execute("""INSERT OR IGNORE INTO crm_messages
                    (conversation_id,external_message_id,direction,message_type,body,media_url,mime_type,duration_seconds,sender_name,
                     author_type,author_label,source_channel,delivery_status,message_at)
                    VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (conversation_id, external_id, direction, kind, body,
                     media_url, mime_type, duration_seconds, name, author_type, author_label, source_channel,
                     "Enviada" if from_me else "Recebida", message_at))
                imported_messages += result.rowcount
        return imported_contacts, imported_conversations, imported_messages

    def import_evolution_contacts(self, contacts: list) -> int:
        imported = 0
        with connect() as db:
            patient_by_phone = {}
            for patient in db.execute("SELECT id,name,phone FROM patients WHERE phone IS NOT NULL AND TRIM(phone)<>''").fetchall():
                normalized = self.crm_phone(patient["phone"])
                if normalized:
                    patient_by_phone[normalized] = (patient["id"], patient["name"])
            for contact in contacts:
                if not isinstance(contact, dict):
                    continue
                jid = str(contact.get("remoteJid") or contact.get("id") or contact.get("jid") or "")
                if not jid or jid.endswith("@g.us") or jid.endswith("@broadcast"):
                    continue
                phone = self.crm_phone(jid.split("@")[0])
                if not phone:
                    continue
                patient_id, patient_name = patient_by_phone.get(phone, (None, None))
                contact_name = str(contact.get("pushName") or contact.get("name") or contact.get("notify") or phone).strip()
                name = patient_name or contact_name or phone
                profile_picture = contact.get("profilePicUrl") or contact.get("profilePictureUrl") or contact.get("picture")
                existing = db.execute("SELECT id FROM crm_contacts WHERE phone=?", (phone,)).fetchone()
                db.execute("""INSERT INTO crm_contacts(patient_id,name,phone,profile_picture_url) VALUES(?,?,?,?)
                              ON CONFLICT(phone) DO UPDATE SET
                              patient_id=COALESCE(excluded.patient_id,crm_contacts.patient_id),
                              name=CASE WHEN excluded.patient_id IS NOT NULL OR crm_contacts.name=crm_contacts.phone THEN excluded.name ELSE crm_contacts.name END,
                              profile_picture_url=COALESCE(excluded.profile_picture_url,crm_contacts.profile_picture_url),
                              updated_at=datetime('now','localtime')""",
                           (patient_id, name, phone, profile_picture))
                if not existing:
                    imported += 1
        return imported

    def run_evolution_history_sync(self) -> None:
        global EVOLUTION_HISTORY_SYNC_STATUS
        sync_upper_bound = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        try:
            with EVOLUTION_HISTORY_SYNC_LOCK:
                EVOLUTION_HISTORY_SYNC_STATUS["phase"] = "Preparando banco de dados..."
            with connect() as db:
                removed = db.execute("DELETE FROM crm_messages WHERE datetime(message_at) < datetime(?)", (EVOLUTION_HISTORY_CUTOFF,)).rowcount
                db.execute("DELETE FROM crm_conversations WHERE NOT EXISTS (SELECT 1 FROM crm_messages m WHERE m.conversation_id=crm_conversations.id)")
            with EVOLUTION_HISTORY_SYNC_LOCK:
                EVOLUTION_HISTORY_SYNC_STATUS["older_messages_removed"] = removed
            with EVOLUTION_HISTORY_SYNC_LOCK:
                EVOLUTION_HISTORY_SYNC_STATUS["phase"] = "Consultando canais da Evolution..."
            payload = self.evolution_api_request("/instance/fetchInstances")
            raw_instances = payload if isinstance(payload, list) else payload.get("instances") or payload.get("data") or []
            instances = [self.evolution_instance_summary(item) for item in raw_instances if isinstance(item, dict)]
            instances = [item for item in instances if item.get("name")]
            with EVOLUTION_HISTORY_SYNC_LOCK:
                EVOLUTION_HISTORY_SYNC_STATUS["instances_total"] = len(instances)
            for instance in instances:
                instance_name = instance["name"]
                try:
                    with EVOLUTION_HISTORY_SYNC_LOCK:
                        EVOLUTION_HISTORY_SYNC_STATUS["phase"] = f"Carregando contatos de {instance_name}..."
                    with connect() as db:
                        db.execute("""INSERT INTO crm_channels(instance_name,display_name,active,connection_status)
                                      VALUES(?,?,1,?) ON CONFLICT(instance_name) DO UPDATE SET active=1,
                                      connection_status=excluded.connection_status,updated_at=datetime('now','localtime')""",
                                   (instance_name, instance_name, "Conectado" if instance["connected"] else "Desconectado"))
                        channel_settings = db.execute("SELECT sync_enabled,sync_from_date FROM crm_channels WHERE instance_name=?", (instance_name,)).fetchone()
                    if channel_settings and not channel_settings["sync_enabled"]:
                        continue
                    channel_cutoff = max(EVOLUTION_HISTORY_CUTOFF, f"{(channel_settings['sync_from_date'] if channel_settings else '2026-07-20')} 00:00:00")
                    try:
                        self.configure_existing_evolution_webhook(instance_name)
                    except RuntimeError:
                        pass
                    # The CRM is not an address-book mirror. Contacts are only
                    # created when a message/history exists for that number.
                    contact_names = {}
                    page, seen_page = 1, set()
                    while page <= 1000:
                        with EVOLUTION_HISTORY_SYNC_LOCK:
                            EVOLUTION_HISTORY_SYNC_STATUS["phase"] = f"Importando mensagens de {instance_name} — página {page}..."
                        message_payload = self.evolution_api_request(
                            f"/chat/findMessages/{quote(instance_name, safe='')}", "POST",
                            {"where": {}, "page": page, "offset": 200})
                        records = self.evolution_list(message_payload, "messages", "records", "items")
                        if not records:
                            break
                        signature = tuple(str(((item.get("key") or {}).get("id")) or item.get("id") or "") for item in records[:5] if isinstance(item, dict))
                        if signature and signature in seen_page:
                            break
                        if signature:
                            seen_page.add(signature)
                        eligible_records = []
                        known_times = []
                        for record in records:
                            if not isinstance(record, dict):
                                continue
                            raw_time = record.get("messageTimestamp") or record.get("timestamp") or record.get("createdAt")
                            message_time = self.evolution_message_time(raw_time)
                            if raw_time:
                                known_times.append(message_time)
                            if channel_cutoff <= message_time <= sync_upper_bound:
                                eligible_records.append(record)
                        counts = self.import_evolution_messages(instance_name, eligible_records, contact_names)
                        with EVOLUTION_HISTORY_SYNC_LOCK:
                            EVOLUTION_HISTORY_SYNC_STATUS["contacts"] += counts[0]
                            EVOLUTION_HISTORY_SYNC_STATUS["conversations"] += counts[1]
                            EVOLUTION_HISTORY_SYNC_STATUS["messages"] += counts[2]
                        container = message_payload.get("messages") if isinstance(message_payload, dict) and isinstance(message_payload.get("messages"), dict) else message_payload
                        total_pages = int(container.get("pages") or container.get("totalPages") or 0) if isinstance(container, dict) else 0
                        descending_page = len(known_times) >= 2 and known_times[0] >= known_times[-1]
                        reached_cutoff = descending_page and min(known_times) < channel_cutoff
                        if reached_cutoff or (total_pages and page >= total_pages) or (not total_pages and len(records) < 200):
                            break
                        page += 1
                except Exception as error:
                    with EVOLUTION_HISTORY_SYNC_LOCK:
                        EVOLUTION_HISTORY_SYNC_STATUS["errors"].append({"instance": instance_name, "error": str(error)[:240]})
                finally:
                    with EVOLUTION_HISTORY_SYNC_LOCK:
                        EVOLUTION_HISTORY_SYNC_STATUS["instances_done"] += 1
        except Exception as error:
            with EVOLUTION_HISTORY_SYNC_LOCK:
                EVOLUTION_HISTORY_SYNC_STATUS["errors"].append({"instance": "Evolution", "error": str(error)[:240]})
        finally:
            with EVOLUTION_HISTORY_SYNC_LOCK:
                EVOLUTION_HISTORY_SYNC_STATUS["running"] = False
                EVOLUTION_HISTORY_SYNC_STATUS["phase"] = "Sincronização concluída"
                EVOLUTION_HISTORY_SYNC_STATUS["finished_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def start_evolution_history_sync(self) -> None:
        if not self.require_crm_feature("integrations"): return
        global EVOLUTION_HISTORY_SYNC_STATUS
        if not self.require_crc_access():
            return
        with connect() as db:
            if not self.crm_can_manage_automation(db):
                return self.send_json({"error": "Somente um responsÃ¡vel autorizado pode sincronizar os canais."}, HTTPStatus.FORBIDDEN)
        with EVOLUTION_HISTORY_SYNC_LOCK:
            if EVOLUTION_HISTORY_SYNC_STATUS["running"]:
                return self.send_json(dict(EVOLUTION_HISTORY_SYNC_STATUS), HTTPStatus.ACCEPTED)
            EVOLUTION_HISTORY_SYNC_STATUS = {
                "running": True, "instances_total": 0, "instances_done": 0,
                "contacts": 0, "conversations": 0, "messages": 0, "older_messages_removed": 0, "errors": [],
                "since": EVOLUTION_HISTORY_CUTOFF[:10],
                "phase": "Iniciando sincronização segura...",
                "started_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "finished_at": None,
            }
        threading.Thread(target=self.run_evolution_history_sync, daemon=True).start()
        self.send_json(dict(EVOLUTION_HISTORY_SYNC_STATUS), HTTPStatus.ACCEPTED)

    def get_evolution_sync_status(self) -> None:
        if not self.require_crm_feature("integrations"): return
        if not self.require_crc_access():
            return
        with EVOLUTION_HISTORY_SYNC_LOCK:
            status = json.loads(json.dumps(EVOLUTION_HISTORY_SYNC_STATUS))
        self.send_json(status)

    def connect_evolution_instance(self, payload: dict) -> None:
        if not self.require_crm_feature("integrations"): return
        if not self.require_crc_access(): return
        with connect() as db:
            if not self.crm_can_manage_automation(db):
                return self.send_json({"error": "Somente um responsÃ¡vel autorizado pode conectar nÃºmeros."}, HTTPStatus.FORBIDDEN)
        instance_name = str(payload.get("instance_name") or "Teste-CRM-IEA").strip()
        display_name = str(payload.get("display_name") or instance_name).strip()
        if not re.fullmatch(r"[A-Za-z0-9_-]{2,64}", instance_name):
            return self.send_json({"error": "Use apenas letras, números, hífen ou sublinhado no nome da instância."}, HTTPStatus.BAD_REQUEST)
        with connect() as db:
            existing_channel = db.execute(
                "SELECT evolution_base_url,evolution_api_key FROM crm_channels WHERE instance_name=?",
                (instance_name,),
            ).fetchone()
        configured_url, configured_key = self.evolution_credentials()
        channel_base_url = (existing_channel["evolution_base_url"] if existing_channel else None) or configured_url
        channel_api_key = (
            decrypt_integration_secret(existing_channel["evolution_api_key"])
            if existing_channel and existing_channel["evolution_api_key"] else configured_key
        )
        try:
            try:
                connected = self.evolution_api_request(
                    f"/instance/connect/{instance_name}", base_url=channel_base_url, api_key=channel_api_key
                )
            except RuntimeError as error:
                if "404" not in str(error): raise
                self.evolution_api_request(
                    "/instance/create", "POST",
                    {"instanceName": instance_name, "qrcode": True, "integration": "WHATSAPP-BAILEYS"},
                    base_url=channel_base_url, api_key=channel_api_key,
                )
                connected = self.evolution_api_request(
                    f"/instance/connect/{instance_name}", base_url=channel_base_url, api_key=channel_api_key
                )
            if not EVOLUTION_WEBHOOK_TOKEN:
                return self.send_json({"error": "Configure o segredo exclusivo do webhook Evolution antes de conectar o canal."}, HTTPStatus.SERVICE_UNAVAILABLE)
            webhook_url = f"{PUBLIC_APP_URL}/api/integrations/evolution/webhook?webhook_key={EVOLUTION_WEBHOOK_TOKEN}"
            def configure_webhook() -> None:
                try:
                    self.evolution_api_request(f"/webhook/set/{instance_name}", "POST", {"webhook": {
                        "enabled": True, "url": webhook_url, "webhookByEvents": False, "webhookBase64": False,
                        "events": ["MESSAGES_UPSERT", "MESSAGES_UPDATE", "CONNECTION_UPDATE"]
                    }}, base_url=channel_base_url, api_key=channel_api_key)
                except RuntimeError:
                    pass
            threading.Thread(target=configure_webhook, daemon=True).start()
            qr_code = connected.get("base64") or (connected.get("qrcode") or {}).get("base64")
            state = str(connected.get("state") or connected.get("connectionStatus") or "connecting")
            with connect() as db:
                db.execute("""INSERT INTO crm_channels(instance_name,display_name,active,connection_status)
                              VALUES(?,?,1,?) ON CONFLICT(instance_name) DO UPDATE SET display_name=excluded.display_name,
                              active=1,connection_status=excluded.connection_status,updated_at=datetime('now','localtime')""",
                           (instance_name, display_name, "Conectado" if state.lower() == "open" else "Aguardando QR"))
            self.send_json({"instance_name": instance_name, "state": state, "connected": state.lower() == "open", "qr_code": qr_code})
        except RuntimeError as error:
            self.send_json({"error": str(error)}, HTTPStatus.BAD_GATEWAY)

    def save_crm_channel(self, payload: dict) -> None:
        if not self.require_crm_feature("integrations"): return
        if not self.require_crc_access(): return
        with connect() as db:
            if not self.crm_can_manage_automation(db):
                return self.send_json({"error": "Somente um responsÃ¡vel autorizado pode cadastrar canais."}, HTTPStatus.FORBIDDEN)
        instance = str(payload.get("instance_name") or "").strip()
        display_name = str(payload.get("display_name") or instance).strip()
        base_url = str(payload.get("evolution_base_url") or "").strip().rstrip("/")
        api_key = str(payload.get("evolution_api_key") or "").strip()
        stored_api_key = encrypt_integration_secret(api_key) if api_key else ""
        if not instance: return self.send_json({"error": "Informe o nome da instância."}, HTTPStatus.BAD_REQUEST)
        with connect() as db:
            db.execute("""INSERT INTO crm_channels(instance_name,display_name,phone,evolution_base_url,evolution_api_key,active)
                          VALUES(?,?,?,?,?,?) ON CONFLICT(instance_name) DO UPDATE SET display_name=excluded.display_name,
                          phone=excluded.phone,evolution_base_url=excluded.evolution_base_url,
                          evolution_api_key=CASE WHEN excluded.evolution_api_key<>'' THEN excluded.evolution_api_key ELSE crm_channels.evolution_api_key END,
                          active=excluded.active,updated_at=datetime('now','localtime')""",
                       (instance, display_name, self.crm_phone(payload.get("phone")) or None, base_url or None, stored_api_key, 1 if payload.get("active", True) else 0))
            row = db.execute("SELECT id,instance_name,display_name,phone,active,connection_status FROM crm_channels WHERE instance_name=?", (instance,)).fetchone()
        self.send_json(dict(row), HTTPStatus.CREATED)

    def crm_record_event(self, db, conversation_id: int, event_type: str, details: dict | None = None,
                         actor_user_id: int | None = None, actor_name: str | None = None) -> None:
        actor_user_id = actor_user_id if actor_user_id is not None else self.authenticated_user.get("id")
        actor_name = actor_name or self.authenticated_user.get("name") or "Sistema"
        db.execute("""INSERT INTO crm_conversation_events
                      (conversation_id,event_type,actor_user_id,actor_name,details_json)
                      VALUES(?,?,?,?,?)""",
                   (conversation_id, event_type, actor_user_id, actor_name,
                    json.dumps(details or {}, ensure_ascii=False, separators=(",", ":"))))

    @staticmethod
    def crm_activate_due_returns(db) -> int:
        due = db.execute("""SELECT id FROM crm_conversations
                            WHERE scheduled_return_at IS NOT NULL
                              AND datetime(scheduled_return_at)<=datetime('now','localtime')""").fetchall()
        for row in due:
            db.execute("""UPDATE crm_conversations SET status='Aberta',pipeline_stage='Novo',queue_name='Retorno programado',
                          assigned_user_id=NULL,assigned_at=NULL,resolved_at=NULL,resolved_by_user_id=NULL,
                          queue_entered_at=scheduled_return_at,scheduled_return_at=NULL,reopened_at=datetime('now','localtime'),
                          updated_at=datetime('now','localtime') WHERE id=?""", (row["id"],))
            db.execute("""INSERT INTO crm_conversation_events(conversation_id,event_type,actor_name,details_json)
                          VALUES(?,'return.reopened','Sistema','{}')""", (row["id"],))
        return len(due)

    @staticmethod
    def crm_ai_handoff_reason(text: str) -> str | None:
        normalized = unicodedata.normalize("NFKD", str(text or "")).encode("ascii", "ignore").decode().lower()
        rules = (
            ("intenção de agendamento", ("agendar", "marcar horario", "marcar consulta", "qual horario", "tem vaga")),
            ("objeção comercial", ("muito caro", "desconto", "nao consigo pagar", "valor", "preco")),
            ("dúvida não resolvida", ("nao entendi", "tenho duvida", "pode explicar", "como funciona")),
            ("frustração detectada", ("reclamacao", "insatisfeito", "pessimo", "absurdo", "nao gostei")),
            ("humano solicitado", ("atendente", "falar com pessoa", "falar com humano", "recepcao")),
        )
        for reason, needles in rules:
            if any(needle in normalized for needle in needles):
                return reason
        return None

    def get_crm_quick_replies(self) -> None:
        if not self.require_crc_access(): return
        if not self.require_crm_feature("inbox"):
            return
        with connect() as db:
            rows = db.execute("""SELECT id,title,content,category,active,created_at,updated_at
                                 FROM crm_quick_replies WHERE active=1 ORDER BY category,title COLLATE NOCASE""").fetchall()
        self.send_json({"items": [dict(row) for row in rows]})

    def create_crm_quick_reply(self, payload: dict) -> None:
        if not self.require_crc_access(): return
        if not self.require_crm_feature("inbox"):
            return
        title = str(payload.get("title") or "").strip()[:80]
        content = str(payload.get("content") or "").strip()[:2000]
        category = str(payload.get("category") or "Geral").strip()[:50] or "Geral"
        if not title or not content:
            return self.send_json({"error": "Informe o título e o texto da resposta rápida."}, HTTPStatus.BAD_REQUEST)
        with connect() as db:
            cur = db.execute("""INSERT INTO crm_quick_replies(title,content,category,created_by_user_id)
                                VALUES(?,?,?,?)""", (title, content, category, self.authenticated_user["id"]))
            row = db.execute("SELECT * FROM crm_quick_replies WHERE id=?", (cur.lastrowid,)).fetchone()
        self.send_json(dict(row), HTTPStatus.CREATED)

    def update_crm_quick_reply(self, reply_id: int, payload: dict) -> None:
        if not self.require_crc_access(): return
        if not self.require_crm_feature("inbox"):
            return
        with connect() as db:
            current = db.execute("SELECT * FROM crm_quick_replies WHERE id=?", (reply_id,)).fetchone()
            if not current: return self.send_json({"error": "Resposta rápida não encontrada."}, HTTPStatus.NOT_FOUND)
            title = str(payload.get("title", current["title"]) or "").strip()[:80]
            content = str(payload.get("content", current["content"]) or "").strip()[:2000]
            category = str(payload.get("category", current["category"]) or "Geral").strip()[:50]
            active = 1 if payload.get("active", current["active"]) else 0
            if not title or not content: return self.send_json({"error": "Título e texto são obrigatórios."}, HTTPStatus.BAD_REQUEST)
            db.execute("""UPDATE crm_quick_replies SET title=?,content=?,category=?,active=?,
                          updated_at=datetime('now','localtime') WHERE id=?""", (title, content, category, active, reply_id))
        self.send_json({"updated": True, "id": reply_id})

    def delete_crm_quick_reply(self, reply_id: int) -> None:
        if not self.require_crc_access(): return
        if not self.require_crm_feature("inbox"):
            return
        with connect() as db:
            changed = db.execute("DELETE FROM crm_quick_replies WHERE id=?", (reply_id,)).rowcount
        if not changed: return self.send_json({"error": "Resposta rápida não encontrada."}, HTTPStatus.NOT_FOUND)
        self.send_json({"deleted": True, "id": reply_id})

    def get_crm_conversation_timeline(self, conversation_id: int) -> None:
        if not self.require_crc_access(): return
        if not self.require_crm_any_feature(CRM_WORKSPACE_FEATURES):
            return
        scope_sql, scope_params = self.crm_channel_scope_clause("ch")
        with connect() as db:
            current = db.execute("SELECT contact_id,channel_id FROM crm_conversations WHERE id=?", (conversation_id,)).fetchone()
            if not current: return self.send_json({"error": "Conversa não encontrada."}, HTTPStatus.NOT_FOUND)
            if not self.crm_channel_allowed(db, current["channel_id"]):
                return self.send_json({"error": "Você não possui acesso a este canal."}, HTTPStatus.FORBIDDEN)
            conversations = db.execute("""SELECT cv.id,cv.status,cv.resolution_reason,cv.scheduled_return_at,
                                      cv.created_at,cv.resolved_at,ch.display_name AS channel_name
                                      FROM crm_conversations cv JOIN crm_channels ch ON ch.id=cv.channel_id
                                      WHERE cv.contact_id=? AND {scope_sql}
                                      ORDER BY cv.updated_at DESC""".format(scope_sql=scope_sql),
                                      (current["contact_id"], *scope_params)).fetchall()
            events = db.execute("""SELECT ev.id,ev.conversation_id,ev.event_type,ev.actor_name,ev.details_json,ev.created_at,
                               ch.display_name AS channel_name FROM crm_conversation_events ev
                               JOIN crm_conversations cv ON cv.id=ev.conversation_id JOIN crm_channels ch ON ch.id=cv.channel_id
                               WHERE cv.contact_id=? AND {scope_sql}
                               ORDER BY ev.id DESC LIMIT 300""".format(scope_sql=scope_sql),
                               (current["contact_id"], *scope_params)).fetchall()
        self.send_json({"conversations": [dict(r) for r in conversations], "events": [dict(r) for r in events]})

    def require_crm_n8n_manager(self) -> bool:
        if not self.require_crc_access():
            return False
        if not self.require_crm_feature("integrations"):
            return False
        with connect() as db:
            if not self.crm_can_manage_automation(db):
                self.send_json(
                    {"error": "A supervisão de automações não está liberada para o seu acesso."},
                    HTTPStatus.FORBIDDEN,
                )
                return False
        return True

    def crm_n8n_config(self) -> dict | None:
        with connect() as db:
            persistent = db.execute(
                "SELECT id,api_base_url,api_token,active,updated_at FROM crm_n8n_config WHERE id=1"
            ).fetchone()
            row = db.execute("""SELECT id,name,description,api_base_url,api_token,active,updated_at
                                FROM api_integrations
                                WHERE lower(name) IN ('n8n','n8n crm','n8n · automações e ia')
                                ORDER BY CASE WHEN lower(name)='n8n' THEN 0 ELSE 1 END,id LIMIT 1""").fetchone()
        if persistent and persistent["api_base_url"] and persistent["api_token"]:
            return {
                "id": persistent["id"],
                "name": "n8n",
                "description": "Controla fluxos e eventos do CRM.",
                "api_base_url": persistent["api_base_url"],
                "api_token": decrypt_integration_secret(persistent["api_token"]),
                "active": persistent["active"],
                "updated_at": persistent["updated_at"],
            }
        config = dict(row) if row else None
        if config and config.get("api_base_url") and config.get("api_token"):
            config["api_token"] = decrypt_integration_secret(config["api_token"])
            self.persist_crm_n8n_config(
                str(config["api_base_url"]), str(config["api_token"]), int(config.get("active") or 0)
            )
            return config
        try:
            stored = json.loads(N8N_CONFIG_PATH.read_text(encoding="utf-8"))
        except (OSError, ValueError, TypeError):
            return config
        if not isinstance(stored, dict):
            return config
        api_base_url = str(stored.get("api_base_url") or "").strip().rstrip("/")
        api_token = str(decrypt_integration_secret(stored.get("api_token")) or "").strip()
        if not api_base_url or not api_token:
            return config
        recovered = {
            "id": config.get("id") if config else None,
            "name": "n8n",
            "description": "Controla fluxos, execuções e eventos de automação do CRM.",
            "api_base_url": api_base_url,
            "api_token": api_token,
            "active": 1 if stored.get("active", True) else 0,
            "updated_at": stored.get("updated_at"),
        }
        self.persist_crm_n8n_config(api_base_url, api_token, int(recovered["active"]))
        return recovered

    def persist_crm_n8n_config(self, api_base_url: str, api_token: str, active: int) -> None:
        saved_at = datetime.now(CLINIC_TIMEZONE).isoformat(timespec="seconds")
        encrypted_token = encrypt_integration_secret(api_token)
        with connect() as db:
            db.execute(
                """INSERT INTO crm_n8n_config(id,api_base_url,api_token,active,updated_at)
                   VALUES(1,?,?,?,?)
                   ON CONFLICT(id) DO UPDATE SET api_base_url=excluded.api_base_url,
                     api_token=excluded.api_token,active=excluded.active,
                     updated_at=excluded.updated_at""",
                (api_base_url.rstrip("/"), encrypted_token, int(bool(active)), saved_at),
            )
        DATA.mkdir(parents=True, exist_ok=True)
        temporary_path = N8N_CONFIG_PATH.with_suffix(".tmp")
        stored = {
            "api_base_url": api_base_url.rstrip("/"),
            "api_token": encrypted_token,
            "active": bool(active),
            "updated_at": saved_at,
        }
        temporary_path.write_text(
            json.dumps(stored, ensure_ascii=False, separators=(",", ":")),
            encoding="utf-8",
        )
        try:
            os.chmod(temporary_path, 0o600)
        except OSError:
            pass
        os.replace(temporary_path, N8N_CONFIG_PATH)
        try:
            os.chmod(N8N_CONFIG_PATH, 0o600)
        except OSError:
            pass
        with N8N_OVERVIEW_CACHE_LOCK:
            N8N_OVERVIEW_CACHE["payload"] = None
            N8N_OVERVIEW_CACHE["updated_at"] = 0.0

    def crm_n8n_request(self, path: str, *, method: str = "GET", payload: dict | None = None,
                        timeout: int = 12) -> dict:
        config = self.crm_n8n_config()
        if not config or not config.get("active"):
            raise RuntimeError("A integração n8n ainda não está ativa.")
        base_url = str(config.get("api_base_url") or "").strip().rstrip("/")
        api_key = str(config.get("api_token") or "").strip()
        if not base_url or not api_key:
            raise RuntimeError("Configure a URL e a chave oficial do n8n.")
        url = f"{base_url}{path if path.startswith('/') else '/' + path}"
        body = json.dumps(payload or {}, ensure_ascii=False).encode("utf-8") if payload is not None else None
        request = Request(url, data=body, method=method, headers={
            "Accept": "application/json",
            "Content-Type": "application/json",
            "X-N8N-API-KEY": api_key,
        })
        try:
            with urlopen(request, timeout=timeout) as response:
                raw = response.read().decode("utf-8", errors="replace")
                return json.loads(raw) if raw.strip() else {}
        except HTTPError as error:
            detail = error.read().decode("utf-8", errors="replace")[:400]
            raise RuntimeError(f"n8n respondeu {error.code}: {detail or error.reason}") from error
        except (URLError, TimeoutError, OSError) as error:
            raise RuntimeError(f"Não foi possível alcançar o n8n: {error}") from error
        except json.JSONDecodeError as error:
            raise RuntimeError("O n8n respondeu em um formato inesperado.") from error

    def get_crm_n8n_config(self) -> None:
        if not self.require_crm_n8n_manager():
            return
        config = self.crm_n8n_config()
        self.send_json({
            "configured": bool(config and config.get("api_base_url") and config.get("api_token")),
            "active": bool(config and config.get("active")),
            "api_base_url": config.get("api_base_url") if config else "",
            "token_configured": bool(config and config.get("api_token")),
            "updated_at": config.get("updated_at") if config else None,
        })

    def save_crm_n8n_config(self, payload: dict) -> None:
        if not self.require_crm_n8n_manager():
            return
        api_base_url_input = str(payload.get("api_base_url") or "").strip()
        parsed_base_url = urlparse(api_base_url_input)
        api_base_url = (
            f"{parsed_base_url.scheme}://{parsed_base_url.netloc}"
            if parsed_base_url.scheme in {"http", "https"} and parsed_base_url.netloc
            else ""
        )
        api_token = str(payload.get("api_token") or "").strip()
        active = 1 if payload.get("active", True) else 0
        if not api_base_url:
            return self.send_json({"error": "Informe a URL pública da instância n8n."}, HTTPStatus.BAD_REQUEST)
        current = self.crm_n8n_config()
        token = api_token or (str(current.get("api_token") or "") if current else "")
        if not token:
            return self.send_json({"error": "Informe a chave criada em Settings > n8n API."}, HTTPStatus.BAD_REQUEST)
        try:
            request = Request(f"{api_base_url}/api/v1/workflows?limit=1", headers={
                "Accept": "application/json",
                "X-N8N-API-KEY": token,
            })
            with urlopen(request, timeout=12) as response:
                test_payload = json.loads(response.read().decode("utf-8", errors="replace") or "{}")
        except HTTPError as error:
            detail = error.read().decode("utf-8", errors="replace")[:300]
            return self.send_json({"error": f"O n8n recusou a configuração ({error.code}): {detail or error.reason}"}, HTTPStatus.BAD_GATEWAY)
        except (URLError, TimeoutError, OSError, json.JSONDecodeError) as error:
            return self.send_json({"error": f"Não foi possível validar essa instância do n8n: {error}"}, HTTPStatus.BAD_GATEWAY)
        try:
            self.persist_crm_n8n_config(api_base_url, token, active)
        except OSError as error:
            return self.send_json({
                "error": f"Não foi possível proteger a configuração persistente do n8n: {error}"
            }, HTTPStatus.INTERNAL_SERVER_ERROR)
        with connect() as db:
            existing = db.execute("""SELECT id FROM api_integrations
                                     WHERE lower(name) IN ('n8n','n8n crm','n8n · automações e ia')
                                     ORDER BY id LIMIT 1""").fetchone()
            if existing:
                db.execute("""UPDATE api_integrations SET name='n8n',description=?,api_base_url=?,
                              api_token=?,active=?,updated_at=datetime('now','localtime'),updated_by=?
                              WHERE id=?""", (
                    "Controla fluxos, execuções e eventos de automação do CRM.",
                    api_base_url, encrypt_integration_secret(token), active, self.authenticated_user["id"], existing["id"],
                ))
            else:
                db.execute("""INSERT INTO api_integrations
                              (name,description,api_base_url,api_token,active,sync_interval_seconds,updated_by)
                              VALUES('n8n',?,?,?,?,60,?)""", (
                    "Controla fluxos, execuções e eventos de automação do CRM.",
                    api_base_url, encrypt_integration_secret(token), active, self.authenticated_user["id"],
                ))
        self.send_json({
            "saved": True,
            "connected": True,
            "workflows_visible": len(test_payload.get("data") or []),
            "message": "n8n conectado com segurança.",
        })

    def get_crm_n8n_overview(self, query: dict) -> None:
        if not self.require_crm_n8n_manager():
            return
        try:
            limit = max(10, min(100, int(str(query.get("limit", ["50"])[0]))))
        except (TypeError, ValueError):
            limit = 50
        force_refresh = str(query.get("refresh", [""])[0]).strip().lower() in {"1", "true", "yes"}
        # Em navegações entre telas, reaproveita por dois minutos a última
        # auditoria bem-sucedida. Assim os fluxos não desaparecem se o n8n
        # estiver momentaneamente lento, e o botão "Atualizar" continua
        # podendo buscar uma versão nova quando necessário.
        with N8N_OVERVIEW_CACHE_LOCK:
            cached_payload = N8N_OVERVIEW_CACHE.get("payload")
            cached_at = float(N8N_OVERVIEW_CACHE.get("updated_at") or 0)
        if not force_refresh and isinstance(cached_payload, dict) and time.monotonic() - cached_at < 120:
            response_payload = dict(cached_payload)
            response_payload["cached"] = True
            return self.send_json(response_payload)
        try:
            # A listagem de execuções do n8n pode ser muito grande mesmo com
            # paginação. Ela não pode impedir a auditoria dos fluxos de abrir.
            # Os detalhes de execução são carregados sob demanda em cada fluxo.
            workflows_payload = self.crm_n8n_request(f"/api/v1/workflows?limit={limit}")
        except RuntimeError as error:
            return self.send_json({"error": str(error)}, HTTPStatus.BAD_GATEWAY)
        executions_payload = {"data": [], "nextCursor": None}
        workflows_raw = workflows_payload.get("data") or []
        workflow_names = {str(item.get("id")): str(item.get("name") or "Fluxo sem nome") for item in workflows_raw}
        workflows = [{
            "id": str(item.get("id") or ""),
            "name": str(item.get("name") or "Fluxo sem nome"),
            "active": bool(item.get("active")),
            "updated_at": item.get("updatedAt"),
            "created_at": item.get("createdAt"),
            "tags": [str(tag.get("name") or "") for tag in (item.get("tags") or []) if tag.get("name")],
        } for item in workflows_raw]
        executions = []
        success_count = error_count = running_count = 0
        for item in executions_payload.get("data") or []:
            status = str(item.get("status") or ("success" if item.get("finished") else "running")).lower()
            if status == "success":
                success_count += 1
            elif status in {"error", "crashed", "failed"}:
                error_count += 1
            elif status in {"running", "waiting", "new"}:
                running_count += 1
            workflow_id = str(item.get("workflowId") or item.get("workflowData", {}).get("id") or "")
            executions.append({
                "id": str(item.get("id") or ""),
                "workflow_id": workflow_id,
                "workflow_name": workflow_names.get(workflow_id) or str(item.get("workflowData", {}).get("name") or "Fluxo"),
                "status": status,
                "started_at": item.get("startedAt"),
                "stopped_at": item.get("stoppedAt"),
                "mode": item.get("mode"),
            })
        workflow_stats: dict[str, dict] = {}
        for execution in executions:
            stats = workflow_stats.setdefault(execution["workflow_id"], {
                "success_count": 0,
                "failure_count": 0,
                "running_count": 0,
                "last_execution_at": None,
                "last_execution_status": None,
            })
            if execution["status"] == "success":
                stats["success_count"] += 1
            elif execution["status"] in {"error", "crashed", "failed"}:
                stats["failure_count"] += 1
            elif execution["status"] in {"running", "waiting", "new"}:
                stats["running_count"] += 1
            if not stats["last_execution_at"] and execution.get("started_at"):
                stats["last_execution_at"] = execution["started_at"]
                stats["last_execution_status"] = execution["status"]
        with connect() as db:
            settings_rows = db.execute(
                """SELECT workflow_id,workflow_kind,manual_enabled,requires_confirmation,
                          max_items,test_mode,webhook_path,source_label,channel_label
                   FROM crm_n8n_workflow_settings"""
            ).fetchall()
        settings_by_workflow = {str(row["workflow_id"]): dict(row) for row in settings_rows}
        integration_labels = {
            "googlesheets": "Google Sheets",
            "googledrive": "Google Drive",
            "postgres": "PostgreSQL",
            "mysql": "MySQL",
            "httprequest": "API HTTP",
            "webhook": "Webhook",
            "openai": "OpenAI",
            "anthropic": "Anthropic",
            "gemini": "Gemini",
            "evolution": "Evolution",
            "whatsapp": "WhatsApp",
        }
        for raw_workflow, workflow in zip(workflows_raw, workflows):
            nodes = raw_workflow.get("nodes") or []
            classification = self.classify_crm_n8n_workflow(nodes)
            trigger_names = []
            integrations = []
            frequency = "Sob demanda"
            for node in nodes:
                node_type = str(node.get("type") or "").lower()
                node_name = str(node.get("name") or "")
                if "scheduletrigger" in node_type:
                    trigger_names.append("Horário")
                    frequency = self.describe_crm_n8n_schedule(
                        node.get("parameters") if isinstance(node.get("parameters"), dict) else {}
                    )
                elif "webhook" in node_type:
                    trigger_names.append("Webhook")
                elif "manualtrigger" in node_type:
                    trigger_names.append("Manual")
                elif "trigger" in node_type:
                    trigger_names.append(node_name or "Evento")
                searchable = f"{node_type} {node_name}".lower()
                for key, label in integration_labels.items():
                    if key in searchable and label not in integrations:
                        integrations.append(label)
            workflow_id = workflow["id"]
            workflow.update(workflow_stats.get(workflow_id, {
                "success_count": 0,
                "failure_count": 0,
                "running_count": 0,
                "last_execution_at": None,
                "last_execution_status": None,
            }))
            workflow["trigger_types"] = list(dict.fromkeys(trigger_names)) or ["Automático"]
            workflow["frequency"] = frequency
            workflow["integrations"] = integrations
            workflow["classification"] = classification["kind"]
            workflow["settings"] = settings_by_workflow.get(workflow_id, {
                "workflow_kind": classification["kind"],
                "manual_enabled": 0,
                "requires_confirmation": 1,
                "max_items": 25,
                "test_mode": 1,
                "webhook_path": None,
                "source_label": "",
                "channel_label": "",
            })
        response_payload = {
            "configured": True,
            "workflows": workflows,
            "executions": executions,
            "summary": {
                "workflows_total": len(workflows),
                "workflows_active": sum(1 for item in workflows if item["active"]),
                "executions_total": len(executions),
                "success": success_count,
                "errors": error_count,
                "running": running_count,
            },
            "next_workflows_cursor": workflows_payload.get("nextCursor"),
            "next_executions_cursor": executions_payload.get("nextCursor"),
        }
        with N8N_OVERVIEW_CACHE_LOCK:
            N8N_OVERVIEW_CACHE["payload"] = response_payload
            N8N_OVERVIEW_CACHE["updated_at"] = time.monotonic()
        self.send_json(response_payload)

    @staticmethod
    def classify_crm_n8n_workflow(nodes: list[dict]) -> dict:
        trigger_types = [str(node.get("type") or "").lower() for node in nodes
                         if "trigger" in str(node.get("type") or "").lower()
                         or str(node.get("type") or "").lower().endswith(".webhook")]
        has_manual = any("manualtrigger" in item for item in trigger_types)
        has_schedule = any("scheduletrigger" in item for item in trigger_types)
        has_webhook = any("webhook" in item for item in trigger_types)
        if has_webhook and not has_manual and not has_schedule:
            kind = "response"
        elif has_schedule and not has_webhook:
            kind = "scheduled"
        elif has_manual and not has_schedule and not has_webhook:
            kind = "manual"
        elif sum((has_manual, has_schedule, has_webhook)) > 1:
            kind = "hybrid"
        else:
            kind = "automatic"
        return {
            "kind": kind,
            "has_manual_trigger": has_manual,
            "has_schedule_trigger": has_schedule,
            "has_webhook_trigger": has_webhook,
        }

    @staticmethod
    def describe_crm_n8n_schedule(parameters: dict) -> str:
        intervals = parameters.get("rule", {}).get("interval", [])
        if not isinstance(intervals, list) or not intervals:
            return "Agendado no n8n"
        descriptions = []
        units = {
            "seconds": ("segundo", "segundos", "secondsInterval"),
            "minutes": ("minuto", "minutos", "minutesInterval"),
            "hours": ("hora", "horas", "hoursInterval"),
            "days": ("dia", "dias", "daysInterval"),
            "weeks": ("semana", "semanas", "weeksInterval"),
            "months": ("mês", "meses", "monthsInterval"),
        }
        for rule in intervals:
            if not isinstance(rule, dict):
                continue
            field = str(rule.get("field") or "").lower()
            if field == "cronexpression":
                descriptions.append(f"Cron: {str(rule.get('expression') or '').strip()}")
                continue
            singular, plural, key = units.get(field, (field or "intervalo", field or "intervalos", "interval"))
            try:
                amount = int(rule.get(key) or 1)
            except (TypeError, ValueError):
                amount = 1
            descriptions.append(f"A cada {amount} {singular if amount == 1 else plural}")
        return " · ".join(filter(None, descriptions)) or "Agendado no n8n"

    def get_crm_n8n_workflow_detail(self, workflow_id: str) -> None:
        if not self.require_crm_n8n_manager():
            return
        if not re.fullmatch(r"[A-Za-z0-9_-]+", workflow_id):
            return self.send_json({"error": "Fluxo inválido."}, HTTPStatus.BAD_REQUEST)
        try:
            workflow = self.crm_n8n_request(f"/api/v1/workflows/{quote(workflow_id)}")
            executions_payload = self.crm_n8n_request(
                f"/api/v1/executions?{urlencode({'workflowId': workflow_id, 'limit': 30, 'includeData': 'false'})}"
            )
        except RuntimeError as error:
            return self.send_json({"error": str(error)}, HTTPStatus.BAD_GATEWAY)
        nodes = workflow.get("nodes") or []
        classification = self.classify_crm_n8n_workflow(nodes)
        node_summary = []
        webhook_nodes = []
        for node in nodes:
            node_type = str(node.get("type") or "")
            parameters = node.get("parameters") if isinstance(node.get("parameters"), dict) else {}
            item = {
                "name": str(node.get("name") or "Nó"),
                "type": node_type,
                "disabled": bool(node.get("disabled")),
            }
            node_summary.append(item)
            if "webhook" in node_type.lower():
                webhook_nodes.append({
                    "name": item["name"],
                    "path": parameters.get("path"),
                    "method": parameters.get("httpMethod") or "GET",
                })
        with connect() as db:
            setting = db.execute(
                """SELECT workflow_kind,manual_enabled,webhook_path,webhook_method,
                          source_label,channel_label,requires_confirmation,max_items,test_mode,updated_at
                   FROM crm_n8n_workflow_settings WHERE workflow_id=?""",
                (workflow_id,),
            ).fetchone()
            local_runs = db.execute(
                """SELECT id,run_key,status,mode,total_items,processed_items,sent_items,
                          delivered_items,replied_items,failed_items,appointment_items,
                          handoff_items,started_at,finished_at
                   FROM crm_n8n_runs WHERE workflow_id=?
                   ORDER BY id DESC LIMIT 20""",
                (workflow_id,),
            ).fetchall()
            versions = db.execute(
                """SELECT v.id,v.action,v.created_at,u.name AS created_by_name
                   FROM crm_n8n_workflow_versions v
                   LEFT JOIN users u ON u.id=v.created_by_user_id
                   WHERE v.workflow_id=?
                   ORDER BY v.id DESC LIMIT 10""",
                (workflow_id,),
            ).fetchall()
        executions = []
        for execution in executions_payload.get("data") or []:
            status = str(execution.get("status") or ("success" if execution.get("finished") else "running")).lower()
            executions.append({
                "id": str(execution.get("id") or ""),
                "status": status,
                "mode": execution.get("mode"),
                "started_at": execution.get("startedAt"),
                "stopped_at": execution.get("stoppedAt"),
            })
        failed_nodes = []
        failed_execution = next(
            (item for item in executions if item["status"] in {"error", "failed", "crashed"} and item["id"]),
            None,
        )
        if failed_execution:
            try:
                failed_detail = self.crm_n8n_request(
                    f"/api/v1/executions/{quote(failed_execution['id'])}?includeData=true"
                )
                run_data = (
                    failed_detail.get("data", {})
                    .get("resultData", {})
                    .get("runData", {})
                )
                if isinstance(run_data, dict):
                    for node_name, attempts in run_data.items():
                        if not isinstance(attempts, list):
                            continue
                        for attempt in attempts:
                            error = attempt.get("error") if isinstance(attempt, dict) else None
                            if not isinstance(error, dict):
                                continue
                            failed_nodes.append({
                                "name": str(node_name),
                                "message": str(error.get("message") or error.get("description") or "Falha no nó"),
                                "execution_id": failed_execution["id"],
                            })
                            break
            except RuntimeError:
                pass
        self.send_json({
            "id": workflow_id,
            "name": str(workflow.get("name") or "Fluxo"),
            "active": bool(workflow.get("active")),
            "created_at": workflow.get("createdAt"),
            "updated_at": workflow.get("updatedAt"),
            "classification": classification,
            "nodes_total": len(nodes),
            "nodes": node_summary,
            "webhooks": webhook_nodes,
            "settings": dict(setting) if setting else {
                "workflow_kind": classification["kind"],
                "manual_enabled": 0,
                "webhook_path": None,
                "webhook_method": "POST",
                "source_label": "",
                "channel_label": "",
                "requires_confirmation": 1,
                "max_items": 25,
                "test_mode": 1,
                "updated_at": None,
            },
            "executions": executions,
            "runs": [dict(row) for row in local_runs],
            "failed_nodes": failed_nodes,
            "versions": [dict(row) for row in versions],
        })

    def save_crm_n8n_workflow_settings(self, workflow_id: str, payload: dict) -> None:
        if not self.require_crm_n8n_manager():
            return
        if not re.fullmatch(r"[A-Za-z0-9_-]+", workflow_id):
            return self.send_json({"error": "Fluxo inválido."}, HTTPStatus.BAD_REQUEST)
        workflow_kind = str(payload.get("workflow_kind") or "automatic").strip().lower()
        if workflow_kind not in {"automatic", "scheduled", "response", "manual", "hybrid", "critical"}:
            return self.send_json({"error": "Classificação de fluxo inválida."}, HTTPStatus.BAD_REQUEST)
        webhook_path = str(payload.get("webhook_path") or "").strip().strip("/")
        manual_enabled = 1 if payload.get("manual_enabled") else 0
        if manual_enabled and not webhook_path:
            return self.send_json({"error": "Informe o webhook seguro usado para executar este fluxo."}, HTTPStatus.BAD_REQUEST)
        if webhook_path and not re.fullmatch(r"[A-Za-z0-9_./-]+", webhook_path):
            return self.send_json({"error": "Caminho de webhook inválido."}, HTTPStatus.BAD_REQUEST)
        webhook_method = str(payload.get("webhook_method") or "POST").strip().upper()
        if webhook_method not in {"POST", "GET"}:
            return self.send_json({"error": "Método de webhook inválido."}, HTTPStatus.BAD_REQUEST)
        source_label = str(payload.get("source_label") or "").strip()[:120]
        channel_label = str(payload.get("channel_label") or "").strip()[:120]
        try:
            max_items = max(1, min(5000, int(payload.get("max_items") or 25)))
        except (TypeError, ValueError):
            return self.send_json({"error": "Limite de contatos inválido."}, HTTPStatus.BAD_REQUEST)
        try:
            workflow = self.crm_n8n_request(f"/api/v1/workflows/{quote(workflow_id)}")
        except RuntimeError as error:
            return self.send_json({"error": str(error)}, HTTPStatus.BAD_GATEWAY)
        workflow_name = str(workflow.get("name") or "Fluxo")
        with connect() as db:
            db.execute(
                """INSERT INTO crm_n8n_workflow_settings
                   (workflow_id,workflow_name,workflow_kind,manual_enabled,webhook_path,
                    webhook_method,source_label,channel_label,requires_confirmation,max_items,
                    test_mode,updated_by_user_id,updated_at)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?,?,datetime('now','localtime'))
                   ON CONFLICT(workflow_id) DO UPDATE SET
                    workflow_name=excluded.workflow_name,workflow_kind=excluded.workflow_kind,
                    manual_enabled=excluded.manual_enabled,webhook_path=excluded.webhook_path,
                    webhook_method=excluded.webhook_method,source_label=excluded.source_label,
                    channel_label=excluded.channel_label,
                    requires_confirmation=excluded.requires_confirmation,max_items=excluded.max_items,
                    test_mode=excluded.test_mode,updated_by_user_id=excluded.updated_by_user_id,
                    updated_at=datetime('now','localtime')""",
                (
                    workflow_id, workflow_name, workflow_kind, manual_enabled, webhook_path or None,
                    webhook_method, source_label, channel_label,
                    1 if payload.get("requires_confirmation", True) else 0,
                    max_items, 1 if payload.get("test_mode", True) else 0,
                    self.authenticated_user["id"],
                ),
            )
        self.send_json({"saved": True, "workflow_id": workflow_id, "workflow_name": workflow_name})

    def run_crm_n8n_workflow(self, workflow_id: str, payload: dict) -> None:
        if not self.require_crm_n8n_manager():
            return
        if not re.fullmatch(r"[A-Za-z0-9_-]+", workflow_id):
            return self.send_json({"error": "Fluxo inválido."}, HTTPStatus.BAD_REQUEST)
        with connect() as db:
            setting = db.execute(
                """SELECT * FROM crm_n8n_workflow_settings WHERE workflow_id=?""",
                (workflow_id,),
            ).fetchone()
        if not setting or not setting["manual_enabled"] or not setting["webhook_path"]:
            return self.send_json(
                {"error": "Este fluxo ainda não foi liberado para execução manual pelo CRM."},
                HTTPStatus.CONFLICT,
            )
        mode = str(payload.get("mode") or ("test" if setting["test_mode"] else "production")).lower()
        if mode not in {"test", "production"}:
            return self.send_json({"error": "Modo de execução inválido."}, HTTPStatus.BAD_REQUEST)
        if mode == "production" and setting["requires_confirmation"] and not payload.get("confirmed"):
            return self.send_json({"error": "Confirme explicitamente a execução em produção."}, HTTPStatus.CONFLICT)
        try:
            requested_limit = int(payload.get("limit") or setting["max_items"])
        except (TypeError, ValueError):
            return self.send_json({"error": "Limite de contatos inválido."}, HTTPStatus.BAD_REQUEST)
        limit = max(1, min(int(setting["max_items"]), requested_limit))
        request_key = str(payload.get("request_key") or "").strip()
        run_key = request_key if re.fullmatch(r"[A-Za-z0-9_.:-]{12,120}", request_key) else f"crm-{secrets.token_urlsafe(18)}"
        event_token = secrets.token_urlsafe(32)
        request_payload = {
            "crm_run_key": run_key,
            "crm_workflow_id": workflow_id,
            "crm_mode": mode,
            "crm_limit": limit,
            "crm_requested_by": self.authenticated_user["name"],
            "crm_source": str(setting["source_label"] or ""),
            "crm_channel": str(setting["channel_label"] or ""),
            "crm_event_token": event_token,
            "parameters": payload.get("parameters") if isinstance(payload.get("parameters"), dict) else {},
        }
        stored_request_payload = {
            key: value for key, value in request_payload.items() if key != "crm_event_token"
        }
        stored_request_payload["crm_event_token_hash"] = hashlib.sha256(
            event_token.encode("utf-8")
        ).hexdigest()
        with connect() as db:
            try:
                cursor = db.execute(
                    """INSERT INTO crm_n8n_runs
                       (run_key,workflow_id,workflow_name,mode,status,requested_by_user_id,
                        request_payload_json,started_at,updated_at)
                       VALUES(?,?,?,?,?,?,?,datetime('now','localtime'),datetime('now','localtime'))""",
                    (
                        run_key, workflow_id, setting["workflow_name"], mode, "requested",
                        self.authenticated_user["id"],
                        json.dumps(stored_request_payload, ensure_ascii=False, separators=(",", ":")),
                    ),
                )
            except IntegrityError:
                existing = db.execute("SELECT id,status FROM crm_n8n_runs WHERE run_key=?", (run_key,)).fetchone()
                return self.send_json({
                    "duplicate": True,
                    "run_id": existing["id"],
                    "status": existing["status"],
                })
            run_id = cursor.lastrowid
        config = self.crm_n8n_config()
        base_url = str(config.get("api_base_url") or "").rstrip("/")
        webhook_url = f"{base_url}/webhook/{str(setting['webhook_path']).lstrip('/')}"
        method = str(setting["webhook_method"] or "POST").upper()
        body = json.dumps(request_payload, ensure_ascii=False).encode("utf-8") if method == "POST" else None
        if method == "GET":
            webhook_url = f"{webhook_url}?{urlencode({key: value for key, value in request_payload.items() if key != 'parameters'})}"
        request = Request(
            webhook_url,
            data=body,
            method=method,
            headers={"Accept": "application/json", "Content-Type": "application/json"},
        )
        try:
            with urlopen(request, timeout=30) as response:
                raw = response.read().decode("utf-8", errors="replace")
                response_payload = json.loads(raw) if raw.strip().startswith(("{", "[")) else {}
            execution_id = str(
                response_payload.get("execution_id")
                or response_payload.get("executionId")
                or ""
            ) if isinstance(response_payload, dict) else ""
            with connect() as db:
                db.execute(
                    """UPDATE crm_n8n_runs SET status='running',n8n_execution_id=?,
                       updated_at=datetime('now','localtime') WHERE id=?""",
                    (execution_id or None, run_id),
                )
        except (HTTPError, URLError, TimeoutError, OSError, json.JSONDecodeError) as error:
            with connect() as db:
                db.execute(
                    """UPDATE crm_n8n_runs SET status='failed',failed_items=1,
                       finished_at=datetime('now','localtime'),updated_at=datetime('now','localtime')
                       WHERE id=?""",
                    (run_id,),
                )
            return self.send_json({"error": f"Não foi possível iniciar o fluxo: {error}", "run_id": run_id}, HTTPStatus.BAD_GATEWAY)
        self.send_json({
            "started": True,
            "run_id": run_id,
            "run_key": run_key,
            "mode": mode,
            "limit": limit,
            "execution_id": execution_id or None,
        })

    def get_crm_n8n_runs(self, query: dict) -> None:
        if not self.require_crm_n8n_manager():
            return
        workflow_id = str(query.get("workflow_id", [""])[0]).strip()
        clauses = []
        params = []
        if workflow_id:
            clauses.append("r.workflow_id=?")
            params.append(workflow_id)
        where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
        with connect() as db:
            rows = db.execute(
                f"""SELECT r.*,u.name AS requested_by_name
                    FROM crm_n8n_runs r
                    LEFT JOIN users u ON u.id=r.requested_by_user_id
                    {where} ORDER BY r.id DESC LIMIT 100""",
                params,
            ).fetchall()
        self.send_json({"items": [dict(row) for row in rows]})

    def get_crm_n8n_run_detail(self, run_id: int) -> None:
        if not self.require_crm_n8n_manager():
            return
        with connect() as db:
            run = db.execute(
                """SELECT r.*,u.name AS requested_by_name FROM crm_n8n_runs r
                   LEFT JOIN users u ON u.id=r.requested_by_user_id WHERE r.id=?""",
                (run_id,),
            ).fetchone()
            if not run:
                return self.send_json({"error": "Execução não encontrada."}, HTTPStatus.NOT_FOUND)
            events = db.execute(
                """SELECT id,event_key,execution_id,campaign_id,patient_name,phone,channel_name,
                          event_type,outcome,external_message_id,occurred_at,received_at
                   FROM crm_n8n_patient_events WHERE run_id=?
                   ORDER BY id DESC LIMIT 1000""",
                (run_id,),
            ).fetchall()
        self.send_json({"run": dict(run), "events": [dict(row) for row in events]})

    def get_crm_n8n_patient_events(self, query: dict) -> None:
        if not self.require_crm_n8n_manager():
            return
        search = str(query.get("search", [""])[0]).strip().lower()[:120]
        event_type = str(query.get("event_type", [""])[0]).strip().lower()[:80]
        try:
            limit = max(20, min(1000, int(query.get("limit", ["250"])[0] or 250)))
        except (TypeError, ValueError):
            limit = 250
        clauses = []
        params: list = []
        if search:
            clauses.append(
                """(lower(COALESCE(e.patient_name,'')) LIKE ?
                    OR lower(COALESCE(e.phone,'')) LIKE ?
                    OR lower(COALESCE(e.campaign_id,'')) LIKE ?
                    OR lower(COALESCE(r.workflow_name,'')) LIKE ?)"""
            )
            value = f"%{search}%"
            params.extend([value, value, value, value])
        if event_type:
            clauses.append("lower(e.event_type) LIKE ?")
            params.append(f"%{event_type}%")
        where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
        with connect() as db:
            rows = db.execute(
                f"""SELECT e.id,e.run_id,e.workflow_id,e.execution_id,e.conversation_id,
                           e.campaign_id,e.patient_name,e.phone,e.channel_name,e.event_type,
                           e.outcome,e.external_message_id,e.occurred_at,e.received_at,
                           r.workflow_name,r.status AS run_status
                    FROM crm_n8n_patient_events e
                    LEFT JOIN crm_n8n_runs r ON r.id=e.run_id
                    {where}
                    ORDER BY e.id DESC LIMIT ?""",
                [*params, limit],
            ).fetchall()
            summary = db.execute(
                """SELECT COUNT(*) AS total,
                          COUNT(DISTINCT COALESCE(NULLIF(phone,''),patient_name)) AS patients,
                          SUM(CASE WHEN event_type LIKE '%sent%' THEN 1 ELSE 0 END) AS sent,
                          SUM(CASE WHEN event_type LIKE '%deliver%' OR event_type LIKE '%read%' THEN 1 ELSE 0 END) AS delivered,
                          SUM(CASE WHEN event_type LIKE '%repl%' OR event_type LIKE '%response%' THEN 1 ELSE 0 END) AS replied,
                          SUM(CASE WHEN event_type LIKE '%fail%' OR lower(COALESCE(outcome,''))='failed' THEN 1 ELSE 0 END) AS failed,
                          SUM(CASE WHEN event_type LIKE '%appointment%' OR event_type LIKE '%agend%' THEN 1 ELSE 0 END) AS appointments,
                          SUM(CASE WHEN event_type LIKE '%handoff%' OR event_type LIKE '%human%' THEN 1 ELSE 0 END) AS handoffs
                   FROM crm_n8n_patient_events"""
            ).fetchone()
        self.send_json({"items": [dict(row) for row in rows], "summary": dict(summary), "limit": limit})

    @staticmethod
    def crm_campaign_label(value: str | None) -> str:
        raw = str(value or "").strip()
        normalized = unicodedata.normalize("NFKD", raw).encode("ascii", "ignore").decode("ascii").lower()
        if "zero carie" in normalized:
            return "Zero Cárie"
        if "julho laranja" in normalized:
            return "Julho Laranja"
        if "aniversar" in normalized:
            return "Aniversariantes"
        if "primeira consulta" in normalized:
            return "Primeira consulta"
        if "sem agendamento" in normalized:
            return "Sem agendamento"
        if "confirmacao" in normalized or "clinicorp" in normalized:
            return "Confirmação de agenda"
        return raw or "Automação sem campanha"

    @staticmethod
    def crm_appointment_source(payload: dict) -> str | None:
        """Classifica apenas uma origem explicitamente enviada pela integração.

        Nunca inferimos IA ou humano pelo texto da conversa. Isso evita atribuir
        agendamentos de forma incorreta quando uma automação só repassou a conversa.
        """
        raw = next((payload.get(field) for field in (
            "appointment_source", "scheduled_by", "booking_source",
            "handled_by_type", "actor_type", "agent_type",
        ) if str(payload.get(field) or "").strip()), "")
        normalized = unicodedata.normalize("NFKD", str(raw)).encode("ascii", "ignore").decode("ascii").lower().strip()
        if normalized in {"ai", "ia", "bot", "automation", "automacao", "assistant", "assistente"}:
            return "IA"
        if normalized in {"human", "humano", "agent", "atendente", "crc", "reception", "recepcao", "user", "usuario"}:
            return "Humano"
        return None

    def get_crm_campaigns(self, query: dict) -> None:
        """Visão operacional, alimentada apenas pelos eventos reais dos workflows."""
        if not self.require_crc_access():
            return
        if not self.require_crm_feature("campaigns"):
            return
        try:
            days = max(1, min(180, int(query.get("days", ["30"])[0] or 30)))
        except (TypeError, ValueError):
            days = 30
        event_scope_sql, event_scope_params = self.crm_event_channel_scope_clause("e")
        with connect() as db:
            rows = db.execute(
                """SELECT COALESCE(NULLIF(campaign_id,''),NULLIF(flow_name,''),'Automação sem campanha') AS campaign,
                          COUNT(DISTINCT COALESCE(NULLIF(phone,''),event_key)) AS patients,
                          SUM(CASE WHEN event_type IN ('message.sent','message.ai.sent') THEN 1 ELSE 0 END) AS sent,
                          SUM(CASE WHEN event_type IN ('message.delivered','message.read') THEN 1 ELSE 0 END) AS delivered,
                          SUM(CASE WHEN event_type IN ('patient.replied','message.received') THEN 1 ELSE 0 END) AS replies,
                          SUM(CASE WHEN event_type IN ('ai.handoff.requested','human.required','opportunity.detected') THEN 1 ELSE 0 END) AS handoffs,
                          COUNT(DISTINCT CASE WHEN event_type IN ('appointment.confirmed','appointment.scheduled','appointment.created','patient.scheduled','booking.confirmed')
                                                    OR lower(COALESCE(outcome,'')) IN ('scheduled','booked','agendado','agendada')
                                              THEN COALESCE(NULLIF(phone,''),event_key) END) AS appointments,
                          COUNT(DISTINCT CASE WHEN (event_type IN ('appointment.confirmed','appointment.scheduled','appointment.created','patient.scheduled','booking.confirmed')
                                                    OR lower(COALESCE(outcome,'')) IN ('scheduled','booked','agendado','agendada'))
                                                   AND lower(COALESCE(appointment_source,''))='ia'
                                              THEN COALESCE(NULLIF(phone,''),event_key) END) AS appointments_ai,
                          COUNT(DISTINCT CASE WHEN (event_type IN ('appointment.confirmed','appointment.scheduled','appointment.created','patient.scheduled','booking.confirmed')
                                                    OR lower(COALESCE(outcome,'')) IN ('scheduled','booked','agendado','agendada'))
                                                   AND lower(COALESCE(appointment_source,''))='humano'
                                              THEN COALESCE(NULLIF(phone,''),event_key) END) AS appointments_human,
                          COUNT(DISTINCT CASE WHEN (event_type IN ('appointment.confirmed','appointment.scheduled','appointment.created','patient.scheduled','booking.confirmed')
                                                    OR lower(COALESCE(outcome,'')) IN ('scheduled','booked','agendado','agendada'))
                                                   AND COALESCE(appointment_source,'')=''
                                              THEN COALESCE(NULLIF(phone,''),event_key) END) AS appointments_unclassified,
                          SUM(CASE WHEN event_type LIKE '%fail%' OR lower(COALESCE(outcome,''))='failed' THEN 1 ELSE 0 END) AS failures,
                          MAX(COALESCE(occurred_at,received_at)) AS last_event_at
                   FROM crm_n8n_patient_events e
                   WHERE {event_scope_sql}
                     AND datetime(COALESCE(occurred_at,received_at)) >= datetime('now','localtime', ?)
                   GROUP BY COALESCE(NULLIF(campaign_id,''),NULLIF(flow_name,''),'Automação sem campanha')
                   ORDER BY last_event_at DESC""".format(event_scope_sql=event_scope_sql),
                (*event_scope_params, f"-{days} days"),
            ).fetchall()
        campaigns = []
        for row in rows:
            item = dict(row)
            item["name"] = self.crm_campaign_label(item.pop("campaign", ""))
            # Conversion is based on unique patients impacted, so the same
            # patient cannot inflate a campaign simply by generating retries.
            patients = int(item.get("patients") or 0)
            item["appointment_rate"] = round((int(item.get("appointments") or 0) / patients) * 100, 1) if patients else 0
            campaigns.append(item)
        self.send_json({"days": days, "items": campaigns})

    def get_crm_n8n_versions(self) -> None:
        if not self.require_crm_n8n_manager():
            return
        with connect() as db:
            rows = db.execute(
                """SELECT v.id,v.workflow_id,v.workflow_name,v.action,v.created_at,
                          u.name AS created_by_name
                   FROM crm_n8n_workflow_versions v
                   LEFT JOIN users u ON u.id=v.created_by_user_id
                   ORDER BY v.id DESC LIMIT 200"""
            ).fetchall()
        self.send_json({"items": [dict(row) for row in rows]})

    def get_crm_n8n_callback_keys(self) -> None:
        if not self.require_crm_n8n_manager():
            return
        with connect() as db:
            rows = db.execute(
                """SELECT id,label,active,created_at,revoked_at
                   FROM crm_n8n_callback_keys ORDER BY id DESC LIMIT 100"""
            ).fetchall()
        self.send_json({"items": [dict(row) for row in rows]})

    def create_crm_n8n_callback_key(self, payload: dict) -> None:
        if not self.require_crm_n8n_manager():
            return
        label = str(payload.get("label") or "Workflows automáticos").strip()[:100]
        token = secrets.token_urlsafe(40)
        token_hash = hashlib.sha256(token.encode("utf-8")).hexdigest()
        with connect() as db:
            cursor = db.execute(
                """INSERT INTO crm_n8n_callback_keys(label,key_hash,active,created_at)
                   VALUES(?,?,1,datetime('now','localtime'))""",
                (label, token_hash),
            )
        self.send_json({
            "created": True,
            "id": cursor.lastrowid,
            "label": label,
            "token": token,
            "warning": "Esta chave será exibida somente agora. Salve-a como credencial protegida no n8n.",
        })

    def revoke_crm_n8n_callback_key(self, key_id: int) -> None:
        if not self.require_crm_n8n_manager():
            return
        with connect() as db:
            updated = db.execute(
                """UPDATE crm_n8n_callback_keys
                   SET active=0,revoked_at=datetime('now','localtime')
                   WHERE id=? AND active=1""",
                (key_id,),
            ).rowcount
        if not updated:
            return self.send_json({"error": "Chave ativa não encontrada."}, HTTPStatus.NOT_FOUND)
        self.send_json({"revoked": True, "id": key_id})

    def change_crm_n8n_workflow(self, workflow_id: str, action: str) -> None:
        if not self.require_crm_n8n_manager():
            return
        if not re.fullmatch(r"[A-Za-z0-9_-]+", workflow_id):
            return self.send_json({"error": "Fluxo inválido."}, HTTPStatus.BAD_REQUEST)
        try:
            current_workflow = self.crm_n8n_request(f"/api/v1/workflows/{quote(workflow_id)}")
            with connect() as db:
                cursor = db.execute(
                    """INSERT INTO crm_n8n_workflow_versions
                       (workflow_id,workflow_name,action,workflow_json,created_by_user_id,created_at)
                       VALUES(?,?,?,?,?,datetime('now','localtime'))""",
                    (
                        workflow_id,
                        str(current_workflow.get("name") or "Fluxo"),
                        action,
                        json.dumps(current_workflow, ensure_ascii=False, separators=(",", ":")),
                        self.authenticated_user["id"],
                    ),
                )
                backup_id = cursor.lastrowid
            result = self.crm_n8n_request(f"/api/v1/workflows/{workflow_id}/{action}", method="POST", payload={})
        except RuntimeError as error:
            return self.send_json({"error": str(error)}, HTTPStatus.BAD_GATEWAY)
        self.send_json({"updated": True, "workflow": result, "backup_id": backup_id})

    def restore_crm_n8n_workflow_version(self, workflow_id: str, version_id: int) -> None:
        if not self.require_crm_n8n_manager():
            return
        if not re.fullmatch(r"[A-Za-z0-9_-]+", workflow_id):
            return self.send_json({"error": "Fluxo inválido."}, HTTPStatus.BAD_REQUEST)
        with connect() as db:
            version = db.execute(
                """SELECT id,workflow_id,workflow_name,workflow_json
                   FROM crm_n8n_workflow_versions
                   WHERE id=? AND workflow_id=?""",
                (version_id, workflow_id),
            ).fetchone()
        if not version:
            return self.send_json({"error": "Versão preservada não encontrada."}, HTTPStatus.NOT_FOUND)
        try:
            saved_workflow = json.loads(version["workflow_json"])
            current_workflow = self.crm_n8n_request(f"/api/v1/workflows/{quote(workflow_id)}")
        except (TypeError, json.JSONDecodeError):
            return self.send_json({"error": "A cópia preservada está inválida."}, HTTPStatus.CONFLICT)
        except RuntimeError as error:
            return self.send_json({"error": str(error)}, HTTPStatus.BAD_GATEWAY)
        restore_payload = {
            key: saved_workflow.get(key)
            for key in ("name", "nodes", "connections", "settings", "staticData")
            if key in saved_workflow
        }
        if not restore_payload.get("name") or not isinstance(restore_payload.get("nodes"), list):
            return self.send_json({"error": "A cópia não contém a estrutura necessária do workflow."}, HTTPStatus.CONFLICT)
        with connect() as db:
            safety_cursor = db.execute(
                """INSERT INTO crm_n8n_workflow_versions
                   (workflow_id,workflow_name,action,workflow_json,created_by_user_id,created_at)
                   VALUES(?,?,?,?,?,datetime('now','localtime'))""",
                (
                    workflow_id,
                    str(current_workflow.get("name") or version["workflow_name"]),
                    f"antes-restaurar-{version_id}",
                    json.dumps(current_workflow, ensure_ascii=False, separators=(",", ":")),
                    self.authenticated_user["id"],
                ),
            )
            safety_backup_id = safety_cursor.lastrowid
        try:
            result = self.crm_n8n_request(
                f"/api/v1/workflows/{quote(workflow_id)}",
                method="PUT",
                payload=restore_payload,
                timeout=30,
            )
        except RuntimeError as error:
            return self.send_json(
                {"error": str(error), "safety_backup_id": safety_backup_id},
                HTTPStatus.BAD_GATEWAY,
            )
        self.send_json({
            "restored": True,
            "workflow": result,
            "restored_version_id": version_id,
            "safety_backup_id": safety_backup_id,
        })

    def get_crm_integration_health(self) -> None:
        if not self.require_crc_access(): return
        if not self.require_crm_feature("integrations"): return
        with connect() as db:
            channels = db.execute("""SELECT id,display_name,instance_name,connection_status,last_event_at,sync_enabled,
                                  CASE WHEN last_event_at IS NULL OR datetime(last_event_at)<datetime('now','localtime','-15 minutes')
                                       THEN 1 ELSE 0 END AS stale
                                  FROM crm_channels WHERE active=1 ORDER BY display_name""").fetchall()
            failures = db.execute("""SELECT COUNT(*) AS total,datetime(MAX(received_at),'localtime') AS last_failure
                                  FROM crm_webhook_events WHERE processing_status='Falhou'
                                  AND datetime(received_at,'localtime')>=datetime('now','localtime','-1 hour')""").fetchone()
            n8n = db.execute("""SELECT COUNT(CASE WHEN datetime(received_at,'localtime')>=datetime('now','localtime','-24 hours') THEN 1 END) AS events_24h,
                              datetime(MAX(received_at),'localtime') AS last_event FROM crm_automation_events""").fetchone()
        n8n_online = False
        n8n_detail = "Servidor n8n sem resposta"
        try:
            request = Request(f"{N8N_INTERNAL_URL}/healthz", headers={"Accept": "application/json"})
            with urlopen(request, timeout=3) as response:
                n8n_online = 200 <= response.status < 300
                if n8n_online:
                    n8n_detail = f"Servidor online · {int(n8n['events_24h'] or 0)} evento(s) nas últimas 24h"
        except (HTTPError, URLError, TimeoutError, OSError):
            pass
        items = []
        for ch in channels:
            connected = str(ch["connection_status"] or "").strip().lower() == "conectado"
            enabled = bool(ch["sync_enabled"])
            if not connected:
                status, message = "error", "Número desconectado na Evolution"
            elif not enabled:
                status, message = "warning", "Canal pausado somente neste CRM"
            elif ch["last_event_at"]:
                status, message = "online", "Conectado e recebendo eventos"
            else:
                status, message = "online", "Conectado · aguardando o primeiro evento"
            items.append({"type": "Evolution", "name": ch["display_name"], "status": status,
                          "message": message, "last_event_at": ch["last_event_at"],
                          "idle": bool(ch["stale"] and ch["last_event_at"])})
        items.append({"type": "Webhook", "name": "Webhook Evolution", "status": "online" if not failures["total"] else "error",
                      "message": "Sem falhas na última hora" if not failures["total"] else f"{failures['total']} falha(s) na última hora",
                      "last_event_at": failures["last_failure"]})
        items.append({"type": "n8n", "name": "Automações n8n", "status": "online" if n8n_online else "error",
                      "message": n8n_detail, "last_event_at": n8n["last_event"],
                      "events_24h": int(n8n["events_24h"] or 0)})
        self.send_json({"items": items, "healthy": all(i["status"] == "online" for i in items)})

    def get_crm_conversations(self, query: dict) -> None:
        if not self.require_crc_access(): return
        status = str(query.get("status", [""])[0]).strip()
        search = str(query.get("search", [""])[0]).strip()
        view = str(query.get("view", ["workspace"])[0]).strip().lower()
        if not self.require_crm_feature(self.crm_conversation_view_feature(view)):
            return
        channel_id = str(query.get("channel_id", [""])[0]).strip()
        scope_sql, scope_params = self.crm_channel_scope_clause("ch")
        conditions, params = ["ch.active=1", "ch.sync_enabled=1", scope_sql], list(scope_params)
        if status: conditions.append("cv.status=?"); params.append(status)
        if channel_id:
            try:
                conditions.append("ch.id=?")
                params.append(int(channel_id))
            except ValueError:
                pass
        if search: conditions.append("(lower(ct.name) LIKE ? OR ct.phone LIKE ?)"); params.extend([f"%{search.lower()}%", f"%{self.crm_phone(search) or search}%"])
        if view == "queue":
            conditions.append("ct.is_internal=0 AND cv.status<>'Resolvida' AND cv.assigned_user_id IS NULL AND cv.queue_entered_at IS NOT NULL")
        elif view == "mine":
            conditions.append("ct.is_internal=0 AND cv.status<>'Resolvida' AND cv.assigned_user_id=?")
            params.append(self.authenticated_user["id"])
        elif view == "operational":
            conditions.append("ct.is_internal=0 AND cv.status<>'Resolvida' AND (cv.assigned_user_id IS NOT NULL OR (cv.assigned_user_id IS NULL AND cv.queue_entered_at IS NOT NULL))")
        elif view == "active":
            conditions.append("cv.status<>'Resolvida'")
        elif view == "internal":
            # Conversas do time não entram na fila externa, mas precisam
            # permanecer acessíveis no CRM depois de uma mensagem enviada.
            conditions.append("ct.is_internal=1 AND cv.status<>'Resolvida'")
        elif view == "resolved":
            conditions.append("cv.status='Resolvida'")
        else:
            conditions.append("ct.is_internal=0 AND cv.status<>'Resolvida' AND (cv.assigned_user_id=? OR (cv.assigned_user_id IS NULL AND cv.queue_entered_at IS NOT NULL))")
            params.append(self.authenticated_user["id"])
        # A visão "operational" (funil/pipeline) só usa id, name, channel,
        # pipeline_stage, assigned_to e tag_names no front-end (ver loadPipeline
        # em crm-whatsapp.html) — as 5 subconsultas correlacionadas abaixo
        # rodavam para cada uma das até 500 linhas mesmo assim, multiplicando
        # o custo da tela mais lenta do CRM sem nenhum ganho visível. Elas só
        # são executadas fora dessa visão, onde o valor é realmente exibido.
        skip_heavy_fields = view == "operational"
        journey_count_sql = "0" if skip_heavy_fields else "(SELECT COUNT(*) FROM crm_conversations sibling WHERE sibling.contact_id=cv.contact_id)"
        campaign_name_sql = "cv.automation_flow" if skip_heavy_fields else """COALESCE((SELECT NULLIF(ae.campaign_id,'') FROM crm_automation_events ae
                                 WHERE ae.conversation_id=cv.id AND NULLIF(ae.campaign_id,'') IS NOT NULL
                                 ORDER BY ae.id DESC LIMIT 1),cv.automation_flow)"""
        automation_last_event_sql = "NULL" if skip_heavy_fields else "(SELECT ae.event_type FROM crm_automation_events ae WHERE ae.conversation_id=cv.id ORDER BY ae.id DESC LIMIT 1)"
        automation_last_event_at_sql = "NULL" if skip_heavy_fields else "(SELECT ae.received_at FROM crm_automation_events ae WHERE ae.conversation_id=cv.id ORDER BY ae.id DESC LIMIT 1)"
        snippet_sql = "''" if skip_heavy_fields else "COALESCE((SELECT m.body FROM crm_messages m WHERE m.conversation_id=cv.id ORDER BY datetime(m.message_at) DESC,m.id DESC LIMIT 1),'')"
        with connect() as db:
            self.crm_activate_due_returns(db)
            total = db.execute(f"""SELECT COUNT(*) FROM crm_conversations cv
                                  JOIN crm_contacts ct ON ct.id=cv.contact_id
                                  JOIN crm_channels ch ON ch.id=cv.channel_id
                                  WHERE {' AND '.join(conditions)}""", params).fetchone()[0]
            rows = db.execute(f"""SELECT cv.id,cv.status,cv.priority,cv.queue_name,cv.pipeline_stage,cv.internal_note,
                       cv.assigned_user_id,cv.assigned_at,cv.queue_entered_at,cv.first_response_at,cv.resolved_by_user_id,
                       cv.automation_state,cv.automation_flow,cv.automation_turns,cv.handoff_reason,
                       cv.unread_count,cv.last_direction,cv.last_message_at,cv.created_at,cv.resolved_at,
                       cv.resolution_reason,cv.scheduled_return_at,cv.reopened_at,ch.sla_minutes,
                       ct.id AS contact_id,ct.name,ct.phone,ct.patient_id,ct.profile_picture_url,ct.is_internal,
                       ch.id AS channel_id,ch.instance_name,ch.display_name AS channel_name,ch.phone AS channel_phone,
                       ROUND(MAX(0,(julianday('now','localtime')-julianday(COALESCE(cv.queue_entered_at,cv.created_at)))*1440),1) AS waiting_minutes,
                       CASE WHEN cv.queue_entered_at IS NOT NULL AND cv.assigned_user_id IS NULL
                                  AND (julianday('now','localtime')-julianday(cv.queue_entered_at))*1440>ch.sla_minutes THEN 1 ELSE 0 END AS sla_overdue,
                       {journey_count_sql} AS journey_count,
                       u.name AS assigned_to, COALESCE(u.service_sector,'CRC') AS assigned_sector,
                       CASE WHEN u.id IS NULL THEN NULL ELSE u.name || ' · ' || COALESCE(NULLIF(u.service_sector,''),'CRC') END AS assigned_label,
                       resolver.name AS resolved_by,
                       {campaign_name_sql} AS campaign_name,
                       {automation_last_event_sql} AS automation_last_event,
                       {automation_last_event_at_sql} AS automation_last_event_at,
                       COALESCE((SELECT GROUP_CONCAT(t.name, '||') FROM crm_conversation_tags ctt
                                 JOIN crm_tags t ON t.id=ctt.tag_id WHERE ctt.conversation_id=cv.id),'') AS tag_names,
                       {snippet_sql} AS snippet
                FROM crm_conversations cv JOIN crm_contacts ct ON ct.id=cv.contact_id JOIN crm_channels ch ON ch.id=cv.channel_id
                LEFT JOIN users u ON u.id=cv.assigned_user_id
                LEFT JOIN users resolver ON resolver.id=cv.resolved_by_user_id WHERE {' AND '.join(conditions)}
                ORDER BY COALESCE(cv.last_message_at,cv.created_at) DESC LIMIT 500""", params).fetchall()
        items = []
        seen_phones = set()
        for row in rows:
            item = dict(row)
            phone_key = self.crm_phone(item.get("phone"))
            if phone_key and phone_key in seen_phones:
                continue
            if phone_key:
                seen_phones.add(phone_key)
            # Sempre exponha a rota protegida do CRM. Ela devolve a foto que
            # está em cache ou tenta obtê-la sob demanda na Evolution; quando
            # o perfil não possui foto, o front preserva as iniciais.
            if item.get("contact_id") and self.crm_phone(item.get("phone")):
                item["profile_picture_url"] = f"/api/crm/contacts/{item['contact_id']}/profile-photo"
            items.append(item)
        self.send_json({"items": items, "total": len(items), "raw_total": total, "view": view})

    def get_crm_agents(self) -> None:
        if not self.require_crc_access(): return
        if not self.require_crm_any_feature((*CRM_WORKSPACE_FEATURES, "management")):
            return
        scope_sql, scope_params = self.crm_channel_id_scope_clause("cv.channel_id")
        with connect() as db:
            rows = db.execute("""SELECT u.id,u.name,u.email,COALESCE(NULLIF(u.service_sector,''),'CRC') AS service_sector,COALESCE(u.crm_channel_scope_enabled,0) AS crm_channel_scope_enabled,
                    (SELECT STRING_AGG(cuc.channel_id::text, ',' ORDER BY cuc.channel_id) FROM crm_user_channels cuc WHERE cuc.user_id=u.id AND cuc.can_reply=1) AS crm_channel_ids,
                    COUNT(DISTINCT CASE WHEN COALESCE(ct.is_internal,0)=0 AND cv.status<>'Resolvida' THEN ct.id END) AS active_count,
                    COUNT(DISTINCT CASE WHEN COALESCE(ct.is_internal,0)=0 AND date(cv.resolved_at)=CURRENT_DATE THEN ct.id END) AS resolved_today
                FROM users u LEFT JOIN crm_conversations cv ON cv.assigned_user_id=u.id AND {scope_sql}
                LEFT JOIN crm_contacts ct ON ct.id=cv.contact_id
                WHERE u.access_role='crc' AND u.active=1 AND COALESCE(u.crm_operational_agent,1)=1
                GROUP BY u.id,u.name,u.email,u.crm_channel_scope_enabled ORDER BY u.name COLLATE NOCASE""".format(scope_sql=scope_sql), scope_params).fetchall()
        self.send_json({"items": [dict(row) for row in rows], "current_user_id": self.authenticated_user["id"]})

    def get_crm_tags(self) -> None:
        if not self.require_crc_access(): return
        if not self.require_crm_any_feature(CRM_WORKSPACE_FEATURES):
            return
        with connect() as db:
            rows = db.execute("SELECT id,name,color FROM crm_tags ORDER BY name COLLATE NOCASE").fetchall()
        self.send_json({"items": [dict(row) for row in rows]})

    def create_crm_tag(self, payload: dict) -> None:
        if not self.require_crc_access(): return
        if not self.require_crm_any_feature(CRM_WORKSPACE_FEATURES):
            return
        name = str(payload.get("name") or "").strip()
        color = str(payload.get("color") or "#8696a0").strip()
        if not name or len(name) > 40:
            return self.send_json({"error": "Informe uma etiqueta com até 40 caracteres."}, HTTPStatus.BAD_REQUEST)
        if not re.fullmatch(r"#[0-9a-fA-F]{6}", color):
            color = "#8696a0"
        with connect() as db:
            db.execute("INSERT OR IGNORE INTO crm_tags(name,color) VALUES(?,?)", (name, color))
            row = db.execute("SELECT id,name,color FROM crm_tags WHERE name=? COLLATE NOCASE", (name,)).fetchone()
        self.send_json(dict(row), HTTPStatus.CREATED)

    @staticmethod
    def crm_goal_actuals(db, user_id: int, start_date: date, end_date: date) -> dict:
        patient_type_sql = "COALESCE(NULLIF(TRIM(COALESCE(r.patient_type,'')),''),CASE WHEN r.category='Primeira consulta' THEN 'Primeira consulta' ELSE '' END)"
        row = db.execute(f"""SELECT
                COUNT(*) AS attendances,
                COUNT(*) FILTER (WHERE {patient_type_sql}='Primeira consulta') AS first_total,
                COUNT(*) FILTER (WHERE {patient_type_sql}='Primeira consulta' AND r.outcome='Agendou') AS first_converted,
                COUNT(*) FILTER (WHERE {patient_type_sql}='Retorno s/ Tratamento') AS recurring_total,
                COUNT(*) FILTER (WHERE {patient_type_sql}='Retorno s/ Tratamento' AND r.outcome='Agendou') AS recurring_converted,
                COUNT(*) FILTER (WHERE {patient_type_sql}='Retorno s/ Tratamento' AND r.is_recovery<>0 AND r.outcome='Agendou') AS recoveries
              FROM crm_service_resolutions r
              JOIN crm_contacts ct ON ct.id=r.contact_id
             WHERE r.resolved_by_user_id=? AND ct.is_internal=0
               AND date(r.resolved_at) BETWEEN ? AND ?""",
            (user_id, start_date.isoformat(), end_date.isoformat()),
        ).fetchone()
        values = {key: int(row[key] or 0) for key in (
            "attendances", "first_total", "first_converted", "recurring_total",
            "recurring_converted", "recoveries",
        )}
        values["first_consultations"] = values["first_converted"]
        return values

    def crm_goal_dashboard_data(self, db, user_id: int, month_value: str) -> dict:
        month_start, month_end = crm_goal_month_bounds(month_value)
        today = datetime.now(CLINIC_TIMEZONE).date()
        month_actuals = self.crm_goal_actuals(db, user_id, month_start, month_end)
        if month_start.year == today.year and month_start.month == today.month:
            day_actuals = self.crm_goal_actuals(db, user_id, today, today)
        else:
            day_actuals = {key: 0 for key in month_actuals}
        remaining_days = crm_open_days_remaining(today, month_start)
        goal_rows = db.execute(
            """SELECT id,metric_key,monthly_target,daily_target,celebration_enabled,celebration_message
                 FROM crm_goals WHERE user_id=? AND month_start=?""",
            (user_id, month_start.isoformat()),
        ).fetchall()
        configured = {row["metric_key"]: dict(row) for row in goal_rows}
        items = []
        for metric_key, label in CRM_GOAL_METRICS.items():
            goal = configured.get(metric_key, {})
            month_progress = crm_goal_progress(
                goal.get("monthly_target", 0), month_actuals.get(metric_key, 0), remaining_days,
            )
            day_progress = crm_goal_progress(
                goal.get("daily_target", 0), day_actuals.get(metric_key, 0), 1,
            )
            items.append({
                "metric_key": metric_key,
                "label": label,
                "goal_id": goal.get("id"),
                "monthly": month_progress,
                "daily": day_progress,
                "celebration_enabled": bool(goal.get("celebration_enabled", 1)),
                "celebration_message": goal.get("celebration_message") or "",
            })
        user = db.execute(
            "SELECT id,name,email FROM users WHERE id=? AND access_role='crc'", (user_id,),
        ).fetchone()
        history = db.execute(
            """SELECT metric_key,achievement_type,period_key,target_value,realized_value,message,achieved_at
                 FROM crm_goal_achievements WHERE user_id=?
                ORDER BY achieved_at DESC,id DESC LIMIT 20""", (user_id,),
        ).fetchall()
        return {
            "month": month_value,
            "month_label": f"{CRM_MONTH_NAMES[month_start.month]} {month_start.year}",
            "user": dict(user) if user else {"id": user_id, "name": "Atendente"},
            "schedule": {
                "weekdays": "Segunda a sexta, 08h às 18h",
                "saturday": "Sábado, 08h às 12h",
                "remaining_open_days": remaining_days,
            },
            "items": items,
            "conversion": {
                "first_consultation": {
                    "converted": month_actuals["first_converted"],
                    "opportunities": month_actuals["first_total"],
                    "percentage": round(month_actuals["first_converted"] / month_actuals["first_total"] * 100, 1)
                    if month_actuals["first_total"] else 0,
                },
                "recurring": {
                    "converted": month_actuals["recurring_converted"],
                    "opportunities": month_actuals["recurring_total"],
                    "percentage": round(month_actuals["recurring_converted"] / month_actuals["recurring_total"] * 100, 1)
                    if month_actuals["recurring_total"] else 0,
                },
            },
            "history": [dict(row) for row in history],
        }

    def get_crm_goals(self, query: dict | None = None) -> None:
        if not self.require_crc_access():
            return
        query = query or {}
        month_value = str((query.get("month") or [datetime.now(CLINIC_TIMEZONE).strftime("%Y-%m")])[0]).strip()
        can_configure = self.can_manage_crm(self.authenticated_user)
        has_requested_user = bool(query.get("user_id"))
        try:
            requested_user_id = int((query.get("user_id") or [self.authenticated_user["id"]])[0])
            crm_goal_month_bounds(month_value)
        except (TypeError, ValueError):
            return self.send_json({"error": "Informe um mês e uma atendente válidos."}, HTTPStatus.BAD_REQUEST)
        user_id = requested_user_id if can_configure else int(self.authenticated_user["id"])
        with connect() as db:
            agents = db.execute(
                "SELECT id,name FROM users WHERE access_role='crc' AND active=1 AND COALESCE(crm_operational_agent,1)=1 ORDER BY name COLLATE NOCASE"
            ).fetchall() if can_configure else []
            if can_configure and not has_requested_user and agents:
                current_is_operational = db.execute(
                    "SELECT 1 FROM users WHERE id=? AND COALESCE(crm_operational_agent,1)=1",
                    (user_id,),
                ).fetchone()
                if not current_is_operational:
                    user_id = int(agents[0]["id"])
            user = db.execute(
                "SELECT id FROM users WHERE id=? AND access_role='crc' AND active=1 AND COALESCE(crm_operational_agent,1)=1", (user_id,),
            ).fetchone()
            if not user:
                return self.send_json({"error": "Atendente do CRC não encontrada."}, HTTPStatus.NOT_FOUND)
            payload = self.crm_goal_dashboard_data(db, user_id, month_value)
        payload["can_configure"] = can_configure
        payload["agents"] = [dict(row) for row in agents]
        self.send_json(payload)

    def crm_evaluate_goal_achievements(self, db, user_id: int, source_resolution_id: int | None = None) -> list[dict]:
        today = datetime.now(CLINIC_TIMEZONE).date()
        month_value = today.strftime("%Y-%m")
        month_start, month_end = crm_goal_month_bounds(month_value)
        goals = db.execute(
            """SELECT id,metric_key,monthly_target,daily_target,celebration_enabled,celebration_message
                 FROM crm_goals WHERE user_id=? AND month_start=?""",
            (user_id, month_start.isoformat()),
        ).fetchall()
        if not goals:
            return []
        user = db.execute("SELECT name FROM users WHERE id=?", (user_id,)).fetchone()
        user_name = user["name"] if user else "Atendente"
        month_actuals = self.crm_goal_actuals(db, user_id, month_start, month_end)
        day_actuals = self.crm_goal_actuals(db, user_id, today, today)
        created = []

        def record(goal_id, metric_key, achievement_type, period_key, target, realized, custom_message=""):
            if not target or realized < target:
                return
            label = CRM_GOAL_METRICS.get(metric_key, "Todas as metas")
            period_label = "diária" if achievement_type == "daily" else "mensal"
            message = (
                f"{custom_message.strip()[:140]} Resultado: {realized} de {target}." if custom_message else
                f"Parabéns, {user_name}! Meta {period_label} de {label} alcançada: {realized} de {target}."
            )
            cursor = db.execute(
                """INSERT OR IGNORE INTO crm_goal_achievements(
                       goal_id,user_id,metric_key,achievement_type,period_key,target_value,
                       realized_value,message,source_resolution_id)
                     VALUES(?,?,?,?,?,?,?,?,?)""",
                (goal_id, user_id, metric_key, achievement_type, period_key, target,
                 realized, message, source_resolution_id),
            )
            if cursor.rowcount:
                created.append({
                    "metric_key": metric_key, "achievement_type": achievement_type,
                    "target": target, "realized": realized, "message": message,
                })

        enabled_goals = []
        for goal in goals:
            metric_key = goal["metric_key"]
            if metric_key not in CRM_GOAL_METRICS or not goal["celebration_enabled"]:
                continue
            enabled_goals.append(goal)
            record(goal["id"], metric_key, "monthly", month_value,
                   int(goal["monthly_target"] or 0), month_actuals.get(metric_key, 0),
                   goal["celebration_message"] or "")
            record(goal["id"], metric_key, "daily", today.isoformat(),
                   int(goal["daily_target"] or 0), day_actuals.get(metric_key, 0),
                   goal["celebration_message"] or "")

        for achievement_type, period_key, actuals, target_column in (
            ("monthly", month_value, month_actuals, "monthly_target"),
            ("daily", today.isoformat(), day_actuals, "daily_target"),
        ):
            complete = len(enabled_goals) == len(CRM_GOAL_METRICS) and all(
                int(goal[target_column] or 0) > 0 and
                actuals.get(goal["metric_key"], 0) >= int(goal[target_column] or 0)
                for goal in enabled_goals
            )
            if complete:
                target_sum = sum(int(goal[target_column] or 0) for goal in enabled_goals)
                realized_sum = sum(actuals.get(goal["metric_key"], 0) for goal in enabled_goals)
                record(None, "all_goals", achievement_type, period_key, target_sum, realized_sum,
                       f"Parabéns, {user_name}! Todas as metas {('do dia' if achievement_type == 'daily' else 'do mês')} foram alcançadas.")
        return created

    def save_crm_goals(self, payload: dict) -> None:
        if not self.require_crc_access():
            return
        if not self.can_manage_crm(self.authenticated_user):
            return self.send_json(
                {"error": "Somente um administrador do CRM pode configurar metas."},
                HTTPStatus.FORBIDDEN,
            )
        month_value = str(payload.get("month") or "").strip()
        try:
            user_id = int(payload.get("user_id") or 0)
            month_start, _ = crm_goal_month_bounds(month_value)
        except (TypeError, ValueError):
            return self.send_json({"error": "Selecione o mês e a atendente."}, HTTPStatus.BAD_REQUEST)
        raw_goals = payload.get("goals") or {}
        if not isinstance(raw_goals, dict):
            return self.send_json({"error": "As metas informadas são inválidas."}, HTTPStatus.BAD_REQUEST)
        normalized = {}
        try:
            for metric_key in CRM_GOAL_METRICS:
                item = raw_goals.get(metric_key) or {}
                monthly_target = int(item.get("monthly_target") or 0)
                daily_target = int(item.get("daily_target") or 0)
                if not 0 <= monthly_target <= 100000 or not 0 <= daily_target <= 10000:
                    raise ValueError
                normalized[metric_key] = {
                    "monthly_target": monthly_target,
                    "daily_target": daily_target,
                    "celebration_enabled": 1 if item.get("celebration_enabled", True) else 0,
                    "celebration_message": str(item.get("celebration_message") or "").strip()[:180] or None,
                }
        except (TypeError, ValueError):
            return self.send_json({"error": "Use apenas números inteiros positivos nas metas."}, HTTPStatus.BAD_REQUEST)
        with connect() as db:
            user = db.execute(
                "SELECT id FROM users WHERE id=? AND access_role='crc' AND active=1 AND COALESCE(crm_operational_agent,1)=1", (user_id,),
            ).fetchone()
            if not user:
                return self.send_json({"error": "Atendente do CRC não encontrada."}, HTTPStatus.NOT_FOUND)
            for metric_key, item in normalized.items():
                db.execute(
                    """INSERT INTO crm_goals(user_id,month_start,metric_key,monthly_target,daily_target,
                                              celebration_enabled,celebration_message,created_by_user_id)
                         VALUES(?,?,?,?,?,?,?,?)
                         ON CONFLICT(user_id,month_start,metric_key) DO UPDATE SET
                           monthly_target=excluded.monthly_target,daily_target=excluded.daily_target,
                           celebration_enabled=excluded.celebration_enabled,
                           celebration_message=excluded.celebration_message,
                           updated_at=datetime('now','localtime')""",
                    (user_id, month_start.isoformat(), metric_key, item["monthly_target"], item["daily_target"],
                     item["celebration_enabled"], item["celebration_message"], self.authenticated_user["id"]),
                )
            achievements = self.crm_evaluate_goal_achievements(db, user_id)
            dashboard = self.crm_goal_dashboard_data(db, user_id, month_value)
        dashboard["can_configure"] = True
        dashboard["achievements"] = achievements
        self.send_json(dashboard)

    def get_crm_metrics(self, query: dict | None = None) -> None:
        if not self.require_crc_access(): return
        if not self.require_crm_any_feature(("inbox", "queue", "management")):
            return
        query = query or {}
        period = str(query.get("period", ["today"])[0] or "today").strip().lower()
        today = datetime.now().date()
        if period == "7d":
            start_date, end_date = today - timedelta(days=6), today
        elif period == "custom":
            try:
                start_date = datetime.strptime(str(query.get("start", [""])[0]), "%Y-%m-%d").date()
                end_date = datetime.strptime(str(query.get("end", [""])[0]), "%Y-%m-%d").date()
            except ValueError:
                return self.send_json({"error": "Informe o período personalizado completo."}, HTTPStatus.BAD_REQUEST)
            if start_date > end_date or (end_date - start_date).days > 366:
                return self.send_json({"error": "O período informado é inválido ou muito extenso."}, HTTPStatus.BAD_REQUEST)
        else:
            period, start_date, end_date = "today", today, today
        try:
            channel_id = int(query.get("channel_id", ["0"])[0] or 0)
        except (TypeError, ValueError):
            channel_id = 0
        scope_sql, scope_params = self.crm_channel_scope_clause("ch")
        metric_conditions = ["ch.active=1", "ch.sync_enabled=1", "ct.is_internal=0", scope_sql,
                             "date(COALESCE(cv.resolved_at,cv.last_message_at,cv.created_at)) BETWEEN ? AND ?"]
        metric_params = [*scope_params, start_date.isoformat(), end_date.isoformat()]
        if channel_id:
            metric_conditions.append("ch.id=?")
            metric_params.append(channel_id)
        metric_where = " AND ".join(metric_conditions)
        with connect() as db:
            self.crm_activate_due_returns(db)
            summary = db.execute(f"""SELECT
                COUNT(DISTINCT CASE WHEN cv.status<>'Resolvida' AND (cv.assigned_user_id IS NOT NULL OR (cv.assigned_user_id IS NULL AND cv.queue_entered_at IS NOT NULL)) THEN ct.id END) AS active,
                COUNT(DISTINCT CASE WHEN cv.status<>'Resolvida' AND cv.assigned_user_id IS NULL AND cv.queue_entered_at IS NOT NULL THEN ct.id END) AS waiting,
                COUNT(DISTINCT CASE WHEN cv.status<>'Resolvida' AND cv.assigned_user_id IS NOT NULL THEN ct.id END) AS in_service,
                COUNT(DISTINCT CASE WHEN cv.status<>'Resolvida' AND cv.assigned_user_id=? THEN ct.id END) AS mine,
                COUNT(DISTINCT CASE WHEN date(cv.resolved_at) BETWEEN ? AND ? THEN ct.id END) AS resolved_today,
                COUNT(DISTINCT CASE WHEN date(cv.resolved_at) BETWEEN ? AND ? AND cv.resolved_by_user_id=? THEN ct.id END) AS resolved_by_me_today,
                COUNT(DISTINCT CASE WHEN cv.status<>'Resolvida' AND cv.assigned_user_id IS NULL AND cv.queue_entered_at IS NOT NULL THEN ct.id END) AS unread,
                COALESCE(SUM(CASE WHEN cv.status<>'Resolvida' AND cv.assigned_user_id IS NULL AND cv.queue_entered_at IS NOT NULL THEN cv.unread_count ELSE 0 END),0) AS unread_messages
              FROM crm_conversations cv
              JOIN crm_contacts ct ON ct.id=cv.contact_id
              JOIN crm_channels ch ON ch.id=cv.channel_id
              WHERE {metric_where}""",
              (self.authenticated_user["id"], start_date.isoformat(), end_date.isoformat(),
               start_date.isoformat(), end_date.isoformat(), self.authenticated_user["id"], *metric_params)).fetchone()
            sla = db.execute(f"""SELECT
                COUNT(DISTINCT CASE WHEN cv.status<>'Resolvida' AND cv.assigned_user_id IS NULL AND cv.queue_entered_at IS NOT NULL
                  AND (julianday('now','localtime')-julianday(cv.queue_entered_at))*1440>ch.sla_minutes THEN ct.id END) AS overdue,
                ROUND(COALESCE(AVG(CASE WHEN cv.first_response_at IS NOT NULL AND cv.queue_entered_at IS NOT NULL
                  AND datetime(cv.first_response_at)>=datetime(cv.queue_entered_at)
                  THEN (julianday(cv.first_response_at)-julianday(cv.queue_entered_at))*1440 END),0),1) AS avg_first_response_minutes,
                ROUND(COALESCE(AVG(CASE WHEN cv.resolved_at IS NOT NULL AND cv.assigned_at IS NOT NULL
                  AND date(cv.resolved_at) BETWEEN ? AND ?
                  AND datetime(cv.resolved_at)>=datetime(cv.assigned_at)
                  THEN (julianday(cv.resolved_at)-julianday(cv.assigned_at))*1440 END),0),1) AS avg_resolution_minutes
              FROM crm_conversations cv JOIN crm_contacts ct ON ct.id=cv.contact_id
              JOIN crm_channels ch ON ch.id=cv.channel_id
              WHERE {metric_where}""", (start_date.isoformat(), end_date.isoformat(), *metric_params)).fetchone()
            agent_scope_sql, agent_scope_params = self.crm_channel_id_scope_clause("cv.channel_id")
            agent_performance = db.execute(f"""SELECT u.id,u.name,COALESCE(NULLIF(u.service_sector,''),'CRC') AS service_sector,
                COUNT(DISTINCT CASE WHEN cv.status<>'Resolvida' AND date(COALESCE(cv.last_message_at,cv.created_at)) BETWEEN ? AND ? THEN ct.id END) AS active,
                COUNT(DISTINCT CASE WHEN date(cv.resolved_at) BETWEEN ? AND ? THEN ct.id END) AS resolved_today,
                ROUND(COALESCE(AVG(CASE WHEN cv.first_response_at IS NOT NULL AND cv.queue_entered_at IS NOT NULL
                  AND date(cv.first_response_at) BETWEEN ? AND ?
                  AND datetime(cv.first_response_at)>=datetime(cv.queue_entered_at)
                  THEN (julianday(cv.first_response_at)-julianday(cv.queue_entered_at))*1440 END),0),1) AS avg_first_response_minutes
              FROM users u LEFT JOIN crm_conversations cv ON cv.assigned_user_id=u.id AND {agent_scope_sql}
              LEFT JOIN crm_contacts ct ON ct.id=cv.contact_id
              LEFT JOIN crm_channels ch ON ch.id=cv.channel_id
              WHERE u.access_role='crc' AND u.active=1 AND COALESCE(u.crm_operational_agent,1)=1 AND (cv.id IS NULL OR ct.is_internal=0)
                AND (?=0 OR ch.id=?)
              GROUP BY u.id,u.name,u.service_sector ORDER BY resolved_today DESC,u.name""",
              (start_date.isoformat(), end_date.isoformat(), start_date.isoformat(), end_date.isoformat(),
               start_date.isoformat(), end_date.isoformat(), *agent_scope_params, channel_id, channel_id)).fetchall()
            volume_scope_sql, volume_scope_params = self.crm_channel_id_scope_clause("cv.channel_id")
            volume_channel_sql = ""
            volume_params: list = [start_date.isoformat(), end_date.isoformat(), *volume_scope_params]
            if channel_id:
                volume_channel_sql = " AND cv.channel_id=?"
                volume_params.append(channel_id)
            if period == "today":
                volume = db.execute(f"""SELECT strftime('%H',m.message_at) AS bucket,
                                      strftime('%Hh',m.message_at) AS label,COUNT(*) AS total
                                      FROM crm_messages m JOIN crm_conversations cv ON cv.id=m.conversation_id
                                      WHERE m.direction='inbound' AND date(m.message_at) BETWEEN ? AND ?
                                        AND {volume_scope_sql} {volume_channel_sql}
                                      GROUP BY 1,2 ORDER BY bucket""", volume_params).fetchall()
            else:
                volume = db.execute(f"""SELECT date(m.message_at) AS bucket,
                                      strftime('%d/%m',m.message_at) AS label,COUNT(*) AS total
                                      FROM crm_messages m JOIN crm_conversations cv ON cv.id=m.conversation_id
                                      WHERE m.direction='inbound' AND date(m.message_at) BETWEEN ? AND ?
                                        AND {volume_scope_sql} {volume_channel_sql}
                                      GROUP BY 1,2 ORDER BY bucket""", volume_params).fetchall()
        with EVOLUTION_CHAT_SYNC_LOCK:
            sync_status = dict(EVOLUTION_CHAT_SYNC_STATUS)
        summary_payload = dict(summary)
        summary_payload.update(dict(sla))
        self.send_json({"summary": summary_payload, "volume": [dict(row) for row in volume],
                        "agents": [dict(row) for row in agent_performance], "chat_sync": sync_status,
                        "filter": {"period": period, "start": start_date.isoformat(), "end": end_date.isoformat(), "channel_id": channel_id}})

    def get_crm_messages(self, conversation_id: int, query: dict | None = None) -> None:
        if not self.require_crc_access(): return
        if not self.require_crm_any_feature(CRM_WORKSPACE_FEATURES):
            return
        query = query or {}
        try:
            after_id = max(0, int(query.get("after_id", ["0"])[0] or 0))
            limit = max(20, min(300, int(query.get("limit", ["200"])[0] or 200)))
        except (TypeError, ValueError):
            return self.send_json({"error": "Parâmetros de paginação inválidos."}, HTTPStatus.BAD_REQUEST)
        with connect() as db:
            scoped_conversation = db.execute("SELECT channel_id FROM crm_conversations WHERE id=?", (conversation_id,)).fetchone()
            if not scoped_conversation:
                return self.send_json({"error": "Conversa não encontrada."}, HTTPStatus.NOT_FOUND)
            if not self.crm_channel_allowed(db, scoped_conversation["channel_id"]):
                return self.send_json({"error": "Você não possui acesso a este canal."}, HTTPStatus.FORBIDDEN)
            if after_id:
                rows = db.execute("""SELECT id,external_message_id,direction,message_type,body,media_url,mime_type,duration_seconds,sender_name,
                                          sent_by_user_id,author_type,author_label,source_channel,delivery_status,message_at
                                   FROM crm_messages WHERE conversation_id=? AND id>?
                                   ORDER BY id LIMIT ?""", (conversation_id, after_id, limit)).fetchall()
            else:
                rows = db.execute("""SELECT * FROM (
                    SELECT id,external_message_id,direction,message_type,body,media_url,mime_type,duration_seconds,sender_name,
                           sent_by_user_id,author_type,author_label,source_channel,delivery_status,message_at
                    FROM crm_messages WHERE conversation_id=? ORDER BY id DESC LIMIT ?
                ) recent ORDER BY id""", (conversation_id, limit)).fetchall()
            # A simples pré-visualização não pode retirar um contato da fila.
            # Só zeramos o não lido quando a conversa já foi assumida pelo usuário atual.
            db.execute("""UPDATE crm_conversations SET unread_count=0
                          WHERE id=? AND assigned_user_id=?""",
                       (conversation_id, self.authenticated_user["id"]))
        items = []
        for row in rows:
            item = dict(row)
            if item["message_type"] == "audio" and not item.get("duration_seconds"):
                cached_match = re.fullmatch(r"/api/crm/media/([a-f0-9]{32}\.(?:webm|ogg|oga|mp4|m4a))", str(item.get("media_url") or ""))
                if cached_match:
                    media_file = CRM_MEDIA_DIR / cached_match.group(1)
                    if media_file.is_file():
                        item["duration_seconds"] = crm_audio_duration_seconds(media_file.read_bytes())
                        if item["duration_seconds"]:
                            with connect() as duration_db:
                                duration_db.execute("UPDATE crm_messages SET duration_seconds=? WHERE id=?", (item["duration_seconds"], item["id"]))
            if item["message_type"] in {"audio", "image", "video", "document", "sticker"} and item["external_message_id"]:
                if not str(item["media_url"] or "").startswith("/api/crm/media/"):
                    item["media_url"] = f"/api/crm/messages/{item['id']}/media"
            items.append(item)
        self.send_json({"items": items, "after_id": after_id, "limit": limit})

    @staticmethod
    def crm_media_extension(mime_type: str | None, file_name: str | None = None) -> str:
        allowed = {"webm", "ogg", "oga", "mp4", "m4a", "jpg", "jpeg", "png", "webp", "pdf", "doc", "docx", "xls", "xlsx", "ppt", "pptx", "txt", "zip"}
        suffix = Path(str(file_name or "")).suffix.lower().lstrip(".")
        if suffix in allowed:
            return suffix
        clean_mime = str(mime_type or "application/octet-stream").split(";", 1)[0].strip().lower()
        known = {
            "audio/ogg": "ogg", "audio/opus": "ogg", "audio/webm": "webm",
            "audio/mp4": "m4a", "audio/mpeg": "mp4", "video/mp4": "mp4",
            "image/jpeg": "jpg", "image/png": "png", "image/webp": "webp",
            "application/pdf": "pdf", "application/msword": "doc",
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document": "docx",
            "application/vnd.ms-excel": "xls",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "xlsx",
            "application/vnd.ms-powerpoint": "ppt",
            "application/vnd.openxmlformats-officedocument.presentationml.presentation": "pptx",
            "text/plain": "txt", "application/zip": "zip", "application/x-zip-compressed": "zip",
        }
        return known.get(clean_mime, "bin")

    @staticmethod
    def crm_media_content_valid(mime_type: str, content: bytes) -> bool:
        """Confere o conteúdo real antes de retransmitir um anexo pelo WhatsApp."""
        if not content:
            return False
        signatures = {
            "image/jpeg": lambda data: data.startswith(b"\xff\xd8\xff"),
            "image/png": lambda data: data.startswith(b"\x89PNG\r\n\x1a\n"),
            "image/webp": lambda data: len(data) >= 12 and data.startswith(b"RIFF") and data[8:12] == b"WEBP",
            "video/mp4": lambda data: len(data) >= 12 and data[4:8] == b"ftyp",
            "application/pdf": lambda data: data.startswith(b"%PDF-"),
            "application/msword": lambda data: data.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"),
            "application/vnd.ms-excel": lambda data: data.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"),
            "application/vnd.ms-powerpoint": lambda data: data.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"),
        }
        if mime_type in signatures:
            return signatures[mime_type](content)
        if mime_type == "text/plain":
            try:
                content.decode("utf-8-sig")
                return b"\x00" not in content
            except UnicodeDecodeError:
                return False
        zip_requirements = {
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document": "word/document.xml",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "xl/workbook.xml",
            "application/vnd.openxmlformats-officedocument.presentationml.presentation": "ppt/presentation.xml",
        }
        if mime_type in {*zip_requirements, "application/zip", "application/x-zip-compressed"}:
            try:
                with zipfile.ZipFile(io.BytesIO(content)) as archive:
                    members = archive.infolist()
                    names = {member.filename.replace("\\", "/") for member in members}
                    if not members or len(members) > 1000:
                        return False
                    if any(name.startswith("/") or ".." in Path(name).parts for name in names):
                        return False
                    uncompressed = sum(member.file_size for member in members)
                    compressed = max(1, sum(member.compress_size for member in members))
                    if uncompressed > 50 * 1024 * 1024 or (uncompressed > 5 * 1024 * 1024 and uncompressed / compressed > 100):
                        return False
                    required = zip_requirements.get(mime_type)
                    return not required or ("[Content_Types].xml" in names and required in names)
            except (zipfile.BadZipFile, OSError, ValueError):
                return False
        return False

    def get_crm_message_media(self, message_id: int) -> None:
        if not self.require_crc_access():
            return
        if not self.require_crm_any_feature(CRM_WORKSPACE_FEATURES):
            return
        with connect() as db:
            row = db.execute("""SELECT m.id,m.external_message_id,m.media_url,m.mime_type,m.direction,
                ch.id AS channel_id,ch.instance_name,ct.phone
                FROM crm_messages m
                JOIN crm_conversations cv ON cv.id=m.conversation_id
                JOIN crm_contacts ct ON ct.id=cv.contact_id
                JOIN crm_channels ch ON ch.id=cv.channel_id
                WHERE m.id=? AND m.message_type IN ('audio','image','video','document','sticker')""",
                (message_id,),
            ).fetchone()
            if row and not self.crm_channel_allowed(db, row["channel_id"]):
                return self.send_json({"error": "Você não possui acesso a esta mídia."}, HTTPStatus.FORBIDDEN)
        if not row or not row["external_message_id"]:
            return self.send_json({"error": "Mídia não encontrada."}, HTTPStatus.NOT_FOUND)
        cached = str(row["media_url"] or "")
        cached_match = re.fullmatch(r"/api/crm/media/([a-f0-9]{32}\.(?:webm|ogg|oga|mp4|m4a|jpg|jpeg|png|webp|pdf|doc|docx|xls|xlsx|ppt|pptx|txt|zip|bin))", cached)
        if cached_match:
            return self.get_crm_media(cached_match.group(1))
        try:
            full_message = None
            with connect() as db:
                event_rows = db.execute("""SELECT payload_json FROM crm_webhook_events
                    WHERE payload_json LIKE ? ORDER BY id DESC LIMIT 25""",
                    (f'%{row["external_message_id"]}%',),
                ).fetchall()
            def find_message(value):
                if isinstance(value, dict):
                    key = value.get("key")
                    if isinstance(key, dict) and str(key.get("id") or "") == str(row["external_message_id"]):
                        return value
                    for child in value.values():
                        found = find_message(child)
                        if found: return found
                elif isinstance(value, list):
                    for child in value:
                        found = find_message(child)
                        if found: return found
                return None
            for event_row in event_rows:
                try:
                    full_message = find_message(json.loads(event_row["payload_json"] or "{}"))
                except (TypeError, ValueError, json.JSONDecodeError):
                    full_message = None
                if full_message:
                    break
            if not full_message:
                phone = self.crm_phone(row["phone"])
                remote_jid = f"{phone}@s.whatsapp.net" if phone else ""
                full_message = {"key": {
                    "id": row["external_message_id"],
                    "remoteJid": remote_jid,
                    "fromMe": row["direction"] == "outbound",
                }}
            payload = self.evolution_api_request(
                f"/chat/getBase64FromMediaMessage/{quote(row['instance_name'], safe='')}",
                "POST",
                {"message": full_message, "convertToMp4": False},
            )
            encoded = str(payload.get("base64") or "") if isinstance(payload, dict) else ""
            if encoded.startswith("data:") and "," in encoded:
                encoded = encoded.split(",", 1)[1]
            content = base64.b64decode(encoded, validate=True)
            if not content or len(content) > 40 * 1024 * 1024:
                raise ValueError("tamanho de mídia inválido")
            mime_type = str(payload.get("mimetype") or row["mime_type"] or "application/octet-stream").split(";", 1)[0]
            extension = self.crm_media_extension(mime_type, payload.get("fileName"))
            digest = hashlib.md5(str(row["external_message_id"]).encode(), usedforsecurity=False).hexdigest()
            file_name = f"{digest}.{extension}"
            CRM_MEDIA_DIR.mkdir(parents=True, exist_ok=True)
            (CRM_MEDIA_DIR / file_name).write_bytes(content)
            duration_seconds = crm_audio_duration_seconds(content) if mime_type.startswith("audio/") else None
            with connect() as db:
                db.execute("UPDATE crm_messages SET media_url=?,mime_type=?,duration_seconds=COALESCE(?,duration_seconds) WHERE id=?",
                           (f"/api/crm/media/{file_name}", mime_type, duration_seconds, message_id))
            return self.get_crm_media(file_name)
        except (RuntimeError, ValueError, TypeError) as error:
            # Mídia remota pode expirar ou deixar de existir na Evolution. Registra a
            # causa sem expor URL, telefone ou credenciais no navegador.
            print(
                f"[crm-media] falha ao obter mídia da mensagem {message_id}: "
                f"{type(error).__name__}: {error}",
                flush=True,
            )
            return self.send_json({"error": f"Não foi possível carregar a mídia: {error}"}, HTTPStatus.BAD_GATEWAY)

    def get_crm_media(self, file_name: str) -> None:
        if not self.require_crc_access():
            return
        if not self.require_crm_any_feature(CRM_WORKSPACE_FEATURES):
            return
        if not re.fullmatch(r"[a-f0-9]{32}\.(?:webm|ogg|oga|mp4|m4a|jpg|jpeg|png|webp|pdf|doc|docx|xls|xlsx|ppt|pptx|txt|zip|bin)", file_name):
            return self.send_json({"error": "Arquivo de mídia inválido."}, HTTPStatus.BAD_REQUEST)
        with connect() as db:
            media_channels = db.execute("""SELECT DISTINCT cv.channel_id
                FROM crm_messages m
                JOIN crm_conversations cv ON cv.id=m.conversation_id
                WHERE m.media_url=?""", (f"/api/crm/media/{file_name}",)).fetchall()
            if not media_channels:
                return self.send_json({"error": "Mídia não encontrada."}, HTTPStatus.NOT_FOUND)
            if not any(self.crm_channel_allowed(db, row["channel_id"]) for row in media_channels):
                return self.send_json({"error": "Você não possui acesso a esta mídia."}, HTTPStatus.FORBIDDEN)
        target = (CRM_MEDIA_DIR / file_name).resolve()
        if target.parent != CRM_MEDIA_DIR.resolve() or not target.is_file():
            return self.send_json({"error": "Mídia não encontrada."}, HTTPStatus.NOT_FOUND)
        content = target.read_bytes()
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", mimetypes.guess_type(target.name)[0] or "application/octet-stream")
        self.send_header("Content-Length", str(len(content)))
        self.send_header("Content-Disposition", f'inline; filename="{file_name}"')
        self.send_header("Cache-Control", "private, max-age=86400")
        self.send_security_headers()
        self.end_headers()
        self.wfile.write(content)

    def evolution_webhook_authorized(self, query: dict) -> bool:
        supplied = self.headers.get("X-Webhook-Token") or self.headers.get("Authorization", "").removeprefix("Bearer ") or str(query.get("webhook_key", [""])[0])
        return bool(EVOLUTION_WEBHOOK_TOKEN and supplied and hmac.compare_digest(str(supplied), EVOLUTION_WEBHOOK_TOKEN))

    def shared_integration_authorized(self, query: dict) -> bool:
        supplied = self.headers.get("X-Integration-Key") or self.headers.get("Authorization", "").removeprefix("Bearer ") or str(query.get("token", [""])[0])
        return bool(INTEGRATION_TOKEN and supplied and hmac.compare_digest(str(supplied), INTEGRATION_TOKEN))

    def crm_n8n_run_event_authorized(self, payload: dict) -> bool:
        """Autoriza callbacks de uma execução iniciada pelo próprio painel.

        O segredo é efêmero, enviado apenas ao webhook do n8n e persistido no
        banco somente como SHA-256. O workflow não recebe o token global das
        demais integrações da clínica.
        """
        run_key = str(payload.get("crm_run_key") or payload.get("run_key") or "").strip()
        supplied = str(payload.get("crm_event_token") or "").strip()
        if not run_key or not supplied:
            return False
        with connect() as db:
            row = db.execute(
                "SELECT request_payload_json FROM crm_n8n_runs WHERE run_key=?",
                (run_key,),
            ).fetchone()
        if not row:
            return False
        try:
            stored = json.loads(row["request_payload_json"] or "{}")
        except (TypeError, json.JSONDecodeError):
            return False
        expected = str(stored.get("crm_event_token_hash") or "")
        actual = hashlib.sha256(supplied.encode("utf-8")).hexdigest()
        return bool(expected and hmac.compare_digest(actual, expected))

    def crm_n8n_callback_authorized(self) -> bool:
        """Valida callbacks automáticos do n8n sem reutilizar segredos globais."""
        supplied = str(self.headers.get("X-CRM-N8N-Token") or "").strip()
        if not supplied:
            return False
        supplied_hash = hashlib.sha256(supplied.encode("utf-8")).hexdigest()
        with connect() as db:
            rows = db.execute(
                "SELECT key_hash FROM crm_n8n_callback_keys WHERE active=1"
            ).fetchall()
        return any(
            hmac.compare_digest(supplied_hash, str(row["key_hash"]))
            for row in rows
        )

    def mark_crm_ai_message(self, payload: dict, query: dict) -> None:
        """Registra autoria da automação independentemente da ordem dos webhooks.

        O n8n deve chamar este endpoint após o envio pela Evolution usando o
        mesmo ID externo retornado pela API. Se o webhook da mensagem ainda
        não chegou, a atribuição fica aguardando e será aplicada na importação.
        """
        if not self.shared_integration_authorized(query):
            return self.send_json({"error": "Integração não autorizada."}, HTTPStatus.UNAUTHORIZED)
        external_id = str(payload.get("external_message_id") or payload.get("message_id") or payload.get("id") or "").strip()
        label = str(payload.get("author_label") or payload.get("agent_name") or "Assistente IA").strip()[:80]
        if not external_id:
            return self.send_json({"error": "Informe o ID externo da mensagem."}, HTTPStatus.BAD_REQUEST)
        with connect() as db:
            db.execute("""INSERT INTO crm_message_attributions(external_message_id,author_type,author_label)
                          VALUES(?,'ai',?) ON CONFLICT(external_message_id) DO UPDATE SET
                          author_type='ai',author_label=excluded.author_label""", (external_id, label))
            updated = db.execute("""UPDATE crm_messages SET author_type='ai',author_label=?
                                    WHERE external_message_id=?""", (label, external_id)).rowcount
        self.send_json({"attributed": True, "message_found": bool(updated), "external_message_id": external_id})

    def receive_crm_handoff(self, payload: dict, query: dict) -> None:
        """Recebe do n8n/IA uma transferência explícita para a fila humana."""
        if not self.shared_integration_authorized(query):
            return self.send_json({"error": "Integração não autorizada."}, HTTPStatus.UNAUTHORIZED)
        instance = str(payload.get("instance") or payload.get("instance_name") or "").strip()
        phone = self.crm_phone(payload.get("phone") or payload.get("number"))
        reason = str(payload.get("reason") or "Transferido pela IA").strip()[:120]
        if not instance or not phone:
            return self.send_json({"error": "Informe a instância e o telefone."}, HTTPStatus.BAD_REQUEST)
        with connect() as db:
            row = db.execute("""SELECT cv.id FROM crm_conversations cv
                JOIN crm_channels ch ON ch.id=cv.channel_id
                JOIN crm_contacts ct ON ct.id=cv.contact_id
                WHERE ch.instance_name=? AND ct.phone=?""", (instance, phone)).fetchone()
            if not row:
                return self.send_json({"error": "Conversa não encontrada para transferência."}, HTTPStatus.NOT_FOUND)
            db.execute("""UPDATE crm_conversations SET status='Aberta',pipeline_stage='Novo',queue_name=?,
                          assigned_user_id=NULL,assigned_at=NULL,resolved_at=NULL,unread_count=GREATEST(unread_count,1),
                          resolved_by_user_id=NULL,last_direction='inbound',queue_entered_at=datetime('now','localtime'),
                          automation_state='handoff',handoff_reason=?,
                          updated_at=datetime('now','localtime') WHERE id=?""", (reason, reason, row["id"]))
        self.send_json({"queued": True, "conversation_id": row["id"]})

    def receive_crm_automation_event(self, payload: dict, query: dict) -> None:
        """Entrada idempotente para monitorar campanhas e decisões do n8n/IA."""
        if not (
            self.shared_integration_authorized(query)
            or self.crm_n8n_run_event_authorized(payload)
            or self.crm_n8n_callback_authorized()
        ):
            return self.send_json({"error":"Integração não autorizada."},HTTPStatus.UNAUTHORIZED)
        event_type = str(payload.get("event_type") or payload.get("event") or "").strip().lower()
        event_key = str(payload.get("event_id") or payload.get("event_key") or "").strip()
        instance = str(payload.get("instance") or payload.get("instance_name") or "").strip()
        phone = self.crm_phone(payload.get("phone") or payload.get("number"))
        if not event_type or not event_key:
            return self.send_json({"error":"Informe event_id e event_type."},HTTPStatus.BAD_REQUEST)
        safe_payload = dict(payload)
        safe_payload.pop("crm_event_token", None)
        raw = json.dumps(safe_payload,ensure_ascii=False,sort_keys=True,separators=(",",":"))
        with connect() as db:
            conversation = None
            if instance and phone:
                conversation = db.execute("""SELECT cv.id,ct.id AS contact_id,ct.name AS contact_name,
                    ch.display_name AS channel_name FROM crm_conversations cv
                    JOIN crm_channels ch ON ch.id=cv.channel_id JOIN crm_contacts ct ON ct.id=cv.contact_id
                    WHERE ch.instance_name=? AND ct.phone=?""",(instance,phone)).fetchone()
            conversation_id = conversation["id"] if conversation else None
            try:
                db.execute("""INSERT INTO crm_automation_events(event_key,conversation_id,campaign_id,flow_name,event_type,outcome,payload_json)
                              VALUES(?,?,?,?,?,?,?)""",(event_key,conversation_id,str(payload.get("campaign_id") or "")[:100] or None,
                              str(payload.get("flow_name") or "")[:120] or None,event_type,str(payload.get("outcome") or "")[:80] or None,raw))
            except IntegrityError:
                return self.send_json({"received":True,"duplicate":True})
            workflow_id = str(payload.get("workflow_id") or payload.get("crm_workflow_id") or "").strip()[:100]
            execution_id = str(payload.get("execution_id") or payload.get("executionId") or "").strip()[:100]
            run_key = str(payload.get("run_key") or payload.get("crm_run_key") or "").strip()[:160]
            flow_name = str(payload.get("flow_name") or payload.get("workflow_name") or "")[:120] or None
            run = None
            if run_key:
                run = db.execute("SELECT id FROM crm_n8n_runs WHERE run_key=?", (run_key,)).fetchone()
            if not run and execution_id:
                run = db.execute(
                    "SELECT id FROM crm_n8n_runs WHERE n8n_execution_id=? ORDER BY id DESC LIMIT 1",
                    (execution_id,),
                ).fetchone()
            if not run and workflow_id and phone:
                run = db.execute(
                    """
                    SELECT r.id
                    FROM crm_n8n_runs r
                    JOIN crm_n8n_patient_events ev ON ev.run_id=r.id
                    WHERE r.workflow_id=? AND ev.phone=?
                    ORDER BY r.id DESC
                    LIMIT 1
                    """,
                    (workflow_id, phone),
                ).fetchone()
            if not run and workflow_id and execution_id:
                generated_run_key = f"n8n:{execution_id}"
                db.execute(
                    """INSERT OR IGNORE INTO crm_n8n_runs
                       (run_key,workflow_id,workflow_name,n8n_execution_id,mode,status,request_payload_json)
                       VALUES(?,?,?,?,?,'running','{}')""",
                    (generated_run_key, workflow_id, flow_name or "Fluxo n8n", execution_id, "automatic"),
                )
                run = db.execute("SELECT id FROM crm_n8n_runs WHERE run_key=?", (generated_run_key,)).fetchone()
            run_id = run["id"] if run else None
            patient_name = str(
                payload.get("patient_name")
                or payload.get("name")
                or (conversation["contact_name"] if conversation else "")
            ).strip()[:160] or None
            channel_name = str(
                payload.get("channel")
                or payload.get("channel_name")
                or (conversation["channel_name"] if conversation else "")
            ).strip()[:120] or None
            occurred_at = str(
                payload.get("occurred_at")
                or payload.get("timestamp")
                or payload.get("created_at")
                or ""
            ).strip()[:40] or None
            external_message_id = str(
                payload.get("external_message_id")
                or payload.get("message_id")
                or ""
            ).strip()[:180] or None
            appointment_source = self.crm_appointment_source(payload)
            try:
                db.execute(
                    """INSERT INTO crm_n8n_patient_events
                       (event_key,run_id,workflow_id,execution_id,conversation_id,campaign_id,
                        contact_id,patient_name,phone,channel_name,event_type,outcome,
                        appointment_source,external_message_id,details_json,occurred_at)
                       VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (
                        event_key, run_id, workflow_id or None, execution_id or None,
                        conversation_id, str(payload.get("campaign_id") or "")[:100] or None,
                        conversation["contact_id"] if conversation else None,
                        patient_name, phone or None, channel_name, event_type,
                        str(payload.get("outcome") or "")[:80] or None,
                        appointment_source, external_message_id, raw, occurred_at,
                    ),
                )
            except IntegrityError:
                pass
            if run_id:
                if event_type in {"workflow.finished", "campaign.finished"}:
                    run_status = "success"
                    finished_sql = "datetime('now','localtime')"
                elif event_type in {"workflow.failed", "campaign.failed"}:
                    run_status = "failed"
                    finished_sql = "datetime('now','localtime')"
                else:
                    run_status = "running"
                    finished_sql = "finished_at"
                db.execute(
                    f"""UPDATE crm_n8n_runs SET
                        status=?,
                        n8n_execution_id=COALESCE(NULLIF(?,''),n8n_execution_id),
                        total_items=(SELECT COUNT(DISTINCT COALESCE(NULLIF(phone,''),event_key))
                                     FROM crm_n8n_patient_events WHERE run_id=?
                                     AND event_type NOT IN
                                       ('workflow.started','workflow.finished','workflow.failed',
                                        'campaign.started','campaign.finished','campaign.failed')),
                        processed_items=(SELECT COUNT(DISTINCT COALESCE(NULLIF(phone,''),event_key))
                                         FROM crm_n8n_patient_events WHERE run_id=?
                                         AND event_type IN ('patient.selected','contact.loaded','item.processed')),
                        sent_items=(SELECT COUNT(DISTINCT COALESCE(NULLIF(phone,''),event_key))
                                    FROM crm_n8n_patient_events WHERE run_id=?
                                    AND event_type IN ('message.sent','message.ai.sent')),
                        delivered_items=(SELECT COUNT(DISTINCT COALESCE(NULLIF(phone,''),event_key))
                                         FROM crm_n8n_patient_events WHERE run_id=?
                                         AND event_type IN ('message.delivered','message.read')),
                        replied_items=(SELECT COUNT(DISTINCT COALESCE(NULLIF(phone,''),event_key))
                                       FROM crm_n8n_patient_events WHERE run_id=?
                                       AND event_type IN ('patient.replied','message.received')),
                        failed_items=(SELECT COUNT(DISTINCT COALESCE(NULLIF(phone,''),event_key))
                                      FROM crm_n8n_patient_events WHERE run_id=?
                                      AND event_type IN ('message.failed','item.failed','workflow.failed','campaign.failed')),
                        appointment_items=(SELECT COUNT(DISTINCT COALESCE(NULLIF(phone,''),event_key))
                                           FROM crm_n8n_patient_events WHERE run_id=?
                                           AND event_type='appointment.confirmed'),
                        handoff_items=(SELECT COUNT(DISTINCT COALESCE(NULLIF(phone,''),event_key))
                                       FROM crm_n8n_patient_events WHERE run_id=?
                                       AND event_type IN ('ai.handoff.requested','human.required','opportunity.detected')),
                        finished_at={finished_sql},
                        updated_at=datetime('now','localtime')
                        WHERE id=?""",
                    (
                        run_status, execution_id, run_id, run_id, run_id, run_id,
                        run_id, run_id, run_id, run_id, run_id,
                    ),
                )
            if conversation_id:
                if event_type in {"message.ai.sent","ai.started","campaign.started"}:
                    db.execute("""UPDATE crm_conversations SET automation_state='ai_active',automation_flow=COALESCE(?,automation_flow),
                                  automation_turns=automation_turns+CASE WHEN ?='message.ai.sent' THEN 1 ELSE 0 END,
                                  updated_at=datetime('now','localtime') WHERE id=?""",(flow_name,event_type,conversation_id))
                elif event_type in {"ai.handoff.requested","human.required","opportunity.detected"}:
                    reason = str(payload.get("reason") or payload.get("outcome") or "Intervenção humana solicitada")[:160]
                    db.execute("""UPDATE crm_conversations SET automation_state='handoff',handoff_reason=?,status='Aberta',
                                  pipeline_stage='Novo',assigned_user_id=NULL,assigned_at=NULL,
                                  queue_entered_at=COALESCE(queue_entered_at,datetime('now','localtime')),
                                  updated_at=datetime('now','localtime') WHERE id=?""",(reason,conversation_id))
                elif event_type in {"appointment.confirmed","conversation.closed","campaign.finished"}:
                    db.execute("""UPDATE crm_conversations SET automation_state='completed',automation_flow=COALESCE(?,automation_flow),
                                  updated_at=datetime('now','localtime') WHERE id=?""",(flow_name,conversation_id))
        self.send_json({"received":True,"conversation_id":conversation_id,"event_type":event_type})

    def receive_evolution_webhook(self, payload: dict, query: dict) -> None:
        if not self.evolution_webhook_authorized(query):
            return self.send_json({"error": "Webhook não autorizado."}, HTTPStatus.UNAUTHORIZED)
        event_type = str(payload.get("event") or payload.get("type") or "evento").lower()
        instance = str(payload.get("instance") or payload.get("instanceName") or "Evolution").strip()
        data = payload.get("data") if isinstance(payload.get("data"), dict) else payload
        key = data.get("key") if isinstance(data.get("key"), dict) else {}
        external_id = str(key.get("id") or data.get("id") or "").strip()
        raw = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
        event_key = external_id or hashlib.sha256(raw.encode()).hexdigest()
        configured_url, configured_key = self.evolution_credentials()
        with connect() as db:
            try: db.execute("INSERT INTO crm_webhook_events(event_key,instance_name,event_type,payload_json) VALUES(?,?,?,?)", (event_key,instance,event_type,raw))
            except IntegrityError: return self.send_json({"received": True, "duplicate": True})
            try:
                db.execute("""INSERT INTO crm_channels(instance_name,display_name,evolution_base_url,evolution_api_key,connection_status,last_event_at)
                              VALUES(?,?,?,?,'Conectado',datetime('now','localtime'))
                              ON CONFLICT(instance_name) DO UPDATE SET connection_status='Conectado',last_event_at=datetime('now','localtime'),
                              evolution_base_url=COALESCE(crm_channels.evolution_base_url,excluded.evolution_base_url),evolution_api_key=COALESCE(crm_channels.evolution_api_key,excluded.evolution_api_key)""",
                           (instance,instance,configured_url or None,
                            encrypt_integration_secret(configured_key) if configured_key else None))
                channel_settings = db.execute("SELECT id,sync_enabled FROM crm_channels WHERE instance_name=?", (instance,)).fetchone()
                if channel_settings and not channel_settings["sync_enabled"]:
                    db.execute("""UPDATE crm_webhook_events SET processing_status='Ignorado',
                                  error_message='Canal pausado no CRM',processed_at=datetime('now','localtime')
                                  WHERE event_key=?""", (event_key,))
                    return self.send_json({"received": True, "processed": False, "reason": "channel_sync_disabled"})
                if "message" not in event_type:
                    db.execute("UPDATE crm_webhook_events SET processing_status='Ignorado',processed_at=datetime('now','localtime') WHERE event_key=?", (event_key,)); return self.send_json({"received":True,"processed":False})
                if event_type == "messages.update":
                    message_external_id = str(data.get("keyId") or data.get("messageId") or external_id).strip()
                    delivery_status = str(data.get("status") or "Atualizada").strip().title()
                    if message_external_id:
                        db.execute("UPDATE crm_messages SET delivery_status=? WHERE external_message_id=?", (delivery_status, message_external_id))
                    db.execute("UPDATE crm_webhook_events SET processing_status='Processado',processed_at=datetime('now','localtime') WHERE event_key=?", (event_key,))
                    return self.send_json({"received": True, "processed": True, "kind": "delivery_update"})
                remote = str(key.get("remoteJid") or data.get("remoteJid") or data.get("from") or "")
                remote_alt = str(key.get("remoteJidAlt") or data.get("remoteJidAlt") or "")
                phone = self.crm_phone(remote.split("@")[0]) or self.crm_phone(remote_alt.split("@")[0])
                if not phone or remote.endswith("@g.us"):
                    db.execute("UPDATE crm_webhook_events SET processing_status='Ignorado',error_message='Contato sem telefone individual',processed_at=datetime('now','localtime') WHERE event_key=?", (event_key,)); return self.send_json({"received":True,"processed":False})
                message = data.get("message") if isinstance(data.get("message"), dict) else {}
                body = str(message.get("conversation") or (message.get("extendedTextMessage") or {}).get("text") or data.get("text") or "")
                kind, media_url, mime_type, duration_seconds = "text", None, None, None
                for candidate, media_kind in (("imageMessage","image"),("audioMessage","audio"),("videoMessage","video"),("documentMessage","document")):
                    if isinstance(message.get(candidate), dict):
                        media=message[candidate]; kind=media_kind; body=str(media.get("caption") or media.get("fileName") or f"[{media_kind}]"); media_url=media.get("url"); mime_type=media.get("mimetype")
                        if media_kind == "audio":
                            try: duration_seconds=float(media.get("seconds") or 0) or None
                            except (TypeError,ValueError): duration_seconds=None
                        break
                from_me = bool(key.get("fromMe")); direction = "outbound" if from_me else "inbound"
                name = str(data.get("pushName") or phone).strip(); patient_id=None
                timestamp=data.get("messageTimestamp")
                message_at=self.evolution_message_time(timestamp) if timestamp else datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                if message_at < EVOLUTION_HISTORY_CUTOFF:
                    db.execute("""UPDATE crm_webhook_events SET processing_status='Ignorado',
                                  error_message='Mensagem anterior ao corte operacional',
                                  processed_at=datetime('now','localtime') WHERE event_key=?""", (event_key,))
                    return self.send_json({"received": True, "processed": False, "reason": "before_cutoff"})
                for patient in db.execute("SELECT id,name,phone FROM patients WHERE phone IS NOT NULL AND TRIM(phone)<>''").fetchall():
                    if self.crm_phone(patient["phone"])==phone: patient_id,name=patient["id"],patient["name"]; break
                db.execute("""INSERT INTO crm_contacts(patient_id,name,phone) VALUES(?,?,?) ON CONFLICT(phone) DO UPDATE SET
                              patient_id=COALESCE(excluded.patient_id,crm_contacts.patient_id),name=CASE WHEN excluded.patient_id IS NOT NULL THEN excluded.name ELSE crm_contacts.name END,updated_at=datetime('now','localtime')""", (patient_id,name,phone))
                contact_id=db.execute("SELECT id FROM crm_contacts WHERE phone=?",(phone,)).fetchone()[0]; channel_id=db.execute("SELECT id FROM crm_channels WHERE instance_name=?",(instance,)).fetchone()[0]
                db.execute("""INSERT INTO crm_conversations(channel_id,contact_id,status,last_message_at,last_direction,unread_count,pipeline_stage,queue_entered_at)
                              VALUES(?,?,'Aberta',?,?,?,'Novo',CASE WHEN ?='inbound' THEN ? ELSE NULL END)
                              ON CONFLICT(channel_id,contact_id) DO UPDATE SET
                              assigned_user_id=CASE WHEN crm_conversations.status='Resolvida' AND excluded.last_direction='inbound'
                                  AND (crm_conversations.resolved_at IS NULL OR datetime(excluded.last_message_at)>datetime(crm_conversations.resolved_at))
                                  THEN NULL ELSE crm_conversations.assigned_user_id END,
                              queue_name=CASE WHEN crm_conversations.status='Resolvida' AND excluded.last_direction='inbound'
                                  AND (crm_conversations.resolved_at IS NULL OR datetime(excluded.last_message_at)>datetime(crm_conversations.resolved_at))
                                  THEN 'Entrada' ELSE crm_conversations.queue_name END,
                              pipeline_stage=CASE WHEN crm_conversations.status='Resolvida' AND excluded.last_direction='inbound'
                                  AND (crm_conversations.resolved_at IS NULL OR datetime(excluded.last_message_at)>datetime(crm_conversations.resolved_at))
                                  THEN 'Novo' ELSE crm_conversations.pipeline_stage END,
                              status=CASE WHEN excluded.last_direction='inbound' AND
                                  (crm_conversations.status<>'Resolvida' OR crm_conversations.resolved_at IS NULL OR
                                   datetime(excluded.last_message_at)>datetime(crm_conversations.resolved_at))
                                  THEN 'Aberta' ELSE crm_conversations.status END,
                              resolved_at=CASE WHEN excluded.last_direction='inbound' AND
                                  (crm_conversations.status<>'Resolvida' OR crm_conversations.resolved_at IS NULL OR
                                   datetime(excluded.last_message_at)>datetime(crm_conversations.resolved_at))
                                  THEN NULL ELSE crm_conversations.resolved_at END,
                              resolved_by_user_id=CASE WHEN crm_conversations.status='Resolvida' AND excluded.last_direction='inbound'
                                  AND (crm_conversations.resolved_at IS NULL OR datetime(excluded.last_message_at)>datetime(crm_conversations.resolved_at))
                                  THEN NULL ELSE crm_conversations.resolved_by_user_id END,
                              queue_entered_at=CASE
                                  WHEN crm_conversations.status='Resolvida' AND excluded.last_direction='inbound'
                                       AND (crm_conversations.resolved_at IS NULL OR datetime(excluded.last_message_at)>datetime(crm_conversations.resolved_at))
                                  THEN excluded.last_message_at
                                  WHEN crm_conversations.queue_entered_at IS NULL AND crm_conversations.assigned_user_id IS NULL
                                       AND excluded.last_direction='inbound' THEN excluded.last_message_at
                                  ELSE crm_conversations.queue_entered_at END,
                              last_message_at=CASE WHEN crm_conversations.last_message_at IS NULL OR datetime(excluded.last_message_at)>datetime(crm_conversations.last_message_at)
                                                   THEN excluded.last_message_at ELSE crm_conversations.last_message_at END,
                              last_direction=CASE WHEN crm_conversations.last_message_at IS NULL OR datetime(excluded.last_message_at)>datetime(crm_conversations.last_message_at)
                                                  THEN excluded.last_direction ELSE crm_conversations.last_direction END,
                              unread_count=CASE
                                  WHEN crm_conversations.status='Resolvida' AND crm_conversations.resolved_at IS NOT NULL
                                       AND datetime(excluded.last_message_at)<=datetime(crm_conversations.resolved_at) THEN 0
                                  ELSE crm_conversations.unread_count+excluded.unread_count END""",
                           (channel_id,contact_id,message_at,direction,0 if from_me else 1,direction,message_at))
                conversation_id=db.execute("SELECT id FROM crm_conversations WHERE channel_id=? AND contact_id=?",(channel_id,contact_id)).fetchone()[0]
                message_external_id = external_id or event_key
                attribution = db.execute("SELECT author_type,author_label FROM crm_message_attributions WHERE external_message_id=?", (message_external_id,)).fetchone()
                author_type = attribution["author_type"] if attribution else ("external" if from_me else "patient")
                author_label = attribution["author_label"] if attribution else ("Enviado fora do CRM" if from_me else name)
                source_channel = str(data.get("source") or "evolution").strip()[:40]
                db.execute("""INSERT OR IGNORE INTO crm_messages
                    (conversation_id,external_message_id,direction,message_type,body,media_url,mime_type,duration_seconds,sender_name,
                     author_type,author_label,source_channel,delivery_status,message_at)
                    VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (conversation_id,message_external_id,direction,kind,body,media_url,mime_type,duration_seconds,name,
                     author_type,author_label,source_channel,"Enviada" if from_me else "Recebida",message_at))
                if not from_me:
                    handoff_reason = self.crm_ai_handoff_reason(body)
                    automation = db.execute("SELECT automation_state FROM crm_conversations WHERE id=?", (conversation_id,)).fetchone()
                    if handoff_reason and automation and automation["automation_state"] in {"ai_active", "manual", "external"}:
                        db.execute("""UPDATE crm_conversations SET automation_state='handoff',handoff_reason=?,status='Aberta',
                                  pipeline_stage='Novo',assigned_user_id=NULL,assigned_at=NULL,
                                  queue_entered_at=COALESCE(queue_entered_at,?),updated_at=datetime('now','localtime')
                                  WHERE id=?""", (handoff_reason, message_at, conversation_id))
                        db.execute("""INSERT INTO crm_conversation_events(conversation_id,event_type,actor_name,details_json)
                                  VALUES(?,'ai.handoff','IA',?)""",
                                   (conversation_id, json.dumps({"reason": handoff_reason}, ensure_ascii=False)))
                db.execute("UPDATE crm_webhook_events SET processing_status='Processado',processed_at=datetime('now','localtime') WHERE event_key=?",(event_key,))
            except Exception as error:
                db.execute("UPDATE crm_webhook_events SET processing_status='Falhou',error_message=?,processed_at=datetime('now','localtime') WHERE event_key=?",(str(error)[:500],event_key)); return self.send_json({"error":"Falha ao processar webhook."},HTTPStatus.BAD_REQUEST)
        self.send_json({"received":True,"processed":True})

    def send_crm_message(self, conversation_id: int, payload: dict) -> None:
        if not self.require_crc_access(): return
        if not self.require_crm_feature("inbox"):
            return
        message_type = str(payload.get("message_type") or "text").strip().lower()
        body = str(payload.get("text") or "").strip()
        audio_bytes = b""
        media_bytes = b""
        original_file_name = ""
        mime_type = None
        media_url = None
        duration_seconds = None
        if message_type == "audio":
            audio_base64 = str(payload.get("audio_base64") or "").strip()
            mime_type = str(payload.get("mime_type") or "audio/webm").split(";", 1)[0].strip().lower()
            allowed_audio_types = {"audio/webm", "audio/ogg", "audio/mp4", "audio/mpeg", "audio/x-m4a"}
            if mime_type not in allowed_audio_types:
                return self.send_json({"error": "Formato de áudio não suportado."}, HTTPStatus.BAD_REQUEST)
            try:
                audio_bytes = base64.b64decode(audio_base64, validate=True)
            except (ValueError, TypeError):
                return self.send_json({"error": "A gravação de áudio está corrompida."}, HTTPStatus.BAD_REQUEST)
            if not audio_bytes:
                return self.send_json({"error": "Grave um áudio antes de enviar."}, HTTPStatus.BAD_REQUEST)
            if len(audio_bytes) > 8 * 1024 * 1024:
                return self.send_json({"error": "O áudio ultrapassa o limite de 8 MB."}, HTTPStatus.REQUEST_ENTITY_TOO_LARGE)
            body = "Áudio"
            try:
                audio_bytes = convert_crm_audio_to_ogg(audio_bytes)
            except RuntimeError as error:
                return self.send_json({"error": str(error)}, HTTPStatus.UNPROCESSABLE_ENTITY)
            mime_type = "audio/ogg"
            duration_seconds = crm_audio_duration_seconds(audio_bytes)
        elif message_type in {"image", "video", "document"}:
            encoded_media = str(payload.get("media_base64") or "").strip()
            mime_type = str(payload.get("mime_type") or "application/octet-stream").split(";", 1)[0].strip().lower()
            original_file_name = Path(str(payload.get("file_name") or "")).name[:180]
            expected_group = {
                "image": {"image/jpeg", "image/png", "image/webp"},
                "video": {"video/mp4"},
                "document": {
                    "application/pdf", "application/msword",
                    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                    "application/vnd.ms-excel",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "application/vnd.ms-powerpoint",
                    "application/vnd.openxmlformats-officedocument.presentationml.presentation",
                    "text/plain", "application/zip", "application/x-zip-compressed",
                },
            }[message_type]
            if mime_type not in expected_group:
                return self.send_json({"error": "Formato de arquivo não suportado."}, HTTPStatus.BAD_REQUEST)
            try:
                media_bytes = base64.b64decode(encoded_media, validate=True)
            except (ValueError, TypeError):
                return self.send_json({"error": "O arquivo enviado está corrompido."}, HTTPStatus.BAD_REQUEST)
            if not media_bytes:
                return self.send_json({"error": "Selecione um arquivo antes de enviar."}, HTTPStatus.BAD_REQUEST)
            if len(media_bytes) > 12 * 1024 * 1024:
                return self.send_json({"error": "O arquivo ultrapassa o limite de 12 MB."}, HTTPStatus.REQUEST_ENTITY_TOO_LARGE)
            if not self.crm_media_content_valid(mime_type, media_bytes):
                return self.send_json(
                    {"error": "O conteúdo do arquivo não corresponde ao formato informado."},
                    HTTPStatus.BAD_REQUEST,
                )
            if not original_file_name:
                original_file_name = {
                    "image/jpeg": "imagem.jpg", "image/png": "imagem.png", "image/webp": "imagem.webp",
                    "video/mp4": "video.mp4", "application/pdf": "documento.pdf",
                    "application/msword": "documento.doc",
                    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": "documento.docx",
                    "application/vnd.ms-excel": "planilha.xls",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "planilha.xlsx",
                    "application/vnd.ms-powerpoint": "apresentacao.ppt",
                    "application/vnd.openxmlformats-officedocument.presentationml.presentation": "apresentacao.pptx",
                    "text/plain": "documento.txt", "application/zip": "arquivo.zip",
                    "application/x-zip-compressed": "arquivo.zip",
                }[mime_type]
            body = body or original_file_name
        elif message_type != "text":
            return self.send_json({"error": "Tipo de mensagem não suportado."}, HTTPStatus.BAD_REQUEST)
        elif not body:
            return self.send_json({"error":"Digite uma mensagem."},HTTPStatus.BAD_REQUEST)
        with connect() as db:
            row=db.execute("""SELECT cv.channel_id,ct.id AS contact_id,ct.phone,ct.is_internal,ch.instance_name,ch.display_name,ch.evolution_base_url,ch.evolution_api_key,
                                      cv.assigned_user_id,u.name AS assigned_to
                               FROM crm_conversations cv
                              JOIN crm_contacts ct ON ct.id=cv.contact_id JOIN crm_channels ch ON ch.id=cv.channel_id
                               LEFT JOIN users u ON u.id=cv.assigned_user_id
                               WHERE cv.id=? AND ch.active=1 AND ch.sync_enabled=1""",(conversation_id,)).fetchone()
            if not row: return self.send_json({"error":"Conversa ou canal não encontrado."},HTTPStatus.NOT_FOUND)
            if not self.crm_channel_allowed(db, row["channel_id"], "reply"):
                return self.send_json({"error":"Você não possui permissão para responder por este número."},HTTPStatus.FORBIDDEN)
            if not row["is_internal"]:
                active_owner = self.crm_contact_active_owner(db, row["contact_id"])
                if active_owner and active_owner["assigned_user_id"] != self.authenticated_user["id"]:
                    return self.reject_crm_contact_owner_conflict(active_owner)
            if row["assigned_user_id"] and row["assigned_user_id"] != self.authenticated_user["id"]:
                return self.send_json({"error":f"Este atendimento já está atribuído a {row['assigned_to']}. Transfira antes de responder."},HTTPStatus.CONFLICT)
            if not row["assigned_user_id"] and not row["is_internal"]:
                return self.send_json({"error":"Inicie o atendimento antes de enviar mensagens para um contato externo."},HTTPStatus.CONFLICT)
            configured_url, configured_key = self.evolution_credentials()
            base_url = row["evolution_base_url"] or configured_url
            api_key = (
                decrypt_integration_secret(row["evolution_api_key"])
                if row["evolution_api_key"] else configured_key
            )
            if not base_url or not api_key: return self.send_json({"error":"Configure a URL e a chave da Evolution neste canal antes de enviar."},HTTPStatus.CONFLICT)
            try:
                connection_state = self.crm_evolution_connection_state(row["instance_name"], base_url, api_key)
            except RuntimeError as error:
                return self.send_json({"error":f"Não foi possível verificar a conexão do canal {row['display_name']}: {error}"},HTTPStatus.BAD_GATEWAY)
            if connection_state != "open":
                db.execute("UPDATE crm_channels SET connection_status='Desconectado',updated_at=datetime('now','localtime') WHERE id=?", (row["channel_id"],))
                return self.send_json({
                    "error": f"O WhatsApp do canal {row['display_name']} está desconectado. Reconecte o número por QR Code em Integrações antes de enviar.",
                    "code": "WHATSAPP_DISCONNECTED",
                }, HTTPStatus.CONFLICT)
            db.execute("UPDATE crm_channels SET connection_status='Conectado',updated_at=datetime('now','localtime') WHERE id=?", (row["channel_id"],))
            instance_path = quote(str(row["instance_name"]), safe="")
            if message_type == "audio":
                evolution_path = f"/message/sendWhatsAppAudio/{instance_path}"
                evolution_payload = {"number": f"55{row['phone']}", "audio": base64.b64encode(audio_bytes).decode("ascii"), "encoding": False}
            elif message_type in {"image", "video", "document"}:
                evolution_path = f"/message/sendMedia/{instance_path}"
                evolution_payload = {
                    "number": f"55{row['phone']}",
                    "mediatype": message_type,
                    "mimetype": mime_type,
                    "caption": body,
                    "media": base64.b64encode(media_bytes).decode("ascii"),
                    "fileName": original_file_name,
                }
            else:
                evolution_path = f"/message/sendText/{instance_path}"
                evolution_payload = {"number": f"55{row['phone']}", "text": body}
            request=Request(f"{base_url.rstrip('/')}{evolution_path}",data=json.dumps(evolution_payload).encode(),method="POST",headers={"Content-Type":"application/json","apikey":api_key})
            try:
                request_timeout = 60 if message_type != "text" else 20
                with urlopen(request,timeout=request_timeout) as response: response_data=json.loads(response.read().decode() or "{}")
            except HTTPError as error:
                evolution_error = error.read().decode(errors="replace")[:500]
                print(f"[crm-send] Evolution HTTP {error.code} na conversa {conversation_id}: {evolution_error}", flush=True)
                if "ConnectionClosed" in evolution_error or "connection closed" in evolution_error.lower():
                    db.execute("UPDATE crm_channels SET connection_status='Desconectado',updated_at=datetime('now','localtime') WHERE id=?", (row["channel_id"],))
                    return self.send_json({
                        "error": f"O WhatsApp do canal {row['display_name']} foi desconectado. Reconecte o número por QR Code em Integrações e tente novamente.",
                        "code": "WHATSAPP_DISCONNECTED",
                    }, HTTPStatus.CONFLICT)
                return self.send_json({"error":f"Evolution respondeu {error.code}: {evolution_error}"},HTTPStatus.BAD_GATEWAY)
            except (URLError,TimeoutError) as error: return self.send_json({"error":f"Não foi possível conectar à Evolution: {error}"},HTTPStatus.BAD_GATEWAY)
            external_id=str(((response_data.get("key") or {}).get("id")) or response_data.get("id") or secrets.token_hex(12))
            if message_type == "audio":
                extension = {"audio/webm": "webm", "audio/ogg": "ogg", "audio/mp4": "mp4", "audio/mpeg": "m4a", "audio/x-m4a": "m4a"}[mime_type]
                CRM_MEDIA_DIR.mkdir(parents=True, exist_ok=True)
                file_name = f"{secrets.token_hex(16)}.{extension}"
                target = CRM_MEDIA_DIR / file_name
                temporary = target.with_suffix(target.suffix + ".tmp")
                temporary.write_bytes(audio_bytes)
                temporary.replace(target)
                media_url = f"/api/crm/media/{file_name}"
            elif message_type in {"image", "video", "document"}:
                extension = self.crm_media_extension(mime_type, original_file_name)
                CRM_MEDIA_DIR.mkdir(parents=True, exist_ok=True)
                file_name = f"{secrets.token_hex(16)}.{extension}"
                target = CRM_MEDIA_DIR / file_name
                temporary = target.with_suffix(target.suffix + ".tmp")
                temporary.write_bytes(media_bytes)
                temporary.replace(target)
                media_url = f"/api/crm/media/{file_name}"
            service_sector = str(self.authenticated_user.get("service_sector") or "CRC").strip()
            author_label = f"{self.authenticated_user['name']} · {service_sector}"
            message_id=db.execute("""INSERT INTO crm_messages
                (conversation_id,external_message_id,direction,message_type,body,media_url,mime_type,duration_seconds,sender_name,sent_by_user_id,
                 author_type,author_label,source_channel,delivery_status,message_at)
                VALUES(?,?,'outbound',?,?,?,?,?,?,?,'human',?,'crm',?,datetime('now','localtime'))""",
                (conversation_id,external_id,message_type,body,media_url,mime_type,duration_seconds,self.authenticated_user["name"],self.authenticated_user["id"],
                 author_label,"Enviada")).lastrowid
            if row["is_internal"]:
                db.execute("""UPDATE crm_conversations SET status='Aberta',pipeline_stage='Novo',
                              assigned_user_id=NULL,assigned_at=NULL,queue_entered_at=NULL,first_response_at=NULL,
                              unread_count=0,last_direction='outbound',last_message_at=datetime('now','localtime'),updated_at=datetime('now','localtime')
                              WHERE id=?""", (conversation_id,))
            else:
                db.execute("""UPDATE crm_conversations SET status='Aberta',pipeline_stage='Em atendimento',
                              assigned_user_id=?,assigned_at=COALESCE(assigned_at,datetime('now','localtime')),
                              first_response_at=COALESCE(first_response_at,datetime('now','localtime')),
                              automation_state='paused',
                              unread_count=0,last_direction='outbound',last_message_at=datetime('now','localtime'),updated_at=datetime('now','localtime')
                              WHERE id=?""",(self.authenticated_user["id"],conversation_id))
                self.crm_record_event(db, conversation_id, "message.sent", {"message_type": message_type})
            result=db.execute("SELECT id,conversation_id,direction,message_type,body,media_url,mime_type,duration_seconds,sender_name,author_type,author_label,source_channel,delivery_status,message_at FROM crm_messages WHERE id=?",(message_id,)).fetchone()
        self.send_json(dict(result),HTTPStatus.CREATED)

    def update_crm_contact(self, contact_id: int, payload: dict) -> None:
        if not self.require_crc_access():
            return
        if not self.require_crm_feature("contacts"):
            return
        if "is_internal" not in payload:
            return self.send_json({"error": "Informe se este e um contato interno."}, HTTPStatus.BAD_REQUEST)
        is_internal = 1 if payload.get("is_internal") else 0
        with connect() as db:
            updated = db.execute("""UPDATE crm_contacts SET is_internal=?,updated_at=datetime('now','localtime')
                                    WHERE id=?""", (is_internal, contact_id)).rowcount
            if not updated:
                return self.send_json({"error": "Contato nao encontrado."}, HTTPStatus.NOT_FOUND)
            if is_internal:
                db.execute("""UPDATE crm_conversations SET assigned_user_id=NULL,assigned_at=NULL,
                              queue_entered_at=NULL,first_response_at=NULL,unread_count=0,pipeline_stage='Novo',
                              updated_at=datetime('now','localtime') WHERE contact_id=?""", (contact_id,))
        self.send_json({"updated": True, "id": contact_id, "is_internal": bool(is_internal)})

    def start_crm_conversation(self, payload: dict) -> None:
        """Open (or reopen) a conversation, optionally without sending a message."""
        if not self.require_crc_access():
            return
        if not self.require_crm_feature("contacts"):
            return
        name = str(payload.get("name") or "").strip()[:160]
        phone = self.crm_phone(payload.get("phone"))
        body = str(payload.get("text") or "").strip()
        open_only = bool(payload.get("open_only"))
        try:
            channel_id = int(payload.get("channel_id") or 0)
        except (TypeError, ValueError):
            channel_id = 0
        if not name:
            return self.send_json({"error": "Selecione um contato."}, HTTPStatus.BAD_REQUEST)
        if not phone:
            return self.send_json({"error": "O contato precisa ter um telefone válido com DDD."}, HTTPStatus.BAD_REQUEST)
        if not channel_id:
            return self.send_json({"error": "Selecione o número que enviará a mensagem."}, HTTPStatus.BAD_REQUEST)
        if not open_only and not body:
            return self.send_json({"error": "Digite a primeira mensagem."}, HTTPStatus.BAD_REQUEST)
        with connect() as db:
            channel = db.execute("SELECT id FROM crm_channels WHERE id=? AND active=1 AND sync_enabled=1", (channel_id,)).fetchone()
            if not channel:
                return self.send_json({"error": "Canal indisponível ou desconectado."}, HTTPStatus.CONFLICT)
            if not self.crm_channel_allowed(db, channel_id, "reply"):
                return self.send_json({"error": "Você não possui permissão para iniciar conversas por este número."}, HTTPStatus.FORBIDDEN)
            db.execute("""INSERT INTO crm_contacts(name,phone) VALUES(?,?)
                          ON CONFLICT(phone) DO UPDATE SET name=CASE WHEN excluded.name<>'' THEN excluded.name ELSE crm_contacts.name END,
                          updated_at=datetime('now','localtime')""", (name, phone))
            contact = db.execute("SELECT id FROM crm_contacts WHERE phone=?", (phone,)).fetchone()
            db.execute("SELECT id FROM crm_contacts WHERE id=? FOR UPDATE", (contact["id"],)).fetchone()
            # Abrir um atendimento manualmente Ã© uma declaraÃ§Ã£o explÃ­cita de
            # atendimento externo. Isso tambÃ©m corrige contatos que tenham sido
            # marcados como internos anteriormente; caso contrÃ¡rio, as rotinas
            # de sincronizaÃ§Ã£o removeriam a atribuiÃ§Ã£o logo depois da abertura.
            db.execute("""UPDATE crm_contacts SET is_internal=0,
                          updated_at=datetime('now','localtime') WHERE id=?""", (contact["id"],))
            active_owner = self.crm_contact_active_owner(db, contact["id"])
            if active_owner and active_owner["assigned_user_id"] != self.authenticated_user["id"]:
                return self.reject_crm_contact_owner_conflict(active_owner)
            existing = db.execute("""SELECT cv.id,cv.channel_id,cv.assigned_user_id,u.name AS assigned_to
                                     FROM crm_conversations cv
                                     LEFT JOIN users u ON u.id=cv.assigned_user_id
                                     WHERE contact_id=? AND status<>'Resolvida'
                                     ORDER BY CASE WHEN assigned_user_id=? THEN 0 ELSE 1 END,
                                              CASE WHEN channel_id=? THEN 0 ELSE 1 END,
                                              datetime(cv.updated_at) DESC,cv.id DESC LIMIT 1""",
                                  (contact["id"], self.authenticated_user["id"], channel_id)).fetchone()
            if existing:
                if existing["assigned_user_id"] and existing["assigned_user_id"] != self.authenticated_user["id"]:
                    return self.send_json({
                        "error": f"Este paciente já está em atendimento com {existing['assigned_to']}. Solicite a transferência antes de acessar.",
                        "code": "CONVERSATION_ASSIGNED_TO_ANOTHER_USER",
                        "conversation_id": int(existing["id"]),
                        "assigned_to": existing["assigned_to"],
                    }, HTTPStatus.CONFLICT)
                db.execute("""UPDATE crm_conversations SET status='Aberta',pipeline_stage='Em atendimento',
                              assigned_user_id=?,assigned_at=COALESCE(assigned_at,datetime('now','localtime')),
                              unread_count=0,resolved_at=NULL,resolved_by_user_id=NULL,
                              automation_state='paused',updated_at=datetime('now','localtime') WHERE id=?""",
                           (self.authenticated_user["id"], existing["id"]))
                conversation = existing
            elif open_only:
                db.execute("""INSERT INTO crm_conversations(channel_id,contact_id,status,pipeline_stage,assigned_user_id,
                                  assigned_at,unread_count,updated_at)
                              VALUES(?,?,'Aberta','Em atendimento',?,datetime('now','localtime'),0,datetime('now','localtime'))
                              ON CONFLICT(channel_id,contact_id) DO UPDATE SET status='Aberta',pipeline_stage='Em atendimento',
                                  assigned_user_id=excluded.assigned_user_id,
                                  assigned_at=COALESCE(crm_conversations.assigned_at,excluded.assigned_at),
                                  unread_count=0,resolved_at=NULL,resolved_by_user_id=NULL,updated_at=datetime('now','localtime')""",
                           (channel_id, contact["id"], self.authenticated_user["id"]))
            else:
                db.execute("""INSERT INTO crm_conversations(channel_id,contact_id,status,pipeline_stage,assigned_user_id,
                                  assigned_at,first_response_at,unread_count,last_direction,updated_at)
                              VALUES(?,?,'Aberta','Em atendimento',?,datetime('now','localtime'),datetime('now','localtime'),0,'outbound',datetime('now','localtime'))
                              ON CONFLICT(channel_id,contact_id) DO UPDATE SET status='Aberta',pipeline_stage='Em atendimento',
                                  assigned_user_id=excluded.assigned_user_id,assigned_at=COALESCE(crm_conversations.assigned_at,excluded.assigned_at),
                                  first_response_at=COALESCE(crm_conversations.first_response_at,excluded.first_response_at),
                                  unread_count=0,resolved_at=NULL,resolved_by_user_id=NULL,updated_at=datetime('now','localtime')""",
                           (channel_id, contact["id"], self.authenticated_user["id"]))
            if not existing:
                conversation = db.execute("SELECT id,channel_id FROM crm_conversations WHERE channel_id=? AND contact_id=?",
                                          (channel_id, contact["id"])).fetchone()
            self.crm_record_event(db, int(conversation["id"]), "conversation.started", {
                "source": "contacts", "open_only": open_only, "reused": bool(existing),
            })
        if open_only:
            return self.send_json({"ok": True, "conversation_id": int(conversation["id"]),
                                   "reused": bool(existing), "channel_id": int(conversation["channel_id"])})
        return self.send_crm_message(int(conversation["id"]), {"text": body})

    def resolve_crm_conversation(self, conversation_id: int, payload: dict) -> None:
        if not self.require_crc_access(): return
        if not self.require_crm_feature("inbox"):
            return
        allowed_categories = {"Primeira consulta", "Controle", "Tratamento", "Orçamento"}
        legacy_outcome_aliases = {
            "Pediu para reagendar": "Retorno",
            "Data específica": "Retorno",
            "Não respondeu": "Novo Contato IA",
            "Não respondeu/atendeu": "Novo Contato IA",
            "Não quer agendar": "Novo Contato IA",
            "Está tratando com outro profissional": "Em tratamento externo",
            "Informação fornecida": "Outros",
            "Contato interno": "Outros",
        }
        allowed_outcomes = {
            "Agendou", "Quer agendar", "Retorno", "Novo Contato IA",
            "Mudou de cidade", "Em tratamento externo", "Desqualificado", "Outros",
        }
        category = str(payload.get("category") or "").strip()
        patient_type = str(payload.get("patient_type") or "").strip()
        recovery_raw = payload.get("is_recovery", False)
        is_recovery = 1 if recovery_raw is True or str(recovery_raw).strip().lower() in {"1", "true", "sim", "yes"} else 0
        outcome = str(payload.get("outcome") or payload.get("reason") or "").strip()
        outcome = legacy_outcome_aliases.get(outcome, outcome)
        # Preserve older integrations while the CRM UI now records this field
        # explicitly. Existing first-consultation categories remain first;
        # the other legacy categories are recurrent returns without treatment.
        if not patient_type:
            patient_type = "Primeira consulta" if category == "Primeira consulta" else "Retorno s/ Tratamento"
        interest = str(payload.get("interest") or "").strip()
        origin = str(payload.get("origin") or "").strip()
        notes = str(payload.get("notes") or "").strip()[:2000]
        professional = str(payload.get("responsible_professional") or "").strip()[:160]
        scheduled_date = str(payload.get("scheduled_date") or "").strip()
        scheduled_time = str(payload.get("scheduled_time") or "").strip()
        schedule_type = str(payload.get("schedule_type") or "").strip().title()
        next_contact_at = str(payload.get("next_contact_at") or "").strip()
        loss_reason = str(payload.get("loss_reason") or "").strip()[:500]
        try:
            attempts = max(0, min(99, int(payload.get("attempts") or 0)))
        except (TypeError, ValueError):
            attempts = 0
        if category not in allowed_categories:
            return self.send_json({"error": "Selecione a categoria do atendimento."}, HTTPStatus.BAD_REQUEST)
        if patient_type not in CRM_PATIENT_TYPES:
            return self.send_json({"error": "Informe se é primeira consulta ou retorno sem tratamento."}, HTTPStatus.BAD_REQUEST)
        if outcome not in allowed_outcomes:
            return self.send_json({"error": "Selecione o resultado do atendimento."}, HTTPStatus.BAD_REQUEST)
        if is_recovery and patient_type != "Retorno s/ Tratamento":
            return self.send_json({"error": "Recuperação de paciente só pode ser marcada em retorno sem tratamento."}, HTTPStatus.BAD_REQUEST)
        if is_recovery and outcome != "Agendou":
            return self.send_json({"error": "A recuperação entra na meta somente quando o paciente agendou."}, HTTPStatus.BAD_REQUEST)
        if outcome == "Agendou":
            if not scheduled_date or not scheduled_time or not professional:
                return self.send_json({"error": "Informe data, horário e profissional do agendamento."}, HTTPStatus.BAD_REQUEST)
            if schedule_type not in {"Agendado", "Programado"}:
                return self.send_json({"error": "Informe se ficou agendado ou programado."}, HTTPStatus.BAD_REQUEST)
        if outcome in {"Quer agendar", "Retorno"} and not next_contact_at:
            return self.send_json({"error": "Informe quando o contato deve retornar."}, HTTPStatus.BAD_REQUEST)
        if outcome in {"Mudou de cidade", "Em tratamento externo", "Desqualificado", "Outros"} and not (loss_reason or notes):
            return self.send_json({"error": "Informe o motivo ou uma observação."}, HTTPStatus.BAD_REQUEST)
        with connect() as db:
            current=db.execute("""SELECT cv.*,ct.id AS contact_id,ct.name AS contact_name,
                                  ch.display_name AS channel_name,u.name AS assigned_to
                                  FROM crm_conversations cv
                                  JOIN crm_contacts ct ON ct.id=cv.contact_id
                                  JOIN crm_channels ch ON ch.id=cv.channel_id
                                  LEFT JOIN users u ON u.id=cv.assigned_user_id WHERE cv.id=?""",
                               (conversation_id,)).fetchone()
            if not current: return self.send_json({"error":"Conversa não encontrada."},HTTPStatus.NOT_FOUND)
            if not self.crm_channel_allowed(db, current["channel_id"], "reply"):
                return self.send_json({"error":"Você não possui permissão para resolver atendimentos deste número."},HTTPStatus.FORBIDDEN)
            db.execute("SELECT id FROM crm_contacts WHERE id=? FOR UPDATE", (current["contact_id"],)).fetchone()
            active_owner = self.crm_contact_active_owner(db, current["contact_id"])
            if active_owner and active_owner["assigned_user_id"] != self.authenticated_user["id"]:
                return self.reject_crm_contact_owner_conflict(active_owner)
            if current["assigned_user_id"] and current["assigned_user_id"] != self.authenticated_user["id"]:
                return self.send_json({"error":f"Este atendimento pertence a {current['assigned_to']}. Transfira antes de resolver."},HTTPStatus.CONFLICT)
            attendance_number = db.execute(
                "SELECT COUNT(*)+1 FROM crm_service_resolutions WHERE conversation_id=?",
                (conversation_id,),
            ).fetchone()[0]
            ai_involved = 1 if current["automation_state"] in {"ai_active", "handoff", "completed"} or current["automation_flow"] else 0
            final_actor = "IA" if str(payload.get("final_actor") or "").strip().upper() == "IA" else "Humano"
            wait_seconds = None
            service_seconds = None
            if current["queue_entered_at"]:
                wait_seconds = db.execute(
                    "SELECT MAX(0,CAST((julianday(COALESCE(?,datetime('now','localtime')))-julianday(?))*86400 AS INTEGER))",
                    (current["first_response_at"], current["queue_entered_at"]),
                ).fetchone()[0]
            if current["assigned_at"]:
                service_seconds = db.execute(
                    "SELECT MAX(0,CAST((julianday('now','localtime')-julianday(?))*86400 AS INTEGER))",
                    (current["assigned_at"],),
                ).fetchone()[0]
            resolution_cursor = db.execute("""INSERT INTO crm_service_resolutions(
                            conversation_id,contact_id,channel_id,attendance_number,patient_type,is_recovery,category,outcome,
                            interest,origin,responsible_professional,notes,scheduled_date,scheduled_time,
                            schedule_type,next_contact_at,attempts,loss_reason,resolved_by_user_id,
                            resolved_by_name,ai_involved,final_actor,campaign_name,workflow_name,
                            wait_seconds,service_seconds,metadata_json)
                          VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                       (conversation_id,current["contact_id"],current["channel_id"],attendance_number,
                        patient_type,is_recovery,category,outcome,interest or None,origin or None,professional or None,notes or None,
                        scheduled_date or None,scheduled_time or None,schedule_type or None,next_contact_at or None,
                        attempts,loss_reason or None,self.authenticated_user["id"],self.authenticated_user["name"],
                        ai_involved,final_actor,current["automation_flow"] or None,current["automation_flow"] or None,
                        wait_seconds,service_seconds,json.dumps({
                            "contact_name": current["contact_name"],
                            "channel_name": current["channel_name"],
                        }, ensure_ascii=False)))
            db.execute("""UPDATE crm_conversations SET status='Resolvida',pipeline_stage='Resolvido',
                               assigned_user_id=COALESCE(assigned_user_id,?),assigned_at=COALESCE(assigned_at,datetime('now','localtime')),
                               resolved_by_user_id=?,resolved_at=datetime('now','localtime'),resolution_reason=?,scheduled_return_at=NULL,
                               unread_count=0,automation_state='completed',
                               updated_at=datetime('now','localtime') WHERE id=?""",
                              (self.authenticated_user["id"],self.authenticated_user["id"],outcome,conversation_id))
            self.crm_record_event(db, conversation_id, "conversation.resolved", {
                "category": category, "patient_type": patient_type, "is_recovery": bool(is_recovery),
                "outcome": outcome, "attendance_number": attendance_number,
            })
            achievements = self.crm_evaluate_goal_achievements(
                db, int(self.authenticated_user["id"]),
                int(resolution_cursor.lastrowid) if resolution_cursor.lastrowid else None,
            )
        self.send_json({"resolved":True,"id":conversation_id,"reason":outcome,
                        "category":category,"patient_type":patient_type,
                        "attendance_number":attendance_number,"achievements":achievements})

    def get_crm_resolution_reports(self, query: dict) -> None:
        if not self.require_crc_access():
            return
        if not self.require_crm_feature("management"):
            return
        legacy_outcome_aliases = {
            "Pediu para reagendar": "Retorno",
            "Data específica": "Retorno",
            "Não respondeu": "Novo Contato IA",
            "Não respondeu/atendeu": "Novo Contato IA",
            "Não quer agendar": "Novo Contato IA",
            "Está tratando com outro profissional": "Em tratamento externo",
            "Informação fornecida": "Outros",
            "Contato interno": "Outros",
        }
        outcome_members = {
            value: [value] + [old for old, new in legacy_outcome_aliases.items() if new == value]
            for value in {
                "Agendou", "Quer agendar", "Retorno", "Novo Contato IA",
                "Mudou de cidade", "Em tratamento externo", "Desqualificado", "Outros",
            }
        }
        period = str((query.get("period") or ["30d"])[0]).strip().lower()
        category = str((query.get("category") or [""])[0]).strip()
        outcome = str((query.get("outcome") or [""])[0]).strip()
        channel_id = str((query.get("channel_id") or [""])[0]).strip()
        agent_id = str((query.get("agent_id") or [""])[0]).strip()
        interest = str((query.get("interest") or [""])[0]).strip()
        origin = str((query.get("origin") or [""])[0]).strip()
        start = str((query.get("start") or [""])[0]).strip()
        end = str((query.get("end") or [""])[0]).strip()
        scope_sql, scope_params = self.crm_channel_id_scope_clause("r.channel_id")
        where, params = [scope_sql], list(scope_params)
        if period == "today":
            where.append("date(r.resolved_at)=CURRENT_DATE")
        elif period == "7d":
            where.append("date(r.resolved_at)>=CURRENT_DATE - INTERVAL '6 days'")
        elif period == "30d":
            where.append("date(r.resolved_at)>=CURRENT_DATE - INTERVAL '29 days'")
        elif period == "custom":
            if start:
                where.append("date(r.resolved_at)>=date(?)")
                params.append(start)
            if end:
                where.append("date(r.resolved_at)<=date(?)")
                params.append(end)
        if category:
            where.append("r.category=?")
            params.append(category)
        if outcome:
            accepted = outcome_members.get(outcome, [outcome])
            where.append("r.outcome IN (" + ",".join("?" for _ in accepted) + ")")
            params.extend(accepted)
        if channel_id.isdigit():
            where.append("r.channel_id=?")
            params.append(int(channel_id))
        if agent_id.isdigit():
            where.append("r.resolved_by_user_id=?")
            params.append(int(agent_id))
        if interest:
            where.append("r.interest=?")
            params.append(interest)
        if origin:
            where.append("r.origin=?")
            params.append(origin)
        clause = (" WHERE " + " AND ".join(where)) if where else ""
        with connect() as db:
            rows = db.execute(f"""SELECT r.*,ct.name AS contact_name,ct.phone,
                                  ch.display_name AS channel_name
                                  FROM crm_service_resolutions r
                                  LEFT JOIN crm_contacts ct ON ct.id=r.contact_id
                                  LEFT JOIN crm_channels ch ON ch.id=r.channel_id
                                  {clause} ORDER BY r.resolved_at DESC LIMIT 500""", params).fetchall()
            report_patient_type_sql = "COALESCE(NULLIF(TRIM(COALESCE(r.patient_type,'')),''),CASE WHEN r.category='Primeira consulta' THEN 'Primeira consulta' ELSE '' END)"
            conversion_totals = db.execute(f"""SELECT
                    COUNT(*) FILTER (WHERE {report_patient_type_sql}='Primeira consulta') AS first_total,
                    COUNT(*) FILTER (WHERE {report_patient_type_sql}='Primeira consulta' AND r.outcome='Agendou') AS first_converted,
                    COUNT(*) FILTER (WHERE {report_patient_type_sql}='Retorno s/ Tratamento') AS recurring_total,
                    COUNT(*) FILTER (WHERE {report_patient_type_sql}='Retorno s/ Tratamento' AND r.outcome='Agendou') AS recurring_converted
                  FROM crm_service_resolutions r {clause}""", params).fetchone()
            data = [dict(row) for row in rows]
            for row in data:
                row["outcome"] = legacy_outcome_aliases.get(row.get("outcome"), row.get("outcome"))
            total = len(data)
            count = lambda field, value: sum(1 for row in data if row.get(field) == value)
            scheduled = count("outcome", "Agendou")
            first_total = int(conversion_totals["first_total"] or 0)
            first_converted = int(conversion_totals["first_converted"] or 0)
            recurring_total = int(conversion_totals["recurring_total"] or 0)
            recurring_converted = int(conversion_totals["recurring_converted"] or 0)
            summary = {
                "total": total,
                "first_consultations": count("category", "Primeira consulta"),
                "controls": count("category", "Controle"),
                "treatments": count("category", "Tratamento"),
                "budgets": count("category", "Orçamento"),
                "scheduled": scheduled,
                "wants_schedule": sum(1 for row in data if row.get("outcome") in {
                    "Quer agendar", "Retorno"
                }),
                "no_response": sum(1 for row in data if row.get("outcome") in {
                    "Novo Contato IA"
                }),
                "losses": sum(1 for row in data if row.get("outcome") in {
                    "Mudou de cidade", "Em tratamento externo", "Desqualificado"
                }),
                "conversion_rate": round((scheduled / total * 100), 1) if total else 0,
                "first_consultation_conversion_rate": round(first_converted / first_total * 100, 1) if first_total else 0,
                "first_consultation_converted": first_converted,
                "first_consultation_opportunities": first_total,
                "recurring_conversion_rate": round(recurring_converted / recurring_total * 100, 1) if recurring_total else 0,
                "recurring_converted": recurring_converted,
                "recurring_opportunities": recurring_total,
                "ai_involved": sum(1 for row in data if row.get("ai_involved")),
                "human_finalized": sum(1 for row in data if row.get("final_actor") == "Humano"),
            }
            def grouped(field):
                values = {}
                for row in data:
                    key = str(row.get(field) or "Não informado")
                    values[key] = values.get(key, 0) + 1
                return [{"label": key, "total": value} for key, value in
                        sorted(values.items(), key=lambda item: (-item[1], item[0]))]
            agents = db.execute("""SELECT id,name FROM users
                                   WHERE access_role='crc' AND active=1 AND COALESCE(crm_operational_agent,1)=1 ORDER BY name""").fetchall()
            channel_scope_sql, channel_scope_params = self.crm_channel_scope_clause("ch")
            channels = db.execute(f"""SELECT ch.id,ch.display_name FROM crm_channels ch
                                     WHERE {channel_scope_sql} ORDER BY ch.display_name""", channel_scope_params).fetchall()
        self.send_json({
            "summary": summary,
            "by_category": grouped("category"),
            "by_outcome": grouped("outcome"),
            "by_agent": grouped("resolved_by_name"),
            "by_channel": grouped("channel_name"),
            "by_professional": grouped("responsible_professional"),
            "by_interest": grouped("interest"),
            "by_origin": grouped("origin"),
            "rows": data,
            "filters": {
                "agents": [dict(row) for row in agents],
                "channels": [dict(row) for row in channels],
            },
        })

    def get_crm_patient_control(self, query: dict) -> None:
        """Read-only, filterable ledger of CRM attendances already finalized."""
        if not self.require_crc_access():
            return
        if not self.require_crm_feature("management"):
            return
        aliases = {
            "Pediu para reagendar": "Retorno",
            "Data específica": "Retorno",
            "Data especÃ­fica": "Retorno",
            "Não respondeu": "Novo Contato IA",
            "NÃ£o respondeu": "Novo Contato IA",
            "Não respondeu/atendeu": "Novo Contato IA",
            "NÃ£o respondeu/atendeu": "Novo Contato IA",
            "Não quer agendar": "Novo Contato IA",
            "NÃ£o quer agendar": "Novo Contato IA",
            "Está tratando com outro profissional": "Em tratamento externo",
            "EstÃ¡ tratando com outro profissional": "Em tratamento externo",
            "Informação fornecida": "Outros",
            "InformaÃ§Ã£o fornecida": "Outros",
            "Contato interno": "Outros",
        }
        period = str((query.get("period") or ["30d"])[0]).strip().lower()
        search = str((query.get("search") or [""])[0]).strip()
        category = str((query.get("category") or [""])[0]).strip()
        outcome = str((query.get("outcome") or [""])[0]).strip()
        interest = str((query.get("interest") or [""])[0]).strip()
        origin = str((query.get("origin") or [""])[0]).strip()
        professional = str((query.get("professional") or [""])[0]).strip()
        actor = str((query.get("actor") or [""])[0]).strip()
        scheduled = str((query.get("scheduled") or [""])[0]).strip().lower()
        channel_id = str((query.get("channel_id") or [""])[0]).strip()
        agent_id = str((query.get("agent_id") or [""])[0]).strip()
        start = str((query.get("start") or [""])[0]).strip()
        end = str((query.get("end") or [""])[0]).strip()
        try:
            page = max(1, int((query.get("page") or ["1"])[0]))
            per_page = min(500, max(10, int((query.get("per_page") or ["50"])[0])))
        except (TypeError, ValueError):
            page, per_page = 1, 50
        scope_sql, scope_params = self.crm_channel_id_scope_clause("r.channel_id")
        where, params = [scope_sql], list(scope_params)
        if period == "today":
            where.append("date(r.resolved_at)=CURRENT_DATE")
        elif period == "7d":
            where.append("date(r.resolved_at)>=CURRENT_DATE - INTERVAL '6 days'")
        elif period == "30d":
            where.append("date(r.resolved_at)>=CURRENT_DATE - INTERVAL '29 days'")
        elif period == "custom":
            if start:
                where.append("date(r.resolved_at)>=date(?)")
                params.append(start)
            if end:
                where.append("date(r.resolved_at)<=date(?)")
                params.append(end)
        if search:
            where.append("(LOWER(COALESCE(ct.name,'')) LIKE LOWER(?) OR REGEXP_REPLACE(COALESCE(ct.phone,''),'[^0-9]','','g') LIKE ?)")
            params.extend([f"%{search}%", f"%{''.join(ch for ch in search if ch.isdigit())}%"])
        for column, value in (
            ("r.category", category), ("r.interest", interest), ("r.origin", origin),
            ("r.responsible_professional", professional), ("r.final_actor", actor),
        ):
            if value:
                where.append(f"{column}=?")
                params.append(value)
        if outcome:
            accepted = [outcome] + [old for old, new in aliases.items() if new == outcome]
            where.append("r.outcome IN (" + ",".join("?" for _ in accepted) + ")")
            params.extend(accepted)
        if scheduled == "yes":
            where.append("NULLIF(TRIM(COALESCE(r.scheduled_date,'')),'') IS NOT NULL")
        elif scheduled == "no":
            where.append("NULLIF(TRIM(COALESCE(r.scheduled_date,'')),'') IS NULL")
        if channel_id.isdigit():
            where.append("r.channel_id=?")
            params.append(int(channel_id))
        if agent_id.isdigit():
            where.append("r.resolved_by_user_id=?")
            params.append(int(agent_id))
        clause = (" WHERE " + " AND ".join(where)) if where else ""
        offset = (page - 1) * per_page
        with connect() as db:
            total = int(db.execute(
                f"""SELECT COUNT(*) FROM crm_service_resolutions r
                    LEFT JOIN crm_contacts ct ON ct.id=r.contact_id {clause}""", params
            ).fetchone()[0])
            rows = db.execute(f"""SELECT r.*,ct.name AS contact_name,ct.phone,
                                  ch.display_name AS channel_name
                                  FROM crm_service_resolutions r
                                  LEFT JOIN crm_contacts ct ON ct.id=r.contact_id
                                  LEFT JOIN crm_channels ch ON ch.id=r.channel_id
                                  {clause}
                                  ORDER BY r.resolved_at DESC,r.id DESC
                                  LIMIT ? OFFSET ?""", [*params, per_page, offset]).fetchall()
            filtered = db.execute(f"""SELECT
                    COUNT(*) AS total,
                    COUNT(*) FILTER (WHERE NULLIF(TRIM(COALESCE(r.scheduled_date,'')),'') IS NOT NULL) AS scheduled,
                    COUNT(*) FILTER (WHERE r.ai_involved<>0) AS ai_involved,
                    COUNT(*) FILTER (WHERE r.final_actor='Humano') AS human_finalized
                FROM crm_service_resolutions r
                LEFT JOIN crm_contacts ct ON ct.id=r.contact_id {clause}""", params).fetchone()
            agents = db.execute(f"""SELECT DISTINCT r.resolved_by_user_id AS id,r.resolved_by_name AS name
                                   FROM crm_service_resolutions r WHERE {scope_sql}
                                   ORDER BY r.resolved_by_name""", scope_params).fetchall()
            channel_scope_sql, channel_scope_params = self.crm_channel_scope_clause("ch")
            channels = db.execute(f"""SELECT DISTINCT ch.id,ch.display_name
                                     FROM crm_service_resolutions r JOIN crm_channels ch ON ch.id=r.channel_id
                                     WHERE {channel_scope_sql} ORDER BY ch.display_name""", channel_scope_params).fetchall()
            dimensions = {}
            for key, column in (
                ("categories", "category"), ("outcomes", "outcome"), ("interests", "interest"),
                ("origins", "origin"), ("professionals", "responsible_professional"),
            ):
                values = db.execute(
                    f"""SELECT DISTINCT r.{column} AS value FROM crm_service_resolutions r
                        WHERE {scope_sql} AND NULLIF(TRIM(COALESCE(r.{column},'')),'') IS NOT NULL
                        ORDER BY r.{column}""", scope_params
                ).fetchall()
                dimensions[key] = [aliases.get(row["value"], row["value"]) for row in values]
                dimensions[key] = sorted(set(dimensions[key]))
        data = [dict(row) for row in rows]
        for row in data:
            row["outcome"] = aliases.get(row.get("outcome"), row.get("outcome"))
        self.send_json({
            "rows": data,
            "pagination": {
                "page": page, "per_page": per_page, "total": total,
                "pages": max(1, (total + per_page - 1) // per_page),
            },
            "summary": {
                "total": int(filtered["total"] or 0),
                "scheduled": int(filtered["scheduled"] or 0),
                "ai_involved": int(filtered["ai_involved"] or 0),
                "human_finalized": int(filtered["human_finalized"] or 0),
            },
            "filters": {
                **dimensions,
                "agents": [dict(row) for row in agents],
                "channels": [dict(row) for row in channels],
            },
        })

    def update_crm_conversation(self, conversation_id: int, payload: dict) -> None:
        if not self.require_crc_access(): return
        if not self.require_crm_any_feature(CRM_WORKSPACE_FEATURES):
            return
        with connect() as db:
            current = db.execute("SELECT * FROM crm_conversations WHERE id=?", (conversation_id,)).fetchone()
            if not current: return self.send_json({"error":"Conversa não encontrada."},HTTPStatus.NOT_FOUND)
            db.execute("SELECT id FROM crm_contacts WHERE id=? FOR UPDATE", (current["contact_id"],)).fetchone()
            if not self.crm_channel_allowed(db, current["channel_id"]):
                return self.send_json({"error":"Você não possui acesso a este canal."},HTTPStatus.FORBIDDEN)
            priority = str(payload.get("priority", current["priority"]) or "Normal").strip().title()
            queue = str(payload.get("queue_name", current["queue_name"]) or "Entrada").strip()
            stage = str(payload.get("pipeline_stage", current["pipeline_stage"]) or "Novo").strip()
            note = str(payload.get("internal_note", current["internal_note"]) or "").strip()
            if priority not in {"Baixa","Normal","Alta"}: return self.send_json({"error":"Prioridade inválida."},HTTPStatus.BAD_REQUEST)
            if stage not in {"Novo","Em atendimento","Aguardando cliente","Resolvido"}:
                return self.send_json({"error":"Etapa inválida."}, HTTPStatus.BAD_REQUEST)
            assigned_user_id = current["assigned_user_id"]
            transfer_contact_ownership = False
            if "assigned_user_id" in payload:
                assigned_user_id = self.authenticated_user["id"] if payload["assigned_user_id"] == "me" else payload["assigned_user_id"]
                if current["assigned_user_id"] and current["assigned_user_id"] != self.authenticated_user["id"]:
                    owner = db.execute("SELECT name FROM users WHERE id=?", (current["assigned_user_id"],)).fetchone()
                    return self.send_json(
                        {"error": f"Este atendimento pertence a {owner['name'] if owner else 'outra atendente'}. Somente ela pode transferi-lo."},
                        HTTPStatus.CONFLICT,
                    )
                if assigned_user_id is not None:
                    assignee = db.execute("SELECT id FROM users WHERE id=? AND access_role='crc' AND active=1", (assigned_user_id,)).fetchone()
                    if not assignee: return self.send_json({"error":"Atendente inválida ou inativa."},HTTPStatus.BAD_REQUEST)
                    if not self.crm_channel_allowed(db, current["channel_id"], None, assigned_user_id):
                        return self.send_json({"error":"A atendente escolhida não possui acesso a este canal."},HTTPStatus.CONFLICT)
                    active_owner = self.crm_contact_active_owner(db, current["contact_id"])
                    if active_owner and active_owner["assigned_user_id"] != assigned_user_id:
                        if current["assigned_user_id"] == self.authenticated_user["id"] and assigned_user_id != self.authenticated_user["id"]:
                            transfer_contact_ownership = True
                        else:
                            return self.reject_crm_contact_owner_conflict(active_owner)
                    if stage == current["pipeline_stage"] and stage == "Novo": stage = "Em atendimento"
            scheduled_return_at = current["scheduled_return_at"]
            if "scheduled_return_at" in payload:
                raw_return = str(payload.get("scheduled_return_at") or "").strip()
                if raw_return:
                    try:
                        parsed_return = datetime.fromisoformat(raw_return.replace("T", " "))
                    except ValueError:
                        return self.send_json({"error": "Informe data e horário válidos para o retorno."}, HTTPStatus.BAD_REQUEST)
                    if parsed_return <= datetime.now():
                        return self.send_json({"error": "O retorno programado deve estar no futuro."}, HTTPStatus.BAD_REQUEST)
                    scheduled_return_at = parsed_return.strftime("%Y-%m-%d %H:%M:%S")
                    stage, status, assigned_user_id = "Aguardando cliente", "Resolvida", None
                else:
                    scheduled_return_at = None
            status = "Resolvida" if stage == "Resolvido" else str(payload.get("status", current["status"]) or "Aberta")
            if scheduled_return_at:
                status = "Resolvida"
            resolved_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S") if status == "Resolvida" else None
            db.execute("""UPDATE crm_conversations SET priority=?,queue_name=?,pipeline_stage=?,internal_note=?,
                          assigned_user_id=?,assigned_at=CASE WHEN ? IS NOT NULL AND (? IS NULL OR ?<>?) THEN datetime('now','localtime') ELSE assigned_at END,
                          automation_state=CASE WHEN ? IS NOT NULL THEN 'paused' ELSE automation_state END,
                          status=?,resolved_at=?,scheduled_return_at=?,resolved_by_user_id=CASE WHEN ?='Resolvida' THEN ? ELSE resolved_by_user_id END,
                          updated_at=datetime('now','localtime') WHERE id=?""",
                       (priority, queue, stage, note, assigned_user_id, assigned_user_id,
                        current["assigned_user_id"], current["assigned_user_id"], assigned_user_id,
                        assigned_user_id,
                        status, resolved_at, scheduled_return_at, status, self.authenticated_user["id"], conversation_id))
            if scheduled_return_at and scheduled_return_at != current["scheduled_return_at"]:
                self.crm_record_event(db, conversation_id, "return.scheduled", {"scheduled_return_at": scheduled_return_at})
            if assigned_user_id != current["assigned_user_id"]:
                event_type = "conversation.assigned" if current["assigned_user_id"] is None else "conversation.transferred"
                self.crm_record_event(db, conversation_id, event_type, {"from_user_id": current["assigned_user_id"], "to_user_id": assigned_user_id})
            if transfer_contact_ownership:
                siblings = db.execute("""SELECT id,assigned_user_id FROM crm_conversations
                                          WHERE contact_id=? AND id<>? AND status<>'Resolvida'""",
                                      (current["contact_id"], conversation_id)).fetchall()
                db.execute("""UPDATE crm_conversations SET assigned_user_id=?,assigned_at=datetime('now','localtime'),
                              pipeline_stage=CASE WHEN pipeline_stage='Novo' THEN 'Em atendimento' ELSE pipeline_stage END,
                              automation_state='paused',updated_at=datetime('now','localtime')
                              WHERE contact_id=? AND id<>? AND status<>'Resolvida'""",
                           (assigned_user_id, current["contact_id"], conversation_id))
                for sibling in siblings:
                    if sibling["assigned_user_id"] != assigned_user_id:
                        self.crm_record_event(db, sibling["id"], "conversation.transferred", {
                            "from_user_id": sibling["assigned_user_id"], "to_user_id": assigned_user_id,
                            "source": "patient_ownership_transfer",
                        })
            if isinstance(payload.get("tag_names"), list):
                names = list(dict.fromkeys(str(name).strip()[:40] for name in payload["tag_names"] if str(name).strip()))
                db.execute("DELETE FROM crm_conversation_tags WHERE conversation_id=?", (conversation_id,))
                for name in names:
                    db.execute("INSERT OR IGNORE INTO crm_tags(name,color) VALUES(?, '#8696a0')", (name,))
                    tag_id = db.execute("SELECT id FROM crm_tags WHERE name=? COLLATE NOCASE", (name,)).fetchone()[0]
                    db.execute("INSERT OR IGNORE INTO crm_conversation_tags(conversation_id,tag_id) VALUES(?,?)", (conversation_id, tag_id))
        self.send_json({"updated":True,"id":conversation_id})

    def claim_crm_conversation(self, conversation_id: int) -> None:
        """Atomically assign an open external conversation to the current CRC user."""
        if not self.require_crc_access():
            return
        if not self.require_crm_feature("inbox"):
            return
        with connect() as db:
            current = db.execute("""SELECT cv.channel_id,cv.contact_id,cv.assigned_user_id,cv.status,ct.is_internal,u.name AS assigned_to
                                   FROM crm_conversations cv
                                   JOIN crm_contacts ct ON ct.id=cv.contact_id
                                   LEFT JOIN users u ON u.id=cv.assigned_user_id
                                   WHERE cv.id=?""", (conversation_id,)).fetchone()
            if not current:
                return self.send_json({"error": "Conversa não encontrada."}, HTTPStatus.NOT_FOUND)
            db.execute("SELECT id FROM crm_contacts WHERE id=? FOR UPDATE", (current["contact_id"],)).fetchone()
            if not self.crm_channel_allowed(db, current["channel_id"], "reply"):
                return self.send_json({"error": "Você não possui permissão para atender por este canal."}, HTTPStatus.FORBIDDEN)
            if current["is_internal"]:
                return self.send_json({"claimed": False, "internal": True, "id": conversation_id})
            if current["status"] != "Aberta":
                return self.send_json({"error": "Somente atendimentos abertos podem ser assumidos."}, HTTPStatus.CONFLICT)
            active_owner = self.crm_contact_active_owner(db, current["contact_id"])
            if active_owner and active_owner["assigned_user_id"] != self.authenticated_user["id"]:
                return self.reject_crm_contact_owner_conflict(active_owner)
            if current["assigned_user_id"] == self.authenticated_user["id"]:
                return self.send_json({"claimed": True, "already_owned": True, "id": conversation_id})
            if current["assigned_user_id"] is not None:
                return self.send_json(
                    {"error": f"Este atendimento já está atribuído a {current['assigned_to'] or 'outra atendente'}."},
                    HTTPStatus.CONFLICT,
                )

            result = db.execute("""UPDATE crm_conversations
                                      SET assigned_user_id=?,
                                          assigned_at=COALESCE(assigned_at,datetime('now','localtime')),
                                          pipeline_stage=CASE WHEN pipeline_stage='Novo' THEN 'Em atendimento' ELSE pipeline_stage END,
                                          automation_state='paused',
                                          updated_at=datetime('now','localtime')
                                    WHERE id=? AND assigned_user_id IS NULL AND status='Aberta'""",
                                (self.authenticated_user["id"], conversation_id))
            if result.rowcount != 1:
                owner = db.execute("""SELECT u.name FROM crm_conversations cv
                                      LEFT JOIN users u ON u.id=cv.assigned_user_id WHERE cv.id=?""",
                                   (conversation_id,)).fetchone()
                return self.send_json(
                    {"error": f"Este atendimento acabou de ser assumido por {(owner['name'] if owner else None) or 'outra atendente'}."},
                    HTTPStatus.CONFLICT,
                )
            self.crm_record_event(
                db,
                conversation_id,
                "conversation.assigned",
                {"from_user_id": None, "to_user_id": self.authenticated_user["id"], "source": "conversation.opened"},
            )
        self.send_json({"claimed": True, "id": conversation_id})

    def get_crm_webhook_events(self) -> None:
        if not self.require_crm_feature("integrations"): return
        if not self.require_crc_access(): return
        scope_sql, scope_params = self.crm_channel_scope_clause("ch")
        with connect() as db:
            rows=db.execute(f"""SELECT ev.id,ev.instance_name,ev.event_type,ev.processing_status,ev.error_message,
                datetime(ev.received_at,'localtime') AS received_at,
                CASE WHEN ev.processed_at IS NULL THEN NULL ELSE datetime(ev.processed_at,'localtime') END AS processed_at
                FROM crm_webhook_events ev JOIN crm_channels ch ON ch.instance_name=ev.instance_name
                WHERE {scope_sql} ORDER BY ev.id DESC LIMIT 100""",scope_params).fetchall()
        self.send_json({"items":[dict(row) for row in rows]})

    def get_crm_automation_events(self) -> None:
        if not self.require_crm_feature("integrations"): return
        if not self.require_crc_access(): return
        scope_sql, scope_params = self.crm_channel_scope_clause("ch")
        with connect() as db:
            rows = db.execute(f"""SELECT ae.id,ae.event_key,ae.flow_name,ae.event_type,ae.outcome,
                datetime(ae.received_at,'localtime') AS received_at,
                ct.name AS contact_name,ch.display_name AS channel_name
                FROM crm_automation_events ae
                LEFT JOIN crm_conversations cv ON cv.id=ae.conversation_id
                LEFT JOIN crm_contacts ct ON ct.id=cv.contact_id
                LEFT JOIN crm_channels ch ON ch.id=cv.channel_id
                WHERE ae.conversation_id IS NULL OR ({scope_sql})
                ORDER BY ae.id DESC LIMIT 100""", scope_params).fetchall()
        self.send_json({"items": [dict(row) for row in rows]})

    def get_patients(self, query: dict) -> None:
        search = query.get("search", [""])[0].strip()
        status = query.get("status", [""])[0].strip()
        attention = query.get("attention", [""])[0].strip()
        crc_stage = query.get("crc_stage", [""])[0].strip()
        month = query.get("month", [""])[0].strip()
        professional = query.get("professional", [""])[0].strip()
        next_schedule = query.get("next_schedule", [""])[0].strip()
        sort = query.get("sort", ["last_visit_asc"])[0].strip()
        page = max(1, int(query.get("page", ["1"])[0] or 1))
        per_page = min(100, max(10, int(query.get("per_page", ["15"])[0] or 15)))
        clauses, params = [], []
        if self.authenticated_user["access_role"] not in {"admin", "crc"}:
            clauses.append("EXISTS (SELECT 1 FROM patient_assignments own_pa WHERE own_pa.patient_id = p.id AND own_pa.professional_id = ? AND own_pa.journey_status = 'Ativo')")
            portfolio_id = (
                self.authenticated_user.get("linked_professional_id")
                if self.authenticated_user["access_role"] == "asb"
                else self.authenticated_user["professional_id"]
            )
            params.append(portfolio_id)
        elif self.authenticated_user["access_role"] == "crc":
            clauses.append("f.crc_status IN ('Aguardando contato', 'Em atendimento', 'Jornada compartilhada')")
            if crc_stage in {"Aguardando contato", "Em atendimento", "Jornada compartilhada"}:
                clauses.append("f.crc_status = ?")
                params.append(crc_stage)
        if professional and self.authenticated_user["access_role"] in {"admin", "crc"}:
            professional_ids = [item for item in professional.split(",") if item.isdigit()]
            if professional_ids:
                clauses.append(f"pa.professional_id IN ({','.join('?' for _ in professional_ids)})")
                params.extend(professional_ids)
        if search:
            # LIKE puro é case-insensitive no SQLite (comportamento herdado do
            # dev original), mas não no PostgreSQL — sem lower() dos dois
            # lados, uma busca por "maria" deixa de encontrar "Maria" (bug
            # confirmado em testes).
            clauses.append("(lower(p.name) LIKE ? OR lower(COALESCE(f.procedure_name, '')) LIKE ? OR EXISTS (SELECT 1 FROM procedures ps WHERE ps.patient_id = p.id AND lower(ps.name) LIKE ?) OR EXISTS (SELECT 1 FROM patient_relationships rl WHERE rl.patient_id = p.id AND lower(rl.related_name) LIKE ?))")
            search_lower = search.lower()
            params.extend([f"%{search_lower}%", f"%{search_lower}%", f"%{search_lower}%", f"%{search_lower}%"])
        if status == "resolved":
            clauses.append("date(f.resolved_at) = date('now', 'localtime')")
        elif status:
            clauses.append("p.status = ?")
            params.append(status)
        if month:
            clauses.append("substr(f.last_visit, 1, 7) = ?")
            params.append(month)
        if next_schedule == "scheduled":
            clauses.append("f.next_appointment IS NOT NULL AND f.next_appointment_type = 'Agendado' AND date(f.next_appointment) >= date('now', 'localtime')")
        elif next_schedule == "programmed":
            clauses.append("f.next_appointment IS NOT NULL AND f.next_appointment_type = 'Programado'")
        elif next_schedule == "unscheduled":
            clauses.append("f.next_appointment IS NULL OR date(f.next_appointment) < date('now', 'localtime')")
        elif next_schedule == "today":
            clauses.append("date(f.next_appointment) = date('now', 'localtime')")
        elif next_schedule == "next_7":
            clauses.append("date(f.next_appointment) BETWEEN date('now', 'localtime') AND date('now', 'localtime', '+7 days')")
        elif next_schedule == "next_30":
            clauses.append("date(f.next_appointment) BETWEEN date('now', 'localtime') AND date('now', 'localtime', '+30 days')")
        elif next_schedule == "overdue":
            clauses.append("date(f.next_appointment) < date('now', 'localtime')")
        if attention == "contact":
            contact_clause = "p.status != 'Inativo' AND (f.next_appointment IS NULL OR date(f.next_appointment) < date('now', 'localtime')) AND julianday('now', 'localtime') - julianday(CASE WHEN f.next_appointment IS NOT NULL AND date(f.next_appointment) < date('now', 'localtime') THEN f.next_appointment ELSE f.last_visit END) >= 60"
            if self.authenticated_user["access_role"] != "crc":
                contact_clause += " AND (f.resolved_at IS NULL OR date(f.resolved_at) != date('now', 'localtime'))"
            clauses.append(contact_clause)
        elif attention == "unscheduled":
            clauses.append("p.status != 'Inativo' AND (f.next_appointment IS NULL OR date(f.next_appointment) < date('now', 'localtime'))")
        elif attention == "treatment":
            clauses.append("EXISTS (SELECT 1 FROM procedures px WHERE px.patient_id = p.id AND px.stage != 'Concluído')")
        elif attention == "inactive":
            clauses.append("p.status = 'Inativo'")
        elif attention == "resolved":
            clauses.append("date(f.resolved_at) = date('now', 'localtime')")
        where = f" WHERE {' AND '.join(clauses)}" if clauses else ""
        sort_options = {
            "last_visit_asc": "f.last_visit ASC, p.name ASC",
            "last_visit_desc": "f.last_visit DESC, p.name ASC",
            "name_asc": "p.name COLLATE NOCASE ASC",
            "name_desc": "p.name COLLATE NOCASE DESC",
            "value_desc": "potential_value_cents DESC, p.name ASC",
            "value_asc": "potential_value_cents ASC, p.name ASC",
            "next_asc": "f.next_appointment IS NULL, f.next_appointment ASC, p.name ASC",
            "next_desc": "f.next_appointment IS NULL, f.next_appointment DESC, p.name ASC",
        }
        order_by = sort_options.get(sort, sort_options["last_visit_asc"])
        with connect() as db:
            total = db.execute(
                "SELECT COUNT(*) FROM patients p JOIN patient_followup f ON f.patient_id = p.id JOIN patient_assignments pa ON pa.patient_id=p.id AND pa.is_primary=1" + where,
                params,
            ).fetchone()[0]
            rows = db.execute(
                patient_select() + where + f" ORDER BY {order_by} LIMIT ? OFFSET ?",
                [*params, per_page, (page - 1) * per_page],
            ).fetchall()
        self.send_json({"items": [dict(row) for row in rows], "total": total, "page": page, "per_page": per_page})

    def get_patient(self, patient_id: int) -> None:
        if not self.require_patient_access(patient_id, self.authenticated_user):
            return
        with connect() as db:
            row = db.execute(patient_select() + " WHERE p.id = ?", (patient_id,)).fetchone()
            if not row:
                return self.send_json({"error": "Paciente não encontrado"}, HTTPStatus.NOT_FOUND)
            history = db.execute(
                "SELECT event_type, description, created_at FROM patient_events WHERE patient_id = ? ORDER BY created_at DESC, id DESC LIMIT 20",
                (patient_id,),
            ).fetchall()
            procedures = db.execute(
                "SELECT id, name, value_cents, discount_cents, stage, notes FROM procedures WHERE patient_id = ? ORDER BY CASE stage WHEN 'Concluído' THEN 1 ELSE 0 END, id",
                (patient_id,),
            ).fetchall()
            clinical = db.execute(
                "SELECT particularities, health_change, health_condition, health_care FROM patient_clinical_profile WHERE patient_id = ?",
                (patient_id,),
            ).fetchone()
            relationships = db.execute(
                "SELECT id, relationship_type, related_name, connection FROM patient_relationships WHERE patient_id = ? ORDER BY id",
                (patient_id,),
            ).fetchall()
            observations = db.execute(
                "SELECT id, note, created_at, updated_at, author_name FROM patient_visit_notes WHERE patient_id = ? ORDER BY datetime(created_at) DESC, id DESC",
                (patient_id,),
            ).fetchall()
            journey = db.execute("""
                SELECT pa.professional_id, pr.name, pa.is_primary, pa.journey_status, pa.assigned_at,
                       pa.origin_professional_id,
                       pa.forward_reason, pa.stage_status, pa.stage_note, pa.stage_updated_at,
                       CAST(julianday('now', 'localtime') - julianday(pa.assigned_at) AS INTEGER) AS days_shared,
                       origin.name AS forwarded_by
                FROM patient_assignments pa
                JOIN professionals pr ON pr.id = pa.professional_id
                LEFT JOIN professionals origin ON origin.id = pa.origin_professional_id
                WHERE pa.patient_id = ? AND (pa.is_primary=1 OR pa.journey_status='Ativo')
                ORDER BY pa.is_primary DESC, pa.assigned_at ASC
            """, (patient_id,)).fetchall()
        payload = dict(row)
        payload["history"] = [dict(item) for item in history]
        payload["procedures"] = [dict(item) for item in procedures]
        payload.update(dict(clinical) if clinical else {"particularities": None, "health_change": 0, "health_condition": None, "health_care": None})
        payload["relationships"] = [dict(item) for item in relationships]
        payload["observations"] = [dict(item) for item in observations]
        payload["journey"] = [dict(item) for item in journey]
        self.send_json(payload)

    def get_journey_professionals(self) -> None:
        source_professional_id = self.portfolio_professional_id(self.authenticated_user) or -1
        with connect() as db:
            rows = db.execute("""
                SELECT DISTINCT pr.id, pr.name,
                       (SELECT STRING_AGG(s.name, ', ' ORDER BY s.name) FROM professional_specialties ps JOIN specialties s ON s.id=ps.specialty_id WHERE ps.professional_id=pr.id AND s.active=1) AS specialties,
                       CASE WHEN pr.photo_data IS NOT NULL THEN 1 ELSE 0 END AS has_photo
                FROM professionals pr
                JOIN users u ON u.professional_id = pr.id
                WHERE pr.active=1 AND u.active=1 AND u.access_role IN ('owner', 'professional')
                  AND pr.id != ?
                 ORDER BY pr.name
            """, (source_professional_id,)).fetchall()
        items = []
        for row in rows:
            item = dict(row)
            item["photo_url"] = f"/api/professionals/{item['id']}/photo" if item.pop("has_photo") else None
            items.append(item)
        self.send_json({"items": items})

    def get_journeys(self) -> None:
        if self.authenticated_user["access_role"] in {"admin", "crc"}:
            return self.send_json({"items": []})
        source_professional_id = self.portfolio_professional_id(self.authenticated_user)
        with connect() as db:
            rows = db.execute("""
                SELECT DISTINCT p.id, p.name, p.status, f.last_visit, f.next_appointment,
                       source.name AS forwarded_by, target.name AS forwarded_to,
                       pa.forward_reason, pa.assigned_at
                FROM patient_assignments pa
                JOIN patients p ON p.id=pa.patient_id
                JOIN patient_followup f ON f.patient_id=p.id
                JOIN professionals target ON target.id=pa.professional_id
                LEFT JOIN professionals source ON source.id=pa.origin_professional_id
                WHERE pa.journey_status='Ativo' AND (
                    pa.professional_id=? OR pa.origin_professional_id=?
                ) AND (pa.is_primary=0 OR pa.origin_professional_id IS NOT NULL)
                ORDER BY datetime(pa.assigned_at) DESC, p.name COLLATE NOCASE
            """, (source_professional_id, source_professional_id)).fetchall()
        self.send_json({"items": [dict(row) for row in rows]})

    def forward_patient(self, patient_id: int, payload: dict) -> None:
        if self.authenticated_user["access_role"] not in {"owner", "professional", "asb"}:
            return self.send_json({"error": "Somente dentistas e ASBs vinculadas podem encaminhar pacientes."}, HTTPStatus.FORBIDDEN)
        if not self.require_patient_access(patient_id, self.authenticated_user):
            return
        source_professional_id = self.portfolio_professional_id(self.authenticated_user)
        if not source_professional_id:
            return self.send_json({"error": "Este acesso não possui uma carteira profissional vinculada."}, HTTPStatus.CONFLICT)
        targets = payload.get("professional_ids")
        if isinstance(targets, list):
            targets = list(dict.fromkeys(str(target).strip() for target in targets if str(target).strip()))
            if not targets:
                return self.send_json({"error": "Selecione ao menos um profissional."}, HTTPStatus.BAD_REQUEST)
            if len(targets) == 1:
                payload = {**payload, "professional_id": targets[0]}
            else:
                for target in targets:
                    self.forward_patient(patient_id, {"professional_id": target, "reason": payload.get("reason"), "_batch": True})
                return self.get_patient(patient_id)
        target_id = int(payload.get("professional_id") or 0)
        reason = str(payload.get("reason") or "").strip()
        if not target_id or target_id == source_professional_id:
            return self.send_json({"error": "Selecione outra profissional para o encaminhamento."}, HTTPStatus.BAD_REQUEST)
        with connect() as db:
            if self.patient_edit_lock_conflict(db, patient_id):
                return
            target = db.execute("""
                SELECT pr.id, pr.name FROM professionals pr JOIN users u ON u.professional_id=pr.id
                WHERE pr.id=? AND pr.active=1 AND u.active=1 AND u.access_role IN ('owner','professional')
            """, (target_id,)).fetchone()
            patient = db.execute("SELECT name FROM patients WHERE id=?", (patient_id,)).fetchone()
            if not target or not patient:
                return self.send_json({"error": "Profissional ou paciente não encontrado."}, HTTPStatus.NOT_FOUND)
            db.execute("""
                INSERT INTO patient_assignments (patient_id, professional_id, is_primary, journey_status, origin_professional_id, forward_reason, completed_at)
                VALUES (?, ?, 0, 'Ativo', ?, ?, NULL)
                ON CONFLICT(patient_id, professional_id) DO UPDATE SET
                    journey_status='Ativo', origin_professional_id=excluded.origin_professional_id,
                    forward_reason=excluded.forward_reason, assigned_at=CURRENT_TIMESTAMP, completed_at=NULL,
                    stage_status='Aguardando início', stage_note=NULL, stage_updated_at=NULL
            """, (patient_id, target_id, source_professional_id, reason or None))
            db.execute("UPDATE patient_followup SET crc_status='Jornada compartilhada', crc_started_at=NULL, crc_completed_at=NULL WHERE patient_id=?", (patient_id,))
            description = f"Encaminhado por {self.authenticated_user['name']} para {target['name']}"
            if reason:
                description += f": {reason}"
            db.execute("INSERT INTO patient_events (patient_id, event_type, description) VALUES (?, 'Encaminhamento', ?)", (patient_id, description))
        if not payload.get("_batch"):
            self.get_patient(patient_id)

    def save_journey_stage(self, patient_id: int, payload: dict) -> None:
        if self.authenticated_user["access_role"] in {"crc", "admin"}:
            return self.send_json({"error": "Somente o dentista responsável pela etapa pode registrar sua evolução."}, HTTPStatus.FORBIDDEN)
        if not self.require_patient_access(patient_id, self.authenticated_user):
            return
        professional_id = self.portfolio_professional_id(self.authenticated_user)
        stage_status = str(payload.get("stage_status") or "Em atendimento").strip()
        note = str(payload.get("note") or "").strip()
        if stage_status not in {"Em atendimento", "Concluído"}:
            return self.send_json({"error": "Etapa inválida."}, HTTPStatus.BAD_REQUEST)
        if not note:
            return self.send_json({"error": "Registre a evolução desta etapa."}, HTTPStatus.BAD_REQUEST)
        with connect() as db:
            if self.patient_edit_lock_conflict(db, patient_id):
                return
            assignment = db.execute("SELECT is_primary FROM patient_assignments WHERE patient_id=? AND professional_id=?", (patient_id, professional_id)).fetchone()
            if not assignment or assignment["is_primary"]:
                return self.send_json({"error": "Crie uma etapa somente após receber um encaminhamento."}, HTTPStatus.FORBIDDEN)
            db.execute("UPDATE patient_assignments SET stage_status=?, stage_note=?, stage_updated_at=datetime('now','localtime'), completed_at=CASE WHEN ?='Concluído' THEN datetime('now','localtime') ELSE NULL END WHERE patient_id=? AND professional_id=?", (stage_status, note, stage_status, patient_id, professional_id))
            self.save_visit_observation(db, patient_id, note, self.authenticated_user["name"])
            remaining = db.execute("SELECT COUNT(*) FROM patient_assignments WHERE patient_id=? AND is_primary=0 AND journey_status='Ativo' AND stage_status!='Concluído'", (patient_id,)).fetchone()[0]
            crc_status = "Aguardando contato" if remaining == 0 else "Jornada compartilhada"
            db.execute("UPDATE patient_followup SET crc_status=? WHERE patient_id=?", (crc_status, patient_id))
            db.execute("INSERT INTO patient_events (patient_id, event_type, description) VALUES (?, 'Etapa da jornada', ?)", (patient_id, f"{self.authenticated_user['name']}: {stage_status}"))
        self.get_patient(patient_id)

    def journey_assignment_owned_by_current_professional(self, db, patient_id: int, professional_id: int):
        current_professional_id = self.portfolio_professional_id(self.authenticated_user)
        if not current_professional_id or self.authenticated_user["access_role"] not in {"owner", "professional", "asb"}:
            return None
        return db.execute("""
            SELECT pa.*, target.name AS target_name
            FROM patient_assignments pa
            JOIN professionals target ON target.id=pa.professional_id
            WHERE pa.patient_id=? AND pa.professional_id=? AND pa.is_primary=0
              AND pa.origin_professional_id=? AND pa.journey_status='Ativo'
        """, (patient_id, professional_id, current_professional_id)).fetchone()

    def update_journey_assignment(self, patient_id: int, professional_id: int, payload: dict) -> None:
        if not self.require_patient_access(patient_id, self.authenticated_user):
            return
        reason = str(payload.get("reason") or "").strip()
        with connect() as db:
            if self.patient_edit_lock_conflict(db, patient_id):
                return
            assignment = self.journey_assignment_owned_by_current_professional(db, patient_id, professional_id)
            if not assignment:
                return self.send_json({"error": "Somente quem realizou o encaminhamento pode editá-lo."}, HTTPStatus.FORBIDDEN)
            db.execute("UPDATE patient_assignments SET forward_reason=? WHERE patient_id=? AND professional_id=?", (reason or None, patient_id, professional_id))
            description = f"Encaminhamento para {assignment['target_name']} editado por {self.authenticated_user['name']}"
            if reason:
                description += f": {reason}"
            db.execute("INSERT INTO patient_events (patient_id, event_type, description) VALUES (?, 'Encaminhamento editado', ?)", (patient_id, description))
        self.get_patient(patient_id)

    def delete_journey_assignment(self, patient_id: int, professional_id: int) -> None:
        if not self.require_patient_access(patient_id, self.authenticated_user):
            return
        with connect() as db:
            if self.patient_edit_lock_conflict(db, patient_id):
                return
            assignment = self.journey_assignment_owned_by_current_professional(db, patient_id, professional_id)
            if not assignment:
                return self.send_json({"error": "Somente quem realizou o encaminhamento pode excluí-lo."}, HTTPStatus.FORBIDDEN)
            db.execute("""
                UPDATE patient_assignments
                SET journey_status='Cancelado', completed_at=datetime('now','localtime'), stage_status='Cancelado'
                WHERE patient_id=? AND professional_id=?
            """, (patient_id, professional_id))
            remaining = db.execute("SELECT COUNT(*) FROM patient_assignments WHERE patient_id=? AND is_primary=0 AND journey_status='Ativo'", (patient_id,)).fetchone()[0]
            if remaining == 0:
                db.execute("UPDATE patient_followup SET crc_status='Aguardando contato' WHERE patient_id=? AND crc_status='Jornada compartilhada'", (patient_id,))
            db.execute("INSERT INTO patient_events (patient_id, event_type, description) VALUES (?, 'Encaminhamento excluído', ?)", (patient_id, f"Encaminhamento para {assignment['target_name']} excluído por {self.authenticated_user['name']}"))
        self.get_patient(patient_id)

    def resolve_patient(self, patient_id: int, payload: dict) -> None:
        if not self.require_patient_access(patient_id, self.authenticated_user):
            return
        resolved = bool(payload.get("resolved", True))
        if not resolved:
            return self.send_json({"error": "A liberação para edição é feita somente pela área Verificados."}, HTTPStatus.FORBIDDEN)
        with connect() as db:
            if self.patient_edit_lock_conflict(db, patient_id):
                return
            exists = db.execute("SELECT id FROM patients WHERE id = ?", (patient_id,)).fetchone()
            if not exists:
                return self.send_json({"error": "Paciente não encontrado"}, HTTPStatus.NOT_FOUND)
            db.execute(
                """UPDATE patient_followup
                   SET resolved_at = datetime('now', 'localtime'),
                       crc_status = 'Aguardando contato',
                       crc_started_at = NULL,
                       crc_completed_at = NULL
                   WHERE patient_id = ?""",
                (patient_id,),
            )
            if resolved:
                db.execute("""
                    INSERT INTO daily_resolutions (patient_id, resolution_date, resolved_at, reopened_at)
                    VALUES (?, date('now', 'localtime'), datetime('now', 'localtime'), NULL)
                    ON CONFLICT(patient_id, resolution_date)
                    DO UPDATE SET resolved_at = datetime('now', 'localtime'), reopened_at = NULL
                """, (patient_id,))
                self.queue_crc_export(db, patient_id)
            else:
                db.execute("""
                    UPDATE daily_resolutions
                    SET reopened_at = datetime('now', 'localtime')
                    WHERE patient_id = ? AND resolution_date = CAST(date('now', 'localtime') AS TEXT)
                """, (patient_id,))
            db.execute(
                "INSERT INTO patient_events (patient_id, event_type, description) VALUES (?, ?, ?)",
                (patient_id, "Resolvido" if resolved else "Reaberto", "Acompanhamento concluído no dia" if resolved else "Paciente retornou à fila de acompanhamento"),
            )
            db.execute(
                "DELETE FROM patient_edit_locks WHERE patient_id=? AND user_id=?",
                (patient_id, self.authenticated_user["id"]),
            )
        self.get_patient(patient_id)

    def integration_authorized(self) -> bool:
        supplied = self.headers.get("X-Integration-Key", "")
        if not INTEGRATION_TOKEN or not supplied or not hmac.compare_digest(supplied, INTEGRATION_TOKEN):
            self.send_json({"error": "Integração não autorizada"}, HTTPStatus.UNAUTHORIZED)
            return False
        return True

    @staticmethod
    def queue_crc_export(db, patient_id: int) -> None:
        snapshot = db.execute("""
            SELECT p.name, p.phone, f.last_visit, pr.name AS professional_name,
                   COALESCE((SELECT note FROM patient_visit_notes vn WHERE vn.patient_id=p.id ORDER BY datetime(vn.created_at) DESC, vn.id DESC LIMIT 1), '') AS observation_text
            FROM patients p
            JOIN patient_followup f ON f.patient_id=p.id
            JOIN patient_assignments pa ON pa.patient_id=p.id AND pa.is_primary=1
            JOIN professionals pr ON pr.id=pa.professional_id
            WHERE p.id=?
        """, (patient_id,)).fetchone()
        if not snapshot:
            return
        last_visit = snapshot["last_visit"] or "sem-data"
        export_key = f"crc-{patient_id}-{last_visit}"
        db.execute("""
            INSERT OR IGNORE INTO crc_export_queue
            (patient_id, export_key, patient_name, phone, last_visit, professional_name, observation_text, status)
            VALUES (?, ?, ?, ?, ?, ?, ?, 'Pendente')
        """, (patient_id, export_key, snapshot["name"], snapshot["phone"], snapshot["last_visit"], snapshot["professional_name"], snapshot["observation_text"]))

    def claim_crc_exports(self, payload: dict) -> None:
        if not self.integration_authorized():
            return
        limit = min(100, max(1, int(payload.get("limit") or 25)))
        claim_token = secrets.token_urlsafe(24)
        with connect() as db:
            # "BEGIN IMMEDIATE" é sintaxe exclusiva do SQLite (lock antecipado de
            # escrita); no PostgreSQL o equivalente para evitar que dois
            # workers reivindiquem o mesmo item é "FOR UPDATE SKIP LOCKED" na
            # seleção abaixo — mais seguro sob concorrência que o original.
            ids = [row[0] for row in db.execute("SELECT id FROM crc_export_queue WHERE status IN ('Pendente','Falhou') ORDER BY id LIMIT ? FOR UPDATE SKIP LOCKED", (limit,)).fetchall()]
            if ids:
                placeholders = ",".join("?" for _ in ids)
                db.execute(f"UPDATE crc_export_queue SET status='Em processamento', claim_token=?, claimed_at=datetime('now','localtime'), attempts=attempts+1, last_error=NULL WHERE id IN ({placeholders})", [claim_token, *ids])
            rows = db.execute("""SELECT export_key, patient_name AS "Nome", phone AS "telefone", last_visit AS "Data da ultima consulta", professional_name AS "Profissional", observation_text AS "Observação do Retorno", status AS "Enviado", message_created AS "Mensagem Criada", claim_token FROM crc_export_queue WHERE claim_token=? ORDER BY id""", (claim_token,)).fetchall()
        self.send_json({"items": [dict(row) for row in rows]})

    def ack_crc_export(self, payload: dict) -> None:
        if not self.integration_authorized():
            return
        export_key = str(payload.get("export_key") or "").strip()
        claim_token = str(payload.get("claim_token") or "").strip()
        success = bool(payload.get("success"))
        message_created = str(payload.get("message_created") or "").strip()
        if not export_key or not claim_token:
            return self.send_json({"error": "export_key e claim_token são obrigatórios"}, HTTPStatus.BAD_REQUEST)
        with connect() as db:
            row = db.execute("SELECT id FROM crc_export_queue WHERE export_key=? AND claim_token=? AND status='Em processamento'", (export_key, claim_token)).fetchone()
            if not row:
                return self.send_json({"error": "Item não encontrado ou já confirmado"}, HTTPStatus.CONFLICT)
            if success:
                db.execute("UPDATE crc_export_queue SET status='Exportado', exported_at=datetime('now','localtime'), message_created=?, message_created_at=CASE WHEN ?<>'' THEN datetime('now','localtime') ELSE NULL END WHERE id=?", (message_created, message_created, row[0]))
            else:
                db.execute("UPDATE crc_export_queue SET status='Falhou', last_error=?, claim_token=NULL WHERE id=?", (str(payload.get("error") or "Falha não informada")[:500], row[0]))
        self.send_json({"acknowledged": True, "export_key": export_key})

    def reopen_patient_with_password(self, patient_id: int, payload: dict) -> None:
        if self.authenticated_user["access_role"] not in {"owner", "professional"}:
            return self.send_json({"error": "Somente o dentista responsável pode liberar alterações."}, HTTPStatus.FORBIDDEN)
        if not self.require_patient_access(patient_id, self.authenticated_user):
            return
        password = str(payload.get("password") or "")
        with connect() as db:
            user = db.execute("SELECT password_hash, password_salt FROM users WHERE id = ?", (self.authenticated_user["id"],)).fetchone()
            password_valid = user and user["password_hash"] and user["password_salt"] and hmac.compare_digest(self.password_digest(password, user["password_salt"]), user["password_hash"])
            if not password_valid:
                self.record_security_event(db, "patient_reopen_denied", self.request_ip(), self.authenticated_user["id"], f"Paciente {patient_id}")
                return self.send_json({"error": "Senha inválida."}, HTTPStatus.UNAUTHORIZED)
            locked = db.execute("SELECT 1 FROM patient_followup WHERE patient_id = ? AND date(resolved_at) = date('now', 'localtime')", (patient_id,)).fetchone()
            if not locked:
                return self.send_json({"error": "Este paciente não está bloqueado na área Verificados de hoje."}, HTTPStatus.CONFLICT)
            db.execute("UPDATE patient_followup SET resolved_at = NULL WHERE patient_id = ?", (patient_id,))
            db.execute("UPDATE daily_resolutions SET reopened_at = datetime('now', 'localtime') WHERE patient_id = ? AND resolution_date = CAST(date('now', 'localtime') AS TEXT)", (patient_id,))
            db.execute("INSERT INTO patient_events (patient_id, event_type, description) VALUES (?, 'Liberado', 'Ficha liberada pelo dentista na área Verificados')", (patient_id,))
            self.record_security_event(db, "patient_reopened", self.request_ip(), self.authenticated_user["id"], f"Paciente {patient_id}")
        self.send_json({"released": True, "patient_id": patient_id})

    def add_visit_observation(self, patient_id: int, payload: dict) -> None:
        if not self.require_patient_access(patient_id, self.authenticated_user):
            return
        note = str(payload.get("note") or "").strip()
        if not note:
            return self.send_json({"error": "Escreva uma observação antes de salvar."}, HTTPStatus.BAD_REQUEST)
        with connect() as db:
            if self.patient_edit_lock_conflict(db, patient_id):
                return
            patient = db.execute("""
                SELECT p.id, p.name, CASE WHEN date(f.resolved_at) = date('now', 'localtime') THEN 1 ELSE 0 END AS locked
                FROM patients p JOIN patient_followup f ON f.patient_id = p.id WHERE p.id = ?
            """, (patient_id,)).fetchone()
            if not patient:
                return self.send_json({"error": "Paciente não encontrado"}, HTTPStatus.NOT_FOUND)
            if patient["locked"]:
                return self.send_json({"error": "Paciente verificado está bloqueado. Reabra o acompanhamento para editar."}, HTTPStatus.CONFLICT)
            self.save_visit_observation(db, patient_id, note, self.authenticated_user["name"])
            db.execute("INSERT INTO patient_events (patient_id, event_type, description) VALUES (?, 'Observação', 'Observação da última consulta registrada')", (patient_id,))
        self.get_patient(patient_id)

    def can_manage_visit_observation(self, note) -> bool:
        return str(note["author_name"] or "").strip() == str(self.authenticated_user.get("name") or "").strip()

    def update_visit_observation(self, patient_id: int, observation_id: int, payload: dict) -> None:
        if self.authenticated_user["access_role"] == "crc":
            return self.send_json({"error": "A CRC não pode alterar observações clínicas."}, HTTPStatus.FORBIDDEN)
        if not self.require_patient_access(patient_id, self.authenticated_user):
            return
        note_text = str(payload.get("note") or "").strip()
        if not note_text:
            return self.send_json({"error": "Escreva uma observação antes de salvar."}, HTTPStatus.BAD_REQUEST)
        with connect() as db:
            if self.patient_edit_lock_conflict(db, patient_id):
                return
            note = db.execute("SELECT id, patient_id, author_name FROM patient_visit_notes WHERE id=? AND patient_id=?", (observation_id, patient_id)).fetchone()
            if not note:
                return self.send_json({"error": "Observação não encontrada."}, HTTPStatus.NOT_FOUND)
            if not self.can_manage_visit_observation(note):
                return self.send_json({"error": "Somente quem registrou esta observação pode alterá-la."}, HTTPStatus.FORBIDDEN)
            locked = db.execute("SELECT 1 FROM patient_followup WHERE patient_id=? AND date(resolved_at)=date('now','localtime')", (patient_id,)).fetchone()
            if locked:
                return self.send_json({"error": "Paciente verificado está bloqueado. Reabra o acompanhamento para editar."}, HTTPStatus.CONFLICT)
            db.execute("UPDATE patient_visit_notes SET note=?, updated_at=datetime('now','localtime') WHERE id=?", (note_text, observation_id))
            db.execute("INSERT INTO patient_events (patient_id, event_type, description) VALUES (?, 'Observação editada', 'Observação de retorno atualizada')", (patient_id,))
        self.get_patient(patient_id)

    def delete_visit_observation(self, patient_id: int, observation_id: int) -> None:
        if self.authenticated_user["access_role"] == "crc":
            return self.send_json({"error": "A CRC não pode excluir observações clínicas."}, HTTPStatus.FORBIDDEN)
        if not self.require_patient_access(patient_id, self.authenticated_user):
            return
        with connect() as db:
            if self.patient_edit_lock_conflict(db, patient_id):
                return
            note = db.execute("SELECT id, patient_id, author_name FROM patient_visit_notes WHERE id=? AND patient_id=?", (observation_id, patient_id)).fetchone()
            if not note:
                return self.send_json({"error": "Observação não encontrada."}, HTTPStatus.NOT_FOUND)
            if not self.can_manage_visit_observation(note):
                return self.send_json({"error": "Somente quem registrou esta observação pode excluí-la."}, HTTPStatus.FORBIDDEN)
            locked = db.execute("SELECT 1 FROM patient_followup WHERE patient_id=? AND date(resolved_at)=date('now','localtime')", (patient_id,)).fetchone()
            if locked:
                return self.send_json({"error": "Paciente verificado está bloqueado. Reabra o acompanhamento para editar."}, HTTPStatus.CONFLICT)
            db.execute("DELETE FROM patient_visit_notes WHERE id=?", (observation_id,))
            db.execute("INSERT INTO patient_events (patient_id, event_type, description) VALUES (?, 'Observação excluída', 'Observação de retorno removida')", (patient_id,))
        self.get_patient(patient_id)

    @staticmethod
    def patient_status_values(value) -> tuple[str, str | None]:
        value_key = str(value or "").strip().casefold()
        for status in STANDARD_PATIENT_STATUSES:
            if status.casefold() == value_key:
                return status, None
        return "Consulta", None

    @staticmethod
    def next_appointment_values(payload: dict) -> tuple[str | None, str | None]:
        next_appointment = str(payload.get("next_appointment") or "").strip() or None
        if not next_appointment:
            return None, None
        appointment_type = str(payload.get("next_appointment_type") or "").strip()
        if appointment_type not in {"Agendado", "Programado"}:
            raise ValueError("Informe se a próxima consulta está agendada ou programada")
        return next_appointment, appointment_type

    @staticmethod
    def route_programmed_appointment_to_crc(
        db,
        patient_id: int,
        next_appointment_type: str | None,
        previous_type: str | None = None,
        actor_name: str | None = None,
    ) -> None:
        if next_appointment_type == "Programado":
            if previous_type == "Programado":
                return
            db.execute(
                """UPDATE patient_followup
                   SET crc_status='Aguardando contato',
                       crc_started_at=NULL,
                       crc_completed_at=NULL
                   WHERE patient_id=?""",
                (patient_id,),
            )
            description = "Retorno programado encaminhado automaticamente para a CRC"
        elif previous_type == "Programado" and next_appointment_type == "Agendado":
            db.execute(
                """UPDATE patient_followup
                   SET crc_status='Retorno concluído',
                       crc_completed_at=datetime('now', 'localtime')
                   WHERE patient_id=?""",
                (patient_id,),
            )
            description = "Retorno programado convertido em agendamento"
        else:
            return
        if actor_name:
            description = f"{description} por {actor_name}"
        db.execute(
            "INSERT INTO patient_events (patient_id, event_type, description) VALUES (?, 'Agendamento', ?)",
            (patient_id, description),
        )

    @staticmethod
    def ensure_patient_status(db, value) -> None:
        return None

    def normalized_procedures(self, payload: dict) -> list[dict]:
        result = []
        allowed_stages = {"Indicado", "Aprovado", "Agendado", "Em andamento", "Concluído"}
        for item in payload.get("procedures", []):
            name = str(item.get("name", "")).strip()
            if not name:
                continue
            try:
                value_cents = max(0, int(item.get("value_cents") or 0))
                discount_cents = max(0, int(item.get("discount_cents") or 0))
            except (TypeError, ValueError):
                raise ValueError("Valor ou desconto do procedimento inválido")
            if discount_cents > value_cents:
                raise ValueError("O desconto não pode ser maior que o valor do procedimento")
            stage = item.get("stage") or "Indicado"
            if stage not in allowed_stages:
                raise ValueError("Etapa de procedimento inválida")
            result.append({"name": name, "value_cents": value_cents, "discount_cents": discount_cents, "stage": stage, "notes": item.get("notes") or None})
        return result

    def replace_procedures(self, db, patient_id: int, procedures: list[dict]) -> None:
        db.execute("DELETE FROM procedures WHERE patient_id = ?", (patient_id,))
        db.executemany(
            "INSERT INTO procedures (patient_id, name, value_cents, discount_cents, stage, notes) VALUES (?, ?, ?, ?, ?, ?)",
            [(patient_id, item["name"], item["value_cents"], item["discount_cents"], item["stage"], item["notes"]) for item in procedures],
        )

    def normalized_relationships(self, payload: dict) -> list[dict]:
        result = []
        for item in payload.get("relationships", []):
            relationship_type = str(item.get("relationship_type", "")).strip()
            related_name = str(item.get("related_name", "")).strip()
            connection = str(item.get("connection", "")).strip()
            if not relationship_type and not related_name:
                continue
            if not relationship_type or not related_name:
                raise ValueError("Informe o vínculo e o nome da pessoa relacionada")
            result.append({"relationship_type": relationship_type, "related_name": related_name, "connection": connection})
        return result

    def save_clinical_profile(self, db, patient_id: int, payload: dict) -> None:
        db.execute("""
            INSERT INTO patient_clinical_profile
                (patient_id, particularities, health_change, health_condition, health_care, updated_at)
            VALUES (?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(patient_id) DO UPDATE SET
                particularities=excluded.particularities,
                health_change=excluded.health_change,
                health_condition=excluded.health_condition,
                health_care=excluded.health_care,
                updated_at=CURRENT_TIMESTAMP
        """, (
            patient_id,
            str(payload.get("particularities") or "").strip() or None,
            1 if payload.get("health_change") else 0,
            str(payload.get("health_condition") or "").strip() or None,
            str(payload.get("health_care") or "").strip() or None,
        ))

    def replace_relationships(self, db, patient_id: int, relationships: list[dict]) -> None:
        db.execute("DELETE FROM patient_relationships WHERE patient_id = ?", (patient_id,))
        db.executemany(
            "INSERT INTO patient_relationships (patient_id, relationship_type, related_name, connection) VALUES (?, ?, ?, ?)",
            [(patient_id, item["relationship_type"], item["related_name"], item["connection"] or None) for item in relationships],
        )

    def remember_action(self, db, value) -> None:
        description = str(value or "").strip()
        if description:
            db.execute("INSERT OR IGNORE INTO action_templates (description) VALUES (?)", (description,))

    @staticmethod
    def save_visit_observation(db, patient_id: int, value, author_name: str | None = None) -> None:
        note = str(value or "").strip()
        if note:
            db.execute("INSERT INTO patient_visit_notes (patient_id, note, author_name) VALUES (?, ?, ?)", (patient_id, note, str(author_name or "").strip() or None))

    def create_patient(self, payload: dict) -> None:
        name = str(payload.get("name", "")).strip()
        last_visit = payload.get("last_visit")
        if not name or not last_visit:
            return self.send_json({"error": "Nome e última consulta são obrigatórios"}, HTTPStatus.BAD_REQUEST)
        try:
            procedures = self.normalized_procedures(payload)
            relationships = self.normalized_relationships(payload)
            next_appointment, next_appointment_type = self.next_appointment_values(payload)
        except ValueError as error:
            return self.send_json({"error": str(error)}, HTTPStatus.BAD_REQUEST)
        if not self.can_manage_crc_fields(self.authenticated_user):
            procedures = []
            payload = {**payload, "phone": None, "reference": None, "last_contact": None}
        status = str(payload.get("status") or "Consulta").strip() or "Consulta"
        base_status, custom_status = self.patient_status_values(status)
        with connect() as db:
            professional_id = self.authenticated_user["professional_id"] if self.authenticated_user["access_role"] != "admin" else db.execute("SELECT id FROM professionals WHERE is_owner = 1 ORDER BY id LIMIT 1").fetchone()[0]
            external_id = db.execute("SELECT COALESCE(MAX(CAST(external_id AS INTEGER)), 0) + 1 FROM patients").fetchone()[0]
            patient_id = db.execute(
                "INSERT INTO patients (external_id, name, phone, reference, status, notes) VALUES (?, ?, ?, ?, ?, ?)",
                (external_id, name, payload.get("phone") or None, payload.get("reference") or None, base_status, payload.get("notes") or None),
            ).lastrowid
            db.execute("INSERT INTO patient_assignments (patient_id, professional_id, is_primary) VALUES (?, ?, 1)", (patient_id, professional_id))
            db.execute(
                "INSERT INTO patient_followup (patient_id, last_visit, next_appointment, next_appointment_type, last_contact, next_action, custom_status) VALUES (?, ?, ?, ?, ?, ?, ?)",
                (patient_id, last_visit, next_appointment, next_appointment_type, payload.get("last_contact") or None, payload.get("next_action") or None, custom_status),
            )
            self.route_programmed_appointment_to_crc(
                db, patient_id, next_appointment_type, actor_name=self.authenticated_user["name"]
            )
            self.ensure_patient_status(db, status)
            self.replace_procedures(db, patient_id, procedures)
            self.save_clinical_profile(db, patient_id, payload)
            self.replace_relationships(db, patient_id, relationships)
            self.remember_action(db, payload.get("next_action"))
            self.save_visit_observation(db, patient_id, payload.get("observation"), self.authenticated_user["name"])
            db.execute("INSERT INTO patient_events (patient_id, event_type, description) VALUES (?, 'Cadastro', 'Paciente incluído na carteira da Dra. Dulce')", (patient_id,))
        self.get_patient(patient_id)

    def update_patient(self, patient_id: int, payload: dict) -> None:
        if not self.require_patient_access(patient_id, self.authenticated_user):
            return
        status = str(payload.get("status") or "Consulta").strip() or "Consulta"
        base_status, custom_status = self.patient_status_values(status)
        name = str(payload.get("name", "")).strip()
        if not name or not payload.get("last_visit"):
            return self.send_json({"error": "Nome e última consulta são obrigatórios"}, HTTPStatus.BAD_REQUEST)
        try:
            procedures = self.normalized_procedures(payload)
            relationships = self.normalized_relationships(payload)
            next_appointment, next_appointment_type = self.next_appointment_values(payload)
        except ValueError as error:
            return self.send_json({"error": str(error)}, HTTPStatus.BAD_REQUEST)
        with connect() as db:
            if self.patient_edit_lock_conflict(db, patient_id):
                return
            exists = db.execute("""
                SELECT p.id, p.name, f.next_appointment_type,
                       CASE WHEN date(f.resolved_at) = date('now', 'localtime') THEN 1 ELSE 0 END AS locked
                FROM patients p JOIN patient_followup f ON f.patient_id = p.id WHERE p.id = ?
            """, (patient_id,)).fetchone()
            if not exists:
                return self.send_json({"error": "Paciente não encontrado"}, HTTPStatus.NOT_FOUND)
            if exists["locked"]:
                return self.send_json({"error": "Paciente verificado está bloqueado. Reabra o acompanhamento para editar."}, HTTPStatus.CONFLICT)
            if name != exists["name"]:
                return self.send_json({"error": "O nome do paciente não pode ser alterado."}, HTTPStatus.FORBIDDEN)
            if not self.can_manage_crc_fields(self.authenticated_user):
                protected = db.execute("SELECT p.phone, p.reference, f.last_contact FROM patients p JOIN patient_followup f ON f.patient_id=p.id WHERE p.id=?", (patient_id,)).fetchone()
                if ((payload.get("phone") or None) != protected["phone"] or (payload.get("reference") or None) != protected["reference"] or (payload.get("last_contact") or None) != protected["last_contact"]):
                    return self.send_json({"error": "Contato e referência são controlados pela CRC"}, HTTPStatus.FORBIDDEN)
                saved_procedures = [dict(row) for row in db.execute("SELECT name, value_cents, discount_cents, stage, notes FROM procedures WHERE patient_id=? ORDER BY id", (patient_id,)).fetchall()]
                requested_procedures = [{key: item[key] for key in ("name", "value_cents", "discount_cents", "stage", "notes")} for item in procedures]
                if saved_procedures != requested_procedures:
                    return self.send_json({"error": "Procedimentos são controlados pela CRC"}, HTTPStatus.FORBIDDEN)
            db.execute(
                """UPDATE patients SET name=?, phone=?, reference=?, status=?, notes=?, updated_at=CURRENT_TIMESTAMP WHERE id=?""",
                (
                    name, payload.get("phone") or None, payload.get("reference") or None,
                    base_status, payload.get("notes") or None, patient_id,
                ),
            )
            db.execute(
                """UPDATE patient_followup SET last_visit=?, next_appointment=?, next_appointment_type=?, procedure_name=NULL, last_contact=?, next_action=?, custom_status=? WHERE patient_id=?""",
                (
                    payload.get("last_visit"), next_appointment, next_appointment_type,
                    payload.get("last_contact") or None, payload.get("next_action") or None, custom_status, patient_id,
                ),
            )
            self.route_programmed_appointment_to_crc(
                db,
                patient_id,
                next_appointment_type,
                exists["next_appointment_type"],
                self.authenticated_user["name"],
            )
            self.ensure_patient_status(db, status)
            self.replace_procedures(db, patient_id, procedures)
            self.save_clinical_profile(db, patient_id, payload)
            self.replace_relationships(db, patient_id, relationships)
            self.remember_action(db, payload.get("next_action"))
            self.save_visit_observation(db, patient_id, payload.get("observation"), self.authenticated_user["name"])
            db.execute(
                "INSERT INTO patient_events (patient_id, event_type, description) VALUES (?, 'Atualização', ?)",
                (patient_id, "Ficha de acompanhamento atualizada pela Dra. Dulce"),
            )
        self.get_patient(patient_id)

    def update_crc_patient(self, patient_id: int, payload: dict) -> None:
        if self.authenticated_user["access_role"] != "crc":
            return self.send_json({"error": "Acesso CRC necessário."}, HTTPStatus.FORBIDDEN)
        if not self.require_patient_access(patient_id, self.authenticated_user):
            return
        try:
            procedures = self.normalized_procedures(payload)
            next_appointment, next_appointment_type = self.next_appointment_values(payload)
        except ValueError as error:
            return self.send_json({"error": str(error)}, HTTPStatus.BAD_REQUEST)
        if next_appointment_type == "Agendado" and not next_appointment:
            return self.send_json({"error": "Informe a data confirmada do agendamento."}, HTTPStatus.BAD_REQUEST)
        crc_status = str(payload.get("crc_status") or "Aguardando contato").strip()
        if crc_status not in {"Aguardando contato", "Em atendimento", "Jornada compartilhada", "Retorno concluído"}:
            return self.send_json({"error": "Etapa CRC inválida."}, HTTPStatus.BAD_REQUEST)
        with connect() as db:
            exists = db.execute(
                """SELECT p.id, f.next_appointment_type
                   FROM patients p
                   JOIN patient_followup f ON f.patient_id=p.id
                   WHERE p.id = ?""",
                (patient_id,),
            ).fetchone()
            if not exists:
                return self.send_json({"error": "Paciente não encontrado."}, HTTPStatus.NOT_FOUND)
            db.execute(
                "UPDATE patients SET phone=?, reference=?, updated_at=CURRENT_TIMESTAMP WHERE id=?",
                (str(payload.get("phone") or "").strip() or None, str(payload.get("reference") or "").strip() or None, patient_id),
            )
            db.execute("""
                UPDATE patient_followup
                SET last_contact = ?,
                    next_appointment = ?,
                    next_appointment_type = ?,
                    crc_status = ?,
                    crc_started_at = CASE
                        WHEN ? = 'Em atendimento' THEN COALESCE(crc_started_at, datetime('now', 'localtime'))
                        ELSE crc_started_at END,
                    crc_completed_at = CASE
                        WHEN ? = 'Retorno concluído' THEN datetime('now', 'localtime')
                        ELSE NULL END
                WHERE patient_id = ?
            """, (
                payload.get("last_contact") or None,
                next_appointment,
                next_appointment_type,
                crc_status,
                crc_status,
                crc_status,
                patient_id,
            ))
            if next_appointment_type == "Agendado":
                crc_status = "Retorno concluído"
                db.execute(
                    """UPDATE patient_followup
                       SET crc_status='Retorno concluído',
                           crc_completed_at=datetime('now', 'localtime')
                       WHERE patient_id=?""",
                    (patient_id,),
                )
                db.execute(
                    "INSERT INTO patient_events (patient_id, event_type, description) VALUES (?, 'Agendamento', ?)",
                    (patient_id, f"Agendamento confirmado pela CRC por {self.authenticated_user['name']}"),
                )
            self.replace_procedures(db, patient_id, procedures)
            db.execute("INSERT INTO patient_events (patient_id, event_type, description) VALUES (?, 'CRC', ?)", (patient_id, f"CRC atualizou o retorno: {crc_status}"))
        self.get_patient(patient_id)

    def serve_static(self, request_path: str) -> None:
        if request_path in ("", "/"):
            relative = "index.html"
        elif request_path in ("/login", "/login/"):
            relative = "login.html"
        elif request_path in ("/setup", "/setup/"):
            relative = "setup.html"
        elif request_path in ("/forgot-password", "/forgot-password/"):
            relative = "forgot-password.html"
        elif request_path in ("/change-password", "/change-password/"):
            relative = "change-password.html"
        elif request_path in ("/two-factor", "/two-factor/"):
            relative = "two-factor.html"
        elif request_path in ("/two-factor-setup", "/two-factor-setup/"):
            relative = "two-factor-setup.html"
        elif request_path in (ADMIN_ROUTE, f"{ADMIN_ROUTE}/"):
            relative = "admin.html"
        elif request_path in (CRC_ROUTE, f"{CRC_ROUTE}/"):
            relative = "crc.html"
        elif request_path in (f"{CRC_ROUTE}/whatsapp", f"{CRC_ROUTE}/whatsapp/"):
            relative = "crm-whatsapp.html"
        elif request_path in (f"{CRC_ROUTE}/controle-pacientes", f"{CRC_ROUTE}/controle-pacientes/"):
            self.send_response(HTTPStatus.FOUND)
            self.send_header("Location", f"{CRC_ROUTE}/whatsapp?screen=patient-control")
            self.send_header("Cache-Control", "no-store, max-age=0")
            self.end_headers()
            return
        elif request_path in (f"{CRC_ROUTE}/controle-pacientes/embed", f"{CRC_ROUTE}/controle-pacientes/embed/"):
            relative = "crm-patient-control.html"
        elif request_path == "/crm-whatsapp.html":
            return self.send_error(HTTPStatus.NOT_FOUND)
        elif request_path in ("/admin", "/admin/"):
            return self.send_error(HTTPStatus.NOT_FOUND)
        else:
            relative = request_path.lstrip("/")
        target = (PUBLIC / relative).resolve()
        if PUBLIC.resolve() not in target.parents and target != PUBLIC.resolve():
            return self.send_error(HTTPStatus.FORBIDDEN)
        if not target.exists() or not target.is_file():
            target = PUBLIC / "index.html"
        content = target.read_bytes()
        bundled_ui_nonce = None
        if relative == "crm-whatsapp.html":
            # SEC-008: nonce por requisição, injetado nos <script> inline do
            # bundle para permitir remover 'unsafe-inline' da CSP.
            bundled_ui_nonce = secrets.token_urlsafe(16)
            content = content.replace(b"__CSP_NONCE__", bundled_ui_nonce.encode("ascii"))
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", mimetypes.guess_type(target.name)[0] or "application/octet-stream")
        self.send_header("Content-Length", str(len(content)))
        # The bundled CRM is generated as one large JSON payload. Serving a
        # stale HTML or bridge file makes browsers execute a mismatched
        # version, so always revalidate these operational screens.
        if target.suffix == ".html" or target.name == "crm-evolution-bridge.js":
            self.send_header("Cache-Control", "no-store, max-age=0")
        self.send_security_headers(
            allow_bundled_ui=relative == "crm-whatsapp.html",
            allow_same_origin_frame=request_path.startswith(f"{CRC_ROUTE}/controle-pacientes/embed"),
            bundled_ui_nonce=bundled_ui_nonce,
        )
        self.end_headers()
        self.wfile.write(content)


class ClinicServer(ThreadingHTTPServer):
    allow_reuse_address = False


def run_evolution_chat_sync_loop() -> None:
    # Webhooks mantêm a operação em tempo real. Esta reconciliação periódica
    # recupera eventos perdidos e alinha os contadores exibidos pelo WhatsApp.
    time.sleep(5)
    while True:
        try:
            handler = ClinicHandler.__new__(ClinicHandler)
            handler.sync_evolution_chat_state()
        except Exception as error:
            with EVOLUTION_CHAT_SYNC_LOCK:
                EVOLUTION_CHAT_SYNC_STATUS["errors"] = [{"instance": "Evolution", "error": str(error)[:240]}]
        time.sleep(EVOLUTION_CHAT_SYNC_INTERVAL)


if __name__ == "__main__":
    initialize_database()
    threading.Thread(target=run_evolution_chat_sync_loop, daemon=True, name="evolution-chat-state-sync").start()
    host = os.environ.get("HOST", "127.0.0.1")
    port = int(os.environ.get("PORT", "8000"))
    print(f"Instituto Eduardo Ayub — piloto em http://{host}:{port}")
    ClinicServer((host, port), ClinicHandler).serve_forever()
